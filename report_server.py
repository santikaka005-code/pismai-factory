from __future__ import annotations

import json
import hashlib
import hmac
import base64
import math
import mimetypes
import os
import re
import secrets
import threading
import time
import textwrap
import urllib.error
import urllib.request
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

HOST = os.environ.get("HOST") or ("0.0.0.0" if os.environ.get("PORT") else "127.0.0.1")
PORT = int(os.environ.get("PORT", "8787"))
BASE_DIR = Path(__file__).parent
DATA_FILE = Path(__file__).with_name("report_data.json")
COMPANY_NAME = "Pitsamai Frozen Fruits Co., Ltd."
SYSTEM_NAME = "SystemPro by Pitsamai Frozen Fruits"
BRAND_GREEN = "#0F7A3D"
THAI_FONT = "Helvetica"
THAI_FONT_BOLD = "Helvetica-Bold"
TIME_SPECIAL_DAILY_WAGE = 365
DURIAN_GRADES = ("A", "B", "C", "D", "E")
TIME_SPECIAL_WAGE_TABLE = {
    2: 91,
    2.5: 114,
    3: 137,
    3.5: 160,
    4: 183,
    4.5: 205,
    5: 228,
    5.5: 251,
    6: 273,
    6.5: 297,
    7: 319,
    7.5: 342,
    8: 365,
}
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
BACKUP_ACCESS_CODE = os.environ.get("BACKUP_ACCESS_CODE", "1150")
SESSION_SIGNING_SECRET = os.environ.get("SESSION_SIGNING_SECRET") or SUPABASE_SERVICE_ROLE_KEY or secrets.token_hex(32)
ACCOUNTING_COMPANY_KEY = os.environ.get("ACCOUNTING_COMPANY_KEY", "pismai-main")
BACKUP_TABLES = [
    "account_users",
    "employees",
    "time_employees",
    "wage_rates",
    "production_sessions",
    "production_records",
    "time_records",
    "deduction_records",
    "deduction_applications",
    "production_save_queue",
    "production_save_queue_events",
    "time_save_queue",
    "time_save_queue_events",
    "issue_reports",
    "audit_logs",
    "community_posts",
    "secret_messages",
    "inbound_fruits",
    "inbound_fruit_prices",
    "inbound_receipts",
]
QUEUE_BACKUP_TABLES = [
    "production_save_queue", "production_save_queue_events",
    "time_save_queue", "time_save_queue_events",
]
MAIN_CLEAR_TABLES = [
    "inbound_receipts",
    "deduction_applications",
    "deduction_records",
    "time_records",
    "production_records",
    "production_sessions",
    "production_save_queue_events",
    "production_save_queue",
    "time_save_queue_events",
    "time_save_queue",
]
BACKUP_ARCHIVE_BUCKET = "pismai-backup-archives"
SUPABASE_FREE_DATABASE_BYTES = 500 * 1024 * 1024
DATABASE_STORAGE_WARNING_PERCENT = 85
LIVE_STATE_TABLES = {
    "production_sessions",
    "production_records",
    "time_records",
    "audit_logs",
}
ONLINE_USER_TIMEOUT_SECONDS = 45
STORAGE_USAGE_CACHE_SECONDS = 15
online_user_lock = threading.Lock()
storage_usage_lock = threading.Lock()
live_state_sync_lock = threading.Lock()
production_record_insert_lock = threading.Lock()
backup_clear_lock = threading.Lock()
deduction_record_insert_lock = threading.Lock()
time_record_insert_lock = threading.Lock()
production_save_queue_wakeup = threading.Event()
production_save_queue_worker_id = f"render-{os.getpid()}-{secrets.token_hex(4)}"
time_save_queue_wakeup = threading.Event()
time_save_queue_worker_id = f"render-time-{os.getpid()}-{secrets.token_hex(4)}"
time_queue_recovery_lock = threading.Lock()
time_queue_last_recovery_at = 0.0
online_user_sessions: dict[str, dict] = {}
storage_usage_cache: dict = {"expires_at": 0.0, "data": None}
production_record_next_id: int | None = None
deduction_record_next_id: int | None = None
time_record_next_id: int | None = None
SYSTEM_ACCOUNT_PROFILES = {
    "Santi": {
        "username": "Santi",
        "password": os.environ.get("SYSTEM_C7_PASSWORD", ""),
        "fullname": "Santi Khl.",
        "role": "developer",
        "user_level": "C7",
        "status": "Active",
        "created_by": "system",
    }
}
SYSTEM_ACCOUNT_USERNAMES = {username.lower() for username in SYSTEM_ACCOUNT_PROFILES}


def session_token(account: dict) -> str:
    payload = {
        "sub": str(account.get("id", "")),
        "username": str(account.get("username", "")),
        "level": str(account.get("user_level", "C1")),
        "role": str(account.get("role", "")),
        "exp": int(time.time()) + 12 * 60 * 60,
    }
    encoded = base64.urlsafe_b64encode(json.dumps(payload, separators=(",", ":")).encode()).decode().rstrip("=")
    signature = hmac.new(SESSION_SIGNING_SECRET.encode(), encoded.encode(), hashlib.sha256).hexdigest()
    return f"{encoded}.{signature}"


def verify_session_token(token: str) -> dict | None:
    try:
        encoded, signature = token.split(".", 1)
        expected = hmac.new(SESSION_SIGNING_SECRET.encode(), encoded.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected):
            return None
        padded = encoded + "=" * (-len(encoded) % 4)
        payload = json.loads(base64.urlsafe_b64decode(padded.encode()).decode())
        if int(payload.get("exp", 0)) <= int(time.time()):
            return None
        return payload
    except (ValueError, TypeError, json.JSONDecodeError):
        return None


def accounting_actor(handler: BaseHTTPRequestHandler, minimum_level: int = 4) -> dict | None:
    actor = verify_session_token(handler.headers.get("X-Session-Token", ""))
    level = int("".join(filter(str.isdigit, str(actor.get("level", "C1")))) or "1") if actor else 0
    return actor if actor and level >= minimum_level else None


PRODUCTION_SELF_EDIT_WINDOW_SECONDS = 5 * 60


def account_level_number(value: object) -> int:
    match = re.search(r"\d+", str(value or "C1"))
    return int(match.group()) if match else 1


def production_record_owned_by_actor(record: dict, actor: dict, actor_account: dict) -> bool:
    owner = str(record.get("created_by") or "").strip().casefold()
    aliases = {
        str(actor.get("username") or "").strip().casefold(),
        str(actor_account.get("username") or "").strip().casefold(),
        str(actor_account.get("fullname") or "").strip().casefold(),
    }
    aliases.discard("")
    return bool(owner and owner in aliases)


def production_record_within_self_edit_window(
    record: dict,
    actor: dict,
    actor_account: dict,
    now: datetime | None = None,
) -> bool:
    if not production_record_owned_by_actor(record, actor, actor_account):
        return False
    try:
        created_at = datetime.fromisoformat(str(record.get("created_at") or "").replace("Z", "+00:00"))
        if created_at.tzinfo is None:
            created_at = created_at.replace(tzinfo=timezone.utc)
    except (TypeError, ValueError):
        return False
    current_time = now or datetime.now(timezone.utc)
    if current_time.tzinfo is None:
        current_time = current_time.replace(tzinfo=timezone.utc)
    elapsed_seconds = (current_time.astimezone(timezone.utc) - created_at.astimezone(timezone.utc)).total_seconds()
    return -30 <= elapsed_seconds <= PRODUCTION_SELF_EDIT_WINDOW_SECONDS


def secret_room_actor(handler: BaseHTTPRequestHandler) -> dict | None:
    """All signed-in website accounts may use the internal collaboration area."""
    return verify_session_token(handler.headers.get("X-Session-Token", ""))


def inbound_clean_text(value: object, maximum: int) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())[:maximum]


def inbound_actor_name(actor: dict) -> str:
    return inbound_clean_text(actor.get("username"), 120) or "unknown"


def inbound_authorized_actor(handler: BaseHTTPRequestHandler) -> dict | None:
    actor = secret_room_actor(handler)
    return actor if actor and account_level_number(actor.get("level")) >= 5 else None


def inbound_audit(handler: BaseHTTPRequestHandler, actor: dict, action: str, detail: str, metadata: dict) -> tuple[int, object]:
    username = inbound_actor_name(actor)
    return insert_audit_log_compatible({
        "action": action,
        "module": "inbound",
        "description": detail,
        "created_by": username,
        "user_fullname": username,
        "ip_address": handler.client_address[0] if handler.client_address else None,
        "metadata": {"action": action, "module": "inbound", "detail": detail, "username": username, **metadata},
    })


ISSUE_REPORT_CATEGORIES = {"system", "data", "display", "performance", "other"}
ISSUE_REPORT_PRIORITIES = {"normal", "urgent", "blocking"}
ISSUE_REPORT_STATUSES = {"received", "investigating", "resolved"}
ISSUE_REPORT_MAX_ATTACHMENT_BYTES = 2 * 1024 * 1024


def validate_issue_report_payload(payload: dict) -> tuple[dict | None, str | None]:
    title = str(payload.get("title") or "").strip()
    category = str(payload.get("category") or "").strip()
    page_name = str(payload.get("page_name") or "").strip()
    priority = str(payload.get("priority") or "normal").strip()
    description = str(payload.get("description") or "").strip()
    attachment_data = str(payload.get("attachment_data") or "").strip()
    attachment_name = str(payload.get("attachment_name") or "").strip()

    if not title or len(title) > 160:
        return None, "Title must contain 1-160 characters."
    if category not in ISSUE_REPORT_CATEGORIES:
        return None, "Issue category is invalid."
    if not page_name or len(page_name) > 120:
        return None, "Page name must contain 1-120 characters."
    if priority not in ISSUE_REPORT_PRIORITIES:
        return None, "Issue priority is invalid."
    if not description or len(description) > 5000:
        return None, "Description must contain 1-5,000 characters."
    if attachment_data:
        match = re.fullmatch(r"data:(image/(?:png|jpeg));base64,([A-Za-z0-9+/=]+)", attachment_data)
        if not match:
            return None, "Attachment must be a PNG or JPG image."
        try:
            decoded = base64.b64decode(match.group(2), validate=True)
        except (ValueError, TypeError):
            return None, "Attachment data is invalid."
        if len(decoded) > ISSUE_REPORT_MAX_ATTACHMENT_BYTES:
            return None, "Attachment must not exceed 2 MB."
        attachment_type = match.group(1)
    else:
        attachment_name = ""
        attachment_type = ""

    return {
        "title": title,
        "category": category,
        "page_name": page_name,
        "priority": priority,
        "description": description,
        "attachment_name": attachment_name[:255],
        "attachment_type": attachment_type,
        "attachment_data": attachment_data,
    }, None


def validate_accounting_workspace(workspace: object) -> str | None:
    if not isinstance(workspace, dict):
        return "workspace must be an object"
    accounts = workspace.get("accounts", [])
    journals = workspace.get("journals", [])
    if not isinstance(accounts, list) or not isinstance(journals, list):
        return "accounts and journals must be arrays"
    account_ids = {str(row.get("id")) for row in accounts if isinstance(row, dict) and row.get("id")}
    codes = [str(row.get("code", "")).strip().lower() for row in accounts if isinstance(row, dict)]
    if not account_ids or any(not code for code in codes) or len(codes) != len(set(codes)):
        return "chart of accounts contains missing or duplicate account codes"
    for journal in journals:
        if not isinstance(journal, dict) or journal.get("status") not in {"draft", "posted", "reversed"}:
            return "journal status is invalid"
        lines = journal.get("lines", [])
        if not isinstance(lines, list):
            return "journal lines must be an array"
        debit = credit = 0.0
        for line in lines:
            if not isinstance(line, dict) or str(line.get("accountId")) not in account_ids:
                return "journal references an unknown account"
            line_debit = round(float(line.get("debit") or 0), 2)
            line_credit = round(float(line.get("credit") or 0), 2)
            if line_debit < 0 or line_credit < 0 or (line_debit and line_credit):
                return "journal line has an invalid debit or credit"
            debit += line_debit
            credit += line_credit
        if journal.get("status") in {"posted", "reversed"} and (len(lines) < 2 or debit <= 0 or abs(debit - credit) >= 0.005):
            return "posted journal is not balanced"
    return None


def register_thai_font() -> str:
    regular_paths = [
        BASE_DIR / "fonts" / "Sarabun-Regular.ttf",
        BASE_DIR / "fonts" / "NotoSansThai-Regular.ttf",
        BASE_DIR / "fonts" / "NotoSansThai.ttf",
        Path("C:/Windows/Fonts/Sarabun-Regular.ttf"),
        Path("C:/Windows/Fonts/THSarabunNew.ttf"),
        Path("C:/Windows/Fonts/tahoma.ttf"),
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/leelawui.ttf"),
    ]
    bold_paths = [
        BASE_DIR / "fonts" / "Sarabun-Bold.ttf",
        BASE_DIR / "fonts" / "NotoSansThai-Bold.ttf",
        Path("C:/Windows/Fonts/Sarabun-Bold.ttf"),
        Path("C:/Windows/Fonts/THSarabunNew Bold.ttf"),
        Path("C:/Windows/Fonts/tahomabd.ttf"),
        Path("C:/Windows/Fonts/arialbd.ttf"),
        Path("C:/Windows/Fonts/leelauib.ttf"),
    ]
    registered_regular = ""
    for font_path in regular_paths:
        if font_path.exists():
            try:
                pdfmetrics.registerFont(TTFont("ThaiFont", str(font_path)))
                registered_regular = "ThaiFont"
                break
            except Exception:
                continue
    for font_path in bold_paths:
        if font_path.exists():
            try:
                pdfmetrics.registerFont(TTFont("ThaiFontBold", str(font_path)))
                globals()["THAI_FONT_BOLD"] = "ThaiFontBold"
                break
            except Exception:
                continue
    if registered_regular:
        return registered_regular
    return "Helvetica"


THAI_FONT = register_thai_font()


def load_data() -> dict:
    if not DATA_FILE.exists():
        return {"employees": [], "production_records": []}
    return json.loads(DATA_FILE.read_text(encoding="utf-8"))


def save_data(data: dict) -> None:
    DATA_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120_000)
    return f"pbkdf2_sha256$120000${salt}${digest.hex()}"


def verify_password(password: str, password_hash: str) -> bool:
    if not password_hash:
        return False
    parts = password_hash.split("$")
    if len(parts) == 4 and parts[0] == "pbkdf2_sha256":
        try:
            iterations = int(parts[1])
            salt = parts[2]
            expected = parts[3]
        except ValueError:
            return False
        digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
        return hmac.compare_digest(digest.hex(), expected)
    return hmac.compare_digest(password, password_hash)


def account_from_payload(payload: dict, include_password: bool = True) -> dict:
    status = "Active" if payload.get("isActive", payload.get("status", "Active")) not in [False, "false", "Inactive"] else "Inactive"
    account = {
        "username": str(payload.get("username", "")).strip(),
        "fullname": str(payload.get("fullname", "")).strip(),
        "phone": str(payload.get("phone", "")).strip(),
        "role": str(payload.get("role_key") or payload.get("role") or "general_staff").strip(),
        "user_level": str(payload.get("level") or payload.get("user_level") or "C1").strip().upper(),
        "status": status,
        "created_by": str(payload.get("created_by", "")).strip() or None,
    }
    if include_password:
        password = str(payload.get("password", ""))
        if password:
            account["password_hash"] = hash_password(password)
    return account


def account_to_client(account: dict) -> dict:
    role_key = account.get("role") or "general_staff"
    username = str(account.get("username", ""))
    is_system_account = username.lower() in SYSTEM_ACCOUNT_USERNAMES
    return {
        "id": account.get("id"),
        "username": username,
        "fullname": account.get("fullname", ""),
        "phone": account.get("phone") or ("0943913997" if username.lower() == "santi" else ""),
        "role_key": role_key,
        "level": account.get("user_level") or "C1",
        "isActive": account.get("status", "Active") == "Active",
        "is_system": is_system_account,
        "created_at": account.get("created_at"),
        "updated_at": account.get("updated_at") or account.get("created_at"),
    }


def ensure_system_accounts() -> None:
    if not supabase_configured():
        return
    for profile in SYSTEM_ACCOUNT_PROFILES.values():
        if not profile.get("password"):
            continue
        username = profile["username"]
        account = {
            "username": username,
            "password_hash": hash_password(profile["password"]),
            "fullname": profile["fullname"],
            "role": profile["role"],
            "user_level": profile["user_level"],
            "status": profile["status"],
            "created_by": profile["created_by"],
        }
        status, existing = supabase_request(
            "GET",
            f"account_users?username=eq.{quote(username)}&select=id&limit=1",
        )
        if status >= 400:
            continue
        if isinstance(existing, list) and existing:
            supabase_request(
                "PATCH",
                f"account_users?username=eq.{quote(username)}",
                account,
                prefer="return=minimal",
            )
        else:
            supabase_request(
                "POST",
                "account_users",
                account,
                prefer="return=minimal",
            )


def employee_from_payload(payload: dict) -> dict:
    employee = {
        "emp_code": str(payload.get("emp_code", "")).strip(),
        "fullname": str(payload.get("fullname", "")).strip(),
        "department": str(payload.get("department", "")).strip(),
        "position": str(payload.get("position", "")).strip() or "-",
        "pay_group": str(payload.get("pay_group", "")).strip(),
        "status": str(payload.get("status", "Active")).strip() or "Active",
        "note": str(payload.get("note", "")).strip() or None,
        "created_by": str(payload.get("created_by", "")).strip() or None,
    }
    if payload.get("id") not in [None, ""]:
        employee["id"] = payload.get("id")
    return employee


def time_employee_from_payload(payload: dict) -> dict:
    employee_type = str(payload.get("employee_type", "normal")).strip() or "normal"
    employee_type_aliases = {
        "normal": "normal_347",
        "special": "special_365",
    }
    employee_type = employee_type_aliases.get(employee_type, employee_type)
    type_daily_wages = {
        "normal_347": 347,
        "special_365": 365,
        "special_347": 347,
        "special_500": 500,
    }
    daily_wage = payload.get("daily_wage", type_daily_wages.get(employee_type, 347))
    ot_hourly_rate = payload.get("ot_hourly_rate", 50)
    employee = {
        "emp_code": str(payload.get("emp_code", "")).strip(),
        "fullname": str(payload.get("fullname", "")).strip(),
        "employee_type": employee_type,
        "daily_wage": daily_wage,
        "ot_hourly_rate": ot_hourly_rate,
        "status": str(payload.get("status", "Active")).strip() or "Active",
        "note": str(payload.get("note", "")).strip() or None,
        "created_by": str(payload.get("created_by", "")).strip() or None,
    }
    if payload.get("id") not in [None, ""]:
        employee["id"] = payload.get("id")
    return employee


def deduction_from_payload(payload: dict) -> dict:
    start_date = str(payload.get("start_date", "")).strip()
    end_date = str(payload.get("end_date", start_date)).strip() or start_date
    deduction = {
        "employee_kind": str(payload.get("employee_kind", "production")).strip() or "production",
        "employee_id": payload.get("employee_id"),
        "emp_code": str(payload.get("emp_code", "")).strip(),
        "employee_name": str(payload.get("employee_name", "")).strip(),
        "start_date": start_date,
        "end_date": end_date,
        "deduction_type": str(payload.get("deduction_type", "")).strip(),
        "deduction_label": str(payload.get("deduction_label", "")).strip(),
        "amount": payload.get("amount", 0),
        "note": str(payload.get("note", "")).strip() or None,
        "status": str(payload.get("status", "Active")).strip() or "Active",
        "created_by": str(payload.get("created_by", "")).strip() or None,
        "updated_by": str(payload.get("updated_by", "")).strip() or None,
    }
    if payload.get("client_uid") not in [None, ""]:
        deduction["client_uid"] = str(payload.get("client_uid")).strip()
    if payload.get("id") not in [None, ""]:
        deduction["id"] = payload.get("id")
    return deduction


def deduction_application_from_payload(payload: dict) -> dict:
    return {
        "deduction_id": payload.get("deduction_id"),
        "amount": payload.get("amount", 0),
        "note": str(payload.get("note", "")).strip() or None,
    }


def next_table_id(table: str) -> int:
    status, body = supabase_request("GET", f"{table}?select=id&order=id.desc&limit=1")
    if status < 400 and isinstance(body, list) and body:
        try:
            return int(body[0].get("id") or 0) + 1
        except (TypeError, ValueError):
            return 1
    return 1


def supabase_error_text(body: object) -> str:
    if isinstance(body, dict):
        return " ".join(str(value) for value in body.values() if value is not None)
    return str(body or "")


def is_unique_constraint_error(body: object, constraint_name: str) -> bool:
    text = supabase_error_text(body).lower()
    return "duplicate key value" in text and constraint_name.lower() in text


def strip_durian_columns(rows: list[dict]) -> list[dict]:
    return [
        {
            key: value
            for key, value in row.items()
            if key not in ("grade_weights", "grade_rates", "grade_amounts")
        }
        for row in rows
    ]


def has_durian_columns(rows: list[dict]) -> bool:
    return any(
        any(key in row for key in ("grade_weights", "grade_rates", "grade_amounts"))
        for row in rows
    )


def reserve_production_record_ids(count: int, refresh: bool = False) -> int:
    global production_record_next_id
    if refresh or production_record_next_id is None:
        production_record_next_id = next_table_id("production_records")
    next_id = production_record_next_id
    production_record_next_id += count
    return next_id


def assign_production_record_ids(rows: list[dict], refresh: bool = False) -> list[dict]:
    next_id = reserve_production_record_ids(len(rows), refresh=refresh)
    assigned_rows = []
    for index, row in enumerate(rows):
        assigned_rows.append({**row, "id": next_id + index})
    return assigned_rows


def production_record_raw_payload(row: dict) -> dict:
    raw = row.get("raw_payload")
    return raw if isinstance(raw, dict) else row


def production_record_client_uid(row: dict) -> str:
    raw = production_record_raw_payload(row)
    return str(raw.get("client_uid") or row.get("client_uid") or "").strip()


def production_record_batch_uid(row: dict) -> str:
    raw = production_record_raw_payload(row)
    return str(raw.get("batch_uid") or row.get("batch_uid") or "").strip()


def production_record_duplicate_signature(row: dict) -> str:
    raw = production_record_raw_payload(row)
    fruit_type = str(row.get("fruit_type") or raw.get("fruit_type") or "mangosteen")
    pile = production_pile_number(raw) or production_pile_number(row) or ""
    if fruit_type == "durian":
        weights = production_grade_weights(raw if raw.get("grade_weights") is not None else row)
        weight_key = {grade: round(weights[grade], 3) for grade in DURIAN_GRADES}
    else:
        weight_key = {
            "water": round(safe_float(row.get("water_weight", raw.get("water_weight", raw.get("water", 0)))), 3),
            "flower": round(safe_float(row.get("flower_weight", raw.get("flower_weight", raw.get("flower", 0)))), 3),
        }
    return json.dumps(
        [
            str(row.get("record_date") or raw.get("record_date") or raw.get("date") or ""),
            str(row.get("emp_code") or raw.get("emp_code") or ""),
            fruit_type,
            str(pile),
            weight_key,
        ],
        ensure_ascii=False,
        sort_keys=True,
    )


def production_record_weight_total(row: dict) -> float:
    raw = production_record_raw_payload(row)
    fruit_type = str(row.get("fruit_type") or raw.get("fruit_type") or "mangosteen")
    if fruit_type == "durian":
        return production_grade_total(raw if raw.get("grade_weights") is not None else row)
    return (
        safe_float(row.get("water_weight", raw.get("water_weight", raw.get("water", 0))))
        + safe_float(row.get("flower_weight", raw.get("flower_weight", raw.get("flower", 0))))
    )


def production_similarity_ratio(a: float, b: float) -> float:
    larger = max(abs(a), abs(b))
    if larger <= 0:
        return 1.0
    return min(abs(a), abs(b)) / larger


def production_batch_group_key(row: dict) -> str:
    batch_uid = production_record_batch_uid(row)
    if batch_uid:
        return f"batch:{batch_uid}"
    raw = production_record_raw_payload(row)
    return f"time:{row.get('record_time') or raw.get('record_time') or row.get('created_at') or raw.get('created_at') or ''}"


def production_batch_summary(rows: list[dict]) -> dict:
    piles: dict[str, float] = {}
    for row in rows:
        raw = production_record_raw_payload(row)
        pile = str(production_pile_number(raw) or production_pile_number(row) or "")
        if not pile:
            continue
        piles[pile] = piles.get(pile, 0.0) + production_record_weight_total(row)
    return {"piles": piles, "total": sum(piles.values())}


def near_duplicate_batch_error(insert_rows: list[dict], existing_rows: list[dict]) -> dict | None:
    incoming_summary = production_batch_summary(insert_rows)
    incoming_piles = set(incoming_summary["piles"].keys())
    if not incoming_piles or incoming_summary["total"] <= 0:
        return None
    existing_groups: dict[str, list[dict]] = {}
    for row in existing_rows:
        existing_groups.setdefault(production_batch_group_key(row), []).append(row)
    for group_rows in existing_groups.values():
        existing_summary = production_batch_summary(group_rows)
        if set(existing_summary["piles"].keys()) != incoming_piles:
            continue
        similarity = production_similarity_ratio(incoming_summary["total"], existing_summary["total"])
        if similarity < 0.5:
            continue
        same_exact = all(
            abs(incoming_summary["piles"].get(pile, 0.0) - existing_summary["piles"].get(pile, 0.0)) < 0.05
            for pile in incoming_piles
        )
        if same_exact:
            continue
        sample = group_rows[0]
        return {
            "error": "พบข้อมูลเดิมที่ใกล้เคียงมากเกิน 50% ระบบจึงไม่บันทึกซ้ำ กรุณาตรวจรายการเดิมก่อน",
            "duplicate_guard": "near_batch",
            "similarity": round(similarity, 3),
            "existing_ids": [row.get("id") for row in group_rows],
            "record_date": sample.get("record_date"),
            "emp_code": sample.get("emp_code"),
            "fruit_type": sample.get("fruit_type"),
        }
    return None


def production_duplicate_lookup_rows(insert_rows: list[dict]) -> tuple[int, list[dict] | dict]:
    lookup_groups = {
        (
            str(row.get("record_date") or ""),
            str(row.get("emp_code") or ""),
            str(row.get("fruit_type") or "mangosteen"),
        )
        for row in insert_rows
        if row.get("record_date") and row.get("emp_code")
    }
    existing_rows: list[dict] = []
    for record_date, emp_code, fruit_type in lookup_groups:
        status, body = supabase_request(
            "GET",
            "production_records?"
            f"record_date=eq.{quote(record_date)}"
            f"&emp_code=eq.{quote(emp_code)}"
            f"&fruit_type=eq.{quote(fruit_type)}"
            "&select=*",
        )
        if status >= 400:
            return status, {"error": body, "lookup": [record_date, emp_code, fruit_type]}
        if isinstance(body, list):
            existing_rows.extend(row for row in body if isinstance(row, dict))
    return 200, existing_rows


def production_duplicate_response_row(existing_row: dict, incoming_row: dict) -> dict:
    response_row = dict(existing_row)
    incoming_raw = production_record_raw_payload(incoming_row)
    if isinstance(incoming_raw, dict):
        response_row["raw_payload"] = {**incoming_raw, "id": existing_row.get("id")}
    return response_row


def insert_production_records_compatible(insert_rows: list[dict]) -> tuple[int, dict | list | str | None]:
    def post(rows: list[dict]) -> tuple[int, dict | list | str | None]:
        return supabase_request(
            "POST",
            "production_records",
            rows,
            prefer="return=representation",
        )

    lookup_status, existing_rows = production_duplicate_lookup_rows(insert_rows)
    if lookup_status >= 400:
        return lookup_status, existing_rows
    existing_by_uid = {
        production_record_client_uid(row): row
        for row in existing_rows
        if production_record_client_uid(row)
    }
    existing_by_signature = {
        production_record_duplicate_signature(row): row
        for row in existing_rows
    }
    existing_by_batch_uid: dict[str, list[dict]] = {}
    for row in existing_rows:
        batch_uid = production_record_batch_uid(row)
        if batch_uid:
            existing_by_batch_uid.setdefault(batch_uid, []).append(row)
    result_rows: list[dict | None] = [None] * len(insert_rows)
    pending_rows: list[dict] = []
    pending_indexes: list[int] = []
    pending_by_uid: dict[str, int] = {}
    pending_by_signature: dict[str, int] = {}
    duplicate_refs: dict[int, int] = {}
    for index, row in enumerate(insert_rows):
        uid = production_record_client_uid(row)
        signature = production_record_duplicate_signature(row)
        existing = existing_by_uid.get(uid) if uid else None
        existing = existing or existing_by_signature.get(signature)
        if existing:
            result_rows[index] = production_duplicate_response_row(existing, row)
            continue
        pending_ref = pending_by_uid.get(uid) if uid else None
        pending_ref = pending_ref if pending_ref is not None else pending_by_signature.get(signature)
        if pending_ref is not None:
            duplicate_refs[index] = pending_ref
            continue
        pending_by_uid[uid] = index
        pending_by_signature[signature] = index
        pending_indexes.append(index)
        pending_rows.append(row)

    incoming_batch_uids = {production_record_batch_uid(row) for row in insert_rows if production_record_batch_uid(row)}
    for batch_uid in incoming_batch_uids:
        existing_batch_rows = existing_by_batch_uid.get(batch_uid, [])
        if existing_batch_rows and len(existing_batch_rows) >= len(insert_rows):
            sorted_existing = sorted(existing_batch_rows, key=lambda item: (production_pile_number(item) or 0, item.get("id") or 0))
            return 200, [
                production_duplicate_response_row(existing, incoming)
                for existing, incoming in zip(sorted_existing, insert_rows)
            ]

    near_error = near_duplicate_batch_error(insert_rows, existing_rows)
    if near_error:
        return 409, near_error

    if not pending_rows:
        return 200, [row for row in result_rows if row is not None]

    queued_rows = assign_production_record_ids(pending_rows)
    status, body = post(queued_rows)
    if status >= 400 and has_durian_columns(queued_rows) and not is_unique_constraint_error(body, "production_records_pkey"):
        status, body = post(strip_durian_columns(queued_rows))

    for attempt in range(3):
        if not (status >= 400 and is_unique_constraint_error(body, "production_records_pkey")):
            break
        retry_rows = assign_production_record_ids(pending_rows, refresh=(attempt == 0))
        status, body = post(retry_rows)
        if status >= 400 and has_durian_columns(retry_rows) and not is_unique_constraint_error(body, "production_records_pkey"):
            status, body = post(strip_durian_columns(retry_rows))

    if status < 400 and isinstance(body, list):
        for index, row in zip(pending_indexes, body):
            result_rows[index] = row
        for duplicate_index, source_index in duplicate_refs.items():
            source_row = result_rows[source_index]
            if isinstance(source_row, dict):
                result_rows[duplicate_index] = production_duplicate_response_row(source_row, insert_rows[duplicate_index])
        return status, [row for row in result_rows if row is not None]

    return status, body


def production_queue_payload_hash(rows: list[dict]) -> str:
    canonical = json.dumps(rows, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def production_queue_dedupe_key(row: dict) -> str:
    signature = production_record_duplicate_signature(row)
    return hashlib.sha256(signature.encode("utf-8")).hexdigest()


def production_queue_to_client(row: dict, include_payload: bool = False) -> dict:
    result = {
        key: value
        for key, value in row.items()
        if key not in {"payload", "payload_hash", "locked_by"}
    }
    result["queue_no"] = f"Q-{int(row.get('id') or 0):06d}"
    if include_payload:
        result["payload"] = row.get("payload") if isinstance(row.get("payload"), list) else []
    return result


def production_queue_event(
    queue_id: int,
    event_type: str,
    status: str,
    message: str = "",
    actor: str = "system",
    metadata: dict | None = None,
) -> None:
    supabase_request(
        "POST",
        "production_save_queue_events",
        {
            "queue_id": queue_id,
            "event_type": event_type,
            "status": status,
            "message": message or None,
            "actor": actor or "system",
            "metadata": metadata or {},
        },
        prefer="return=minimal",
        timeout_seconds=5,
    )


def update_production_queue(
    queue_id: int,
    values: dict,
    expected_status: str | None = None,
    worker_id: str | None = None,
) -> tuple[int, dict | None]:
    filters = [f"id=eq.{queue_id}"]
    if expected_status:
        filters.append(f"status=eq.{quote(expected_status)}")
    if worker_id:
        filters.append(f"locked_by=eq.{quote(worker_id)}")
    status, body = supabase_request(
        "PATCH",
        f"production_save_queue?{'&'.join(filters)}",
        values,
        prefer="return=representation",
        timeout_seconds=6,
    )
    row = body[0] if status < 400 and isinstance(body, list) and body else None
    return status, row


def production_queue_success_values(rows: list[dict]) -> dict:
    client_rows = [live_state_to_client("production_records", row) for row in rows if isinstance(row, dict)]
    record_ids = [int(row.get("id")) for row in rows if str(row.get("id") or "").isdigit()]
    return {
        "status": "succeeded",
        "result_record_ids": record_ids,
        "result_payload": client_rows,
        "duplicate_details": {},
        "error_code": None,
        "error_message": None,
        "finished_at": datetime.now(timezone.utc).isoformat(),
        "locked_at": None,
        "locked_by": None,
    }


def inspect_production_queue_rows(rows: list[dict]) -> tuple[str, list[dict], dict]:
    lookup_status, existing_rows = production_duplicate_lookup_rows(rows)
    if lookup_status >= 400:
        return "error", [], {"status": lookup_status, "error": existing_rows}
    existing_by_uid = {
        production_record_client_uid(row): row
        for row in existing_rows
        if production_record_client_uid(row)
    }
    existing_by_signature = {
        production_record_duplicate_signature(row): row
        for row in existing_rows
    }
    matched = []
    for incoming in rows:
        uid = production_record_client_uid(incoming)
        existing = existing_by_uid.get(uid) if uid else None
        existing = existing or existing_by_signature.get(production_record_duplicate_signature(incoming))
        if existing:
            matched.append(production_duplicate_response_row(existing, incoming))
    if len(matched) == len(rows):
        return "exact", matched, {"match_count": len(matched)}
    near_error = near_duplicate_batch_error(rows, existing_rows)
    if near_error:
        return "near", matched, near_error
    return "not_found", matched, {
        "match_count": len(matched),
        "expected_count": len(rows),
        "existing_ids": [row.get("id") for row in matched if row.get("id") is not None],
    }


def production_uid_uniqueness_error(rows: list[dict]) -> dict | None:
    for row in rows:
        uid = production_record_client_uid(row)
        if not uid:
            return {"error": "Production record is missing client_uid.", "duplicate_guard": "missing_client_uid"}
        uid_filter_field = quote("raw_payload->>client_uid", safe="")
        status, matches = supabase_request(
            "GET",
            f"production_records?{uid_filter_field}=eq.{quote(uid)}&select=id&order=id.asc&limit=3",
        )
        if status >= 400:
            return {"error": "Could not verify production client_uid uniqueness.", "duplicate_guard": "uid_verify_failed", "status": status}
        ids = [int(item.get("id")) for item in matches if isinstance(item, dict) and str(item.get("id") or "").isdigit()] if isinstance(matches, list) else []
        if len(ids) != 1:
            return {
                "error": "พบ client_uid ซ้ำในฐานข้อมูล ระบบหยุดคิวไว้เพื่อตรวจสอบ",
                "duplicate_guard": "duplicate_client_uid",
                "client_uid": uid,
                "existing_ids": ids,
            }
    return None


def validate_production_queue_rows(rows: list[dict]) -> str | None:
    today = datetime.now(timezone(timedelta(hours=7))).date().isoformat()
    for index, row in enumerate(rows, start=1):
        raw = production_record_raw_payload(row)
        record_date = str(row.get("record_date") or raw.get("record_date") or raw.get("date") or "")
        emp_code = str(row.get("emp_code") or raw.get("emp_code") or "").strip()
        fruit_type = str(row.get("fruit_type") or raw.get("fruit_type") or "mangosteen")
        pile = production_pile_number(raw) or production_pile_number(row)
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", record_date) or record_date > today:
            return f"Queue row {index} has an invalid or future production date."
        if not emp_code or pile is None or pile < 1:
            return f"Queue row {index} has an invalid employee code or pile."
        if fruit_type == "durian":
            weight = production_record_weight_total(row)
            rate = safe_float(raw.get("durian_rate", raw.get("grade_rates", {}).get("A", 0)))
            amount = safe_float(row.get("amount", raw.get("total_amount", raw.get("grand_total", 0))))
            if weight <= 0 or abs(weight * 10 - round(weight * 10)) > 0.001:
                return f"Queue row {index} has an invalid durian weight."
            if rate <= 0 or abs(amount - weight * rate) > 0.05:
                return f"Queue row {index} has an invalid durian rate or amount."
            continue
        water = safe_float(row.get("water_weight", raw.get("water_weight", raw.get("water", 0))))
        flower = safe_float(row.get("flower_weight", raw.get("flower_weight", raw.get("flower", 0))))
        water_rate = safe_float(raw.get("water_rate"))
        flower_rate = safe_float(raw.get("flower_rate"))
        amount = safe_float(row.get("amount", raw.get("total_amount", raw.get("grand_total", 0))))
        if water < 0 or flower < 0:
            return f"Queue row {index} has a negative production weight."
        if any(abs(value * 10 - round(value * 10)) > 0.001 for value in (water, flower)):
            return f"Queue row {index} has more than one weight decimal place."
        if water_rate <= 0 or flower_rate <= 0 or abs(amount - (water * water_rate + flower * flower_rate)) > 0.05:
            return f"Queue row {index} has an invalid rate or amount."
    return None


def finish_production_queue_success(job: dict, rows: list[dict], event_type: str = "succeeded") -> None:
    queue_id = int(job.get("id") or 0)
    status, updated = update_production_queue(
        queue_id,
        production_queue_success_values(rows),
        expected_status="processing" if job.get("status") == "processing" else None,
        worker_id=production_save_queue_worker_id if job.get("status") == "processing" else None,
    )
    if status < 400 and updated:
        production_queue_event(
            queue_id,
            event_type,
            "succeeded",
            f"Saved {len(rows)} production record(s).",
            production_save_queue_worker_id,
            {"record_ids": updated.get("result_record_ids", [])},
        )


def process_production_save_queue_job(job: dict) -> None:
    queue_id = int(job.get("id") or 0)
    rows = job.get("payload") if isinstance(job.get("payload"), list) else []
    if not rows or len(rows) != int(job.get("record_count") or 0):
        update_production_queue(
            queue_id,
            {
                "status": "needs_review",
                "error_code": "invalid_payload",
                "error_message": "Queue payload is missing or record count does not match.",
                "finished_at": datetime.now(timezone.utc).isoformat(),
                "locked_at": None,
                "locked_by": None,
            },
            expected_status="processing",
            worker_id=production_save_queue_worker_id,
        )
        production_queue_event(queue_id, "integrity_failed", "needs_review", "Queue payload validation failed.", production_save_queue_worker_id)
        return
    if not hmac.compare_digest(str(job.get("payload_hash") or ""), production_queue_payload_hash(rows)):
        update_production_queue(
            queue_id,
            {
                "status": "needs_review",
                "error_code": "payload_hash_mismatch",
                "error_message": "Queue payload hash does not match the accepted data.",
                "finished_at": datetime.now(timezone.utc).isoformat(),
                "locked_at": None,
                "locked_by": None,
            },
            expected_status="processing",
            worker_id=production_save_queue_worker_id,
        )
        production_queue_event(queue_id, "integrity_failed", "needs_review", "Queue payload hash mismatch.", production_save_queue_worker_id)
        return

    validation_error = validate_production_queue_rows(rows)
    emp_code = str(rows[0].get("emp_code") or "").strip()
    employee_status, employees = supabase_request(
        "GET",
        f"employees?emp_code=eq.{quote(emp_code)}&status=eq.Active&select=id,emp_code,fullname&limit=1",
        timeout_seconds=5,
    )
    if employee_status >= 400:
        validation_error = validation_error or f"Employee validation failed: {supabase_error_text(employees)}"
    elif not isinstance(employees, list) or not employees:
        validation_error = validation_error or "The employee is missing or inactive in the central database."
    if validation_error:
        update_production_queue(
            queue_id,
            {
                "status": "needs_review",
                "error_code": "validation_failed",
                "error_message": validation_error,
                "finished_at": datetime.now(timezone.utc).isoformat(),
                "locked_at": None,
                "locked_by": None,
            },
            expected_status="processing",
            worker_id=production_save_queue_worker_id,
        )
        production_queue_event(queue_id, "validation_failed", "needs_review", validation_error, production_save_queue_worker_id)
        return

    with production_record_insert_lock:
        status, body = insert_production_records_compatible(rows)
    if status < 400 and isinstance(body, list) and len(body) == len(rows):
        uid_error = production_uid_uniqueness_error(rows)
        if uid_error:
            update_production_queue(
                queue_id,
                {
                    "status": "needs_review",
                    "duplicate_details": uid_error,
                    "error_code": str(uid_error.get("duplicate_guard") or "duplicate_client_uid"),
                    "error_message": str(uid_error.get("error")),
                    "finished_at": datetime.now(timezone.utc).isoformat(),
                    "locked_at": None,
                    "locked_by": None,
                },
                expected_status="processing",
                worker_id=production_save_queue_worker_id,
            )
            production_queue_event(queue_id, "uid_duplicate_review", "needs_review", str(uid_error.get("error")), production_save_queue_worker_id, uid_error)
            return
        finish_production_queue_success(job, body)
        return
    if status >= 400:
        match_type, matched_rows, match_details = inspect_production_queue_rows(rows)
        if match_type == "exact":
            finish_production_queue_success(job, matched_rows, event_type="concurrent_duplicate_resolved")
            return
        if match_type == "near":
            body = match_details
            status = 409
    if status == 409 and isinstance(body, dict):
        update_production_queue(
            queue_id,
            {
                "status": "needs_review",
                "duplicate_details": body,
                "error_code": str(body.get("duplicate_guard") or "possible_duplicate"),
                "error_message": str(body.get("error") or "Possible duplicate production data was found."),
                "finished_at": datetime.now(timezone.utc).isoformat(),
                "locked_at": None,
                "locked_by": None,
            },
            expected_status="processing",
            worker_id=production_save_queue_worker_id,
        )
        production_queue_event(queue_id, "duplicate_review", "needs_review", str(body.get("error") or "Possible duplicate."), production_save_queue_worker_id, body)
        return

    attempt_count = int(job.get("attempt_count") or 1)
    max_attempts = int(job.get("max_attempts") or 3)
    error_message = supabase_error_text(body) or f"Production save failed ({status})."
    if status >= 500 and attempt_count < max_attempts:
        update_production_queue(
            queue_id,
            {
                "status": "queued",
                "next_attempt_at": (datetime.now(timezone.utc) + timedelta(seconds=2)).isoformat(),
                "error_code": "temporary_cloud_error",
                "error_message": error_message,
                "locked_at": None,
                "locked_by": None,
            },
            expected_status="processing",
            worker_id=production_save_queue_worker_id,
        )
        production_queue_event(queue_id, "retry_scheduled", "queued", error_message, production_save_queue_worker_id, {"attempt": attempt_count})
        production_save_queue_wakeup.set()
        return

    update_production_queue(
        queue_id,
        {
            "status": "needs_review",
            "error_code": "save_failed",
            "error_message": error_message,
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "locked_at": None,
            "locked_by": None,
        },
        expected_status="processing",
        worker_id=production_save_queue_worker_id,
    )
    production_queue_event(queue_id, "save_failed", "needs_review", error_message, production_save_queue_worker_id, {"status": status})


def production_save_queue_worker() -> None:
    while True:
        production_save_queue_wakeup.wait(timeout=2)
        production_save_queue_wakeup.clear()
        if not supabase_configured():
            continue
        while True:
            status, body = supabase_request(
                "POST",
                "rpc/claim_next_production_save_queue",
                {"p_worker_id": production_save_queue_worker_id},
                timeout_seconds=6,
            )
            if status >= 400 or not isinstance(body, list) or not body:
                break
            try:
                with backup_clear_lock:
                    process_production_save_queue_job(body[0])
            except Exception as error:
                queue_id = int(body[0].get("id") or 0)
                update_production_queue(
                    queue_id,
                    {
                        "status": "needs_review",
                        "error_code": "worker_exception",
                        "error_message": str(error),
                        "finished_at": datetime.now(timezone.utc).isoformat(),
                        "locked_at": None,
                        "locked_by": None,
                    },
                    expected_status="processing",
                    worker_id=production_save_queue_worker_id,
                )
                production_queue_event(queue_id, "worker_exception", "needs_review", str(error), production_save_queue_worker_id)


def time_queue_to_client(row: dict, include_payload: bool = False) -> dict:
    result = {key: value for key, value in row.items() if key not in {"payload", "payload_hash", "locked_by"}}
    result["queue_no"] = f"TQ-{int(row.get('id') or 0):06d}"
    if include_payload:
        result["payload"] = row.get("payload") if isinstance(row.get("payload"), list) else []
    return result


def time_queue_event(queue_id: int, event_type: str, status: str, message: str, actor: str, metadata: dict | None = None) -> None:
    supabase_request(
        "POST",
        "time_save_queue_events",
        {"queue_id": queue_id, "event_type": event_type, "status": status, "message": message,
         "actor": actor, "metadata": metadata or {}},
        prefer="return=minimal",
        timeout_seconds=5,
    )


def update_time_queue(queue_id: int, values: dict, expected_status: str | None = None) -> tuple[int, dict | None]:
    filters = [f"id=eq.{queue_id}"]
    if expected_status:
        filters.append(f"status=eq.{quote(expected_status)}")
    status, body = supabase_request(
        "PATCH", f"time_save_queue?{'&'.join(filters)}", values,
        prefer="return=representation", timeout_seconds=6,
    )
    return status, body[0] if status < 400 and isinstance(body, list) and body else None


def time_queue_row_key(queue_uid: str, index: int) -> str:
    return hashlib.sha256(f"{queue_uid}:{index}".encode("utf-8")).hexdigest()


def time_queue_existing_rows(rows: list[dict]) -> tuple[int, list[dict] | dict]:
    keys = [str(row.get("queue_dedupe_key") or "") for row in rows if row.get("queue_dedupe_key")]
    matched_rows: list[dict] = []
    if keys:
        status, body = supabase_request(
            "GET",
            f"time_records?queue_dedupe_key=in.({','.join(quote(key) for key in keys)})&select=*",
            timeout_seconds=6,
        )
        if status >= 400:
            return status, body
        if isinstance(body, list):
            matched_rows.extend(row for row in body if isinstance(row, dict))

    matched_keys = {str(row.get("queue_dedupe_key") or "") for row in matched_rows}
    used_ids = {str(row.get("id") or "") for row in matched_rows}
    # A database insert can succeed immediately before the queue status update
    # fails. Older rows may therefore have no dedupe key. An exact business-key
    # match is safe to recover because identical/overlapping shifts are never
    # valid as a second attendance record.
    for row in rows:
        dedupe_key = str(row.get("queue_dedupe_key") or "")
        if dedupe_key and dedupe_key in matched_keys:
            continue
        work_date, emp_code = time_record_identity(row)
        check_in = str(row.get("check_in") or row.get("clock_in") or "").strip()
        check_out = str(row.get("check_out") or row.get("clock_out") or "").strip()
        if not work_date or not emp_code or not check_in or not check_out:
            continue
        exact_status, exact_body = supabase_request(
            "GET",
            "time_records?"
            f"work_date=eq.{quote(work_date)}&emp_code=eq.{quote(emp_code)}&"
            f"check_in=eq.{quote(check_in)}&check_out=eq.{quote(check_out)}&"
            "select=*&order=id.asc&limit=2",
            timeout_seconds=6,
        )
        if exact_status >= 400:
            return exact_status, exact_body
        exact_match = next(
            (
                candidate for candidate in exact_body if isinstance(candidate, dict)
                and str(candidate.get("id") or "") not in used_ids
            ),
            None,
        ) if isinstance(exact_body, list) else None
        if exact_match:
            matched_rows.append(exact_match)
            used_ids.add(str(exact_match.get("id") or ""))
            if dedupe_key:
                matched_keys.add(dedupe_key)
    return 200, matched_rows


def finish_time_queue(job: dict, rows: list[dict], event_type: str = "succeeded") -> None:
    client_rows = [live_state_to_client("time_records", row) for row in rows if isinstance(row, dict)]
    record_ids = [int(row["id"]) for row in rows if str(row.get("id") or "").isdigit()]
    operation = str(job.get("operation") or "insert")
    queue_uid = str(job.get("queue_uid") or "")
    audit_failure: tuple[int, str] | None = None
    for row in rows:
        raw = row.get("raw_payload") if isinstance(row.get("raw_payload"), dict) else {}
        if operation == "update" and not raw.get("audit_required"):
            continue
        audit_filter = quote("metadata->>time_queue_uid", safe="")
        audit_record_filter = quote("metadata->>record_id", safe="")
        audit_status, existing_audit = supabase_request(
            "GET", f"audit_logs?{audit_filter}=eq.{quote(queue_uid)}&{audit_record_filter}=eq.{quote(str(row.get('id') or ''))}&select=id&limit=1",
            timeout_seconds=5,
        )
        if audit_status >= 400:
            audit_failure = (audit_status, supabase_error_text(existing_audit) or "Audit lookup failed.")
            break
        if audit_status < 400 and isinstance(existing_audit, list) and existing_audit:
            continue
        before = raw.get("audit_before") if isinstance(raw.get("audit_before"), dict) else {}
        detail = (
            f"Edited queued time record #{row.get('id')}: "
            f"{before.get('record_date', '')} {before.get('clock_in', '')}-{before.get('clock_out', '')} -> "
            f"{row.get('work_date', '')} {row.get('check_in', '')}-{row.get('check_out', '')}"
            if operation == "update"
            else f"Added queued time record {row.get('emp_code', '')} {row.get('check_in', '')}-{row.get('check_out', '')}"
        )
        audit_status, audit_body = insert_audit_log_compatible(
            {"action": "UPDATE_TIME_RECORD" if operation == "update" else "INSERT_TIME_RECORD",
             "module": "time_records", "description": detail, "created_by": job.get("created_by"),
             "user_fullname": job.get("created_by"),
             "metadata": {"time_queue_uid": queue_uid, "queue_id": job.get("id"), "record_id": row.get("id")}},
        )
        if audit_status >= 400:
            audit_failure = (audit_status, supabase_error_text(audit_body) or "Audit insert failed.")
            break
    values = {
        "status": "succeeded", "result_record_ids": record_ids, "result_payload": client_rows,
        "error_code": "audit_log_pending" if audit_failure else None,
        "error_message": audit_failure[1] if audit_failure else None,
        "finished_at": datetime.now(timezone.utc).isoformat(),
        "locked_at": None, "locked_by": None,
    }
    status, updated = update_time_queue(int(job.get("id") or 0), values, expected_status="processing")
    if status < 400 and updated:
        final_event = "succeeded_audit_pending" if audit_failure else event_type
        event_message = (
            f"Saved {len(rows)} time record(s); audit log is pending: {audit_failure[1]}"
            if audit_failure else f"Saved {len(rows)} time record(s)."
        )
        time_queue_event(int(job["id"]), final_event, "succeeded", event_message, time_save_queue_worker_id,
                         {"record_ids": record_ids, "audit_pending": bool(audit_failure)})


def fail_time_queue(job: dict, code: str, message: str, retryable: bool = False) -> None:
    attempt = int(job.get("attempt_count") or 1)
    can_retry = retryable and attempt < int(job.get("max_attempts") or 3)
    values = {
        "status": "queued" if can_retry else "needs_review",
        "error_code": code, "error_message": message,
        "next_attempt_at": (datetime.now(timezone.utc) + timedelta(seconds=2)).isoformat(),
        "locked_at": None, "locked_by": None,
    }
    if not can_retry:
        values["finished_at"] = datetime.now(timezone.utc).isoformat()
    update_time_queue(int(job.get("id") or 0), values, expected_status="processing")
    time_queue_event(int(job.get("id") or 0), "retry_scheduled" if can_retry else "needs_review",
                     values["status"], message, time_save_queue_worker_id, {"attempt": attempt})
    if can_retry:
        time_save_queue_wakeup.set()


def process_time_save_queue_job(job: dict) -> None:
    rows = job.get("payload") if isinstance(job.get("payload"), list) else []
    if not rows or len(rows) != int(job.get("record_count") or 0):
        fail_time_queue(job, "invalid_payload", "Queue payload is missing or record count does not match.")
        return
    if not hmac.compare_digest(str(job.get("payload_hash") or ""), production_queue_payload_hash(rows)):
        fail_time_queue(job, "payload_hash_mismatch", "Queue payload hash does not match the accepted data.")
        return

    operation = str(job.get("operation") or "insert")
    existing_status, already_saved = time_queue_existing_rows(rows)
    if existing_status >= 500:
        fail_time_queue(job, "lookup_failed", supabase_error_text(already_saved), retryable=True)
        return
    if operation == "insert" and isinstance(already_saved, list) and len(already_saved) == len(rows):
        finish_time_queue(job, already_saved, "idempotent_recovery")
        return

    pending_rows = rows
    recovered_rows = []
    if operation == "insert" and isinstance(already_saved, list):
        recovered_by_key = {str(row.get("queue_dedupe_key") or ""): row for row in already_saved}
        recovered_rows = list(recovered_by_key.values())
        pending_rows = [row for row in rows if str(row.get("queue_dedupe_key") or "") not in recovered_by_key]

    employee_code = str(rows[0].get("emp_code") or "").strip()
    employee_status, employees = supabase_request(
        "GET", f"time_employees?emp_code=eq.{quote(employee_code)}&status=eq.Active&select=id&limit=1",
        timeout_seconds=5,
    )
    if employee_status >= 500:
        fail_time_queue(job, "employee_lookup_failed", supabase_error_text(employees), retryable=True)
        return
    if employee_status >= 400 or not isinstance(employees, list) or not employees:
        fail_time_queue(job, "employee_invalid", "The time employee is missing or inactive.")
        return

    conflict_status, conflict = validate_time_record_conflicts(pending_rows)
    if conflict_status >= 500:
        fail_time_queue(job, "conflict_check_failed", supabase_error_text(conflict), retryable=True)
        return
    if conflict_status >= 400:
        conflict_code = "time_overlap" if conflict_status == 409 else "invalid_time"
        fail_time_queue(job, conflict_code, str((conflict or {}).get("error") or "Time record validation failed."))
        return

    with time_record_insert_lock:
        if operation == "update":
            status, body = update_time_records_compatible(rows)
        else:
            status, body = insert_time_records_compatible(pending_rows)
    expected_saved_count = len(pending_rows)
    if status < 400 and isinstance(body, list) and len(body) == expected_saved_count:
        finish_time_queue(job, [*recovered_rows, *body], "partial_recovery" if recovered_rows else "succeeded")
        return
    if status >= 500:
        fail_time_queue(job, "temporary_cloud_error", supabase_error_text(body), retryable=True)
        return
    fail_time_queue(job, "save_failed", supabase_error_text(body) or f"Time save failed ({status}).")


def recover_stale_time_queues() -> None:
    global time_queue_last_recovery_at
    now_monotonic = time.monotonic()
    with time_queue_recovery_lock:
        if now_monotonic - time_queue_last_recovery_at < 30:
            return
        time_queue_last_recovery_at = now_monotonic
    now_iso = datetime.now(timezone.utc).isoformat()
    stale_lock_iso = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
    # Recover a worker that disappeared after claiming a job. The payload stays in
    # the same queue row, so another worker can safely continue it.
    supabase_request(
        "PATCH",
        f"time_save_queue?status=eq.processing&locked_at=lt.{quote(stale_lock_iso)}",
        {
            "status": "queued",
            "next_attempt_at": now_iso,
            "locked_at": None,
            "locked_by": None,
            "error_code": "stale_worker_recovered",
            "error_message": "The previous worker stopped responding; the queue was recovered automatically.",
        },
        prefer="return=minimal",
        timeout_seconds=6,
    )
    supabase_request(
        "PATCH",
        "time_save_queue?status=eq.processing&locked_at=is.null",
        {
            "status": "queued",
            "next_attempt_at": now_iso,
            "error_code": "stale_worker_recovered",
            "error_message": "A queue without a worker lock was recovered automatically.",
        },
        prefer="return=minimal",
        timeout_seconds=6,
    )


def claim_time_queue_fallback(queue_id: int | None = None) -> tuple[int, dict | None]:
    recover_stale_time_queues()
    now_iso = datetime.now(timezone.utc).isoformat()
    filters = ["status=eq.queued", f"next_attempt_at=lte.{quote(now_iso)}"]
    if queue_id:
        filters.append(f"id=eq.{queue_id}")
    status, body = supabase_request(
        "GET",
        f"time_save_queue?{'&'.join(filters)}&select=*&order=id.asc&limit=1",
        timeout_seconds=6,
    )
    if status >= 400 or not isinstance(body, list) or not body:
        return status, None
    candidate = body[0]
    candidate_id = int(candidate.get("id") or 0)
    attempt_count = int(candidate.get("attempt_count") or 0)
    max_attempts = int(candidate.get("max_attempts") or 3)
    if not candidate_id:
        return 200, None
    if attempt_count >= max_attempts:
        review_status, reviewed = supabase_request(
            "PATCH",
            f"time_save_queue?id=eq.{candidate_id}&status=eq.queued&attempt_count=eq.{attempt_count}",
            {
                "status": "needs_review",
                "error_code": "retry_limit_reached",
                "error_message": "Automatic retries reached the safety limit; the queue data is preserved for review.",
                "finished_at": now_iso,
                "locked_at": None,
                "locked_by": None,
            },
            prefer="return=representation",
            timeout_seconds=6,
        )
        if review_status < 400 and isinstance(reviewed, list) and reviewed:
            time_queue_event(
                candidate_id,
                "retry_limit_reached",
                "needs_review",
                "Automatic retries reached the safety limit.",
                "system",
            )
        return review_status, None
    claim_status, claimed = supabase_request(
        "PATCH",
        f"time_save_queue?id=eq.{candidate_id}&status=eq.queued&attempt_count=eq.{attempt_count}",
        {
            "status": "processing",
            "attempt_count": attempt_count + 1,
            "locked_at": now_iso,
            "locked_by": time_save_queue_worker_id,
            "started_at": candidate.get("started_at") or now_iso,
            "error_code": None,
            "error_message": None,
        },
        prefer="return=representation",
        timeout_seconds=6,
    )
    row = claimed[0] if claim_status < 400 and isinstance(claimed, list) and claimed else None
    return claim_status, row


def process_time_queue_by_id(queue_id: int) -> None:
    status, job = claim_time_queue_fallback(queue_id)
    if status >= 400 or not job:
        time_save_queue_wakeup.set()
        return
    try:
        with backup_clear_lock:
            process_time_save_queue_job(job)
    except Exception as error:
        fail_time_queue(job, "worker_exception", str(error))


def schedule_time_queue_job(queue_id: int) -> None:
    time_save_queue_wakeup.set()
    if queue_id <= 0:
        return
    threading.Thread(
        target=process_time_queue_by_id,
        args=(queue_id,),
        name=f"time-save-queue-{queue_id}",
        daemon=True,
    ).start()


def time_save_queue_worker() -> None:
    while True:
        time_save_queue_wakeup.wait(timeout=2)
        time_save_queue_wakeup.clear()
        if not supabase_configured():
            continue
        while True:
            status, body = supabase_request(
                "POST", "rpc/claim_next_time_save_queue",
                {"p_worker_id": time_save_queue_worker_id}, timeout_seconds=6,
            )
            if status >= 400:
                fallback_status, fallback_job = claim_time_queue_fallback()
                if fallback_status >= 400 or not fallback_job:
                    break
                body = [fallback_job]
            if not isinstance(body, list) or not body:
                break
            try:
                with backup_clear_lock:
                    process_time_save_queue_job(body[0])
            except Exception as error:
                fail_time_queue(body[0], "worker_exception", str(error))


def reserve_deduction_record_id(refresh: bool = False) -> int:
    global deduction_record_next_id
    if refresh or deduction_record_next_id is None:
        deduction_record_next_id = next_table_id("deduction_records")
    next_id = deduction_record_next_id
    deduction_record_next_id += 1
    return next_id


def deduction_duplicate_lookup(deduction: dict) -> tuple[int, list[dict] | dict]:
    client_uid = str(deduction.get("client_uid") or "").strip()
    if client_uid:
        status, body = supabase_request(
            "GET",
            f"deduction_records?client_uid=eq.{quote(client_uid)}&select=*&limit=1",
        )
        if status >= 400:
            text = supabase_error_text(body).lower()
            if "client_uid" not in text and "column" not in text and "schema cache" not in text:
                return status, {"error": body}
        elif isinstance(body, list) and body:
            return 200, body
    amount = round(safe_float(deduction.get("amount")), 2)
    status, body = supabase_request(
        "GET",
        "deduction_records?"
        f"employee_kind=eq.{quote(str(deduction.get('employee_kind') or ''))}"
        f"&employee_id=eq.{quote(str(deduction.get('employee_id') or ''))}"
        f"&start_date=eq.{quote(str(deduction.get('start_date') or ''))}"
        f"&end_date=eq.{quote(str(deduction.get('end_date') or ''))}"
        f"&deduction_type=eq.{quote(str(deduction.get('deduction_type') or ''))}"
        f"&deduction_label=eq.{quote(str(deduction.get('deduction_label') or ''))}"
        "&select=*",
    )
    if status >= 400:
        return status, {"error": body}
    matches = [
        row for row in (body if isinstance(body, list) else [])
        if abs(round(safe_float(row.get("amount")), 2) - amount) < 0.005
        and str(row.get("status") or "") == str(deduction.get("status") or "")
    ]
    return 200, matches


def insert_deduction_record_compatible(deduction: dict) -> tuple[int, dict | list | str | None]:
    lookup_status, duplicate_rows = deduction_duplicate_lookup(deduction)
    if lookup_status >= 400:
        return lookup_status, duplicate_rows
    if isinstance(duplicate_rows, list) and duplicate_rows:
        return 200, duplicate_rows[:1]

    insert_row = dict(deduction)
    insert_row["id"] = reserve_deduction_record_id()
    for attempt in range(4):
        status, body = supabase_request(
            "POST",
            "deduction_records",
            insert_row,
            prefer="return=representation",
        )
        if status < 400:
            return status, body
        if not is_unique_constraint_error(body, "deduction_records_pkey"):
            text = supabase_error_text(body).lower()
            if "client_uid" in text and ("column" in text or "schema cache" in text):
                fallback_row = {key: value for key, value in insert_row.items() if key != "client_uid"}
                status, body = supabase_request(
                    "POST",
                    "deduction_records",
                    fallback_row,
                    prefer="return=representation",
                )
                if status < 400 or not is_unique_constraint_error(body, "deduction_records_pkey"):
                    return status, body
            return status, body
        insert_row["id"] = reserve_deduction_record_id(refresh=(attempt == 0))
    return 409, {"error": "Could not allocate a unique deduction record id."}


def reserve_time_record_id(refresh: bool = False) -> int:
    global time_record_next_id
    if refresh or time_record_next_id is None:
        time_record_next_id = next_table_id("time_records")
    next_id = time_record_next_id
    time_record_next_id += 1
    return next_id


def time_value_to_minutes(value: object) -> int | None:
    match = re.fullmatch(r"(\d{1,2}):(\d{2})(?::\d{2})?", str(value or "").strip())
    if not match:
        return None
    hours, minutes = (int(part) for part in match.groups())
    if hours > 23 or minutes > 59:
        return None
    return hours * 60 + minutes


def time_record_interval(clock_in: object, clock_out: object) -> tuple[int, int] | None:
    start = time_value_to_minutes(clock_in)
    end = time_value_to_minutes(clock_out)
    if start is None or end is None:
        return None
    if end <= start:
        end += 24 * 60
    return start, end


def time_record_intervals_overlap(first: dict, second: dict) -> bool:
    first_interval = time_record_interval(
        first.get("check_in") or first.get("clock_in"),
        first.get("check_out") or first.get("clock_out"),
    )
    second_interval = time_record_interval(
        second.get("check_in") or second.get("clock_in"),
        second.get("check_out") or second.get("clock_out"),
    )
    if first_interval is None or second_interval is None:
        return False
    return first_interval[0] < second_interval[1] and second_interval[0] < first_interval[1]


def time_record_identity(row: dict) -> tuple[str, str]:
    raw = row.get("raw_payload") if isinstance(row.get("raw_payload"), dict) else {}
    work_date = row.get("work_date") or row.get("record_date") or raw.get("record_date") or raw.get("date")
    emp_code = row.get("emp_code") or raw.get("emp_code")
    return str(work_date or "").strip(), str(emp_code or "").strip()


def validate_time_record_conflicts(rows: list[dict]) -> tuple[int, dict | None]:
    prepared: list[tuple[dict, str, str, tuple[int, int]]] = []
    for index, row in enumerate(rows):
        work_date, emp_code = time_record_identity(row)
        interval = time_record_interval(
            row.get("check_in") or row.get("clock_in"),
            row.get("check_out") or row.get("clock_out"),
        )
        if not work_date or not emp_code or interval is None:
            return 400, {"error": "ข้อมูลวันที่ รหัสพนักงาน หรือเวลาเข้าออกไม่ถูกต้อง"}
        prepared.append((row, work_date, emp_code, interval))

        for other in rows[:index]:
            if time_record_identity(other) != (work_date, emp_code):
                continue
            if row.get("id") in (None, "") and other.get("id") in (None, "") and time_record_intervals_overlap(row, other):
                return 409, {
                    "error": f"เวลาของพนักงานรหัส {emp_code} วันที่ {work_date} ทับกันภายในชุดที่ส่งมา"
                }

    dates_by_employee: dict[str, set[str]] = {}
    for _row, work_date, emp_code, _interval in prepared:
        dates_by_employee.setdefault(emp_code, set()).add(work_date)

    existing_by_identity: dict[tuple[str, str], list[dict]] = {}
    for emp_code, work_dates in dates_by_employee.items():
        date_list = ",".join(sorted(work_dates))
        status, existing_rows = supabase_request(
            "GET",
            "time_records?"
            f"work_date=in.({date_list})&emp_code=eq.{quote(emp_code)}&"
            "select=id,work_date,emp_code,check_in,check_out,raw_payload",
        )
        if status >= 400:
            return status, {"error": existing_rows}
        for existing in existing_rows if isinstance(existing_rows, list) else []:
            existing_identity = time_record_identity(existing)
            existing_by_identity.setdefault(existing_identity, []).append(existing)

    for row, work_date, emp_code, interval in prepared:
        incoming_id = row.get("id")
        existing_list = existing_by_identity.get((work_date, emp_code), [])
        existing_self = next(
            (
                existing
                for existing in existing_list
                if incoming_id not in (None, "") and str(existing.get("id")) == str(incoming_id)
            ),
            None,
        )
        if existing_self is not None:
            current_interval = time_record_interval(existing_self.get("check_in"), existing_self.get("check_out"))
            if current_interval == interval:
                continue

        for existing in existing_list:
            if incoming_id not in (None, "") and str(existing.get("id")) == str(incoming_id):
                continue
            if time_record_intervals_overlap(row, existing):
                existing_raw = existing.get("raw_payload") if isinstance(existing.get("raw_payload"), dict) else {}
                old_in = existing.get("check_in") or existing_raw.get("clock_in")
                old_out = existing.get("check_out") or existing_raw.get("clock_out")
                return 409, {
                    "error": (
                        f"เวลาของพนักงานรหัส {emp_code} ทับกับรายการเดิม "
                        f"{old_in}-{old_out} วันที่ {work_date}"
                    ),
                    "conflict_id": existing.get("id"),
                }
    return 200, None


def insert_time_records_compatible(rows: list[dict]) -> tuple[int, list | dict | str | None]:
    if not rows:
        return 200, []

    for attempt in range(4):
        insert_rows = []
        for row in rows:
            insert_row = dict(row)
            insert_row["id"] = reserve_time_record_id(refresh=(attempt == 1 and not insert_rows))
            insert_rows.append(insert_row)

        status, body = supabase_request(
            "POST",
            "time_records",
            insert_rows,
            prefer="return=representation",
        )
        if status < 400:
            return status, body
        if not is_unique_constraint_error(body, "time_records_pkey"):
            return status, body

    return 409, {"error": "Could not allocate unique time record ids."}


def update_time_records_compatible(rows: list[dict]) -> tuple[int, list | dict | str | None]:
    updated_rows = []
    for row in rows:
        row_id = row.get("id")
        if row_id in (None, ""):
            return 400, {"error": "time record id is required for update"}
        update_row = {
            key: value for key, value in row.items()
            if key != "id" and not (key == "queue_dedupe_key" and value in (None, ""))
        }
        status, body = supabase_request(
            "PATCH",
            f"time_records?id=eq.{quote(str(row_id))}",
            update_row,
            prefer="return=representation",
        )
        if status >= 400:
            return status, body
        if not isinstance(body, list) or not body:
            return 404, {"error": f"time record {row_id} was not found"}
        updated_rows.extend(body)
    return 200, updated_rows


def ensure_row_id(table: str, row: dict) -> dict:
    if row.get("id") in [None, ""]:
        return {**row, "id": next_table_id(table)}
    return row


def sync_rows_by_id(table: str, rows: list[dict]) -> tuple[int, dict]:
    synced = []
    next_id = None
    seen_production_uids: set[str] = set()
    for row in rows:
        clean_row = dict(row)
        incoming_raw = clean_row.get("raw_payload") if isinstance(clean_row.get("raw_payload"), dict) else {}
        incoming_production_uid = str(incoming_raw.get("client_uid") or "").strip() if table == "production_records" else ""
        if incoming_production_uid:
            if incoming_production_uid in seen_production_uids:
                continue
            seen_production_uids.add(incoming_production_uid)
            uid_filter_field = quote("raw_payload->>client_uid", safe="")
            uid_status, uid_rows = supabase_request(
                "GET",
                f"production_records?{uid_filter_field}=eq.{quote(incoming_production_uid)}&select=*&order=id.asc&limit=2",
            )
            if uid_status >= 400:
                return uid_status, {"error": uid_rows, "table": table, "client_uid": incoming_production_uid}
            if isinstance(uid_rows, list) and uid_rows:
                # A browser cache may carry a stale numeric id. client_uid is
                # the durable identity: return the central row and never mint
                # another id for the same logical production record.
                synced.append(uid_rows[0])
                continue
        if clean_row.get("id") in [None, ""]:
            if next_id is None:
                next_id = next_table_id(table)
            clean_row["id"] = next_id
            next_id += 1
        row_id = clean_row.get("id")
        select_fields = "id,raw_payload" if table in {"production_records", "time_records"} else "id"
        status, existing = supabase_request(
            "GET",
            f"{table}?id=eq.{quote(str(row_id))}&select={select_fields}&limit=1",
        )
        if status >= 400:
            return status, {"error": existing, "table": table}
        has_existing = isinstance(existing, list) and bool(existing)
        incoming_uid = ""
        existing_uid = ""
        if table == "production_records":
            incoming_raw = clean_row.get("raw_payload") if isinstance(clean_row.get("raw_payload"), dict) else {}
            existing_raw = existing[0].get("raw_payload") if has_existing and isinstance(existing[0].get("raw_payload"), dict) else {}
            incoming_uid = str(incoming_raw.get("client_uid") or "")
            existing_uid = str(existing_raw.get("client_uid") or "")
            identity_fields = (
                "record_date",
                "record_time",
                "emp_code",
                "pile_no",
                "fruit_type",
                "created_at",
            )
            incoming_identity = tuple(str(incoming_raw.get(field) or "") for field in identity_fields)
            existing_identity = tuple(str(existing_raw.get(field) or "") for field in identity_fields)
            uid_collision = has_existing and incoming_uid and incoming_uid != existing_uid
            legacy_collision = (
                has_existing
                and not incoming_uid
                and not existing_uid
                and incoming_identity != existing_identity
            )
            if uid_collision or legacy_collision:
                # Two browsers can allocate the same local numeric id. Allocate
                # the next central id explicitly because legacy identity
                # sequences may lag behind rows inserted with browser ids.
                clean_row["id"] = next_table_id(table)
                row_id = clean_row["id"]
                has_existing = False
        elif table == "time_records":
            incoming_raw = clean_row.get("raw_payload") if isinstance(clean_row.get("raw_payload"), dict) else {}
            existing_raw = existing[0].get("raw_payload") if has_existing and isinstance(existing[0].get("raw_payload"), dict) else {}
            incoming_created_at = str(incoming_raw.get("created_at") or clean_row.get("created_at") or "")
            existing_created_at = str(existing_raw.get("created_at") or existing[0].get("created_at") or "") if has_existing else ""
            incoming_identity = (
                str(incoming_raw.get("record_date") or clean_row.get("work_date") or ""),
                str(incoming_raw.get("emp_code") or clean_row.get("emp_code") or ""),
                incoming_created_at,
            )
            existing_identity = (
                str(existing_raw.get("record_date") or (existing[0].get("work_date") if has_existing else "") or ""),
                str(existing_raw.get("emp_code") or (existing[0].get("emp_code") if has_existing else "") or ""),
                existing_created_at,
            )
            if has_existing and incoming_created_at and existing_created_at and incoming_identity != existing_identity:
                clean_row["id"] = next_table_id(table)
                row_id = clean_row["id"]
                has_existing = False
        method = "PATCH" if has_existing else "POST"
        path = f"{table}?id=eq.{quote(str(row_id))}" if method == "PATCH" else table
        status, body = supabase_request(
            method,
            path,
            clean_row,
            prefer="return=representation",
        )
        if status >= 400 and table == "production_records" and any(
            key in clean_row for key in ("grade_weights", "grade_rates", "grade_amounts")
        ):
            # Older Supabase projects may not have the dedicated durian
            # columns yet. The complete record remains durable in raw_payload.
            fallback_row = {
                key: value
                for key, value in clean_row.items()
                if key not in ("grade_weights", "grade_rates", "grade_amounts")
            }
            status, body = supabase_request(
                method,
                path,
                fallback_row,
                prefer="return=representation",
            )
        if status >= 400:
            return status, {"error": body, "table": table, "row": clean_row}
        if isinstance(body, list):
            synced.extend(body)
    return 200, {"synced": synced}


def live_state_row(table: str, payload: dict) -> dict:
    """Convert the browser cache shape to the stable Supabase shape."""
    row_id = payload.get("id")
    if table == "production_sessions":
        row = {
            "session_date": payload.get("session_date") or payload.get("date"),
            "fruit_type": payload.get("fruit_type") or "mangosteen",
            "status": payload.get("status") or "open",
            "created_by": payload.get("created_by"),
            "closed_by": payload.get("closed_by"),
            "opened_at": payload.get("opened_at") or payload.get("start_time") or datetime.utcnow().isoformat() + "Z",
            "closed_at": payload.get("closed_at") or payload.get("end_time") or None,
            "raw_payload": payload,
        }
    elif table == "production_records":
        water = safe_float(payload.get("water_weight", payload.get("water", 0)))
        flower = safe_float(payload.get("flower_weight", payload.get("flower", 0)))
        pile_number = production_pile_number(payload)
        fruit_type = payload.get("fruit_type") or "mangosteen"
        total_weight = production_total_weight(payload) if fruit_type == "durian" else water + flower
        row = {
            "record_date": payload.get("record_date") or payload.get("date"),
            # Legacy browser ids belong to a different local database. Keep
            # them in raw_payload, rather than failing cloud migration on a
            # foreign-key mismatch.
            "session_id": None,
            "employee_id": None,
            "emp_code": payload.get("emp_code"),
            "employee_name": payload.get("employee_name") or payload.get("fullname"),
            "pay_group": payload.get("pay_group"),
            "fruit_type": fruit_type,
            "pile_no": str(pile_number) if pile_number is not None else None,
            "item_type": payload.get("item_type"),
            "water_weight": water,
            "flower_weight": flower,
            "total_weight": total_weight,
            "rate": payload.get("rate", 0) or 0,
            "amount": payload.get("amount", payload.get("total_amount", payload.get("grand_total", 0))) or 0,
            "note": payload.get("note"),
            "created_by": payload.get("created_by"),
            "updated_by": payload.get("updated_by"),
            "created_at": payload.get("created_at") or datetime.utcnow().isoformat() + "Z",
            "updated_at": payload.get("updated_at") or payload.get("created_at") or datetime.utcnow().isoformat() + "Z",
            "raw_payload": payload,
        }
        if row["fruit_type"] == "durian":
            row.update({
                "grade_weights": payload.get("grade_weights") or {},
                "grade_rates": payload.get("grade_rates") or {},
                "grade_amounts": payload.get("grade_amounts") or {},
            })
    elif table == "time_records":
        row = {
            "work_date": payload.get("work_date") or payload.get("record_date") or payload.get("date"),
            "employee_id": None,
            "emp_code": payload.get("emp_code"),
            "employee_name": payload.get("employee_name") or payload.get("fullname"),
            "check_in": payload.get("check_in") or payload.get("clock_in"),
            "check_out": payload.get("check_out") or payload.get("clock_out"),
            "break_minutes": payload.get("break_minutes", 0) or 0,
            "total_minutes": payload.get("total_minutes", payload.get("net_minutes", 0)) or 0,
            "note": payload.get("note"),
            "created_by": payload.get("created_by"),
            "updated_by": payload.get("updated_by"),
            "created_at": payload.get("created_at") or datetime.utcnow().isoformat() + "Z",
            "updated_at": payload.get("updated_at") or payload.get("created_at") or datetime.utcnow().isoformat() + "Z",
            "queue_dedupe_key": payload.get("queue_dedupe_key"),
            "raw_payload": payload,
        }
    else:
        row = {
            "action": payload.get("action") or "UNKNOWN",
            "module": payload.get("module"),
            "description": payload.get("description") or payload.get("detail"),
            "created_by": payload.get("created_by"),
            "user_fullname": payload.get("user_fullname"),
            "created_at": payload.get("created_at") or datetime.utcnow().isoformat() + "Z",
            "metadata": payload,
        }
    if row_id not in [None, ""]:
        row["id"] = row_id
    return row


def live_state_to_client(table: str, row: dict) -> dict:
    raw = row.get("raw_payload") if table != "audit_logs" else row.get("metadata")
    if isinstance(raw, dict):
        if table == "audit_logs":
            return {
                **raw,
                "id": row.get("id", raw.get("id")),
                "action": row.get("action") or raw.get("action"),
                "module": row.get("module") or raw.get("module"),
                "detail": row.get("description") or raw.get("detail") or raw.get("description"),
                "description": row.get("description") or raw.get("description") or raw.get("detail"),
                "created_by": row.get("created_by") or raw.get("created_by"),
                "user_fullname": row.get("user_fullname") or raw.get("user_fullname"),
                "created_at": row.get("created_at") or raw.get("created_at"),
            }
        return {**raw, "id": row.get("id", raw.get("id"))}
    return row


def supabase_configured() -> bool:
    return bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY)


def online_user_snapshot(now: float | None = None) -> dict:
    current_time = now or time.time()
    cutoff = current_time - ONLINE_USER_TIMEOUT_SECONDS
    with online_user_lock:
        expired = [
            client_id
            for client_id, session in online_user_sessions.items()
            if float(session.get("last_seen", 0)) < cutoff
        ]
        for client_id in expired:
            online_user_sessions.pop(client_id, None)
        users = sorted(
            online_user_sessions.values(),
            key=lambda session: str(session.get("fullname") or session.get("username") or ""),
        )
        return {
            "count": len(users),
            "timeout_seconds": ONLINE_USER_TIMEOUT_SECONDS,
            "users": [
                {
                    "username": session.get("username", ""),
                    "fullname": session.get("fullname", ""),
                    "route": session.get("route", ""),
                    "last_seen": session.get("last_seen_iso", ""),
                }
                for session in users
            ],
        }


def register_online_user(payload: dict) -> dict:
    client_id = str(payload.get("client_id") or "").strip()
    username = str(payload.get("username") or "").strip()
    if not client_id:
        client_id = secrets.token_urlsafe(18)
    now = time.time()
    with online_user_lock:
        online_user_sessions[client_id] = {
            "client_id": client_id,
            "username": username,
            "fullname": str(payload.get("fullname") or username or "ผู้ใช้งาน"),
            "route": str(payload.get("route") or ""),
            "last_seen": now,
            "last_seen_iso": datetime.utcnow().isoformat() + "Z",
        }
    snapshot = online_user_snapshot(now)
    snapshot["client_id"] = client_id
    return snapshot


def supabase_request(
    method: str,
    path: str,
    payload: dict | list | None = None,
    prefer: str | None = None,
    timeout_seconds: float = 20,
    extra_headers: dict[str, str] | None = None,
) -> tuple[int, dict | list | str | None]:
    if not supabase_configured():
        return 503, {"error": "Supabase environment variables are not configured."}

    url = f"{SUPABASE_URL}/rest/v1/{path.lstrip('/')}"
    data = None
    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    if extra_headers:
        headers.update(extra_headers)
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    request = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            raw = response.read().decode("utf-8")
            if not raw:
                return response.status, None
            try:
                return response.status, json.loads(raw)
            except json.JSONDecodeError:
                return response.status, raw
    except urllib.error.HTTPError as error:
        raw = error.read().decode("utf-8")
        try:
            body = json.loads(raw)
        except json.JSONDecodeError:
            body = {"error": raw or error.reason}
        return error.code, body
    except Exception as error:
        return 500, {"error": str(error)}


def supabase_get_all(path: str, page_size: int = 100, timeout_seconds: float = 20) -> tuple[int, list | dict | str | None]:
    """Read every PostgREST page so exports never inherit a server row cap."""
    page_size = max(1, min(int(page_size), 1000))
    rows: list = []
    offset = 0
    while True:
        status, body = supabase_request(
            "GET",
            path,
            timeout_seconds=timeout_seconds,
            extra_headers={"Range-Unit": "items", "Range": f"{offset}-{offset + page_size - 1}"},
        )
        if status >= 400:
            return status, body
        if not isinstance(body, list):
            return 502, {"error": "Supabase collection response was not a list."}
        rows.extend(body)
        if len(body) < page_size:
            return status, rows
        offset += len(body)


def insert_audit_log_compatible(audit_row: dict) -> tuple[int, dict | list | str | None]:
    """Write audit rows across legacy schemas and repaired/imported sequences."""
    variants = [
        audit_row,
        {key: value for key, value in audit_row.items() if key != "ip_address"},
        {key: value for key, value in audit_row.items() if key not in {"ip_address", "user_fullname"}},
    ]
    last_status: int = 500
    last_result: dict | list | str | None = {"error": "Audit log insert failed."}
    attempted_shapes: set[tuple[str, ...]] = set()
    for candidate in variants:
        shape = tuple(sorted(candidate))
        if shape in attempted_shapes:
            continue
        attempted_shapes.add(shape)
        last_status, last_result = supabase_request(
            "POST",
            "audit_logs",
            candidate,
            prefer="return=representation",
        )
        if last_status < 400:
            return last_status, last_result

    # Imported audit rows can leave the PostgreSQL identity sequence behind
    # the highest existing id. Fall back to a guarded explicit id so payroll
    # edits are not blocked by a duplicate primary key.
    id_status, id_rows = supabase_request(
        "GET",
        "audit_logs?select=id&order=id.desc&limit=1",
    )
    if id_status < 400 and isinstance(id_rows, list):
        highest_id = max(
            [int(row.get("id") or 0) for row in id_rows if isinstance(row, dict)] or [0]
        )
        for offset in range(1, 6):
            candidate_id = highest_id + offset
            for candidate in variants:
                explicit_candidate = {**candidate, "id": candidate_id}
                last_status, last_result = supabase_request(
                    "POST",
                    "audit_logs",
                    explicit_candidate,
                    prefer="return=representation",
                )
                if last_status < 400:
                    return last_status, last_result
    return last_status, last_result


def account_phone_column_missing(body: dict | list | str | None) -> bool:
    text = json.dumps(body, ensure_ascii=False).lower() if isinstance(body, (dict, list)) else str(body or "").lower()
    return "phone" in text and ("column" in text or "schema cache" in text or "pgrst" in text)


def supabase_account_write(
    method: str,
    path: str,
    account: dict,
    prefer: str = "return=representation",
) -> tuple[int, dict | list | str | None]:
    status, body = supabase_request(method, path, account, prefer=prefer)
    if status >= 400 and "phone" in account and account_phone_column_missing(body):
        fallback_account = {key: value for key, value in account.items() if key != "phone"}
        status, body = supabase_request(method, path, fallback_account, prefer=prefer)
        if status < 400:
            if isinstance(body, list):
                body = [{**row, "phone": account.get("phone", "")} if isinstance(row, dict) else row for row in body]
            elif isinstance(body, dict):
                body = {**body, "phone": account.get("phone", "")}
    return status, body


def supabase_account_bulk_write(
    method: str,
    path: str,
    accounts: list[dict],
    prefer: str,
) -> tuple[int, dict | list | str | None]:
    status, body = supabase_request(method, path, accounts, prefer=prefer)
    if status >= 400 and accounts and account_phone_column_missing(body):
        fallback_accounts = [
            {key: value for key, value in account.items() if key != "phone"}
            for account in accounts
        ]
        status, body = supabase_request(method, path, fallback_accounts, prefer=prefer)
    return status, body


def backup_authorized(handler: BaseHTTPRequestHandler) -> bool:
    provided = handler.headers.get("X-Backup-Code", "")
    return bool(BACKUP_ACCESS_CODE and hmac.compare_digest(provided, BACKUP_ACCESS_CODE))


def read_supabase_backup(tables: list[str] | None = None) -> tuple[int, dict]:
    backup_data = {}
    for table in tables or BACKUP_TABLES:
        status, body = supabase_get_all(f"{table}?select=*&order=id.asc")
        if status >= 400:
            return status, {"error": body, "table": table}
        backup_data[table] = body if isinstance(body, list) else []
    return 200, backup_data


def read_database_storage_usage(force: bool = False) -> tuple[int, dict]:
    now = time.time()
    with storage_usage_lock:
        cached = storage_usage_cache.get("data")
        if not force and cached and now < float(storage_usage_cache.get("expires_at") or 0):
            return 200, cached

    status, body = supabase_request("POST", "rpc/get_database_storage_usage", {})
    if status >= 400:
        return status, {"error": body}
    raw = body[0] if isinstance(body, list) and body else body
    if not isinstance(raw, dict):
        return 502, {"error": "Storage usage response is invalid."}
    used_bytes = max(0, int(raw.get("used_bytes") or 0))
    limit_bytes = max(1, int(raw.get("limit_bytes") or SUPABASE_FREE_DATABASE_BYTES))
    percent = min(100, (used_bytes / limit_bytes) * 100)
    data = {
        **raw,
        "used_bytes": used_bytes,
        "limit_bytes": limit_bytes,
        "remaining_bytes": max(0, limit_bytes - used_bytes),
        "percent": round(percent, 2),
        "warning_percent": DATABASE_STORAGE_WARNING_PERCENT,
        "warning": percent >= DATABASE_STORAGE_WARNING_PERCENT,
    }
    with storage_usage_lock:
        storage_usage_cache["data"] = data
        storage_usage_cache["expires_at"] = now + STORAGE_USAGE_CACHE_SECONDS
    return 200, data


def backup_archive_bytes(payload: dict) -> bytes:
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2).encode("utf-8")


def backup_archive_checksum(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def supabase_storage_request(
    method: str,
    object_path: str,
    content: bytes | None = None,
    timeout_seconds: float = 30,
) -> tuple[int, bytes | dict | str | None]:
    if not supabase_configured():
        return 503, {"error": "Supabase environment variables are not configured."}
    encoded_path = quote(object_path.strip("/"), safe="/")
    url = f"{SUPABASE_URL}/storage/v1/object/{BACKUP_ARCHIVE_BUCKET}/{encoded_path}"
    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
    }
    if content is not None:
        headers["Content-Type"] = "application/json"
        headers["x-upsert"] = "false"
    request = urllib.request.Request(url, data=content, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            raw = response.read()
            if method == "GET":
                return response.status, raw
            if not raw:
                return response.status, None
            try:
                return response.status, json.loads(raw.decode("utf-8"))
            except (UnicodeDecodeError, json.JSONDecodeError):
                return response.status, raw.decode("utf-8", errors="replace")
    except urllib.error.HTTPError as error:
        raw = error.read().decode("utf-8", errors="replace")
        try:
            return error.code, json.loads(raw)
        except json.JSONDecodeError:
            return error.code, {"error": raw or str(error.reason)}
    except Exception as error:
        return 500, {"error": str(error)}


def backup_snapshot_payload(scope: str, actor: str, data: dict) -> dict:
    exported_at = datetime.now(timezone.utc).isoformat()
    row_counts = {
        table: len(rows)
        for table, rows in data.items()
        if isinstance(rows, list)
    }
    snapshot = {
        "exported_at": exported_at,
        "cutoff_at": exported_at,
        "app": "Pismai Factory Wage",
        "version": 3,
        "source": "supabase",
        "scope": scope,
        "created_by": actor,
        "row_counts": row_counts,
        "total_rows": sum(row_counts.values()),
        "data": data,
    }
    logical_bytes = len(json.dumps(snapshot, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
    percent = min(100, (logical_bytes / SUPABASE_FREE_DATABASE_BYTES) * 100)
    snapshot["storage_usage"] = {
        "used_bytes": logical_bytes,
        "limit_bytes": SUPABASE_FREE_DATABASE_BYTES,
        "remaining_bytes": max(0, SUPABASE_FREE_DATABASE_BYTES - logical_bytes),
        "percent": round(percent, 2),
        "warning_percent": DATABASE_STORAGE_WARNING_PERCENT,
        "warning": percent >= DATABASE_STORAGE_WARNING_PERCENT,
        "measurement": "logical_backup_size",
    }
    return snapshot


def backup_snapshot_ids(data: dict, table: str) -> list[int]:
    return sorted({
        int(row.get("id"))
        for row in data.get(table, [])
        if isinstance(row, dict) and str(row.get("id") or "").isdigit() and int(row.get("id")) > 0
    })


def delete_backup_snapshot_rows(data: dict, scope: str) -> tuple[bool, dict, dict | None]:
    cleared: dict[str, int] = {}
    allowed_queue_ids = {
        int(row.get("id"))
        for row in data.get("production_save_queue", [])
        if isinstance(row, dict)
        and str(row.get("id") or "").isdigit()
        and str(row.get("status") or "") in {"succeeded", "cancelled"}
    }
    tables = QUEUE_BACKUP_TABLES if scope == "queue" else MAIN_CLEAR_TABLES
    for table in tables:
        rows = data.get(table, [])
        if table == "production_save_queue":
            ids = sorted(allowed_queue_ids)
        elif table == "production_save_queue_events":
            ids = sorted({
                int(row.get("id"))
                for row in rows
                if isinstance(row, dict)
                and str(row.get("id") or "").isdigit()
                and int(row.get("queue_id") or 0) in allowed_queue_ids
            })
        else:
            ids = backup_snapshot_ids(data, table)
        cleared[table] = 0
        for offset in range(0, len(ids), 100):
            chunk = ids[offset:offset + 100]
            status, deleted = supabase_request(
                "DELETE",
                f"{table}?id=in.({','.join(str(value) for value in chunk)})",
                prefer="return=representation",
                timeout_seconds=30,
            )
            if status >= 400:
                return False, cleared, {"table": table, "error": deleted, "status": status}
            deleted_count = len(deleted) if isinstance(deleted, list) else 0
            cleared[table] += deleted_count
            if deleted_count != len(chunk):
                return False, cleared, {
                    "table": table,
                    "error": "Delete count did not match the archived snapshot.",
                    "expected": len(chunk),
                    "deleted": deleted_count,
                }
    return True, cleared, None


def restore_supabase_backup(data: dict) -> tuple[int, dict]:
    restored = {}
    for table in BACKUP_TABLES:
        rows = data.get(table)
        if not isinstance(rows, list):
            continue
        if not rows:
            restored[table] = 0
            continue
        status, body = supabase_request(
            "POST",
            f"{table}?on_conflict=id",
            rows,
            prefer="resolution=merge-duplicates,return=minimal",
        )
        if status >= 400:
            return status, {"error": body, "table": table}
        restored[table] = len(rows)
    return 200, {"restored": restored}


def find_employee(data: dict, employee_id: int) -> dict | None:
    return next(
        (employee for employee in data.get("employees", []) if employee.get("id") == employee_id),
        None,
    )


def employee_records(data: dict, date: str, employee_id: int) -> list[dict]:
    return [
        record
        for record in data.get("production_records", [])
        if (record.get("record_date") or record.get("date")) == date
        and record.get("employee_id") == employee_id
    ]


def employee_range_records(
    data: dict,
    start_date: str,
    end_date: str,
    employee_id: int,
    fruit_type: str = "all",
) -> list[dict]:
    start, end = sorted([start_date, end_date])
    selected_fruit = str(fruit_type or "all").strip().lower()
    return sorted(
        [
            record
            for record in data.get("production_records", [])
            if record.get("employee_id") == employee_id
            and start <= (record.get("record_date") or record.get("date") or "") <= end
            and (selected_fruit == "all" or (record.get("fruit_type") or "mangosteen") == selected_fruit)
        ],
        key=lambda record: (
            record.get("record_date") or record.get("date") or "",
            record.get("record_time", ""),
        ),
    )


def employee_daily_summaries(records: list[dict]) -> list[dict]:
    summaries: dict[str, dict] = {}
    for record in records:
        record_date = record.get("record_date") or record.get("date") or ""
        water_weight = float(record.get("water_weight", 0) or 0)
        flower_weight = float(record.get("flower_weight", 0) or 0)
        total_amount = float(record.get("total_amount", 0) or 0)
        summary = summaries.setdefault(
            record_date,
            {
                "date": record_date,
                "water_weight": 0.0,
                "flower_weight": 0.0,
                "grade_weights": {grade: 0.0 for grade in DURIAN_GRADES},
                "total_weight": 0.0,
                "total_amount": 0.0,
            },
        )
        summary["water_weight"] += water_weight
        summary["flower_weight"] += flower_weight
        grade_weights = production_grade_weights(record)
        for grade in DURIAN_GRADES:
            summary["grade_weights"][grade] += grade_weights[grade]
        summary["total_weight"] += production_total_weight(record)
        summary["total_amount"] += total_amount

    return [summaries[date] for date in sorted(summaries)]


def daily_records(data: dict, date: str) -> list[dict]:
    return [
        record
        for record in data.get("production_records", [])
        if (record.get("record_date") or record.get("date")) == date
    ]


def range_records(
    data: dict,
    start_date: str,
    end_date: str,
    product: str = "all",
) -> list[dict]:
    start, end = sorted([start_date, end_date])
    product_key = (product or "all").strip()
    records = []
    for record in data.get("production_records", []):
        record_date = record.get("record_date") or record.get("date") or ""
        if not (start <= record_date <= end):
            continue
        if product_key != "all" and (record.get("fruit_type") or "mangosteen") != product_key:
            continue
        records.append(record)
    return records


def pile_summary_rows(records: list[dict]) -> list[dict]:
    summaries: dict[object, dict] = {}
    for record in records:
        pile = production_pile_number(record)
        pile = pile if pile is not None else "-"
        is_durian = (record.get("fruit_type") or "mangosteen") == "durian"
        incoming = 0.0 if is_durian else float(record.get("water_weight", record.get("water", 0)) or 0)
        outgoing = 0.0 if is_durian else float(record.get("flower_weight", record.get("flower", 0)) or 0)
        amount = float(record.get("total_amount", record.get("grand_total", 0)) or 0)
        summary = summaries.setdefault(
            pile,
            {
                "pile": pile,
                "incoming": 0.0,
                "outgoing": 0.0,
                "balance": 0.0,
                "grades": {grade: 0.0 for grade in DURIAN_GRADES},
                "total_weight": 0.0,
                "amount": 0.0,
            },
        )
        summary["incoming"] += incoming
        summary["outgoing"] += outgoing
        summary["balance"] += incoming - outgoing
        grade_weights = production_grade_weights(record)
        for grade in DURIAN_GRADES:
            summary["grades"][grade] += grade_weights[grade]
        summary["total_weight"] += production_total_weight(record)
        summary["amount"] += amount

    def sort_key(item: dict) -> tuple[int, str]:
        pile = str(item["pile"])
        return (0, f"{int(pile):08d}") if pile.isdigit() else (1, pile)

    return sorted(summaries.values(), key=sort_key)


def summary_totals(rows: list[dict]) -> dict:
    return {
        "incoming": sum(row["incoming"] for row in rows),
        "outgoing": sum(row["outgoing"] for row in rows),
        "balance": sum(row["balance"] for row in rows),
        "grades": {grade: sum(row.get("grades", {}).get(grade, 0) for row in rows) for grade in DURIAN_GRADES},
        "total_weight": sum(row.get("total_weight", row["incoming"] + row["outgoing"]) for row in rows),
        "amount": sum(row["amount"] for row in rows),
    }


def format_report_date(value: str) -> str:
    try:
        return datetime.fromisoformat(value).strftime("%d/%m/%Y")
    except ValueError:
        return value or "-"


def format_report_datetime(value: datetime) -> tuple[str, str]:
    return value.strftime("%d/%m/%Y"), value.strftime("%H:%M:%S")


def clean_filename_date(value: str) -> str:
    return value or datetime.now().date().isoformat()


def money(value: float | int | str | None) -> str:
    try:
        return f"{float(value):,.2f}"
    except (TypeError, ValueError):
        return "0.00"


def number(value: float | int | str | None) -> str:
    try:
        return f"{float(value):,.2f}"
    except (TypeError, ValueError):
        return "0.00"


def build_employee_story(data: dict, date: str, employee: dict) -> list:
    records = employee_records(data, date, int(employee["id"]))
    total_water = sum(float(record.get("water_weight", 0)) for record in records)
    total_flower = sum(float(record.get("flower_weight", 0)) for record in records)
    total_amount = sum(float(record.get("total_amount", 0)) for record in records)
    recorder = records[-1].get("created_by", "-") if records else "-"

    title, normal, _, section = pdf_styles()
    payload = {
        "printed_by": "ระบบรายงาน",
        "printed_by_position": "ฝ่ายทรัพยากรบุคคล",
        "print_date": format_report_date(date),
        "print_time": "-",
    }
    story = report_header_story(
        "รายงานผลผลิตและค่าแรงประจำวัน",
        f"ประจำวันที่ {format_report_date(date)}",
        payload,
    )
    story[-1] = Spacer(1, 3 * mm)

    employee_info = Table(
        [[
            Paragraph(f"<b>รหัสพนักงาน</b><br/>{employee.get('emp_code', '-')}", normal),
            Paragraph(f"<b>ชื่อ-นามสกุล</b><br/>{employee.get('fullname', '-')}", normal),
            Paragraph(f"<b>จำนวนรายการ</b><br/>{len(records):,} รายการ", normal),
            Paragraph(f"<b>น้ำหนักรวม</b><br/>{number(total_water + total_flower)} กก.", normal),
            Paragraph(f"<b>ยอดเงินรวม</b><br/>{money(total_amount)} บาท", normal),
        ]],
        colWidths=[48 * mm, 72 * mm, 42 * mm, 50 * mm, 55 * mm],
    )
    employee_info.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F1F8F4")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(BRAND_GREEN)),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CFE3D6")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([employee_info, Spacer(1, 5 * mm), Paragraph("รายละเอียดผลผลิต", section)])

    rows = [
        [
            "กองที่",
            "นน. น้ำ",
            "นน. ดอก",
            "เรทน้ำ",
            "เรทดอก",
            "เงินค่าน้ำ",
            "เงินค่าดอก",
            "นน. รวม",
            "เงินรวม",
        ]
    ]

    for record in sorted(records, key=lambda item: (item.get("pile_no", 0), item.get("record_time", ""))):
        water_weight = float(record.get("water_weight", 0))
        flower_weight = float(record.get("flower_weight", 0))
        rows.append(
            [
                record.get("pile_no", ""),
                number(water_weight),
                number(flower_weight),
                money(record.get("water_rate")),
                money(record.get("flower_rate")),
                money(record.get("water_amount")),
                money(record.get("flower_amount")),
                number(water_weight + flower_weight),
                money(record.get("total_amount")),
            ]
        )

    if len(rows) == 1:
        rows.append(["-", "0.00", "0.00", "-", "-", "0.00", "0.00", "0.00", "0.00"])

    rows.append(
        [
            "รวมทั้งสิ้น",
            number(total_water),
            number(total_flower),
            "",
            "",
            "",
            "",
            number(total_water + total_flower),
            money(total_amount),
        ]
    )

    table = Table(
        rows,
        repeatRows=1,
        colWidths=[20 * mm, 29 * mm, 29 * mm, 25 * mm, 25 * mm, 32 * mm, 32 * mm, 30 * mm, 35 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_GREEN)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), THAI_FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, -1), THAI_FONT),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CFE3D6")),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, -1), (-1, -1), THAI_FONT_BOLD),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#DDEFE4")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F7FAF8")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    story.extend(
        [
            table,
            Spacer(1, 6 * mm),
            Paragraph(f"ผู้บันทึกข้อมูล: {recorder}", normal),
            Spacer(1, 2 * mm),
            Table(
                [
                    [
                        Paragraph("ลงชื่อ ..........................................................<br/>ผู้ตรวจสอบ", normal),
                        Paragraph("ลงชื่อ ..........................................................<br/>พนักงาน", normal),
                    ]
                ],
                colWidths=[130 * mm, 130 * mm],
                style=TableStyle([
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]),
            ),
        ]
    )
    return story


def build_pdf(data: dict, date: str, employee_ids: list[int]) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=12 * mm,
    )
    story = []

    for index, employee_id in enumerate(employee_ids):
        employee = find_employee(data, employee_id)
        if not employee:
            continue
        if index:
            story.append(PageBreak())
        story.extend(build_employee_story(data, date, employee))

    if not story:
        styles = getSampleStyleSheet()
        story = [
            Paragraph(COMPANY_NAME, styles["Title"]),
            Paragraph("No employee report data found.", styles["Heading2"]),
        ]

    doc.build(story)
    return buffer.getvalue()


def _build_employee_range_pdf_legacy(
    data: dict,
    start_date: str,
    end_date: str,
    employee_id: int,
) -> bytes:
    employee = find_employee(data, employee_id)
    records = employee_range_records(data, start_date, end_date, employee_id)
    daily_summaries = employee_daily_summaries(records)
    total_water = sum(summary["water_weight"] for summary in daily_summaries)
    total_flower = sum(summary["flower_weight"] for summary in daily_summaries)
    total_amount = sum(summary["total_amount"] for summary in daily_summaries)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    for style_name in ["Title", "Heading2", "Normal"]:
        styles[style_name].fontName = THAI_FONT
    story = [
        Paragraph(COMPANY_NAME, styles["Title"]),
        Paragraph("รายงานสรุปรายบุคคล", styles["Heading2"]),
        Spacer(1, 5 * mm),
        Paragraph(
            f"ชื่อ: {employee.get('emp_code', '-') if employee else '-'} - "
            f"{employee.get('fullname', '-') if employee else '-'}",
            styles["Normal"],
        ),
        Paragraph(f"ช่วงวันที่: {start_date} ถึง {end_date}", styles["Normal"]),
        Spacer(1, 5 * mm),
    ]

    rows = [
        [
            "วันที่",
            "น้ำหนักดอก",
            "น้ำหนักน้ำ",
            "รวมเป็นเงิน",
        ]
    ]

    for summary in daily_summaries:
        rows.append(
            [
                summary["date"],
                number(summary["flower_weight"]),
                number(summary["water_weight"]),
                money(summary["total_amount"]),
            ]
        )

    if len(rows) == 1:
        rows.append(["-", "0.00", "0.00", "0.00"])

    rows.append(
        [
            "รวม",
            number(total_flower),
            number(total_water),
            money(total_amount),
        ]
    )

    table = Table(rows, repeatRows=1, colWidths=[55 * mm, 55 * mm, 55 * mm, 65 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6F4F1")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("FONTNAME", (0, 0), (-1, -1), THAI_FONT),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F7FAFC")),
            ]
        )
    )

    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def _build_employee_range_pdf_report(
    data: dict,
    start_date: str,
    end_date: str,
    employee_id: int,
    fruit_type: str = "all",
) -> bytes:
    employee = find_employee(data, employee_id)
    records = employee_range_records(data, start_date, end_date, employee_id, fruit_type)
    daily_summaries = employee_daily_summaries(records)
    total_water = sum(item["water_weight"] for item in daily_summaries)
    total_flower = sum(item["flower_weight"] for item in daily_summaries)
    total_weight = sum(item.get("total_weight", item["water_weight"] + item["flower_weight"]) for item in daily_summaries)
    total_amount = sum(item["total_amount"] for item in daily_summaries)
    deduction_rows = deduction_records_for(
        data,
        "production",
        employee_id,
        (employee or {}).get("emp_code"),
        start_date,
        end_date,
    )
    bonus_rows = bonus_records_for(
        data,
        "production",
        employee_id,
        (employee or {}).get("emp_code"),
        start_date,
        end_date,
    )
    deduction_amount = deduction_total(deduction_rows)
    bonus_amount = deduction_total(bonus_rows)
    net_amount = max(0, total_amount + bonus_amount - deduction_amount)
    _, _, normal, section = pdf_styles()
    payload = {
        "printed_by": "ระบบรายงาน",
        "printed_by_position": "ฝ่ายทรัพยากรบุคคล",
        "print_date": format_report_date(datetime.now().strftime("%Y-%m-%d")),
        "print_time": datetime.now().strftime("%H:%M"),
    }

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=10 * mm,
        bottomMargin=13 * mm,
    )
    story = report_header_story(
        "รายงานสรุปผลผลิตรายบุคคล",
        f"ผลไม้ {selected_production_fruit_label({'fruit_type': fruit_type})} | ช่วงวันที่ {format_report_date(start_date)} ถึง {format_report_date(end_date)}",
        payload,
    )
    story[-1] = Spacer(1, 3 * mm)

    info = Table(
        [[
            Paragraph(f"<b>รหัสพนักงาน</b><br/>{employee.get('emp_code', '-') if employee else '-'}", normal),
            Paragraph(f"<b>ชื่อ-นามสกุล</b><br/>{employee.get('fullname', '-') if employee else '-'}", normal),
            Paragraph(f"<b>จำนวนวันทำงาน</b><br/>{len(daily_summaries):,} วัน", normal),
            Paragraph(f"<b>น้ำหนักรวม</b><br/>{number(total_weight)} กก.", normal),
            Paragraph(f"<b>รายได้รวม</b><br/>{money(total_amount)} บาท", normal),
        ]],
        colWidths=[48 * mm, 72 * mm, 42 * mm, 50 * mm, 55 * mm],
    )
    info.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F1F8F4")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(BRAND_GREEN)),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CFE3D6")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.extend([info, Spacer(1, 5 * mm)])
    money_summary = Table(
        [[
            Paragraph(f"<b>รวมก่อนปรับยอด</b><br/>{money(total_amount)} บาท", normal),
            Paragraph(f"<b>เบี้ยขยัน</b><br/>{money(bonus_amount)} บาท", normal),
            Paragraph(f"<b>หัก</b><br/>{money(deduction_amount)} บาท", normal),
            Paragraph(f"<b>สุทธิ</b><br/>{money(net_amount)} บาท", normal),
        ]],
        colWidths=[52 * mm, 52 * mm, 52 * mm, 52 * mm],
    )
    money_summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7ED")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#FDBA74")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#FED7AA")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.extend([money_summary, Spacer(1, 5 * mm)])
    adjustment_rows = [
        ["เงินเพิ่ม", row.get("deduction_label", "เบี้ยขยัน"), money(row.get("amount", 0)), row.get("note") or "-"]
        for row in bonus_rows
    ] + [
        ["รายการหัก", row.get("deduction_label", "-"), money(row.get("amount", 0)), row.get("note") or "-"]
        for row in deduction_rows
    ]
    if adjustment_rows:
        adjustment_table = Table(
            [["ประเภท", "รายการ", "จำนวนเงิน", "หมายเหตุ"]] + adjustment_rows,
            repeatRows=1,
            colWidths=[38 * mm, 64 * mm, 38 * mm, 65 * mm],
        )
        set_pdf_table_style(adjustment_table, 1)
        story.extend([Paragraph("รายการปรับยอด", section), adjustment_table, Spacer(1, 4 * mm)])
    story.extend([Paragraph("สรุปผลผลิตรายวัน", section)])

    total_grades = {grade: sum(item.get("grade_weights", {}).get(grade, 0) for item in daily_summaries) for grade in DURIAN_GRADES}
    is_durian = str(fruit_type or "all").lower() == "durian"
    water_label, flower_label = production_report_weight_labels({"fruit_type": fruit_type})
    if is_durian:
        rows = [["วันที่", "น้ำหนักทุเรียน (กก.)", "รายได้รวม (บาท)"]]
        for item in daily_summaries:
            rows.append([
                format_report_date(item["date"]),
                number(item.get("total_weight", 0)),
                money(item["total_amount"]),
            ])
        if len(rows) == 1:
            rows.append(["-", "0.00", "0.00"])
        rows.append(["รวมทั้งสิ้น", number(total_weight), money(total_amount)])
        column_widths = [60 * mm, 80 * mm, 80 * mm]
    else:
        rows = [["วันที่", f"{water_label} (กก.)", f"{flower_label} (กก.)", "น้ำหนักรวม (กก.)", "รายได้รวม (บาท)"]]
        for item in daily_summaries:
            rows.append([
                format_report_date(item["date"]),
                number(item["water_weight"]),
                number(item["flower_weight"]),
                number(item.get("total_weight", item["water_weight"] + item["flower_weight"])),
                money(item["total_amount"]),
            ])
        if len(rows) == 1:
            rows.append(["-", "0.00", "0.00", "0.00", "0.00"])
        rows.append(["รวมทั้งสิ้น", number(total_water), number(total_flower), number(total_weight), money(total_amount)])
        column_widths = [45 * mm, 52 * mm, 52 * mm, 52 * mm, 66 * mm]

    adjustment_blanks = ["" for _ in range(len(rows[0]) - 2)]
    if bonus_amount:
        rows.append(["เบี้ยขยัน", *adjustment_blanks, money(bonus_amount)])
    if deduction_amount:
        rows.append(["หัก", *adjustment_blanks, money(deduction_amount)])
    if bonus_amount or deduction_amount:
        rows.append(["สุทธิ", *adjustment_blanks, money(net_amount)])

    table = Table(rows, repeatRows=1, colWidths=column_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_GREEN)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), THAI_FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), THAI_FONT),
        ("FONTNAME", (0, -1), (-1, -1), THAI_FONT_BOLD),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CFE3D6")),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F7FAF8")]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#DDEFE4")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.extend([
        table,
        Spacer(1, 8 * mm),
        Table(
            [[
                Paragraph("ลงชื่อ ..........................................................<br/>ผู้จัดทำรายงาน", normal),
                Paragraph("ลงชื่อ ..........................................................<br/>ผู้ตรวจสอบ", normal),
            ]],
            colWidths=[130 * mm, 130 * mm],
            style=TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")]),
        ),
    ])

    def draw_footer(canvas, document):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D6E5DB"))
        canvas.line(15 * mm, 10 * mm, 282 * mm, 10 * mm)
        canvas.setFont(THAI_FONT, 8)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawString(15 * mm, 5.5 * mm, "รายงานสรุปผลผลิตรายบุคคล - เอกสารภายในบริษัท")
        canvas.drawRightString(282 * mm, 5.5 * mm, f"หน้า {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    return buffer.getvalue()


def build_employee_range_pdf(
    data: dict,
    start_date: str,
    end_date: str,
    employee_id: int,
    fruit_type: str = "all",
) -> bytes:
    employee = find_employee(data, employee_id) or {}
    records = employee_range_records(data, start_date, end_date, employee_id, fruit_type)
    daily_summaries = employee_daily_summaries(records)
    record_counts: dict[str, int] = {}
    for record in records:
        record_date = str(record.get("record_date") or record.get("date") or "")
        record_counts[record_date] = record_counts.get(record_date, 0) + 1

    total_water = sum(item["water_weight"] for item in daily_summaries)
    total_flower = sum(item["flower_weight"] for item in daily_summaries)
    total_weight = sum(
        item.get("total_weight", item["water_weight"] + item["flower_weight"])
        for item in daily_summaries
    )
    gross_amount = sum(item["total_amount"] for item in daily_summaries)
    deduction_rows = deduction_records_for(
        data, "production", employee_id, employee.get("emp_code"), start_date, end_date
    )
    bonus_rows = bonus_records_for(
        data, "production", employee_id, employee.get("emp_code"), start_date, end_date
    )
    deduction_amount = deduction_total(deduction_rows)
    bonus_amount = deduction_total(bonus_rows)
    net_amount = max(0, gross_amount + bonus_amount - deduction_amount)
    fruit_label = selected_production_fruit_label({"fruit_type": fruit_type})
    water_label, flower_label = production_report_weight_labels({"fruit_type": fruit_type})
    is_durian = str(fruit_type or "all").lower() == "durian"
    printed_by = str(data.get("printed_by") or "ระบบรายงาน").strip()
    printed_by_position = str(data.get("printed_by_position") or "ฝ่ายทรัพยากรบุคคล").strip()
    printed_by_text = " ".join(value for value in [printed_by, printed_by_position] if value)
    now = datetime.now()

    buffer = BytesIO()
    page_width, page_height = landscape(A4)
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
    margin = 8 * mm
    gap = 7 * mm
    divider_x = page_width / 2
    panel_width = (page_width - (margin * 2) - gap) / 2
    panel_height = page_height - (margin * 2)

    def fit_text(value: object, max_width: float, size: float, bold: bool = False) -> tuple[str, float]:
        font_name = THAI_FONT_BOLD if bold else THAI_FONT
        text_value = str(value or "-")
        fitted_size = size
        while fitted_size > 5.5 and pdfmetrics.stringWidth(text_value, font_name, fitted_size) > max_width:
            fitted_size -= 0.25
        if pdfmetrics.stringWidth(text_value, font_name, fitted_size) <= max_width:
            return text_value, fitted_size
        while text_value and pdfmetrics.stringWidth(f"{text_value}...", font_name, fitted_size) > max_width:
            text_value = text_value[:-1]
        return f"{text_value}...", fitted_size

    def text(x: float, y: float, value: object, size=8, bold=False, fill="#111827", max_width=None):
        shown, shown_size = fit_text(value, max_width, size, bold) if max_width else (str(value), size)
        pdf.setFillColor(colors.HexColor(fill))
        pdf.setFont(THAI_FONT_BOLD if bold else THAI_FONT, shown_size)
        pdf.drawString(x, y, shown)

    def right_text(x: float, y: float, value: object, size=8, bold=False, fill="#111827", max_width=None):
        shown, shown_size = fit_text(value, max_width, size, bold) if max_width else (str(value), size)
        pdf.setFillColor(colors.HexColor(fill))
        pdf.setFont(THAI_FONT_BOLD if bold else THAI_FONT, shown_size)
        pdf.drawRightString(x, y, shown)

    def draw_panel(x: float, y: float, is_copy: bool = False):
        pdf.setStrokeColor(colors.HexColor("#0F7A3D"))
        pdf.setLineWidth(1)
        pdf.roundRect(x, y, panel_width, panel_height, 6, stroke=1, fill=0)
        if is_copy:
            pdf.saveState()
            pdf.translate(x + panel_width / 2, y + panel_height / 2)
            pdf.rotate(30)
            pdf.setFillColor(colors.Color(0.06, 0.48, 0.24, alpha=0.08))
            pdf.setFont(THAI_FONT_BOLD, 58)
            pdf.drawCentredString(0, 0, "สำเนา")
            pdf.restoreState()

        cursor_y = y + panel_height - 12 * mm
        if logo_path.exists():
            try:
                pdf.drawImage(
                    str(logo_path),
                    x + 5 * mm,
                    cursor_y - 3 * mm,
                    width=17 * mm,
                    height=17 * mm,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass
        text(x + 25 * mm, cursor_y + 6 * mm, COMPANY_NAME, 11, True, "#0F7A3D", panel_width - 31 * mm)
        text(x + 25 * mm, cursor_y, "ใบสรุปผลผลิตและค่าแรง", 10, True, "#0F7A3D")
        text(x + 25 * mm, cursor_y - 5 * mm, "ต้นฉบับ" if not is_copy else "สำเนา", 8, True, "#667085")
        pdf.setStrokeColor(colors.HexColor("#B8D8D1"))
        pdf.line(x + 5 * mm, cursor_y - 9 * mm, x + panel_width - 5 * mm, cursor_y - 9 * mm)

        meta_y = cursor_y - 17 * mm
        receipt_number = f"PR-{start_date.replace('-', '')}-{employee.get('emp_code', '-')}"
        text(x + 5 * mm, meta_y, f"เลขที่: {receipt_number}", 7.3, True, max_width=58 * mm)
        text(x + 5 * mm, meta_y - 5 * mm, f"ช่วงวันที่: {format_report_date(start_date)} - {format_report_date(end_date)}", 7.3)
        right_text(x + panel_width - 5 * mm, meta_y, f"ออกโดย: {printed_by_text}", 7.2, max_width=63 * mm)
        right_text(x + panel_width - 5 * mm, meta_y - 5 * mm, now.strftime("%d/%m/%Y %H:%M"), 7.2)

        employee_y = meta_y - 15 * mm
        text(x + 5 * mm, employee_y, f"รหัสพนักงาน: {employee.get('emp_code', '-')}", 8, True, max_width=40 * mm)
        text(x + 48 * mm, employee_y, f"ชื่อ: {employee.get('fullname', '-')}", 8, True, max_width=panel_width - 55 * mm)
        text(x + 5 * mm, employee_y - 5 * mm, f"ผลไม้: {fruit_label}", 7.5)
        right_text(
            x + panel_width - 5 * mm,
            employee_y - 5 * mm,
            f"{len(daily_summaries)} วัน / {len(records)} รายการ",
            7.5,
            True,
        )

        table_y = employee_y - 13 * mm
        if is_durian:
            headers = ["วันที่", "รายการ", "น้ำหนักทุเรียน", "รวมเงิน"]
            col_widths = [29, 21, 39, 34]
        else:
            headers = ["วันที่", "รายการ", water_label, flower_label, "รวม", "รวมเงิน"]
            col_widths = [24, 16, 22, 22, 21, 25]
        inner_width = panel_width - 10 * mm
        scale = inner_width / (sum(col_widths) * mm)
        col_widths = [width * mm * scale for width in col_widths]
        pdf.setFillColor(colors.HexColor("#0F7A3D"))
        pdf.rect(x + 5 * mm, table_y, inner_width, 7 * mm, stroke=0, fill=1)
        col_x = x + 6 * mm
        for header, width in zip(headers, col_widths):
            text(col_x, table_y + 2.2 * mm, header, 6.2, True, "#FFFFFF", width - 2 * mm)
            col_x += width

        row_y = table_y - 6 * mm
        visible_rows = daily_summaries[:10]
        for row_index, item in enumerate(visible_rows):
            if row_index % 2:
                pdf.setFillColor(colors.HexColor("#F3FBF9"))
                pdf.rect(x + 5 * mm, row_y - 1 * mm, inner_width, 5.5 * mm, stroke=0, fill=1)
            pdf.setStrokeColor(colors.HexColor("#D8E5E1"))
            pdf.line(x + 5 * mm, row_y - 1 * mm, x + panel_width - 5 * mm, row_y - 1 * mm)
            if is_durian:
                values = [
                    format_report_date(item["date"]),
                    record_counts.get(item["date"], 0),
                    report_number(item.get("total_weight", 0)),
                    report_number(item["total_amount"], 0),
                ]
            else:
                values = [
                    format_report_date(item["date"]),
                    record_counts.get(item["date"], 0),
                    report_number(item["water_weight"]),
                    report_number(item["flower_weight"]),
                    report_number(item.get("total_weight", 0)),
                    report_number(item["total_amount"], 0),
                ]
            col_x = x + 6 * mm
            for value, width in zip(values, col_widths):
                text(col_x, row_y + 1 * mm, value, 6.2, max_width=width - 2 * mm)
                col_x += width
            row_y -= 5.5 * mm
        if len(daily_summaries) > len(visible_rows):
            text(
                x + 6 * mm,
                row_y + 1 * mm,
                f"มีเพิ่มอีก {len(daily_summaries) - len(visible_rows)} วัน รวมอยู่ในยอดด้านล่าง",
                6.2,
                False,
                "#667085",
            )

        adjustment_y = y + 66 * mm
        adjustments = [
            *[("เงินเพิ่ม", row.get("deduction_label", "เบี้ยขยัน"), row.get("amount", 0)) for row in bonus_rows],
            *[("หัก", row.get("deduction_label", "-"), row.get("amount", 0)) for row in deduction_rows],
        ]
        if adjustments:
            text(x + 6 * mm, adjustment_y, "รายการปรับยอด", 6.8, True, "#344054")
            for index, (kind, label, amount) in enumerate(adjustments[:3]):
                color = "#166534" if kind == "เงินเพิ่ม" else "#B42318"
                text(
                    x + 28 * mm,
                    adjustment_y - (index * 4.2 * mm),
                    f"{kind}: {label} {report_number(amount, 0)} บาท",
                    6.3,
                    False,
                    color,
                    panel_width - 36 * mm,
                )
            if len(adjustments) > 3:
                right_text(
                    x + panel_width - 6 * mm,
                    adjustment_y,
                    f"+{len(adjustments) - 3} รายการ",
                    6.2,
                    True,
                    "#667085",
                )

        summary_y = y + 35 * mm
        pdf.setFillColor(colors.HexColor("#E9F5EE"))
        pdf.roundRect(x + 5 * mm, summary_y, panel_width - 10 * mm, 19 * mm, 4, stroke=0, fill=1)
        text(x + 8 * mm, summary_y + 13 * mm, f"น้ำหนักรวม {report_number(total_weight)} กก.", 7.5, True, "#064E25")
        text(x + 8 * mm, summary_y + 7.5 * mm, f"รวมก่อนปรับ {report_number(gross_amount, 0)} บาท", 7.5, True, "#064E25")
        right_text(x + panel_width - 8 * mm, summary_y + 13 * mm, f"เบี้ยขยัน {report_number(bonus_amount, 0)} บาท", 7.2, True, "#166534")
        right_text(x + panel_width - 8 * mm, summary_y + 7.5 * mm, f"หัก {report_number(deduction_amount, 0)} บาท", 7.2, True, "#B42318")
        right_text(x + panel_width - 8 * mm, summary_y + 2 * mm, f"ยอดรับสุทธิ {report_number(net_amount, 0)} บาท", 10.5, True, "#064E25")

        sign_y = y + 12 * mm
        pdf.setStrokeColor(colors.HexColor("#111827"))
        pdf.line(x + 14 * mm, sign_y + 5 * mm, x + 54 * mm, sign_y + 5 * mm)
        pdf.line(x + panel_width - 54 * mm, sign_y + 5 * mm, x + panel_width - 14 * mm, sign_y + 5 * mm)
        text(x + 25 * mm, sign_y, "ผู้รับเงิน", 7)
        text(x + panel_width - 43 * mm, sign_y, "ผู้จ่ายเงิน", 7)

    draw_panel(margin, margin, False)
    pdf.setDash(3, 3)
    pdf.setStrokeColor(colors.HexColor("#667085"))
    pdf.line(divider_x, margin, divider_x, page_height - margin)
    pdf.setDash()
    draw_panel(divider_x + gap / 2, margin, True)
    pdf.showPage()
    pdf.save()
    return buffer.getvalue()


def _build_daily_excel_legacy(data: dict, date: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily Production"

    title = f"{COMPANY_NAME} - Daily Production Wage Report"
    sheet["A1"] = title
    sheet["A2"] = f"Date: {date}"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"].font = Font(bold=True)

    headers = [
        "ID",
        "Date",
        "Time",
        "Employee Code",
        "Employee Fullname",
        "Pile",
        "Water Weight",
        "Flower Weight",
        "Water Rate",
        "Flower Rate",
        "Water Amount",
        "Flower Amount",
        "Total Weight",
        "Total Amount",
        "Recorder",
        "Status",
    ]
    sheet.append([])
    sheet.append(headers)
    header_row = 4

    employee_by_id = {employee["id"]: employee for employee in data.get("employees", [])}

    for record in sorted(daily_records(data, date), key=lambda item: (item.get("emp_code", ""), item.get("record_time", ""))):
        employee = employee_by_id.get(record.get("employee_id"), {})
        water_weight = float(record.get("water_weight", 0))
        flower_weight = float(record.get("flower_weight", 0))
        sheet.append(
            [
                record.get("id"),
                record.get("record_date"),
                record.get("record_time"),
                record.get("emp_code"),
                employee.get("fullname", ""),
                record.get("pile_no"),
                water_weight,
                flower_weight,
                record.get("water_rate"),
                record.get("flower_rate"),
                record.get("water_amount"),
                record.get("flower_amount"),
                water_weight + flower_weight,
                record.get("total_amount"),
                record.get("created_by"),
                record.get("status"),
            ]
        )

    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E6F4F1")
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 28)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_daily_excel(data: dict, date: str) -> bytes:
    records = sorted(
        daily_records(data, date),
        key=lambda item: (str(item.get("emp_code", "")), str(item.get("record_time", ""))),
    )
    employees = {employee["id"]: employee for employee in data.get("employees", [])}
    total_weight = sum(production_total_weight(record) for record in records)
    total_amount = sum(safe_float(record.get("total_amount")) for record in records)
    employee_count = len({record.get("employee_id") for record in records if record.get("employee_id") is not None})

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "รายงานผลผลิตรายวัน"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A8"

    green = "0F7A3D"
    dark_green = "075E2E"
    light_green = "EAF5EE"
    pale_green = "F7FAF8"
    border_color = "CFE3D6"
    white = "FFFFFF"
    text_color = "1F2937"
    thin = Side(style="thin", color=border_color)
    medium = Side(style="medium", color=green)

    sheet.merge_cells("A1:O1")
    sheet["A1"] = COMPANY_NAME
    sheet["A1"].font = Font(name="Sarabun", bold=True, size=12, color=dark_green)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells("A2:O2")
    sheet["A2"] = "รายงานผลผลิตและค่าแรงประจำวัน"
    sheet["A2"].font = Font(name="Sarabun", bold=True, size=22, color=green)
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.merge_cells("A3:O3")
    sheet["A3"] = f"ประจำวันที่ {format_report_date(date)} | สร้างรายงาน {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sheet["A3"].font = Font(name="Sarabun", size=10, color="667085")

    logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
    if logo_path.exists():
        try:
            logo = ExcelImage(str(logo_path))
            logo.width = 82
            logo.height = 82
            sheet.add_image(logo, "O1")
        except Exception:
            pass

    cards = [
        ("A4:D5", "จำนวนพนักงาน", f"{employee_count:,} คน"),
        ("E4:H5", "จำนวนรายการ", f"{len(records):,} รายการ"),
        ("I4:L5", "น้ำหนักรวม", f"{total_weight:,.2f} กก."),
        ("M4:P5", "ยอดเงินรวม", f"{total_amount:,.2f} บาท"),
    ]
    for cell_range, label, value in cards:
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":")[0]]
        cell.value = f"{label}\n{value}"
        cell.font = Font(name="Sarabun", bold=True, size=12, color=green)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=light_green)
        cell.border = Border(left=medium, right=medium, top=medium, bottom=medium)

    headers = [
        "ลำดับ", "วันที่", "เวลา", "รหัสพนักงาน", "ชื่อ-นามสกุล", "กองที่",
        "นน. น้ำ", "นน. ดอก", "เรทน้ำ", "เรทดอก", "เงินค่าน้ำ", "เงินค่าดอก",
        "น้ำหนักทุเรียน", "นน. รวม", "เงินรวม", "ผู้บันทึก", "สถานะ",
    ]
    header_row = 7
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(name="Sarabun", bold=True, color=white, size=10)
        cell.fill = PatternFill("solid", fgColor=green)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for index, record in enumerate(records, 1):
        row = header_row + index
        employee = employees.get(record.get("employee_id"), {})
        water_weight = safe_float(record.get("water_weight"))
        flower_weight = safe_float(record.get("flower_weight"))
        raw_date = record.get("record_date") or date
        try:
            date_value = datetime.strptime(str(raw_date), "%Y-%m-%d").date()
        except ValueError:
            date_value = str(raw_date)
        values = [
            index,
            date_value,
            record.get("record_time", ""),
            record.get("emp_code") or employee.get("emp_code", ""),
            employee.get("fullname", ""),
            record.get("pile_no", ""),
            water_weight,
            flower_weight,
            safe_float(record.get("water_rate")),
            safe_float(record.get("flower_rate")),
            safe_float(record.get("water_amount")),
            safe_float(record.get("flower_amount")),
            production_grade_text(record) if (record.get("fruit_type") or "mangosteen") == "durian" else "-",
            production_total_weight(record),
            safe_float(record.get("total_amount")),
            record.get("created_by", ""),
            "อนุมัติแล้ว" if str(record.get("status", "")).lower() == "approved" else record.get("status", ""),
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row, column, value)
            cell.font = Font(name="Sarabun", size=10, color=text_color)
            cell.fill = PatternFill("solid", fgColor=white if index % 2 else pale_green)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if column in (4, 5, 13, 16, 17) else "center" if column in (1, 2, 3, 6) else "right",
                vertical="center",
            )
        sheet.cell(row, 2).number_format = "dd/mm/yyyy"
        for column in list(range(7, 13)) + [14, 15]:
            sheet.cell(row, column).number_format = "#,##0.00"
        if str(values[16]) == "อนุมัติแล้ว":
            sheet.cell(row, 17).font = Font(name="Sarabun", bold=True, size=10, color=green)

    total_row = header_row + len(records) + 1
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    sheet.cell(total_row, 1, "รวมทั้งสิ้น")
    sheet.cell(total_row, 7, sum(safe_float(record.get("water_weight")) for record in records))
    sheet.cell(total_row, 8, sum(safe_float(record.get("flower_weight")) for record in records))
    total_grades = {grade: sum(production_grade_weights(record)[grade] for record in records) for grade in DURIAN_GRADES}
    sheet.cell(total_row, 13, grade_totals_text(total_grades))
    sheet.cell(total_row, 14, total_weight)
    sheet.cell(total_row, 15, total_amount)
    for column in range(1, 18):
        cell = sheet.cell(total_row, column)
        cell.font = Font(name="Sarabun", bold=True, size=10, color=dark_green)
        cell.fill = PatternFill("solid", fgColor="DDEFE4")
        cell.border = Border(top=medium, bottom=medium)
        cell.alignment = Alignment(horizontal="right" if column >= 7 else "left", vertical="center")
        if column >= 7:
            cell.number_format = "#,##0.00"

    widths = [8, 13, 10, 15, 25, 9, 12, 12, 11, 11, 14, 14, 28, 13, 15, 20, 14]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.row_dimensions[1].height = 20
    sheet.row_dimensions[2].height = 31
    sheet.row_dimensions[3].height = 19
    sheet.row_dimensions[4].height = 24
    sheet.row_dimensions[5].height = 24
    sheet.row_dimensions[7].height = 28
    for row in range(8, total_row + 1):
        sheet.row_dimensions[row].height = 21

    sheet.auto_filter.ref = f"A7:Q{total_row - 1}"
    sheet.print_title_rows = "1:7"
    sheet.print_area = f"A1:Q{total_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "รายงานผลผลิตและค่าแรงประจำวัน"
    sheet.oddFooter.right.text = "หน้า &P จาก &N"
    sheet.oddFooter.left.text = SYSTEM_NAME
    sheet.sheet_view.zoomScale = 80

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_employee_range_excel(
    data: dict,
    start_date: str,
    end_date: str,
    employee_id: int,
    fruit_type: str = "all",
) -> bytes:
    employee = find_employee(data, employee_id) or {}
    records = employee_range_records(data, start_date, end_date, employee_id, fruit_type)
    daily_summaries = employee_daily_summaries(records)
    deduction_rows = deduction_records_for(
        data,
        "production",
        employee_id,
        employee.get("emp_code"),
        start_date,
        end_date,
    )
    bonus_rows = bonus_records_for(
        data,
        "production",
        employee_id,
        employee.get("emp_code"),
        start_date,
        end_date,
    )
    deduction_amount = deduction_total(deduction_rows)
    bonus_amount = deduction_total(bonus_rows)
    gross_amount = sum(summary["total_amount"] for summary in daily_summaries)
    net_amount = max(0, gross_amount + bonus_amount - deduction_amount)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "สรุปรายบุคคล"
    is_durian = str(fruit_type or "all").lower() == "durian"
    water_label, flower_label = production_report_weight_labels({"fruit_type": fruit_type})
    total_water = sum(summary["water_weight"] for summary in daily_summaries)
    total_flower = sum(summary["flower_weight"] for summary in daily_summaries)
    total_weight = sum(
        summary.get("total_weight", summary["flower_weight"] + summary["water_weight"])
        for summary in daily_summaries
    )
    font_name = "Sarabun"
    dark_green = "064E3B"
    brand_green = "0F8A55"
    pale_green = "D1FAE5"
    light_gray = "F8FAFC"
    border_color = "D9E2EA"
    orange = "C2410C"
    thin = Side(style="thin", color=border_color)
    medium = Side(style="medium", color=dark_green)
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:B3")
    sheet["A1"] = "PF"
    sheet["A1"].fill = PatternFill("solid", fgColor=brand_green)
    sheet["A1"].font = Font(name=font_name, bold=True, size=25, color="FFFFFF")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells("C1:J2")
    sheet["C1"] = COMPANY_NAME
    sheet["C1"].fill = PatternFill("solid", fgColor=dark_green)
    sheet["C1"].font = Font(name=font_name, bold=True, size=20, color="FFFFFF")
    sheet["C1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells("C3:J3")
    sheet["C3"] = "รายงานสรุปผลผลิตรายบุคคล • Employee Production Report"
    sheet["C3"].fill = PatternFill("solid", fgColor=dark_green)
    sheet["C3"].font = Font(name=font_name, size=10, color="D1FAE5")
    sheet["C3"].alignment = Alignment(horizontal="left", vertical="center")
    add_excel_logo(sheet, "I1")

    sheet.merge_cells("A5:J5")
    sheet["A5"] = "ข้อมูลพนักงานและช่วงรายงาน"
    sheet["A5"].fill = PatternFill("solid", fgColor=pale_green)
    sheet["A5"].font = Font(name=font_name, bold=True, size=12, color="065F46")
    sheet["A5"].alignment = Alignment(vertical="center")

    info_rows = [
        ("A6", "รหัสพนักงาน", "B6:C6", str(employee.get("emp_code") or "-")),
        ("D6", "ชื่อ-นามสกุล", "E6:G6", employee.get("fullname") or "-"),
        ("H6", "กลุ่ม", "I6:J6", employee.get("group_name") or employee.get("department") or "-"),
        ("A7", "ช่วงวันที่", "B7:C7", f"{start_date} ถึง {end_date}"),
        ("D7", "ประเภทผลไม้", "E7:G7", selected_production_fruit_label({"fruit_type": fruit_type})),
        ("H7", "จำนวนวันมีรายการ", "I7:J7", len(daily_summaries)),
    ]
    for label_cell, label, value_range, value in info_rows:
        sheet[label_cell] = label
        sheet[label_cell].font = Font(name=font_name, bold=True, size=10, color="64748B")
        sheet.merge_cells(value_range)
        value_cell = sheet[value_range.split(":")[0]]
        value_cell.value = value
        value_cell.font = Font(name=font_name, size=10, color="1E293B")
        value_cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet["B6"].number_format = "@"
    for row in sheet.iter_rows(min_row=6, max_row=7, min_col=1, max_col=10):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=light_gray)
            cell.border = table_border
            if cell.alignment is None:
                cell.alignment = Alignment(vertical="center")

    cards = [
        ("A9:B9", "A10:B11", "วันมีรายการ", len(daily_summaries), "ECFDF5", "047857", '0 "วัน"'),
        ("D9:E9", "D10:E11", "น้ำหนักรวม", total_weight, "EFF6FF", "1D4ED8", '#,##0.00 "กก."'),
        ("G9:H9", "G10:H11", "รายได้ก่อนหัก", gross_amount, "FFF7ED", orange, '#,##0.00 "บาท"'),
        ("I9:J9", "I10:J11", "รับสุทธิ", net_amount, "F0FDF4", "15803D", '#,##0.00 "บาท"'),
    ]
    for label_range, value_range, label, value, fill_color, font_color, number_format in cards:
        sheet.merge_cells(label_range)
        sheet.merge_cells(value_range)
        label_cell = sheet[label_range.split(":")[0]]
        value_cell = sheet[value_range.split(":")[0]]
        label_cell.value = label
        value_cell.value = value
        for row in sheet[label_range.split(":")[0]:value_range.split(":")[1]]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = table_border
        label_cell.font = Font(name=font_name, bold=True, size=10, color="64748B")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.font = Font(name=font_name, bold=True, size=17, color=font_color)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format

    sheet.merge_cells("A13:J13")
    sheet["A13"] = "รายละเอียดผลผลิตรายวัน"
    sheet["A13"].fill = PatternFill("solid", fgColor=dark_green)
    sheet["A13"].font = Font(name=font_name, bold=True, size=12, color="FFFFFF")
    sheet["A13"].alignment = Alignment(vertical="center")

    header_row = 14
    if is_durian:
        table_ranges = [("A14:B14", "วันที่"), ("C14:F14", "น้ำหนักทุเรียน"), ("G14:J14", "รวมเป็นเงิน")]
    else:
        table_ranges = [
            ("A14:B14", "วันที่"),
            ("C14:D14", water_label),
            ("E14:F14", flower_label),
            ("G14:H14", "น้ำหนักรวม"),
            ("I14:J14", "รวมเป็นเงิน"),
        ]
    for cell_range, label in table_ranges:
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":")[0]]
        cell.value = label
        cell.fill = PatternFill("solid", fgColor=brand_green)
        cell.font = Font(name=font_name, bold=True, size=10, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet[cell_range]:
            for part in row:
                part.fill = PatternFill("solid", fgColor=brand_green)
                part.border = table_border

    data_start = 15
    for offset, summary in enumerate(daily_summaries):
        row_index = data_start + offset
        row_ranges = (
            [f"A{row_index}:B{row_index}", f"C{row_index}:F{row_index}", f"G{row_index}:J{row_index}"]
            if is_durian
            else [
                f"A{row_index}:B{row_index}",
                f"C{row_index}:D{row_index}",
                f"E{row_index}:F{row_index}",
                f"G{row_index}:H{row_index}",
                f"I{row_index}:J{row_index}",
            ]
        )
        values = (
            [summary["date"], summary.get("total_weight", 0), summary["total_amount"]]
            if is_durian
            else [
                summary["date"],
                summary["water_weight"],
                summary["flower_weight"],
                summary.get("total_weight", summary["flower_weight"] + summary["water_weight"]),
                summary["total_amount"],
            ]
        )
        for cell_range, value in zip(row_ranges, values):
            sheet.merge_cells(cell_range)
            cell = sheet[cell_range.split(":")[0]]
            cell.value = value
            cell.font = Font(name=font_name, size=10, color="1E293B")
            cell.alignment = Alignment(
                horizontal="center" if cell_range.startswith("A") else "right",
                vertical="center",
            )
            if not cell_range.startswith("A"):
                cell.number_format = "#,##0.00"
            for row in sheet[cell_range]:
                for part in row:
                    part.border = table_border
                    if offset % 2:
                        part.fill = PatternFill("solid", fgColor=light_gray)

    total_row = data_start + len(daily_summaries)
    total_ranges = (
        [(f"A{total_row}:B{total_row}", "รวมทั้งสิ้น"), (f"C{total_row}:F{total_row}", total_weight), (f"G{total_row}:J{total_row}", gross_amount)]
        if is_durian
        else [
            (f"A{total_row}:B{total_row}", "รวมทั้งสิ้น"),
            (f"C{total_row}:D{total_row}", total_water),
            (f"E{total_row}:F{total_row}", total_flower),
            (f"G{total_row}:H{total_row}", total_weight),
            (f"I{total_row}:J{total_row}", gross_amount),
        ]
    )
    for cell_range, value in total_ranges:
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":")[0]]
        cell.value = value
        cell.font = Font(name=font_name, bold=True, size=10, color=dark_green)
        cell.alignment = Alignment(horizontal="right" if not cell_range.startswith("A") else "left", vertical="center")
        if not cell_range.startswith("A"):
            cell.number_format = "#,##0.00"
        for row in sheet[cell_range]:
            for part in row:
                part.fill = PatternFill("solid", fgColor="DDEFE4")
                part.border = Border(top=medium, bottom=medium)

    summary_start = total_row + 2
    sheet.merge_cells(start_row=summary_start, start_column=1, end_row=summary_start, end_column=10)
    sheet.cell(summary_start, 1, "สรุปการจ่ายเงิน")
    sheet.cell(summary_start, 1).fill = PatternFill("solid", fgColor=pale_green)
    sheet.cell(summary_start, 1).font = Font(name=font_name, bold=True, size=12, color="065F46")
    payment_rows = [
        ("รายได้ก่อนหัก", gross_amount, "FFF7ED", orange),
        ("เงินเพิ่ม / เบี้ยขยัน", bonus_amount, "F0FDF4", "15803D"),
        ("รายการหัก", deduction_amount, "FEF2F2", "B91C1C"),
        ("ยอดรับสุทธิ", net_amount, dark_green, "FFFFFF"),
    ]
    for offset, (label, amount, fill_color, font_color) in enumerate(payment_rows, 1):
        row_index = summary_start + offset
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
        sheet.merge_cells(start_row=row_index, start_column=9, end_row=row_index, end_column=10)
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 9, amount)
        for cell in sheet[row_index]:
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = table_border
            cell.font = Font(name=font_name, bold=offset == 4, size=11 if offset == 4 else 10, color=font_color)
            cell.alignment = Alignment(horizontal="right" if cell.column >= 9 else "left", vertical="center")
        sheet.cell(row_index, 9).number_format = '#,##0.00 "บาท"'

    detail_row = summary_start + len(payment_rows) + 2
    details = [
        ("รายการหัก", deduction_rows, "-"),
        ("รายการเงินเพิ่ม / เบี้ยขยัน", bonus_rows, "เบี้ยขยัน"),
    ]
    for title, rows, default_label in details:
        if not rows:
            continue
        sheet.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=10)
        sheet.cell(detail_row, 1, title)
        sheet.cell(detail_row, 1).fill = PatternFill("solid", fgColor=dark_green)
        sheet.cell(detail_row, 1).font = Font(name=font_name, bold=True, size=11, color="FFFFFF")
        detail_row += 1
        for cell_range, label in [
            (f"A{detail_row}:D{detail_row}", "รายการ"),
            (f"E{detail_row}:F{detail_row}", "จำนวนเงิน"),
            (f"G{detail_row}:J{detail_row}", "หมายเหตุ"),
        ]:
            sheet.merge_cells(cell_range)
            cell = sheet[cell_range.split(":")[0]]
            cell.value = label
            cell.fill = PatternFill("solid", fgColor=brand_green)
            cell.font = Font(name=font_name, bold=True, size=10, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for item in rows:
            detail_row += 1
            for cell_range, value in [
                (f"A{detail_row}:D{detail_row}", item.get("deduction_label", default_label)),
                (f"E{detail_row}:F{detail_row}", safe_float(item.get("amount"))),
                (f"G{detail_row}:J{detail_row}", item.get("note") or "-"),
            ]:
                sheet.merge_cells(cell_range)
                cell = sheet[cell_range.split(":")[0]]
                cell.value = value
                cell.font = Font(name=font_name, size=10, color="1E293B")
                cell.border = table_border
                cell.alignment = Alignment(horizontal="right" if cell_range.startswith("E") else "left", vertical="center", wrap_text=True)
                if cell_range.startswith("E"):
                    cell.number_format = '#,##0.00 "บาท"'
        detail_row += 2

    widths = [13, 7, 13, 7, 13, 7, 13, 7, 14, 9]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 22
    for row_index in [1, 2, 3]:
        sheet.row_dimensions[row_index].height = 23
    for row_index in [5, 13, header_row, summary_start]:
        sheet.row_dimensions[row_index].height = 25
    sheet.freeze_panes = "A15"
    sheet.print_title_rows = "1:14"
    sheet.print_area = f"A1:J{sheet.max_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.left.text = SYSTEM_NAME
    sheet.oddFooter.center.text = "รายงานสรุปผลผลิตรายบุคคล"
    sheet.oddFooter.right.text = "หน้า &P จาก &N"
    sheet.sheet_view.zoomScale = 85

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        page_count = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_footer(page_count)
            super().showPage()
        super().save()

    def draw_page_footer(self, page_count: int):
        width, _ = A4
        page_number = self._pageNumber
        self.setStrokeColor(colors.HexColor(BRAND_GREEN))
        self.setLineWidth(0.6)
        self.line(15 * mm, 24 * mm, width - 15 * mm, 24 * mm)
        self.setFont(THAI_FONT_BOLD, 9)
        self.setFillColor(colors.HexColor(BRAND_GREEN))
        self.drawString(16 * mm, 16 * mm, COMPANY_NAME)
        self.setFont(THAI_FONT, 8)
        self.setFillColor(colors.HexColor("#111827"))
        self.drawString(16 * mm, 11 * mm, SYSTEM_NAME)
        self.setFont(THAI_FONT, 9)
        self.drawRightString(width - 16 * mm, 13 * mm, f"Page {page_number} / {page_count}")


def report_number(value: float | int | str | None, digits: int = 2) -> str:
    try:
        numeric = float(value or 0)
    except (TypeError, ValueError):
        numeric = 0.0
    text = f"{numeric:,.{digits}f}"
    return text.rstrip("0").rstrip(".") if "." in text else text


def build_summary_by_pile_pdf(data: dict, payload: dict) -> bytes:
    now = datetime.now()
    start_date = payload.get("start_date", now.date().isoformat())
    end_date = payload.get("end_date", start_date)
    product = payload.get("product", "all")
    product_label = payload.get("product_label") or "ทั้งหมด"
    product_type = payload.get("product_type") or "ทั้งหมด"
    printed_by = payload.get("printed_by") or "System Admin"
    print_date, print_time = format_report_datetime(now)
    rows = pile_summary_rows(range_records(data, start_date, end_date, product))
    totals = summary_totals(rows)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=14 * mm,
        bottomMargin=30 * mm,
    )
    styles = getSampleStyleSheet()
    for style_name in ["Title", "Heading1", "Heading2", "Normal"]:
        styles[style_name].fontName = THAI_FONT
    title_style = styles["Title"].clone("SummaryReportTitle")
    title_style.fontName = THAI_FONT_BOLD
    title_style.fontSize = 17
    title_style.leading = 21
    title_style.textColor = colors.HexColor(BRAND_GREEN)
    title_style.spaceAfter = 2
    subtitle_style = styles["Normal"].clone("SummaryReportSubtitle")
    subtitle_style.fontSize = 10
    subtitle_style.leading = 13
    section_style = styles["Heading2"].clone("SummaryReportSection")
    section_style.fontName = THAI_FONT_BOLD
    section_style.fontSize = 13
    section_style.textColor = colors.HexColor(BRAND_GREEN)
    section_style.spaceAfter = 6

    logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
    logo = Image(str(logo_path), width=22 * mm, height=22 * mm) if logo_path.exists() else Paragraph("SP", title_style)
    title_block = [
        Paragraph(COMPANY_NAME, subtitle_style),
        Paragraph("รายงานสรุปน้ำหนักตามกอง", title_style),
    ]
    meta_rows = [
        ["ช่วงวันที่", f"{format_report_date(start_date)} - {format_report_date(end_date)}"],
        ["สินค้า", product_label],
        ["ประเภทสินค้า", product_type],
        ["ผู้พิมพ์", printed_by],
        ["วันที่พิมพ์", print_date],
        ["เวลา", print_time],
    ]
    meta_table = Table(meta_rows, colWidths=[23 * mm, 37 * mm])
    meta_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), THAI_FONT),
                ("FONTNAME", (0, 0), (0, -1), THAI_FONT_BOLD),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#111827")),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
            ]
        )
    )
    header = Table([[logo, title_block, meta_table]], colWidths=[24 * mm, 96 * mm, 60 * mm])
    header.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    green_line = Table([[""]], colWidths=[180 * mm], rowHeights=[1.1 * mm])
    green_line.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND_GREEN))]))

    table_rows = [["กอง", "น้ำหนักเข้า (ตัน)", "น้ำหนักออก (ตัน)", "น้ำหนักคงเหลือ", "รวมเงิน"]]
    for item in rows:
        table_rows.append(
            [
                f"กอง {item['pile']}",
                report_number(item["incoming"]),
                report_number(item["outgoing"]),
                report_number(item["balance"]),
                report_number(item["amount"], 0),
            ]
        )
    if len(table_rows) == 1:
        table_rows.append(["-", "0", "0", "0", "0"])
    table_rows.append(
        [
            "รวมทั้งหมด",
            report_number(totals["incoming"]),
            report_number(totals["outgoing"]),
            report_number(totals["balance"]),
            report_number(totals["amount"], 0),
        ]
    )

    summary_table = Table(
        table_rows,
        repeatRows=1,
        colWidths=[36 * mm, 36 * mm, 36 * mm, 36 * mm, 36 * mm],
        hAlign="LEFT",
    )
    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_GREEN)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), THAI_FONT_BOLD),
        ("FONTNAME", (0, 1), (-1, -1), THAI_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CDD6DF")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E9F5EE")),
        ("FONTNAME", (0, -1), (-1, -1), THAI_FONT_BOLD),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#064E25")),
    ]
    for row_index in range(1, len(table_rows) - 1):
        if row_index % 2 == 0:
            table_style.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F7FAF8")))
    summary_table.setStyle(TableStyle(table_style))

    note_box = Table(
        [[Paragraph("<b>หมายเหตุ</b><br/>หน่วยน้ำหนัก : ตัน", styles["Normal"])]],
        colWidths=[180 * mm],
    )
    note_box.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), THAI_FONT),
                ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CDD6DF")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFFFFF")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )

    story = [
        header,
        Spacer(1, 6 * mm),
        green_line,
        Spacer(1, 10 * mm),
        Paragraph("สรุปตามกอง (ตัน)", section_style),
        KeepTogether([summary_table, Spacer(1, 8 * mm), note_box]),
    ]
    doc.build(story, canvasmaker=NumberedCanvas)
    return buffer.getvalue()


def build_summary_by_pile_excel(data: dict, payload: dict) -> bytes:
    now = datetime.now()
    start_date = payload.get("start_date", now.date().isoformat())
    end_date = payload.get("end_date", start_date)
    product = payload.get("product", "all")
    product_label = payload.get("product_label") or "ทั้งหมด"
    product_type = payload.get("product_type") or "ทั้งหมด"
    printed_by = payload.get("printed_by") or "System Admin"
    print_date, print_time = format_report_datetime(now)
    rows = pile_summary_rows(range_records(data, start_date, end_date, product))
    totals = summary_totals(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary By Pile"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = "portrait"
    sheet.page_margins.left = 0.45
    sheet.page_margins.right = 0.45
    sheet.page_margins.top = 0.55
    sheet.page_margins.bottom = 0.55

    green_fill = PatternFill("solid", fgColor="0F7A3D")
    light_green_fill = PatternFill("solid", fgColor="E9F5EE")
    zebra_fill = PatternFill("solid", fgColor="F7FAF8")
    thin_green = Side(style="thin", color="0F7A3D")
    thin_gray = Side(style="thin", color="CDD6DF")
    table_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    sarabun = "Sarabun"

    sheet.merge_cells("A1:E1")
    sheet["A1"] = COMPANY_NAME
    sheet["A1"].font = Font(name=sarabun, bold=True, size=11, color="0F7A3D")
    sheet.merge_cells("A2:E2")
    sheet["A2"] = "รายงานสรุปน้ำหนักตามกอง"
    sheet["A2"].font = Font(name=sarabun, bold=True, size=20, color="0F7A3D")
    sheet.merge_cells("A3:E3")
    sheet["A3"] = f"ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)} | สินค้า {product_label} | ประเภทสินค้า {product_type}"
    sheet["A3"].font = Font(name=sarabun, size=10, color="111827")
    sheet.merge_cells("A4:E4")
    sheet["A4"] = f"ผู้พิมพ์ {printed_by} | วันที่พิมพ์ {print_date} | เวลา {print_time}"
    sheet["A4"].font = Font(name=sarabun, size=10, color="111827")
    for cell in sheet[5]:
        cell.fill = green_fill
        cell.border = Border(bottom=thin_green)
    sheet.row_dimensions[5].height = 4

    headers = ["กอง", "น้ำหนักเข้า (ตัน)", "น้ำหนักออก (ตัน)", "น้ำหนักคงเหลือ", "รวมเงิน"]
    header_row = 7
    for column_index, value in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column_index, value)
        cell.fill = green_fill
        cell.font = Font(name=sarabun, bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = table_border

    row_index = header_row + 1
    for item in rows:
        values = [
            item["pile"],
            item["incoming"],
            item["outgoing"],
            item["balance"],
            item["amount"],
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, value)
            cell.font = Font(name=sarabun, size=11)
            cell.border = table_border
            cell.alignment = Alignment(horizontal="left" if column_index == 1 else "right", vertical="center")
            if column_index > 1:
                cell.number_format = "#,##0.00"
            if row_index % 2 == 1:
                cell.fill = zebra_fill
        row_index += 1

    if not rows:
        for column_index, value in enumerate(["-", 0, 0, 0, 0], start=1):
            cell = sheet.cell(row_index, column_index, value)
            cell.font = Font(name=sarabun, size=11)
            cell.border = table_border
            cell.alignment = Alignment(horizontal="left" if column_index == 1 else "right")
            if column_index > 1:
                cell.number_format = "#,##0.00"
        row_index += 1

    total_values = ["รวมทั้งหมด", totals["incoming"], totals["outgoing"], totals["balance"], totals["amount"]]
    for column_index, value in enumerate(total_values, start=1):
        cell = sheet.cell(row_index, column_index, value)
        cell.fill = light_green_fill
        cell.font = Font(name=sarabun, bold=True, color="064E25", size=11)
        cell.border = table_border
        cell.alignment = Alignment(horizontal="left" if column_index == 1 else "right", vertical="center")
        if column_index > 1:
            cell.number_format = "#,##0.00"

    note_row = row_index + 3
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 1, end_column=5)
    note_cell = sheet.cell(note_row, 1, "หมายเหตุ\nหน่วยน้ำหนัก : ตัน")
    note_cell.font = Font(name=sarabun, size=11)
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    note_cell.border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    footer_row = note_row + 5
    sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=3)
    sheet.cell(footer_row, 1, SYSTEM_NAME)
    sheet.cell(footer_row, 1).font = Font(name=sarabun, bold=True, color="0F7A3D", size=10)
    sheet.cell(footer_row, 5, "Page 1 / 1")
    sheet.cell(footer_row, 5).alignment = Alignment(horizontal="right")

    widths = [22, 20, 20, 20, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A7"
    sheet.print_area = f"A1:E{footer_row}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def normalized_range(payload: dict) -> tuple[str, str]:
    today = datetime.now().date().isoformat()
    start_date = payload.get("start_date") or today
    end_date = payload.get("end_date") or start_date
    return tuple(sorted([start_date, end_date]))


def deduction_records_for(payload: dict, employee_kind: str, employee_id: object, emp_code: object, start_date: str, end_date: str) -> list[dict]:
    employee_id_text = str(employee_id or "")
    emp_code_text = str(emp_code or "")
    rows = []
    for record in payload.get("deduction_records", []) or []:
        if str(record.get("deduction_type") or "") == "attendance_bonus":
            continue
        if str(record.get("status") or "Active") != "Active":
            continue
        if str(record.get("employee_kind") or "production") != employee_kind:
            continue
        record_start = str(record.get("start_date") or "")
        record_end = str(record.get("end_date") or record_start)
        if record_start > end_date or record_end < start_date:
            continue
        same_employee = (
            employee_id_text and str(record.get("employee_id") or "") == employee_id_text
        ) or (
            emp_code_text and str(record.get("emp_code") or "") == emp_code_text
        )
        if same_employee:
            rows.append(record)
    return rows


def bonus_records_for(payload: dict, employee_kind: str, employee_id: object, emp_code: object, start_date: str, end_date: str) -> list[dict]:
    employee_id_text = str(employee_id or "")
    emp_code_text = str(emp_code or "")
    rows = []
    for record in payload.get("deduction_records", []) or []:
        if str(record.get("deduction_type") or "") != "attendance_bonus":
            continue
        if str(record.get("status") or "Active") != "Active":
            continue
        if str(record.get("employee_kind") or "production") != employee_kind:
            continue
        record_start = str(record.get("start_date") or "")
        record_end = str(record.get("end_date") or record_start)
        if record_start > end_date or record_end < start_date:
            continue
        same_employee = (
            employee_id_text and str(record.get("employee_id") or "") == employee_id_text
        ) or (
            emp_code_text and str(record.get("emp_code") or "") == emp_code_text
        )
        if same_employee:
            rows.append(record)
    return rows


def deduction_total(records: list[dict]) -> float:
    return sum(safe_float(record.get("amount")) for record in records)


def full_export_records(payload: dict) -> tuple[list[dict], list[dict]]:
    start_date, end_date = normalized_range(payload)
    department = payload.get("department") or "all"
    production_records = [
        record
        for record in payload.get("production_records", [])
        if start_date <= (record.get("record_date") or record.get("date") or "") <= end_date
    ]
    time_records = [
        record
        for record in payload.get("time_records", [])
        if start_date <= (record.get("record_date") or "") <= end_date
        and (department == "all" or record.get("department") == department)
    ]
    return production_records, time_records


def build_time_full_export_excel(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    production_records, time_records = full_export_records(payload)
    employees = payload.get("employees", [])
    printed_by = payload.get("printed_by") or "System Admin"
    department = payload.get("department") or "all"
    now = datetime.now()
    print_date, print_time = format_report_datetime(now)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Overview"
    green_fill = PatternFill("solid", fgColor="0F7A3D")
    light_green_fill = PatternFill("solid", fgColor="E9F5EE")
    zebra_fill = PatternFill("solid", fgColor="F7FAF8")
    thin_gray = Side(style="thin", color="CDD6DF")
    table_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    font_name = "Sarabun"

    def style_header(row):
        for cell in row:
            cell.fill = green_fill
            cell.font = Font(name=font_name, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = table_border

    def style_table(sheet_obj, header_row: int):
        style_header(sheet_obj[header_row])
        for row_index, row in enumerate(sheet_obj.iter_rows(min_row=header_row + 1), start=header_row + 1):
            for cell in row:
                cell.border = table_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.font = Font(name=font_name, size=10)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if row_index % 2 == 0:
                    cell.fill = zebra_fill

    def autosize(sheet_obj):
        for column_cells in sheet_obj.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet_obj.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 34)

    sheet.merge_cells("A1:H1")
    sheet["A1"] = COMPANY_NAME
    sheet["A1"].font = Font(name=font_name, bold=True, size=18, color="0F7A3D")
    sheet.merge_cells("A2:H2")
    sheet["A2"] = "Export รายละเอียดทั้งหมด"
    sheet["A2"].font = Font(name=font_name, bold=True, size=16, color="111827")
    sheet.merge_cells("A3:H3")
    sheet["A3"] = f"ช่วงวันที่ {start_date} ถึง {end_date} | แผนก {department if department != 'all' else 'ทุกแผนก'}"
    sheet.merge_cells("A4:H4")
    sheet["A4"] = f"ผู้ Export {printed_by} | วันที่ {print_date} | เวลา {print_time}"
    for ref in ["A3", "A4"]:
        sheet[ref].font = Font(name=font_name, size=11, color="344054")

    logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
    if logo_path.exists():
        try:
            logo = ExcelImage(str(logo_path))
            logo.width = 70
            logo.height = 70
            sheet.add_image(logo, "I1")
        except Exception:
            pass

    total_weight = sum(production_total_weight(record) for record in production_records)
    total_production_amount = sum(float(record.get("total_amount", record.get("grand_total", 0)) or 0) for record in production_records)
    total_time_minutes = sum(float(record.get("net_minutes", 0) or 0) for record in time_records)
    overview_rows = [
        ["หัวข้อ", "ค่า"],
        ["จำนวนรายการผลผลิต", len(production_records)],
        ["น้ำหนักผลผลิตรวม", total_weight],
        ["ยอดเงินผลผลิตรวม", total_production_amount],
        ["จำนวนรายการเวลา", len(time_records)],
        ["ชั่วโมงทำงานรวม", total_time_minutes / 60],
        ["จำนวนพนักงานทั้งหมด", len(employees)],
    ]
    for row in overview_rows:
        sheet.append(row)
    style_table(sheet, 6)
    autosize(sheet)

    prod_sheet = workbook.create_sheet("Production Details")
    prod_headers = [
        "วันที่",
        "เวลา",
        "รหัสพนักงาน",
        "ชื่อพนักงาน",
        "สินค้า",
        "กอง",
        "น้ำหนักน้ำ",
        "น้ำหนักดอก",
        "น้ำหนักทุเรียน",
        "น้ำหนักรวม",
        "รวมเงิน",
        "ผู้บันทึก",
    ]
    prod_sheet.append(prod_headers)
    for record in sorted(production_records, key=lambda item: ((item.get("record_date") or item.get("date") or ""), item.get("record_time", ""), item.get("emp_code", ""))):
        prod_sheet.append(
            [
                record.get("record_date") or record.get("date") or "",
                record.get("record_time", ""),
                record.get("emp_code", ""),
                record.get("employee_name", ""),
                record.get("fruit_type", "mangosteen"),
                record.get("pile_no") or record.get("pile", ""),
                float(record.get("water_weight", record.get("water", 0)) or 0),
                float(record.get("flower_weight", record.get("flower", 0)) or 0),
                production_grade_text(record) if (record.get("fruit_type") or "mangosteen") == "durian" else "-",
                production_total_weight(record),
                float(record.get("total_amount", record.get("grand_total", 0)) or 0),
                record.get("created_by", ""),
            ]
        )
    style_table(prod_sheet, 1)
    autosize(prod_sheet)

    time_sheet = workbook.create_sheet("Time Details")
    time_headers = ["วันที่", "รหัสพนักงาน", "ชื่อพนักงาน", "แผนก", "เข้า", "ออก", "พัก(นาที)", "สุทธิ(นาที)", "ชั่วโมง", "ผู้บันทึก"]
    time_sheet.append(time_headers)
    for record in sorted(time_records, key=lambda item: (item.get("record_date", ""), item.get("emp_code", ""), item.get("clock_in", ""))):
        net_minutes = float(record.get("net_minutes", 0) or 0)
        time_sheet.append(
            [
                record.get("record_date", ""),
                record.get("emp_code", ""),
                record.get("fullname", ""),
                record.get("department", ""),
                record.get("clock_in", ""),
                record.get("clock_out", ""),
                float(record.get("break_minutes", 0) or 0),
                net_minutes,
                net_minutes / 60,
                record.get("created_by", ""),
            ]
        )
    style_table(time_sheet, 1)
    autosize(time_sheet)

    summary_sheet = workbook.create_sheet("Time Summary")
    summary_sheet.append(["รหัสพนักงาน", "ชื่อพนักงาน", "แผนก", "จำนวนวัน", "จำนวนรายการ", "ชั่วโมงรวม"])
    grouped: dict[str, dict] = {}
    for record in time_records:
        key = str(record.get("employee_id") or record.get("emp_code") or record.get("fullname") or "")
        item = grouped.setdefault(
            key,
            {
                "emp_code": record.get("emp_code", ""),
                "fullname": record.get("fullname", ""),
                "department": record.get("department", ""),
                "days": set(),
                "records": 0,
                "minutes": 0.0,
            },
        )
        item["days"].add(record.get("record_date", ""))
        item["records"] += 1
        item["minutes"] += float(record.get("net_minutes", 0) or 0)
    for item in sorted(grouped.values(), key=lambda value: (value["emp_code"], value["fullname"])):
        summary_sheet.append([item["emp_code"], item["fullname"], item["department"], len(item["days"]), item["records"], item["minutes"] / 60])
    style_table(summary_sheet, 1)
    autosize(summary_sheet)

    deduction_sheet = workbook.create_sheet("Adjustments")
    deduction_sheet.append(["ประเภทรายการ", "ประเภทพนักงาน", "วันที่เริ่ม", "วันที่สิ้นสุด", "รหัส", "ชื่อพนักงาน", "รายการ", "จำนวนเงิน", "หมายเหตุ", "ผู้บันทึก"])
    deduction_rows = [
        record
        for record in payload.get("deduction_records", []) or []
        if str(record.get("status", "Active")) == "Active"
        and str(record.get("start_date", "")) <= end_date
        and str(record.get("end_date", record.get("start_date", ""))) >= start_date
    ]
    for record in sorted(deduction_rows, key=lambda item: (item.get("employee_kind", ""), item.get("start_date", ""), item.get("emp_code", ""))):
        deduction_sheet.append([
            "เบี้ยขยัน" if record.get("deduction_type") == "attendance_bonus" else "รายการหัก",
            "พนักงานเหมาเวลา" if record.get("employee_kind") == "time" else "พนักงานเหมาน้ำหนัก",
            record.get("start_date", ""),
            record.get("end_date", ""),
            record.get("emp_code", ""),
            record.get("employee_name", ""),
            record.get("deduction_label", ""),
            safe_float(record.get("amount")),
            record.get("note", ""),
            record.get("created_by", ""),
        ])
    style_table(deduction_sheet, 1)
    autosize(deduction_sheet)

    for sheet_obj in workbook.worksheets:
        sheet_obj.freeze_panes = "A2"
        sheet_obj.page_setup.paperSize = sheet_obj.PAPERSIZE_A4
        sheet_obj.page_setup.orientation = "landscape"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def time_receipt_groups(payload: dict) -> list[dict]:
    start_date, end_date = normalized_range(payload)
    department = payload.get("department") or "all"
    records = [
        record
        for record in payload.get("time_records", [])
        if start_date <= (record.get("record_date") or "") <= end_date
        and (department == "all" or record.get("department") == department)
    ]
    daily_wage = float(payload.get("daily_wage", 347) or 347)
    standard_hours = float(payload.get("standard_hours", 8) or 8)
    normal_rate = daily_wage / standard_hours
    ot_rate = float(payload.get("ot_hourly_rate", 50) or 50)
    daily_records: dict[tuple[str, str], dict] = {}
    for record in sorted(records, key=lambda item: (item.get("record_date", ""), item.get("clock_in", ""))):
        key = str(record.get("employee_id") or record.get("emp_code") or record.get("fullname") or "")
        date_key = record.get("record_date", "-")
        daily_key = (key, date_key)
        if daily_key not in daily_records:
            daily_records[daily_key] = {
                **record,
                "clock_ins": [],
                "clock_outs": [],
                "net_minutes": 0.0,
            }
        daily_record = daily_records[daily_key]
        daily_record["clock_ins"].append(record.get("clock_in", "-"))
        daily_record["clock_outs"].append(record.get("clock_out", "-"))
        daily_record["net_minutes"] += float(record.get("net_minutes", 0) or 0)

    groups: dict[str, dict] = {}
    for record in daily_records.values():
        key = str(record.get("employee_id") or record.get("emp_code") or record.get("fullname") or "")
        group = groups.setdefault(
            key,
            {
                "employee_id": record.get("employee_id"),
                "emp_code": record.get("emp_code", "-"),
                "fullname": record.get("fullname", "-"),
                "department": record.get("department", "-"),
                "rows": [],
                "normal_hours": 0.0,
                "ot_hours": 0.0,
                "normal_amount": 0.0,
                "ot_amount": 0.0,
                "deduction_amount": 0.0,
                "deductions": [],
                "bonus_amount": 0.0,
                "bonuses": [],
            },
        )
        net_minutes = float(record.get("net_minutes", 0) or 0)
        normal_minutes = min(net_minutes, standard_hours * 60)
        ot_minutes = max(0, net_minutes - normal_minutes)
        normal_hours = normal_minutes / 60
        ot_hours = ot_minutes / 60
        record_daily_wage = float(record.get("daily_wage") or daily_wage)
        record_normal_rate = float(record.get("normal_hourly_rate") or (record_daily_wage / standard_hours))
        record_ot_rate = float(record.get("ot_hourly_rate") or ot_rate)
        # The factory wage table rounds each day's proportional normal wage
        # to the nearest whole baht (0.50 rounds up) before employee totals.
        rounded_half_hour = round(normal_hours * 2) / 2
        if int(record_daily_wage) == TIME_SPECIAL_DAILY_WAGE and rounded_half_hour in TIME_SPECIAL_WAGE_TABLE:
            normal_amount = TIME_SPECIAL_WAGE_TABLE[rounded_half_hour]
        elif normal_hours >= standard_hours:
            normal_amount = record_daily_wage
        else:
            normal_amount = math.floor((normal_hours * record_normal_rate) + 0.5)
        ot_amount = ot_hours * record_ot_rate
        group["normal_hours"] += normal_hours
        group["ot_hours"] += ot_hours
        group["normal_amount"] += normal_amount
        group["ot_amount"] += ot_amount
        group["rows"].append(
            {
                "date": record.get("record_date", "-"),
                "clock_in": " / ".join(record.get("clock_ins", [])),
                "clock_out": " / ".join(record.get("clock_outs", [])),
                "normal_hours": normal_hours,
                "ot_hours": ot_hours,
                "normal_amount": normal_amount,
                "ot_amount": ot_amount,
                "total_amount": normal_amount + ot_amount,
            }
        )
    for group in groups.values():
        group["rows"].sort(key=lambda item: (item["date"], item["clock_in"]))
        group["deductions"] = deduction_records_for(
            payload,
            "time",
            group.get("employee_id"),
            group.get("emp_code"),
            start_date,
            end_date,
        )
        group["deduction_amount"] = deduction_total(group["deductions"])
        group["bonuses"] = bonus_records_for(
            payload,
            "time",
            group.get("employee_id"),
            group.get("emp_code"),
            start_date,
            end_date,
        )
        group["bonus_amount"] = deduction_total(group["bonuses"])
        group["gross_amount"] = group["normal_amount"] + group["ot_amount"]
        group["total_amount"] = max(0, group["gross_amount"] + group["bonus_amount"] - group["deduction_amount"])
    return sorted(groups.values(), key=lambda item: (str(item["emp_code"]), str(item["fullname"])))


def build_time_receipts_pdf(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    groups = time_receipt_groups(payload)
    printed_by = payload.get("printed_by") or "System Admin"
    printed_by_position = payload.get("printed_by_position") or ""
    printed_by_text = f"{printed_by} {printed_by_position}".strip()
    now = datetime.now()
    buffer = BytesIO()
    page_width, page_height = landscape(A4)
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
    margin = 8 * mm
    gap = 7 * mm
    divider_x = page_width / 2
    panel_width = (page_width - (margin * 2) - gap) / 2
    panel_height = page_height - (margin * 2)

    def text(x, y, value, size=8, bold=False, fill="#111827"):
        c.setFillColor(colors.HexColor(fill))
        c.setFont(THAI_FONT_BOLD if bold else THAI_FONT, size)
        c.drawString(x, y, str(value))

    def right_text(x, y, value, size=8, bold=False, fill="#111827"):
        c.setFillColor(colors.HexColor(fill))
        c.setFont(THAI_FONT_BOLD if bold else THAI_FONT, size)
        c.drawRightString(x, y, str(value))

    def draw_panel(x, y, group, is_copy=False):
        c.setStrokeColor(colors.HexColor("#0F7A3D"))
        c.setLineWidth(1)
        c.roundRect(x, y, panel_width, panel_height, 6, stroke=1, fill=0)
        if is_copy:
            c.saveState()
            c.translate(x + panel_width / 2, y + panel_height / 2)
            c.rotate(30)
            c.setFillColor(colors.Color(0.06, 0.48, 0.24, alpha=0.08))
            c.setFont(THAI_FONT_BOLD, 58)
            c.drawCentredString(0, 0, "สำเนา")
            c.restoreState()

        cursor_y = y + panel_height - 12 * mm
        if logo_path.exists():
            try:
                c.drawImage(str(logo_path), x + 5 * mm, cursor_y - 3 * mm, width=17 * mm, height=17 * mm, preserveAspectRatio=True, mask="auto")
            except Exception:
                pass
        text(x + 25 * mm, cursor_y + 6 * mm, COMPANY_NAME, 11, True, "#0F7A3D")
        text(x + 25 * mm, cursor_y, "ใบเสร็จเวลาและค่าแรง", 10, True, "#0F7A3D")
        text(x + 25 * mm, cursor_y - 5 * mm, "ต้นฉบับ" if not is_copy else "สำเนา", 8, True, "#667085")
        c.setStrokeColor(colors.HexColor("#B8D8D1"))
        c.line(x + 5 * mm, cursor_y - 9 * mm, x + panel_width - 5 * mm, cursor_y - 9 * mm)

        meta_y = cursor_y - 17 * mm
        text(x + 5 * mm, meta_y, f"เลขที่: TR-{start_date.replace('-', '')}-{group['emp_code']}", 7.5, True)
        text(x + 5 * mm, meta_y - 5 * mm, f"ช่วงวันที่: {start_date} - {end_date}", 7.5)
        right_text(x + panel_width - 5 * mm, meta_y, f"ออกโดย: {printed_by_text}", 7.5)
        right_text(x + panel_width - 5 * mm, meta_y - 5 * mm, now.strftime("%d/%m/%Y %H:%M"), 7.5)

        employee_y = meta_y - 15 * mm
        text(x + 5 * mm, employee_y, f"รหัสพนักงาน: {group['emp_code']}", 8, True)
        text(x + 48 * mm, employee_y, f"ชื่อ: {group['fullname']}", 8, True)
        text(x + 5 * mm, employee_y - 5 * mm, f"แผนก: {group['department']}", 8)

        table_y = employee_y - 13 * mm
        col_widths = [20, 13, 13, 16, 13, 19, 17, 18]
        headers = ["วันที่", "เข้า", "ออก", "ปกติ", "OT", "เงินปกติ", "เงิน OT", "รวม"]
        c.setFillColor(colors.HexColor("#0F7A3D"))
        c.rect(x + 5 * mm, table_y, panel_width - 10 * mm, 7 * mm, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont(THAI_FONT_BOLD, 6.5)
        col_x = x + 6 * mm
        for header, width in zip(headers, col_widths):
            c.drawString(col_x, table_y + 2.2 * mm, header)
            col_x += width * mm

        row_y = table_y - 6 * mm
        c.setFont(THAI_FONT, 6.2)
        c.setFillColor(colors.HexColor("#111827"))
        for row in group["rows"][:10]:
            c.setStrokeColor(colors.HexColor("#D8E5E1"))
            c.line(x + 5 * mm, row_y - 1 * mm, x + panel_width - 5 * mm, row_y - 1 * mm)
            values = [
                row["date"],
                row["clock_in"],
                row["clock_out"],
                report_number(row["normal_hours"]),
                report_number(row["ot_hours"]),
                report_number(row["normal_amount"], 0),
                report_number(row["ot_amount"], 0),
                report_number(row["total_amount"], 0),
            ]
            col_x = x + 6 * mm
            for value, width in zip(values, col_widths):
                c.drawString(col_x, row_y + 1 * mm, str(value))
                col_x += width * mm
            row_y -= 5.5 * mm
        if len(group["rows"]) > 10:
            text(x + 6 * mm, row_y + 1 * mm, f"มีรายการเพิ่มเติม {len(group['rows']) - 10} รายการ รวมอยู่ในยอดด้านล่าง", 6.2, False, "#667085")

        summary_y = y + 35 * mm
        c.setFillColor(colors.HexColor("#E9F5EE"))
        c.roundRect(x + 5 * mm, summary_y, panel_width - 10 * mm, 18 * mm, 4, stroke=0, fill=1)
        text(x + 8 * mm, summary_y + 12 * mm, f"ชั่วโมงปกติ {report_number(group['normal_hours'])} ชม.", 7.5, True, "#064E25")
        text(x + 8 * mm, summary_y + 7 * mm, f"OT {report_number(group['ot_hours'])} ชม.", 7.5, True, "#064E25")
        text(x + 55 * mm, summary_y + 12 * mm, f"เงินปกติ {report_number(group['normal_amount'], 0)} บาท", 7.5, True, "#064E25")
        text(x + 55 * mm, summary_y + 7 * mm, f"เงิน OT {report_number(group['ot_amount'], 0)} บาท", 7.5, True, "#064E25")
        right_text(x + panel_width - 8 * mm, summary_y + 12 * mm, f"รวม {report_number(group.get('gross_amount', group['total_amount']), 0)} บาท", 8, True, "#064E25")
        right_text(x + panel_width - 8 * mm, summary_y + 7 * mm, f"เบี้ยขยัน {report_number(group.get('bonus_amount', 0), 0)} | หัก {report_number(group.get('deduction_amount', 0), 0)} บาท", 7.2, True, "#166534")
        right_text(x + panel_width - 8 * mm, summary_y + 2 * mm, f"สุทธิ {report_number(group['total_amount'], 0)} บาท", 10.5, True, "#064E25")

        deduction_y = y + 28 * mm
        if group.get("deductions"):
            text(x + 6 * mm, deduction_y, "รายการหัก", 6.6, True, "#B42318")
            for index, deduction in enumerate(group.get("deductions", [])[:3]):
                text(
                    x + 26 * mm,
                    deduction_y - (index * 4 * mm),
                    f"{deduction.get('deduction_label', '-')} {report_number(deduction.get('amount'), 0)} บาท",
                    6.3,
                    False,
                    "#7A271A",
                )

        sign_y = y + 12 * mm
        c.setStrokeColor(colors.HexColor("#111827"))
        c.line(x + 14 * mm, sign_y + 5 * mm, x + 54 * mm, sign_y + 5 * mm)
        c.line(x + panel_width - 54 * mm, sign_y + 5 * mm, x + panel_width - 14 * mm, sign_y + 5 * mm)
        text(x + 25 * mm, sign_y, "ผู้รับเงิน", 7, False)
        text(x + panel_width - 43 * mm, sign_y, "ผู้จ่ายเงิน", 7, False)

    if not groups:
        groups = [{"emp_code": "-", "fullname": "-", "department": "-", "rows": [], "normal_hours": 0, "ot_hours": 0, "normal_amount": 0, "ot_amount": 0, "deduction_amount": 0, "deductions": [], "bonus_amount": 0, "bonuses": [], "gross_amount": 0, "total_amount": 0}]

    for group in groups:
        draw_panel(margin, margin, group, False)
        c.setDash(3, 3)
        c.setStrokeColor(colors.HexColor("#667085"))
        c.line(divider_x, margin, divider_x, page_height - margin)
        c.setDash()
        draw_panel(divider_x + gap / 2, margin, group, True)
        c.showPage()

    c.save()
    return buffer.getvalue()


def safe_float(value: float | int | str | None) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def production_pile_number(record: dict) -> int | None:
    value = record.get("pile_no")
    if value in [None, ""]:
        value = record.get("pile")
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return int(numeric) if numeric.is_integer() and numeric > 0 else None


def selected_production_fruit(payload: dict) -> str:
    fruit_id = str(payload.get("fruit_type") or "all").strip().lower()
    return fruit_id if fruit_id in {"all", "mangosteen", "durian", "mango"} else "all"


def selected_production_fruit_label(payload: dict) -> str:
    return {
        "all": "ทุกผลไม้",
        "mangosteen": "มังคุด",
        "durian": "ทุเรียน",
        "mango": "มะม่วง",
    }[selected_production_fruit(payload)]


def production_report_field_visibility(payload: dict) -> tuple[bool, bool]:
    fruit_id = selected_production_fruit(payload)
    return fruit_id != "durian", fruit_id in {"all", "durian"}


def production_report_weight_labels(payload: dict) -> tuple[str, str]:
    fruit_id = selected_production_fruit(payload)
    if fruit_id == "mango":
        return "มะม่วงฝา", "มะม่วงหั่นเต๋า"
    if fruit_id == "all":
        return "น้ำหนักน้ำ / มะม่วงฝา", "น้ำหนักดอก / มะม่วงหั่นเต๋า"
    return "น้ำหนักน้ำ", "น้ำหนักดอก"


def production_grade_weights(record: dict) -> dict[str, float]:
    source = record.get("grade_weights") if isinstance(record.get("grade_weights"), dict) else {}
    return {grade: safe_float(source.get(grade, source.get(grade.lower(), 0))) for grade in DURIAN_GRADES}


def production_grade_total(record: dict) -> float:
    return sum(production_grade_weights(record).values())


def production_total_weight(record: dict) -> float:
    if (record.get("fruit_type") or "mangosteen") == "durian":
        direct_weight = safe_float(record.get("durian_weight"))
        if direct_weight > 0:
            return direct_weight
        stored_total = safe_float(record.get("total_weight"))
        if stored_total > 0:
            return stored_total
        return production_grade_total(record)
    return safe_float(record.get("water_weight", record.get("water", 0))) + safe_float(record.get("flower_weight", record.get("flower", 0)))


def production_grade_text(record: dict) -> str:
    return report_number(production_total_weight(record))


def grade_totals_text(weights: dict | None) -> str:
    source = weights if isinstance(weights, dict) else {}
    return report_number(sum(safe_float(source.get(grade, 0)) for grade in DURIAN_GRADES))


def minutes_text(value: float | int | str | None) -> str:
    minutes = int(round(safe_float(value)))
    hours, remain = divmod(minutes, 60)
    return f"{hours} ชม. {remain} นาที"


def is_late_time(value: str | None) -> bool:
    if not value:
        return False
    try:
        hour, minute = [int(part) for part in str(value).split(":")[:2]]
        return hour * 60 + minute > 8 * 60
    except (TypeError, ValueError):
        return False


def is_early_out_time(value: str | None) -> bool:
    if not value:
        return False
    try:
        hour, minute = [int(part) for part in str(value).split(":")[:2]]
        return hour * 60 + minute < 17 * 60
    except (TypeError, ValueError):
        return False


def filtered_production_records(payload: dict) -> list[dict]:
    start_date, end_date = normalized_range(payload)
    fruit_id = selected_production_fruit(payload)
    return sorted(
        [
            record
            for record in payload.get("production_records", [])
            if start_date <= (record.get("record_date") or record.get("date") or "") <= end_date
            and (fruit_id == "all" or (record.get("fruit_type") or "mangosteen") == fruit_id)
        ],
        key=lambda record: (
            record.get("record_date") or record.get("date") or "",
            record.get("record_time", ""),
            production_pile_number(record) or 0,
            str(record.get("emp_code") or ""),
        ),
    )


def filtered_time_records(payload: dict) -> list[dict]:
    start_date, end_date = normalized_range(payload)
    department = payload.get("department") or "all"
    return sorted(
        [
            record
            for record in payload.get("time_records", [])
            if start_date <= (record.get("record_date") or "") <= end_date
            and (department == "all" or record.get("department") == department)
        ],
        key=lambda record: (
            record.get("record_date", ""),
            str(record.get("emp_code") or ""),
            record.get("clock_in", ""),
        ),
    )


def export_meta_text(payload: dict) -> str:
    printed_by = payload.get("printed_by") or "System Admin"
    printed_by_position = payload.get("printed_by_position") or ""
    print_date, print_time = format_report_datetime(datetime.now())
    return f"ออกโดย {printed_by} {printed_by_position} | วันที่ {print_date} | เวลา {print_time}".strip()


def add_excel_logo(sheet, cell: str = "H1") -> None:
    logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
    if not logo_path.exists():
        return
    try:
        logo = ExcelImage(str(logo_path))
        logo.width = 72
        logo.height = 72
        sheet.add_image(logo, cell)
    except Exception:
        pass


def style_excel_report_sheet(sheet, header_rows: list[int], widths: list[int]) -> None:
    font_name = "Sarabun"
    green_fill = PatternFill("solid", fgColor="0F7A3D")
    zebra_fill = PatternFill("solid", fgColor="F7FAF8")
    thin_gray = Side(style="thin", color="CDD6DF")
    table_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    for header_row in header_rows:
        for cell in sheet[header_row]:
            if cell.value is None:
                continue
            cell.fill = green_fill
            cell.font = Font(name=font_name, bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = table_border
    for row_index, row in enumerate(sheet.iter_rows(), start=1):
        if row_index in header_rows or row_index < min(header_rows, default=1):
            continue
        for cell in row:
            if cell.value is None:
                continue
            cell.font = Font(name=font_name, size=10)
            cell.border = table_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if row_index % 2 == 0:
                cell.fill = zebra_fill
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.freeze_panes = f"A{max(header_rows) + 1}" if header_rows else None


def build_production_summary_excel(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    records = filtered_production_records(payload)
    rows = pile_summary_rows(records)
    totals = summary_totals(rows)
    total_records = len(records)
    employee_count = len({record.get("employee_id") or record.get("emp_code") or record.get("employee_name") for record in records})

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Production Summary"
    overview.merge_cells("A1:G1")
    overview["A1"] = COMPANY_NAME
    overview["A1"].font = Font(name="Sarabun", bold=True, size=18, color="0F7A3D")
    overview.merge_cells("A2:G2")
    overview["A2"] = "รายงานสรุปผลผลิตและน้ำหนัก"
    overview["A2"].font = Font(name="Sarabun", bold=True, size=16, color="111827")
    overview.merge_cells("A3:G3")
    overview["A3"] = f"ผลไม้ {selected_production_fruit_label(payload)} | ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)}"
    overview.merge_cells("A4:G4")
    overview["A4"] = export_meta_text(payload)
    add_excel_logo(overview, "H1")
    overview.append([])
    overview.append(["หัวข้อ", "ค่า"])
    overview.append(["จำนวนรายการ", total_records])
    overview.append(["จำนวนพนักงาน", employee_count])
    overview.append(["น้ำหนักเข้า", totals["incoming"]])
    overview.append(["น้ำหนักออก", totals["outgoing"]])
    overview.append(["น้ำหนักคงเหลือ", totals["balance"]])
    overview.append(["ยอดเงินรวม", totals["amount"]])
    style_excel_report_sheet(overview, [6], [28, 22, 14, 14, 14, 14, 14, 14])

    pile_sheet = workbook.create_sheet("Pile Summary")
    pile_sheet.append(["กอง", "น้ำหนักเข้า", "น้ำหนักออก", "น้ำหนักคงเหลือ", "ยอดเงิน"])
    for row in rows:
        pile_sheet.append([row["pile"], row["incoming"], row["outgoing"], row["balance"], row["amount"]])
    pile_sheet.append(["รวม", totals["incoming"], totals["outgoing"], totals["balance"], totals["amount"]])
    style_excel_report_sheet(pile_sheet, [1], [18, 18, 18, 18, 18])

    detail_sheet = workbook.create_sheet("Details")
    detail_sheet.append(["วันที่", "เวลา", "รหัสพนักงาน", "ชื่อพนักงาน", "สินค้า", "กอง", "น้ำหนักน้ำ", "น้ำหนักดอก", "ยอดเงิน", "ผู้บันทึก"])
    for record in records:
        detail_sheet.append(
            [
                record.get("record_date") or record.get("date") or "",
                record.get("record_time", ""),
                record.get("emp_code", ""),
                record.get("employee_name", ""),
                record.get("fruit_type", ""),
                record.get("pile_no") or record.get("pile", ""),
                safe_float(record.get("water_weight", record.get("water", 0))),
                safe_float(record.get("flower_weight", record.get("flower", 0))),
                safe_float(record.get("total_amount", record.get("grand_total", 0))),
                record.get("created_by", ""),
            ]
        )
    style_excel_report_sheet(detail_sheet, [1], [14, 12, 16, 24, 16, 10, 14, 14, 14, 20])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def pdf_styles() -> tuple:
    styles = getSampleStyleSheet()
    for style_name in ["Title", "Heading1", "Heading2", "Normal", "BodyText"]:
        styles[style_name].fontName = THAI_FONT
    title = styles["Title"].clone("PismaiReportTitle")
    title.fontName = THAI_FONT_BOLD
    title.fontSize = 17
    title.leading = 22
    title.textColor = colors.HexColor(BRAND_GREEN)
    normal = styles["Normal"].clone("PismaiReportNormal")
    normal.fontName = THAI_FONT
    normal.fontSize = 9
    normal.leading = 12
    section = styles["Heading2"].clone("PismaiReportSection")
    section.fontName = THAI_FONT_BOLD
    section.fontSize = 12
    section.leading = 16
    section.textColor = colors.HexColor(BRAND_GREEN)
    return styles, title, normal, section


def report_header_story(title_text: str, subtitle: str, payload: dict, page_width_mm: int = 267) -> list:
    _, title, normal, _ = pdf_styles()
    logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
    logo = Image(str(logo_path), width=20 * mm, height=20 * mm) if logo_path.exists() else Paragraph("SP", title)
    title_block = [Paragraph(COMPANY_NAME, normal), Paragraph(title_text, title), Paragraph(subtitle, normal)]
    meta = Paragraph(export_meta_text(payload), normal)
    header = Table([[logo, title_block, meta]], colWidths=[24 * mm, (page_width_mm - 94) * mm, 70 * mm])
    header.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    line = Table([[""]], colWidths=[page_width_mm * mm], rowHeights=[1 * mm])
    line.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND_GREEN))]))
    return [header, Spacer(1, 5 * mm), line, Spacer(1, 7 * mm)]


def set_pdf_table_style(table: Table, numeric_from: int = 1) -> None:
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_GREEN)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), THAI_FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, -1), THAI_FONT),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (numeric_from, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CDD6DF")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAF8")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )


def build_production_summary_pdf(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    records = filtered_production_records(payload)
    pile_rows = pile_summary_rows(records)
    totals = summary_totals(pile_rows)
    employee_count = len({record.get("employee_id") or record.get("emp_code") or record.get("employee_name") for record in records})
    _, _, _, section = pdf_styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=14 * mm,
        bottomMargin=16 * mm,
    )
    story = report_header_story(
        "รายงานสรุปผลผลิตและน้ำหนัก",
        f"ผลไม้ {selected_production_fruit_label(payload)} | ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)}",
        payload,
    )
    overview = Table(
        [
            ["จำนวนรายการ", "จำนวนพนักงาน", "น้ำหนักเข้า", "น้ำหนักออก", "คงเหลือ", "ยอดเงินรวม"],
            [
                report_number(len(records), 0),
                report_number(employee_count, 0),
                report_number(totals["incoming"]),
                report_number(totals["outgoing"]),
                report_number(totals["balance"]),
                money(totals["amount"]),
            ],
        ],
        colWidths=[35 * mm, 35 * mm, 42 * mm, 42 * mm, 42 * mm, 42 * mm],
    )
    set_pdf_table_style(overview, 0)
    story += [Paragraph("ภาพรวม", section), overview, Spacer(1, 7 * mm)]

    pile_table_rows = [["กอง", "น้ำหนักเข้า", "น้ำหนักออก", "คงเหลือ", "ยอดเงิน"]]
    for row in pile_rows:
        pile_table_rows.append([row["pile"], report_number(row["incoming"]), report_number(row["outgoing"]), report_number(row["balance"]), money(row["amount"])])
    if len(pile_table_rows) == 1:
        pile_table_rows.append(["-", "0", "0", "0", "0"])
    pile_table_rows.append(["รวม", report_number(totals["incoming"]), report_number(totals["outgoing"]), report_number(totals["balance"]), money(totals["amount"])])
    pile_table = Table(pile_table_rows, repeatRows=1, colWidths=[45 * mm, 48 * mm, 48 * mm, 48 * mm, 48 * mm])
    set_pdf_table_style(pile_table)
    story += [Paragraph("สรุปตามกอง", section), pile_table, Spacer(1, 7 * mm)]

    detail_rows = [["วันที่", "เวลา", "รหัส", "ชื่อพนักงาน", "สินค้า", "กอง", "น้ำหนักน้ำ", "น้ำหนักดอก", "ยอดเงิน"]]
    for record in records:
        detail_rows.append(
            [
                format_report_date(record.get("record_date") or record.get("date") or ""),
                record.get("record_time", ""),
                record.get("emp_code", ""),
                record.get("employee_name", ""),
                record.get("fruit_type", ""),
                record.get("pile_no") or record.get("pile", ""),
                report_number(record.get("water_weight", record.get("water", 0))),
                report_number(record.get("flower_weight", record.get("flower", 0))),
                money(record.get("total_amount", record.get("grand_total", 0))),
            ]
        )
    detail_table = Table(detail_rows, repeatRows=1, colWidths=[24 * mm, 18 * mm, 22 * mm, 45 * mm, 26 * mm, 14 * mm, 30 * mm, 30 * mm, 28 * mm])
    set_pdf_table_style(detail_table, 6)
    story += [Paragraph("รายละเอียดรายการ", section), detail_table]
    doc.build(story)
    return buffer.getvalue()


def selected_export_sections(payload: dict) -> dict:
    sections = payload.get("export_sections") or {}
    return {
        "overview": bool(sections.get("overview", True)),
        "piles": bool(sections.get("piles", True)),
        "details": bool(sections.get("details", True)),
    }


def selected_export_fields(payload: dict, section: str, definitions: list[tuple[str, str, object]]) -> list[tuple[str, str, object]]:
    show_standard_weights, show_grades = production_report_field_visibility(payload)
    definitions = [
        definition
        for definition in definitions
        if (show_standard_weights or definition[0] not in {"water", "flower"})
        and (show_grades or definition[0] != "grades")
    ]
    fields = payload.get("export_fields") or {}
    section_fields = fields.get(section) or {}
    if not section_fields:
        return definitions
    return [definition for definition in definitions if bool(section_fields.get(definition[0]))]


def employee_name_for_record(payload: dict, record: dict) -> str:
    employees = payload.get("employees") or []
    employee_id = str(record.get("employee_id") or "").strip()
    emp_code = str(record.get("emp_code") or "").strip()
    for employee in employees:
        registered_id = str(employee.get("id") or "").strip()
        registered_code = str(employee.get("emp_code") or "").strip()
        if (
            (employee_id and registered_id and registered_id == employee_id)
            or (emp_code and registered_code and registered_code == emp_code)
        ):
            return str(employee.get("fullname") or record.get("employee_name") or record.get("fullname") or "-")
    return str(record.get("employee_name") or record.get("fullname") or "-")


def production_summary_context(payload: dict) -> tuple[str, str, list[dict], list[dict], dict, int]:
    start_date, end_date = normalized_range(payload)
    records = filtered_production_records(payload)
    pile_rows = pile_summary_rows(records)
    totals = summary_totals(pile_rows)
    total_weight = totals["total_weight"]
    totals = {**totals, "total_weight": total_weight}
    employee_count = len(
        {record.get("employee_id") or record.get("emp_code") or record.get("employee_name") for record in records}
    )
    return start_date, end_date, records, pile_rows, totals, employee_count


def build_production_summary_excel(payload: dict) -> bytes:
    start_date, end_date, records, pile_rows, totals, employee_count = production_summary_context(payload)
    sections = selected_export_sections(payload)
    water_label, flower_label = production_report_weight_labels(payload)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "ภาพรวม"
    font_name = "Sarabun"
    dark_green = "075B44"
    brand_green = "0F8A55"
    mint = "D1FAE5"
    pale = "F8FAFC"
    line_color = "D8E2EA"
    thin = Side(style="thin", color=line_color)
    medium = Side(style="medium", color=dark_green)
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def setup_summary_sheet(sheet, subtitle: str, last_column: str, freeze_at: str | None = None):
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells("A1:B3")
        sheet["A1"] = "PF"
        sheet["A1"].fill = PatternFill("solid", fgColor=brand_green)
        sheet["A1"].font = Font(name=font_name, bold=True, size=25, color="FFFFFF")
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(f"C1:{last_column}2")
        sheet["C1"] = COMPANY_NAME
        sheet["C1"].fill = PatternFill("solid", fgColor=dark_green)
        sheet["C1"].font = Font(name=font_name, bold=True, size=20, color="FFFFFF")
        sheet["C1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet.merge_cells(f"C3:{last_column}3")
        sheet["C3"] = subtitle
        sheet["C3"].fill = PatternFill("solid", fgColor=dark_green)
        sheet["C3"].font = Font(name=font_name, size=10, color=mint)
        sheet["C3"].alignment = Alignment(horizontal="left", vertical="center")
        add_excel_logo(sheet, "A1")
        sheet.freeze_panes = freeze_at
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.orientation = "landscape"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.oddFooter.left.text = SYSTEM_NAME
        sheet.oddFooter.center.text = "รายงานสรุปข้อมูลหลัก"
        sheet.oddFooter.right.text = "หน้า &P จาก &N"
        for row_index in (1, 2, 3):
            sheet.row_dimensions[row_index].height = 23

    setup_summary_sheet(overview, "รายงานสรุปข้อมูลหลัก • Production Summary", "L", "A13")
    overview.merge_cells("A5:L5")
    overview["A5"] = (
        f"ผลไม้ {selected_production_fruit_label(payload)}"
        f"  |  ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)}"
        f"  |  {export_meta_text(payload)}"
    )
    overview["A5"].fill = PatternFill("solid", fgColor=mint)
    overview["A5"].font = Font(name=font_name, bold=True, size=10, color=dark_green)
    overview["A5"].alignment = Alignment(vertical="center", wrap_text=True)
    overview.row_dimensions[5].height = 28
    overview_fields = selected_export_fields(
        payload,
        "overview",
        [
            ("totalWeight", "น้ำหนักรวมทั้งหมด (กก.)", totals["total_weight"]),
            ("water", f"{water_label} (กก.)", totals["incoming"]),
            ("flower", f"{flower_label} (กก.)", totals["outgoing"]),
            ("grades", "น้ำหนักทุเรียน", grade_totals_text(totals.get("grades"))),
            ("amount", "ยอดเงินรวม", totals["amount"]),
            ("employees", "พนักงานที่มีรายการ", employee_count),
            ("records", "จำนวนรายการ", len(records)),
        ],
    )
    if sections["overview"] and overview_fields:
        metric_values = {key: (label, value) for key, label, value in overview_fields}
        cards = [
            ("A7:C7", "A8:C9", metric_values.get("totalWeight", ("น้ำหนักรวม", totals["total_weight"])), "ECFDF5", "047857", '#,##0.00 "กก."'),
            ("D7:F7", "D8:F9", metric_values.get("amount", ("ยอดเงินรวม", totals["amount"])), "FFF7ED", "C2410C", '#,##0.00 "บาท"'),
            ("G7:I7", "G8:I9", metric_values.get("employees", ("พนักงาน", employee_count)), "EFF6FF", "1D4ED8", '0 "คน"'),
            ("J7:L7", "J8:L9", metric_values.get("records", ("จำนวนรายการ", len(records))), "F0FDF4", "15803D", '0 "รายการ"'),
        ]
        for label_range, value_range, (label, value), fill_color, font_color, number_format in cards:
            overview.merge_cells(label_range)
            overview.merge_cells(value_range)
            label_cell = overview[label_range.split(":")[0]]
            value_cell = overview[value_range.split(":")[0]]
            label_cell.value = label
            value_cell.value = value
            for row in overview[f"{label_range.split(':')[0]}:{value_range.split(':')[1]}"]:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.border = table_border
            label_cell.font = Font(name=font_name, bold=True, size=10, color="64748B")
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.font = Font(name=font_name, bold=True, size=17, color=font_color)
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.number_format = number_format

    if sections["piles"]:
        overview.merge_cells("A11:L11")
        overview["A11"] = "สรุปตามกอง"
        overview["A11"].fill = PatternFill("solid", fgColor=dark_green)
        overview["A11"].font = Font(name=font_name, bold=True, size=12, color="FFFFFF")
        overview["A11"].alignment = Alignment(vertical="center")

    if sections["piles"]:
        pile_defs = selected_export_fields(
            payload,
            "piles",
            [
                ("pile", "กอง", lambda row: row["pile"]),
                ("water", f"{water_label} (กก.)", lambda row: row["incoming"]),
                ("flower", f"{flower_label} (กก.)", lambda row: row["outgoing"]),
                ("grades", "น้ำหนักทุเรียน", lambda row: grade_totals_text(row.get("grades"))),
                ("total", "รวม (กก.)", lambda row: row.get("total_weight", row["incoming"] + row["outgoing"])),
                ("amount", "รวมเงิน", lambda row: row["amount"]),
            ],
        )
        if pile_defs:
            overview_pile_header = 13
            for column, (_, label, _) in enumerate(pile_defs, 1):
                overview.cell(overview_pile_header, column, label)
            for row in pile_rows:
                overview.append([getter(row) for _, _, getter in pile_defs])
            total_row = {
                "pile": "รวม",
                "incoming": totals["incoming"],
                "outgoing": totals["outgoing"],
                "grades": totals.get("grades", {}),
                "total_weight": totals["total_weight"],
                "amount": totals["amount"],
            }
            overview.append([getter(total_row) for _, _, getter in pile_defs])
            style_excel_report_sheet(overview, [overview_pile_header], [18] * 12)
            for cell in overview[overview.max_row]:
                if cell.value is not None:
                    cell.fill = PatternFill("solid", fgColor=mint)
                    cell.font = Font(name=font_name, bold=True, color=dark_green)
                    cell.border = Border(top=medium, bottom=medium)

            pile_sheet = workbook.create_sheet("สรุปตามกอง")
            setup_summary_sheet(pile_sheet, "สรุปผลผลิตแยกตามกอง", get_column_letter(max(3, len(pile_defs))), "A6")
            pile_sheet.append([])
            pile_sheet.append([label for _, label, _ in pile_defs])
            for row in pile_rows:
                pile_sheet.append([getter(row) for _, _, getter in pile_defs])
            pile_sheet.append([getter(total_row) for _, _, getter in pile_defs])
            style_excel_report_sheet(pile_sheet, [5], [20] * max(1, len(pile_defs)))
            for cell in pile_sheet[pile_sheet.max_row]:
                cell.fill = PatternFill("solid", fgColor=mint)
                cell.font = Font(name=font_name, bold=True, color=dark_green)

    if sections["details"]:
        detail_defs = selected_export_fields(
            payload,
            "details",
            [
                ("date", "วันที่", lambda record: record.get("record_date") or record.get("date") or ""),
                ("time", "เวลา", lambda record: record.get("record_time", "")),
                ("empCode", "รหัสพนักงาน", lambda record: record.get("emp_code", "")),
                ("employeeName", "ชื่อพนักงาน", lambda record: employee_name_for_record(payload, record)),
                ("pile", "กอง", lambda record: production_pile_number(record) or "-"),
                ("water", f"{water_label} (กก.)", lambda record: safe_float(record.get("water_weight", record.get("water", 0)))),
                ("flower", f"{flower_label} (กก.)", lambda record: safe_float(record.get("flower_weight", record.get("flower", 0)))),
                ("grades", "น้ำหนักทุเรียน", lambda record: production_grade_text(record) if (record.get("fruit_type") or "mangosteen") == "durian" else "-"),
                ("total", "น้ำหนักรวม (กก.)", lambda record: production_total_weight(record)),
                ("amount", "รวมเงิน", lambda record: safe_float(record.get("total_amount", record.get("grand_total", 0)))),
                ("createdBy", "ผู้บันทึก", lambda record: record.get("created_by", "")),
            ],
        )
        if detail_defs:
            detail_sheet = workbook.create_sheet("รายละเอียด")
            last_column = get_column_letter(max(3, len(detail_defs)))
            setup_summary_sheet(detail_sheet, "รายละเอียดผลผลิต • ชื่อจากทะเบียนพนักงาน", last_column, "A6")
            detail_sheet.append([])
            detail_sheet.append([label for _, label, _ in detail_defs])
            for record in records:
                detail_sheet.append([getter(record) for _, _, getter in detail_defs])
            detail_widths = {
                "date": 15, "time": 11, "empCode": 17, "employeeName": 31, "pile": 10,
                "water": 16, "flower": 16, "grades": 27, "total": 16, "amount": 17, "createdBy": 18,
            }
            style_excel_report_sheet(
                detail_sheet,
                [5],
                [detail_widths.get(key, 16) for key, _, _ in detail_defs],
            )
            for column, (key, _, _) in enumerate(detail_defs, 1):
                if key == "empCode":
                    for row_index in range(6, detail_sheet.max_row + 1):
                        detail_sheet.cell(row_index, column).number_format = "@"
                if key == "employeeName":
                    for row_index in range(6, detail_sheet.max_row + 1):
                        cell = detail_sheet.cell(row_index, column)
                        cell.font = Font(name=font_name, bold=True, size=10, color=dark_green)
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            detail_sheet.print_title_rows = "1:5"

    for index, width in enumerate([18] * 12, 1):
        overview.column_dimensions[get_column_letter(index)].width = width
    overview.print_area = f"A1:L{overview.max_row}"
    overview.print_title_rows = "1:13" if sections["piles"] else "1:5"
    overview.sheet_view.zoomScale = 80

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_production_summary_pdf(payload: dict) -> bytes:
    start_date, end_date, records, pile_rows, totals, employee_count = production_summary_context(payload)
    sections = selected_export_sections(payload)
    water_label, flower_label = production_report_weight_labels(payload)
    _, _, pdf_normal, section = pdf_styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=31 * mm,
        bottomMargin=18 * mm,
    )
    story = []
    dark_green = colors.HexColor("#075B44")
    brand_green = colors.HexColor("#0F8A55")
    mint = colors.HexColor("#D1FAE5")
    pale = colors.HexColor("#F8FAFC")
    line_color = colors.HexColor("#D8E2EA")
    meta = Table(
        [[
            Paragraph(f"<b>ผลไม้</b><br/>{xml_escape(selected_production_fruit_label(payload))}", pdf_normal),
            Paragraph(f"<b>ช่วงรายงาน</b><br/>{format_report_date(start_date)} - {format_report_date(end_date)}", pdf_normal),
            Paragraph(f"<b>ข้อมูลการพิมพ์</b><br/>{xml_escape(export_meta_text(payload))}", pdf_normal),
        ]],
        colWidths=[64 * mm, 78 * mm, 127 * mm],
    )
    meta.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), mint),
        ("GRID", (0, 0), (-1, -1), 0.35, line_color),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.extend([meta, Spacer(1, 5 * mm)])

    overview_fields = selected_export_fields(
        payload,
        "overview",
        [
            ("totalWeight", "น้ำหนักรวมทั้งหมด (กก.)", report_number(totals["total_weight"])),
            ("water", f"{water_label} (กก.)", report_number(totals["incoming"])),
            ("flower", f"{flower_label} (กก.)", report_number(totals["outgoing"])),
            ("grades", "น้ำหนักทุเรียน", grade_totals_text(totals.get("grades"))),
            ("amount", "ยอดเงินรวม", money(totals["amount"])),
            ("employees", "พนักงานที่มีรายการ", report_number(employee_count, 0)),
            ("records", "จำนวนรายการ", report_number(len(records), 0)),
        ],
    )
    if sections["overview"] and overview_fields:
        field_map = {key: (label, value) for key, label, value in overview_fields}
        preferred = ["totalWeight", "amount", "employees", "records"]
        card_fields = [field_map[key] for key in preferred if key in field_map]
        card_fields.extend(
            (label, value)
            for key, label, value in overview_fields
            if key not in preferred and (label, value) not in card_fields
        )
        card_fields = card_fields[:4]
        card_colors = [
            ("#ECFDF5", "#047857"),
            ("#FFF7ED", "#C2410C"),
            ("#EFF6FF", "#1D4ED8"),
            ("#F0FDF4", "#15803D"),
        ]
        metric_style = getSampleStyleSheet()["BodyText"]
        metric_style.fontName = THAI_FONT
        metric_style.fontSize = 8
        metric_style.leading = 20
        metric_cells = []
        for index, (label, value) in enumerate(card_fields):
            metric_cells.append(Paragraph(
                f"<font color='#64748B'>{xml_escape(str(label))}</font>"
                f"<br/><font size='15' color='{card_colors[index][1]}'><b>{xml_escape(str(value))}</b></font>",
                metric_style,
            ))
        if metric_cells:
            metric_table = Table(
                [metric_cells],
                colWidths=[(269 / len(metric_cells)) * mm] * len(metric_cells),
                rowHeights=[23 * mm],
            )
            metric_commands = [
                ("GRID", (0, 0), (-1, -1), 0.35, line_color),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
            for index, (fill_color, _) in enumerate(card_colors[:len(metric_cells)]):
                metric_commands.append(("BACKGROUND", (index, 0), (index, 0), colors.HexColor(fill_color)))
            metric_table.setStyle(TableStyle(metric_commands))
            story.extend([metric_table, Spacer(1, 5 * mm)])

    if sections["piles"]:
        pile_defs = selected_export_fields(
            payload,
            "piles",
            [
                ("pile", "กอง", lambda row: row["pile"]),
                ("water", f"{water_label} (กก.)", lambda row: report_number(row["incoming"])),
                ("flower", f"{flower_label} (กก.)", lambda row: report_number(row["outgoing"])),
                ("grades", "น้ำหนักทุเรียน", lambda row: grade_totals_text(row.get("grades"))),
                ("total", "รวม (กก.)", lambda row: report_number(row.get("total_weight", row["incoming"] + row["outgoing"]))),
                ("amount", "รวมเงิน", lambda row: money(row["amount"])),
            ],
        )
        if pile_defs:
            pile_table_rows = [[label for _, label, _ in pile_defs]]
            for row in pile_rows:
                pile_table_rows.append([getter(row) for _, _, getter in pile_defs])
            total_row = {"pile": "รวม", "incoming": totals["incoming"], "outgoing": totals["outgoing"], "grades": totals.get("grades", {}), "total_weight": totals["total_weight"], "amount": totals["amount"]}
            pile_table_rows.append([getter(total_row) for _, _, getter in pile_defs])
            col_width = (267 / len(pile_defs)) * mm
            pile_table = Table(pile_table_rows, repeatRows=1, colWidths=[col_width] * len(pile_defs))
            set_pdf_table_style(pile_table, 1)
            pile_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), brand_green),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, pale]),
                ("BACKGROUND", (0, -1), (-1, -1), mint),
                ("TEXTCOLOR", (0, -1), (-1, -1), dark_green),
                ("FONTNAME", (0, -1), (-1, -1), THAI_FONT_BOLD),
                ("LINEABOVE", (0, -1), (-1, -1), 1, dark_green),
                ("GRID", (0, 0), (-1, -1), 0.35, line_color),
            ]))
            story += [Paragraph("สรุปตามกอง", section), pile_table, Spacer(1, 4 * mm)]

    if sections["details"]:
        detail_defs = selected_export_fields(
            payload,
            "details",
            [
                ("date", "วันที่", lambda record: format_report_date(record.get("record_date") or record.get("date") or "")),
                ("time", "เวลา", lambda record: record.get("record_time", "")),
                ("empCode", "รหัสพนักงาน", lambda record: record.get("emp_code", "")),
                ("employeeName", "ชื่อพนักงาน", lambda record: employee_name_for_record(payload, record)),
                ("pile", "กอง", lambda record: production_pile_number(record) or "-"),
                ("water", f"{water_label} (กก.)", lambda record: report_number(record.get("water_weight", record.get("water", 0)))),
                ("flower", f"{flower_label} (กก.)", lambda record: report_number(record.get("flower_weight", record.get("flower", 0)))),
                ("grades", "น้ำหนักทุเรียน", lambda record: production_grade_text(record) if (record.get("fruit_type") or "mangosteen") == "durian" else "-"),
                ("total", "น้ำหนักรวม (กก.)", lambda record: report_number(production_total_weight(record))),
                ("amount", "รวมเงิน", lambda record: money(record.get("total_amount", record.get("grand_total", 0)))),
                ("createdBy", "ผู้บันทึก", lambda record: record.get("created_by", "")),
            ],
        )
        if detail_defs:
            if sections["overview"] or sections["piles"]:
                story.append(PageBreak())
                logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
                logo_flowable = (
                    Image(str(logo_path), width=15 * mm, height=15 * mm)
                    if logo_path.exists()
                    else Paragraph("<b>PF</b>", pdf_normal)
                )
                detail_page_header = Table(
                    [[
                        logo_flowable,
                        Paragraph(
                            f"<font color='#FFFFFF' size='15'><b>{xml_escape(COMPANY_NAME)}</b></font>"
                            "<br/><font color='#D1FAE5'>รายงานสรุปข้อมูลหลัก - รายละเอียดรายการ</font>",
                            pdf_normal,
                        ),
                    ]],
                    colWidths=[22 * mm, 247 * mm],
                    rowHeights=[20 * mm],
                )
                detail_page_header.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), dark_green),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]))
                story.extend([detail_page_header, Spacer(1, 4 * mm)])
            detail_rows = [[label for _, label, _ in detail_defs]]
            for record in records:
                detail_rows.append([getter(record) for _, _, getter in detail_defs])
            width_weights = {
                "date": 25, "time": 17, "empCode": 24, "employeeName": 53, "pile": 14,
                "water": 26, "flower": 26, "grades": 45, "total": 28, "amount": 31, "createdBy": 29,
            }
            raw_widths = [width_weights.get(key, 24) for key, _, _ in detail_defs]
            width_scale = 267 / sum(raw_widths)
            detail_widths = [width * width_scale * mm for width in raw_widths]
            name_column = next(
                (index for index, (key, _, _) in enumerate(detail_defs) if key == "employeeName"),
                None,
            )
            if name_column is not None:
                for row in detail_rows[1:]:
                    row[name_column] = Paragraph(
                        f"<b><font color='#075B44'>{xml_escape(str(row[name_column] or '-'))}</font></b>",
                        pdf_normal,
                    )
            detail_table = Table(detail_rows, repeatRows=1, colWidths=detail_widths)
            set_pdf_table_style(detail_table, max(1, len(detail_defs) - 3))
            detail_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), brand_green),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, pale]),
                ("GRID", (0, 0), (-1, -1), 0.35, line_color),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            if name_column is not None:
                detail_table.setStyle(TableStyle([
                    ("ALIGN", (name_column, 1), (name_column, -1), "LEFT"),
                ]))
            story += [Paragraph("รายละเอียดรายการ - ชื่อพนักงานจากทะเบียน", section), detail_table]

    if not any(sections.values()):
        story.append(Paragraph("ไม่มีฟิลด์ที่เลือกสำหรับรายงานนี้", section))

    def draw_page(canvas_obj, document):
        page_width, page_height = landscape(A4)
        canvas_obj.saveState()
        canvas_obj.setFillColor(dark_green)
        canvas_obj.rect(0, page_height - 22 * mm, page_width, 22 * mm, fill=1, stroke=0)
        logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
        if logo_path.exists():
            try:
                canvas_obj.drawImage(
                    str(logo_path), 12 * mm, page_height - 19 * mm,
                    width=15 * mm, height=15 * mm, preserveAspectRatio=True, mask="auto",
                )
            except Exception:
                pass
        canvas_obj.setFont(THAI_FONT_BOLD, 15)
        canvas_obj.setFillColor(colors.white)
        canvas_obj.drawString(31 * mm, page_height - 10 * mm, COMPANY_NAME)
        canvas_obj.setFont(THAI_FONT, 8)
        canvas_obj.setFillColor(mint)
        canvas_obj.drawString(31 * mm, page_height - 16 * mm, "รายงานสรุปข้อมูลหลัก")
        canvas_obj.setStrokeColor(brand_green)
        canvas_obj.line(14 * mm, 12 * mm, page_width - 14 * mm, 12 * mm)
        canvas_obj.setFont(THAI_FONT, 7.5)
        canvas_obj.setFillColor(colors.HexColor("#475467"))
        canvas_obj.drawString(
            14 * mm, 7 * mm,
            f"ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)} | {export_meta_text(payload)}",
        )
        canvas_obj.drawRightString(page_width - 14 * mm, 7 * mm, f"หน้า {document.page}")
        canvas_obj.restoreState()

    doc.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
    return buffer.getvalue()


def employee_lookup_maps(payload: dict) -> tuple[dict[str, dict], dict[str, dict]]:
    employees = payload.get("employees") or []
    by_id = {str(employee.get("id")): employee for employee in employees}
    by_code = {str(employee.get("emp_code")): employee for employee in employees}
    return by_id, by_code


GROUP_REPORT_PAY_GROUPS = ["เหมาโรงงาน", "เหมา(นนท์)", "เหมาปุ้ย"]
PRODUCTION_WITHHOLDING_TAX_RATE = 0.03
DEFAULT_PRODUCTION_WITHHOLDING_TAX_GROUPS = {"เหมา(นนท์)", "เหมาปุ้ย"}
GROUP_REPORT_LEGACY_MAP = {
    "กลุ่ม A": "เหมาโรงงาน",
    "กลุ่ม B": "เหมา(นนท์)",
    "กลุ่ม C": "เหมาปุ้ย",
    "กลุ่ม D": "เหมาปุ้ย",
}


def normalize_group_report_pay_group(value: str) -> str:
    pay_group = str(value or "").strip()
    if pay_group in GROUP_REPORT_LEGACY_MAP:
        return GROUP_REPORT_LEGACY_MAP[pay_group]
    return pay_group


def production_withholding_tax_groups(payload: dict | None = None) -> set[str]:
    groups = (payload or {}).get("withholding_tax_groups")
    if not isinstance(groups, list):
        return set(DEFAULT_PRODUCTION_WITHHOLDING_TAX_GROUPS)
    return {normalize_group_report_pay_group(group) for group in groups if normalize_group_report_pay_group(group)}


def production_withholding_tax(pay_group: str, amount: float | int | str | None, withholding_groups: set[str] | None = None) -> float:
    groups = withholding_groups if withholding_groups is not None else DEFAULT_PRODUCTION_WITHHOLDING_TAX_GROUPS
    if normalize_group_report_pay_group(pay_group) not in groups:
        return 0.0
    return round(max(0.0, safe_float(amount)) * PRODUCTION_WITHHOLDING_TAX_RATE + 1e-9, 2)


def production_fruit_id(record: dict) -> str:
    return record.get("fruit_type") or "mangosteen"


def production_fruit_label(payload: dict, fruit_id: str) -> str:
    if fruit_id == "all":
        return "ทั้งหมด"
    labels = {
        "mangosteen": "มังคุด",
        "durian": "ทุเรียน",
        "mango": "มะม่วง",
    }
    return (payload.get("fruit_labels") or {}).get(fruit_id) or labels.get(fruit_id, fruit_id)


def group_report_records(payload: dict) -> list[dict]:
    start_date, end_date = normalized_range(payload)
    selected_group = payload.get("pay_group") or "all"
    selected_fruit = payload.get("fruit_type") or "all"
    by_id, by_code = employee_lookup_maps(payload)
    rows = []
    adjusted_employee_keys = set()
    for record in payload.get("production_records", []):
        record_date = record.get("record_date") or record.get("date") or ""
        if not (start_date <= record_date <= end_date):
            continue
        employee = by_id.get(str(record.get("employee_id") or "")) or by_code.get(str(record.get("emp_code") or ""))
        pay_group = normalize_group_report_pay_group((employee or {}).get("pay_group") or record.get("pay_group"))
        fruit_id = production_fruit_id(record)
        if not pay_group:
            continue
        if selected_group != "all" and pay_group != selected_group:
            continue
        if selected_fruit != "all" and fruit_id != selected_fruit:
            continue
        employee_key = str((employee or {}).get("id") or record.get("employee_id") or record.get("emp_code") or "")
        record_deduction = 0
        record_bonus = 0
        if employee_key and employee_key not in adjusted_employee_keys:
            adjusted_employee_keys.add(employee_key)
            record_deduction = deduction_total(
                deduction_records_for(
                    payload,
                    "production",
                    (employee or {}).get("id") or record.get("employee_id"),
                    (employee or {}).get("emp_code") or record.get("emp_code"),
                    start_date,
                    end_date,
                )
            )
            record_bonus = deduction_total(
                bonus_records_for(
                    payload,
                    "production",
                    (employee or {}).get("id") or record.get("employee_id"),
                    (employee or {}).get("emp_code") or record.get("emp_code"),
                    start_date,
                    end_date,
                )
            )
        rows.append(
            {
                **record,
                "record_date": record_date,
                "employee": employee or {},
                "pay_group": pay_group,
                "fruit_type": fruit_id,
                "fruit_label": production_fruit_label(payload, fruit_id),
                "employee_name": record.get("employee_name") or (employee or {}).get("fullname") or "",
                "deduction_amount": record_deduction,
                "bonus_amount": record_bonus,
            }
        )
    return sorted(rows, key=lambda item: (item["pay_group"], item["fruit_label"], item["record_date"], str(item.get("emp_code", ""))))


def summarize_group_report(records: list[dict], mode: str = "group", withholding_groups: set[str] | None = None) -> list[dict]:
    summaries: dict[str, dict] = {}
    for record in records:
        key = record["pay_group"] if mode == "group" else f"{record['pay_group']}__{record['fruit_type']}"
        row = summaries.setdefault(
            key,
            {
                "pay_group": record["pay_group"],
                "fruit_label": record["fruit_label"] if mode == "fruit" else "ทั้งหมด",
                "employees": set(),
                "records": 0,
                "water": 0.0,
                "flower": 0.0,
                "grades": {grade: 0.0 for grade in DURIAN_GRADES},
                "total": 0.0,
                "amount": 0.0,
                "deduction_amount": 0.0,
                "withholding_tax_amount": 0.0,
                "bonus_amount": 0.0,
                "net_amount": 0.0,
            },
        )
        water = safe_float(record.get("water_weight", record.get("water", 0)))
        flower = safe_float(record.get("flower_weight", record.get("flower", 0)))
        row["employees"].add(record.get("employee_id") or record.get("emp_code") or record.get("employee_name") or "")
        row["records"] += 1
        row["water"] += water
        row["flower"] += flower
        grade_weights = production_grade_weights(record)
        for grade in DURIAN_GRADES:
            row["grades"][grade] += grade_weights[grade]
        row["total"] += production_total_weight(record)
        row["amount"] += safe_float(record.get("total_amount", record.get("grand_total", 0)))
        row["deduction_amount"] += safe_float(record.get("deduction_amount"))
        row["bonus_amount"] += safe_float(record.get("bonus_amount"))
        row["withholding_tax_amount"] = production_withholding_tax(row["pay_group"], row["amount"], withholding_groups)
        row["net_amount"] = max(
            0,
            row["amount"] + row["bonus_amount"] - row["deduction_amount"] - row["withholding_tax_amount"],
        )
    return sorted(summaries.values(), key=lambda item: (item["pay_group"], item["fruit_label"]))


def group_report_employee_rows(records: list[dict], withholding_groups: set[str] | None = None) -> list[dict]:
    rows: dict[str, dict] = {}
    for record in records:
        key = str(record.get("employee_id") or record.get("emp_code") or record.get("employee_name") or "")
        row = rows.setdefault(
            key,
            {
                "pay_group": record["pay_group"],
                "emp_code": record.get("emp_code") or record.get("employee", {}).get("emp_code") or "-",
                "fullname": record.get("employee_name") or record.get("employee", {}).get("fullname") or "-",
                "records": 0,
                "water": 0.0,
                "flower": 0.0,
                "grades": {grade: 0.0 for grade in DURIAN_GRADES},
                "total": 0.0,
                "amount": 0.0,
                "deduction_amount": 0.0,
                "withholding_tax_amount": 0.0,
                "bonus_amount": 0.0,
                "net_amount": 0.0,
            },
        )
        water = safe_float(record.get("water_weight", record.get("water", 0)))
        flower = safe_float(record.get("flower_weight", record.get("flower", 0)))
        row["records"] += 1
        row["water"] += water
        row["flower"] += flower
        grade_weights = production_grade_weights(record)
        for grade in DURIAN_GRADES:
            row["grades"][grade] += grade_weights[grade]
        row["total"] += production_total_weight(record)
        row["amount"] += safe_float(record.get("total_amount", record.get("grand_total", 0)))
        row["deduction_amount"] += safe_float(record.get("deduction_amount"))
        row["bonus_amount"] += safe_float(record.get("bonus_amount"))
        row["withholding_tax_amount"] = production_withholding_tax(row["pay_group"], row["amount"], withholding_groups)
        row["net_amount"] = max(
            0,
            row["amount"] + row["bonus_amount"] - row["deduction_amount"] - row["withholding_tax_amount"],
        )
    return sorted(rows.values(), key=lambda item: (item["pay_group"], item["emp_code"]))


def group_report_options(payload: dict) -> dict:
    options = payload.get("export_options") or {}
    return {
        "summary": bool(options.get("summary", True)),
        "fruit": bool(options.get("fruit", True)),
        "employees": bool(options.get("employees", True)),
        "details": bool(options.get("details", False)),
    }


def build_group_report_excel(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    records = group_report_records(payload)
    options = group_report_options(payload)
    withholding_groups = production_withholding_tax_groups(payload)
    group_rows = summarize_group_report(records, "group", withholding_groups)
    fruit_rows = summarize_group_report(records, "fruit", withholding_groups)
    employee_rows = group_report_employee_rows(records, withholding_groups)
    show_standard_weights, show_durian_weight = production_report_field_visibility(payload)
    water_label, flower_label = production_report_weight_labels(payload)

    def weight_headers() -> list[str]:
        return (
            ([water_label, flower_label] if show_standard_weights else [])
            + (["น้ำหนักทุเรียน"] if show_durian_weight else [])
        )

    def summary_weight_values(row: dict) -> list:
        return (
            ([row["water"], row["flower"]] if show_standard_weights else [])
            + ([sum(safe_float((row.get("grades") or {}).get(grade)) for grade in DURIAN_GRADES)] if show_durian_weight else [])
        )

    def detail_weight_values(record: dict) -> list:
        is_durian = production_fruit_id(record) == "durian"
        return (
            ([
                "-" if is_durian else safe_float(record.get("water_weight", record.get("water", 0))),
                "-" if is_durian else safe_float(record.get("flower_weight", record.get("flower", 0))),
            ] if show_standard_weights else [])
            + ([production_total_weight(record) if is_durian else "-"] if show_durian_weight else [])
        )

    def excel_widths(headers: list[str]) -> list[int]:
        preferred = {
            "กลุ่ม": 20,
            "รหัส": 14,
            "ชื่อพนักงาน": 32,
            "ผลไม้": 15,
            "วันที่": 14,
            "กอง": 9,
            "จำนวนคน": 11,
            "รายการ": 11,
            "น้ำหนักน้ำ": 15,
            "น้ำหนักดอก": 15,
            "น้ำหนักทุเรียน": 16,
            "น้ำหนักรวม": 15,
            "รวม": 14,
            "รวมเงิน": 16,
            "เบี้ยขยัน": 14,
            "หักทั่วไป": 14,
            "หัก ณ ที่จ่าย 3%": 18,
            "สุทธิ": 16,
        }
        return [preferred.get(str(header), 15) for header in headers]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Group Report"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = COMPANY_NAME
    sheet["A1"].font = Font(name="Sarabun", bold=True, size=18, color="0F7A3D")
    sheet.merge_cells("A2:H2")
    sheet["A2"] = "รายงานแบบกลุ่ม"
    sheet["A2"].font = Font(name="Sarabun", bold=True, size=16, color="111827")
    sheet.merge_cells("A3:H3")
    sheet["A3"] = f"ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)} | กลุ่ม {payload.get('group_label', 'ทุกกลุ่ม')} | ผลไม้ {payload.get('fruit_label', 'ทั้งหมด')}"
    sheet.merge_cells("A4:H4")
    sheet["A4"] = export_meta_text(payload)
    sheet.append([])
    style_excel_report_sheet(sheet, [1], [18, 18, 18, 18, 18, 18, 18, 18])

    if options["summary"]:
        summary = workbook.create_sheet("Summary By Group")
        summary_headers = ["กลุ่ม", "จำนวนคน", "รายการ", *weight_headers(), "รวม", "รวมเงิน", "เบี้ยขยัน", "หักทั่วไป", "หัก ณ ที่จ่าย 3%", "สุทธิ"]
        summary.append(summary_headers)
        for row in group_rows:
            summary.append([row["pay_group"], len(row["employees"]), row["records"], *summary_weight_values(row), row["total"], row["amount"], row.get("bonus_amount", 0), row.get("deduction_amount", 0), row.get("withholding_tax_amount", 0), row.get("net_amount", row["amount"])])
        style_excel_report_sheet(summary, [1], excel_widths(summary_headers))

    if options["fruit"]:
        fruit = workbook.create_sheet("Group By Fruit")
        fruit_headers = ["กลุ่ม", "ผลไม้", "จำนวนคน", "รายการ", *weight_headers(), "รวม", "รวมเงิน", "เบี้ยขยัน", "หักทั่วไป", "หัก ณ ที่จ่าย 3%", "สุทธิ"]
        fruit.append(fruit_headers)
        for row in fruit_rows:
            fruit.append([row["pay_group"], row["fruit_label"], len(row["employees"]), row["records"], *summary_weight_values(row), row["total"], row["amount"], row.get("bonus_amount", 0), row.get("deduction_amount", 0), row.get("withholding_tax_amount", 0), row.get("net_amount", row["amount"])])
        style_excel_report_sheet(fruit, [1], excel_widths(fruit_headers))

    if options["employees"]:
        employees = workbook.create_sheet("Employees")
        employee_headers = ["กลุ่ม", "รหัส", "ชื่อพนักงาน", "รายการ", *weight_headers(), "รวม", "รวมเงิน", "เบี้ยขยัน", "หักทั่วไป", "หัก ณ ที่จ่าย 3%", "สุทธิ"]
        employees.append(employee_headers)
        for row in employee_rows:
            employees.append([row["pay_group"], row["emp_code"], row["fullname"], row["records"], *summary_weight_values(row), row["total"], row["amount"], row.get("bonus_amount", 0), row.get("deduction_amount", 0), row.get("withholding_tax_amount", 0), row.get("net_amount", row["amount"])])
        style_excel_report_sheet(employees, [1], excel_widths(employee_headers))

    if options["details"]:
        details = workbook.create_sheet("Details")
        detail_headers = ["วันที่", "กลุ่ม", "ผลไม้", "รหัส", "ชื่อพนักงาน", "กอง", *weight_headers(), "น้ำหนักรวม", "รวมเงิน"]
        details.append(detail_headers)
        for record in records:
            details.append([record["record_date"], record["pay_group"], record["fruit_label"], record.get("emp_code", ""), record.get("employee_name", ""), record.get("pile_no") or record.get("pile", ""), *detail_weight_values(record), production_total_weight(record), safe_float(record.get("total_amount", record.get("grand_total", 0)))])
        style_excel_report_sheet(details, [1], excel_widths(detail_headers))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_group_report_pdf(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    records = group_report_records(payload)
    options = group_report_options(payload)
    withholding_groups = production_withholding_tax_groups(payload)
    group_rows = summarize_group_report(records, "group", withholding_groups)
    fruit_rows = summarize_group_report(records, "fruit", withholding_groups)
    employee_rows = group_report_employee_rows(records, withholding_groups)
    show_standard_weights, show_durian_weight = production_report_field_visibility(payload)
    water_label, flower_label = production_report_weight_labels(payload)
    _, _, pdf_normal, section = pdf_styles()
    grade_style = pdf_normal.clone("GroupReportGradeCell")
    grade_style.fontSize = 6
    grade_style.leading = 7
    grade_style.alignment = 1
    cell_style = pdf_normal.clone("GroupReportCell")
    cell_style.fontSize = 6.5
    cell_style.leading = 8
    cell_style.wordWrap = "CJK"
    header_style = cell_style.clone("GroupReportHeaderCell")
    header_style.fontName = THAI_FONT_BOLD
    header_style.textColor = colors.white
    header_style.alignment = 1

    def grade_cell(weights: dict | None):
        values = weights or {}
        total = sum(safe_float(values.get(grade)) for grade in DURIAN_GRADES)
        if total <= 0:
            return "-"
        return Paragraph(report_number(total), grade_style)

    def weight_headers() -> list[str]:
        fruit_id = selected_production_fruit(payload)
        short_standard_headers = {
            "all": ["ช่อง 1", "ช่อง 2"],
            "mango": ["ฝา", "หั่นเต๋า"],
        }.get(fruit_id, ["น้ำ", "ดอก"])
        return (
            (short_standard_headers if show_standard_weights else [])
            + (["ทุเรียน"] if show_durian_weight else [])
        )

    def summary_weight_values(row: dict) -> list:
        return (
            ([report_number(row["water"]), report_number(row["flower"])] if show_standard_weights else [])
            + ([grade_cell(row.get("grades"))] if show_durian_weight else [])
        )

    def detail_weight_values(record: dict) -> list:
        is_durian = production_fruit_id(record) == "durian"
        return (
            ([
                "-" if is_durian else report_number(record.get("water_weight", record.get("water", 0))),
                "-" if is_durian else report_number(record.get("flower_weight", record.get("flower", 0))),
            ] if show_standard_weights else [])
            + ([production_grade_text(record) if is_durian else "-"] if show_durian_weight else [])
        )
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=15 * mm, rightMargin=15 * mm, topMargin=14 * mm, bottomMargin=16 * mm)
    story = report_header_story(
        "รายงานแบบกลุ่ม",
        f"ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)} | กลุ่ม {payload.get('group_label', 'ทุกกลุ่ม')} | ผลไม้ {payload.get('fruit_label', 'ทั้งหมด')}",
        payload,
    )

    def group_column_widths(headers: list[str]) -> list[float]:
        preferred = {
            "กลุ่ม": 25,
            "รหัส": 17,
            "ชื่อ": 50,
            "ผลไม้": 18,
            "วันที่": 21,
            "กอง": 11,
            "คน": 11,
            "รายการ": 13,
            "ช่อง 1": 15,
            "ช่อง 2": 15,
            "ฝา": 15,
            "หั่นเต๋า": 15,
            "น้ำ": 15,
            "ดอก": 15,
            "ทุเรียน": 16,
            "รวม": 16,
            "เงิน": 19,
            "รวมเงิน": 19,
            "เบี้ยขยัน": 19,
            "หักทั่วไป": 19,
            "หัก 3%": 17,
            "สุทธิ": 19,
        }
        widths = [preferred.get(str(header), 17) for header in headers]
        total = sum(widths)
        scale = min(1.0, 267 / total) if total else 1.0
        return [width * scale * mm for width in widths]

    def paragraph_cell(value, style, max_chars: int | None = None):
        if isinstance(value, Paragraph):
            return value
        raw_text = str("-" if value is None or value == "" else value)
        if max_chars and len(raw_text) > max_chars:
            raw_text = "\n".join(
                textwrap.wrap(
                    raw_text,
                    width=max_chars,
                    break_long_words=True,
                    break_on_hyphens=False,
                )
            )
        text = xml_escape(raw_text).replace("\n", "<br/>")
        return Paragraph(text, style)

    def add_table(title: str, headers: list[str], rows: list[list], widths: list[float] | None = None):
        wrapped_headers = [paragraph_cell(value, header_style) for value in headers]
        wrap_limits = {"กลุ่ม": 10, "ชื่อ": 11, "ผลไม้": 9}
        wrapped_rows = [
            [
                paragraph_cell(value, cell_style, wrap_limits.get(str(headers[index])))
                for index, value in enumerate(row)
            ]
            for row in rows
        ]
        if not wrapped_rows:
            wrapped_rows.append([paragraph_cell("-", cell_style) for _ in headers])
        table_rows = [wrapped_headers] + wrapped_rows
        col_widths = widths or group_column_widths(headers)
        table = Table(table_rows, repeatRows=1, colWidths=col_widths)
        set_pdf_table_style(table, 1)
        story.extend([Paragraph(title, section), table, Spacer(1, 7 * mm)])

    if options["summary"]:
        add_table(
            "สรุปตามกลุ่ม",
            ["กลุ่ม", "คน", "รายการ", *weight_headers(), "รวม", "เงิน", "เบี้ยขยัน", "หักทั่วไป", "หัก 3%", "สุทธิ"],
            [[row["pay_group"], len(row["employees"]), row["records"], *summary_weight_values(row), report_number(row["total"]), money(row["amount"]), money(row.get("bonus_amount", 0)), money(row.get("deduction_amount", 0)), money(row.get("withholding_tax_amount", 0)), money(row.get("net_amount", row["amount"]))] for row in group_rows],
        )

    if options["fruit"]:
        add_table(
            "สรุปตามกลุ่มและผลไม้",
            ["กลุ่ม", "ผลไม้", "คน", "รายการ", *weight_headers(), "รวม", "เงิน", "เบี้ยขยัน", "หักทั่วไป", "หัก 3%", "สุทธิ"],
            [[row["pay_group"], row["fruit_label"], len(row["employees"]), row["records"], *summary_weight_values(row), report_number(row["total"]), money(row["amount"]), money(row.get("bonus_amount", 0)), money(row.get("deduction_amount", 0)), money(row.get("withholding_tax_amount", 0)), money(row.get("net_amount", row["amount"]))] for row in fruit_rows],
        )

    if options["employees"]:
        add_table(
            "รายละเอียดพนักงานในกลุ่ม",
            ["กลุ่ม", "รหัส", "ชื่อ", "รายการ", *weight_headers(), "รวม", "เงิน", "เบี้ยขยัน", "หักทั่วไป", "หัก 3%", "สุทธิ"],
            [[row["pay_group"], row["emp_code"], row["fullname"], row["records"], *summary_weight_values(row), report_number(row["total"]), money(row["amount"]), money(row.get("bonus_amount", 0)), money(row.get("deduction_amount", 0)), money(row.get("withholding_tax_amount", 0)), money(row.get("net_amount", row["amount"]))] for row in employee_rows],
        )

    if options["details"]:
        add_table(
            "รายละเอียดรายการ",
            ["วันที่", "กลุ่ม", "ผลไม้", "รหัส", "ชื่อ", "กอง", *weight_headers(), "รวม", "รวมเงิน"],
            [[format_report_date(record["record_date"]), record["pay_group"], record["fruit_label"], record.get("emp_code", ""), record.get("employee_name", ""), record.get("pile_no") or record.get("pile", ""), *detail_weight_values(record), report_number(production_total_weight(record)), money(record.get("total_amount", record.get("grand_total", 0)))] for record in records],
        )

    if len(story) <= 4:
        story.append(Paragraph("ไม่มีส่วนรายงานที่เลือก", section))
    def draw_footer(canvas, document):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D6E5DB"))
        canvas.line(15 * mm, 11 * mm, 282 * mm, 11 * mm)
        canvas.setFont(THAI_FONT, 8)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawString(15 * mm, 6.5 * mm, "เอกสารสร้างจากระบบบริหารจัดการผลผลิต บริษัท พิศมัย ฟรุตส์ จำกัด")
        canvas.drawRightString(282 * mm, 6.5 * mm, f"หน้า {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    return buffer.getvalue()


def time_group_report_options(payload: dict) -> dict:
    options = payload.get("export_options") or {}
    return {
        "summary": bool(options.get("summary", True)),
        "employees": bool(options.get("employees", True)),
        "details": bool(options.get("details", False)),
    }


def build_time_group_report_excel(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    options = time_group_report_options(payload)
    group_rows = payload.get("time_group_rows", []) or []
    employee_rows = payload.get("time_employee_rows", []) or []
    records = payload.get("time_group_records", []) or []
    workbook = Workbook()
    overview = workbook.active
    overview.title = "ภาพรวม"
    font_name = "Sarabun"
    dark_green = "075B44"
    brand_green = "0F8A55"
    mint = "D1FAE5"
    pale = "F8FAFC"
    line_color = "D8E2EA"
    thin = Side(style="thin", color=line_color)
    medium = Side(style="medium", color=dark_green)
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def group_employee_count(row: dict) -> int:
        value = row.get("employees", 0)
        if isinstance(value, (int, float)):
            return int(value)
        group_name = str(row.get("pay_group") or "")
        return len({
            str(item.get("emp_code") or item.get("fullname") or "")
            for item in employee_rows
            if str(item.get("pay_group") or "") == group_name
        })

    total_employees = len({
        str(row.get("emp_code") or row.get("fullname") or "")
        for row in employee_rows
        if row.get("emp_code") or row.get("fullname")
    })
    total_hours = sum(safe_float(row.get("normal_hours")) + safe_float(row.get("ot_hours")) for row in group_rows)
    total_amount = sum(safe_float(row.get("amount")) for row in group_rows)
    total_bonus = sum(safe_float(row.get("bonus_amount")) for row in group_rows)
    total_deduction = sum(safe_float(row.get("deduction_amount")) for row in group_rows)
    total_net = sum(safe_float(row.get("net_amount", row.get("amount"))) for row in group_rows)

    def setup_sheet(sheet, subtitle: str, last_column: str, freeze_at: str | None = None):
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells(f"A1:B3")
        sheet["A1"] = "PF"
        sheet["A1"].fill = PatternFill("solid", fgColor=brand_green)
        sheet["A1"].font = Font(name=font_name, bold=True, size=25, color="FFFFFF")
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(f"C1:{last_column}2")
        sheet["C1"] = COMPANY_NAME
        sheet["C1"].fill = PatternFill("solid", fgColor=dark_green)
        sheet["C1"].font = Font(name=font_name, bold=True, size=20, color="FFFFFF")
        sheet["C1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet.merge_cells(f"C3:{last_column}3")
        sheet["C3"] = subtitle
        sheet["C3"].fill = PatternFill("solid", fgColor=dark_green)
        sheet["C3"].font = Font(name=font_name, size=10, color=mint)
        sheet["C3"].alignment = Alignment(horizontal="left", vertical="center")
        add_excel_logo(sheet, "A1")
        sheet.freeze_panes = freeze_at
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.orientation = "landscape"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.oddFooter.left.text = SYSTEM_NAME
        sheet.oddFooter.center.text = "รายงานแบบกลุ่มตามเวลา"
        sheet.oddFooter.right.text = "หน้า &P จาก &N"
        for row_index in (1, 2, 3):
            sheet.row_dimensions[row_index].height = 23

    setup_sheet(overview, "รายงานแบบกลุ่มตามเวลา • Time Group Report", "L", "A13")
    overview.merge_cells("A5:L5")
    overview["A5"] = (
        f"ช่วงรายงาน {format_report_date(start_date)} - {format_report_date(end_date)}"
        f"  |  กลุ่ม {payload.get('group_label') or 'ทุกกลุ่ม'}  |  {export_meta_text(payload)}"
    )
    overview["A5"].fill = PatternFill("solid", fgColor=mint)
    overview["A5"].font = Font(name=font_name, bold=True, size=10, color=dark_green)
    overview["A5"].alignment = Alignment(vertical="center", wrap_text=True)
    overview.row_dimensions[5].height = 28

    if options["summary"]:
        cards = [
            ("A7:C7", "A8:C9", "จำนวนพนักงาน", total_employees, "ECFDF5", "047857", '0 "คน"'),
            ("D7:F7", "D8:F9", "ชั่วโมงทำงาน", total_hours, "EFF6FF", "1D4ED8", '#,##0.0 "ชม."'),
            ("G7:I7", "G8:I9", "รายได้ก่อนหัก", total_amount, "FFF7ED", "C2410C", '#,##0.00 "บาท"'),
            ("J7:L7", "J8:L9", "ยอดรับสุทธิ", total_net, "F0FDF4", "15803D", '#,##0.00 "บาท"'),
        ]
        for label_range, value_range, label, value, fill_color, font_color, number_format in cards:
            overview.merge_cells(label_range)
            overview.merge_cells(value_range)
            label_cell = overview[label_range.split(":")[0]]
            value_cell = overview[value_range.split(":")[0]]
            label_cell.value = label
            value_cell.value = value
            for row in overview[f"{label_range.split(':')[0]}:{value_range.split(':')[1]}"]:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.border = table_border
            label_cell.font = Font(name=font_name, bold=True, size=10, color="64748B")
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.font = Font(name=font_name, bold=True, size=17, color=font_color)
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.number_format = number_format

        overview.merge_cells("A11:L11")
        overview["A11"] = "สรุปตามกลุ่ม"
        overview["A11"].fill = PatternFill("solid", fgColor=dark_green)
        overview["A11"].font = Font(name=font_name, bold=True, size=12, color="FFFFFF")
        overview["A11"].alignment = Alignment(vertical="center")
        overview.append([])
        overview.append([
            "กลุ่ม", "พนักงาน", "รายการ", "เวลาสุทธิ", "ชม.ปกติ", "OT",
            "ค่าแรงปกติ", "ค่า OT", "ก่อนหัก", "เงินเพิ่ม", "รายการหัก", "สุทธิ",
        ])
        for row in group_rows:
            overview.append([
                row.get("pay_group", "-"),
                group_employee_count(row),
                row.get("records", 0),
                minutes_text(row.get("net_minutes", 0)),
                safe_float(row.get("normal_hours")),
                safe_float(row.get("ot_hours")),
                safe_float(row.get("normal_amount")),
                safe_float(row.get("ot_amount")),
                safe_float(row.get("amount")),
                safe_float(row.get("bonus_amount")),
                safe_float(row.get("deduction_amount")),
                safe_float(row.get("net_amount", row.get("amount"))),
            ])
        summary_header = 13
        summary_total = overview.max_row + 1
        overview.cell(summary_total, 1, "รวมทั้งหมด")
        overview.merge_cells(start_row=summary_total, start_column=1, end_row=summary_total, end_column=3)
        overview.cell(summary_total, 5, sum(safe_float(row.get("normal_hours")) for row in group_rows))
        overview.cell(summary_total, 6, sum(safe_float(row.get("ot_hours")) for row in group_rows))
        overview.cell(summary_total, 7, sum(safe_float(row.get("normal_amount")) for row in group_rows))
        overview.cell(summary_total, 8, sum(safe_float(row.get("ot_amount")) for row in group_rows))
        overview.cell(summary_total, 9, total_amount)
        overview.cell(summary_total, 10, total_bonus)
        overview.cell(summary_total, 11, total_deduction)
        overview.cell(summary_total, 12, total_net)
        for cell in overview[summary_header]:
            cell.fill = PatternFill("solid", fgColor=brand_green)
            cell.font = Font(name=font_name, bold=True, size=9, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = table_border
        for row_index in range(summary_header + 1, summary_total):
            for cell in overview[row_index]:
                cell.font = Font(name=font_name, size=9, color="1E293B")
                cell.border = table_border
                cell.alignment = Alignment(
                    horizontal="right" if cell.column >= 2 else "left",
                    vertical="center",
                    wrap_text=True,
                )
                if row_index % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor=pale)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
        for cell in overview[summary_total]:
            cell.fill = PatternFill("solid", fgColor=mint)
            cell.font = Font(name=font_name, bold=True, size=9, color=dark_green)
            cell.border = Border(top=medium, bottom=medium)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
        overview.row_dimensions[summary_header].height = 28
        overview.print_title_rows = "1:13"
    else:
        overview.merge_cells("A7:L9")
        overview["A7"] = "ไม่ได้เลือกส่วนสรุปตามกลุ่ม"
        overview["A7"].font = Font(name=font_name, italic=True, color="64748B")
        overview["A7"].alignment = Alignment(horizontal="center", vertical="center")

    if options["employees"]:
        employees = workbook.create_sheet("สรุปพนักงาน")
        setup_sheet(employees, "สรุปรายบุคคลแยกตามกลุ่ม", "K", "A6")
        employees.append([])
        employees.append([
            "กลุ่ม", "รหัสพนักงาน", "ชื่อ-นามสกุล", "วันทำงาน", "เวลาสุทธิ",
            "ชม.ปกติ", "OT", "ก่อนหัก", "เงินเพิ่ม", "รายการหัก", "สุทธิ",
        ])
        for row in employee_rows:
            employees.append([
                row.get("pay_group", "-"),
                str(row.get("emp_code") or "-"),
                row.get("fullname", "-"),
                row.get("records", 0),
                minutes_text(row.get("net_minutes", 0)),
                safe_float(row.get("normal_hours")),
                safe_float(row.get("ot_hours")),
                safe_float(row.get("amount")),
                safe_float(row.get("bonus_amount")),
                safe_float(row.get("deduction_amount")),
                safe_float(row.get("net_amount", row.get("amount"))),
            ])
        style_excel_report_sheet(employees, [5], [20, 15, 28, 12, 16, 14, 12, 16, 15, 15, 16])
        employees["B6"].number_format = "@"
        employees.print_title_rows = "1:5"

    if options["details"]:
        details = workbook.create_sheet("รายละเอียดเวลา")
        setup_sheet(details, "รายละเอียดเวลาเข้า-ออก", "J", "A6")
        details.append([])
        details.append(["วันที่", "กลุ่ม", "รหัสพนักงาน", "ชื่อ-นามสกุล", "เวลาเข้า", "เวลาออก", "เวลาสุทธิ", "อัตรา OT", "ยอดเงิน", "สถานะ"])
        for record in records:
            status_parts = []
            if is_late_time(record.get("clock_in")):
                status_parts.append("มาสาย")
            if is_early_out_time(record.get("clock_out")):
                status_parts.append("ออกก่อน")
            details.append([
                record.get("record_date", ""),
                record.get("employee_type_label", ""),
                str(record.get("emp_code") or ""),
                record.get("fullname", ""),
                record.get("clock_in", ""),
                record.get("clock_out", ""),
                minutes_text(record.get("net_minutes", 0)),
                safe_float(record.get("ot_hourly_rate")),
                safe_float(record.get("total_amount")),
                " / ".join(status_parts) or "ปกติ",
            ])
        style_excel_report_sheet(details, [5], [14, 20, 15, 27, 12, 12, 16, 13, 16, 16])
        details["C6"].number_format = "@"
        details.print_title_rows = "1:5"

    widths = [20, 12, 12, 16, 14, 12, 16, 15, 16, 15, 15, 16]
    for index, width in enumerate(widths, 1):
        overview.column_dimensions[get_column_letter(index)].width = width
    overview.print_area = f"A1:L{overview.max_row}"
    overview.page_setup.orientation = "landscape"
    overview.page_setup.paperSize = overview.PAPERSIZE_A4
    overview.sheet_properties.pageSetUpPr.fitToPage = True
    overview.page_setup.fitToWidth = 1
    overview.page_setup.fitToHeight = 0
    overview.oddFooter.left.text = SYSTEM_NAME
    overview.oddFooter.center.text = "รายงานแบบกลุ่มตามเวลา"
    overview.oddFooter.right.text = "หน้า &P จาก &N"
    overview.sheet_view.zoomScale = 80

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_time_group_report_pdf(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    options = time_group_report_options(payload)
    group_rows = payload.get("time_group_rows", []) or []
    employee_rows = payload.get("time_employee_rows", []) or []
    records = payload.get("time_group_records", []) or []
    _, _, pdf_normal, section = pdf_styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=31 * mm,
        bottomMargin=18 * mm,
    )
    story = []
    dark_green = colors.HexColor("#075B44")
    brand_green = colors.HexColor("#0F8A55")
    mint = colors.HexColor("#D1FAE5")
    pale = colors.HexColor("#F8FAFC")
    line_color = colors.HexColor("#D8E2EA")

    def group_employee_count(row: dict) -> int:
        value = row.get("employees", 0)
        if isinstance(value, (int, float)):
            return int(value)
        group_name = str(row.get("pay_group") or "")
        return len({
            str(item.get("emp_code") or item.get("fullname") or "")
            for item in employee_rows
            if str(item.get("pay_group") or "") == group_name
        })

    total_employees = len({
        str(row.get("emp_code") or row.get("fullname") or "")
        for row in employee_rows
        if row.get("emp_code") or row.get("fullname")
    })
    total_normal_hours = sum(safe_float(row.get("normal_hours")) for row in group_rows)
    total_ot_hours = sum(safe_float(row.get("ot_hours")) for row in group_rows)
    total_hours = total_normal_hours + total_ot_hours
    total_amount = sum(safe_float(row.get("amount")) for row in group_rows)
    total_bonus = sum(safe_float(row.get("bonus_amount")) for row in group_rows)
    total_deduction = sum(safe_float(row.get("deduction_amount")) for row in group_rows)
    total_net = sum(safe_float(row.get("net_amount", row.get("amount"))) for row in group_rows)

    meta = Table(
        [[
            Paragraph(
                f"<b>ช่วงรายงาน</b><br/>{format_report_date(start_date)} - {format_report_date(end_date)}",
                pdf_normal,
            ),
            Paragraph(
                f"<b>ตัวกรอง</b><br/>{xml_escape(str(payload.get('group_label') or 'ทุกกลุ่ม'))}",
                pdf_normal,
            ),
            Paragraph(f"<b>ข้อมูลการพิมพ์</b><br/>{xml_escape(export_meta_text(payload))}", pdf_normal),
        ]],
        colWidths=[78 * mm, 68 * mm, 121 * mm],
    )
    meta.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), mint),
        ("BOX", (0, 0), (-1, -1), 0.4, line_color),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, line_color),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.extend([meta, Spacer(1, 5 * mm)])

    if options["summary"]:
        metric_style = getSampleStyleSheet()["BodyText"]
        metric_style.fontName = THAI_FONT
        metric_style.fontSize = 8
        metric_style.leading = 21
        metrics = [
            Paragraph(f"<font color='#64748B'>พนักงาน</font><br/><font size='16' color='#047857'><b>{total_employees:,} คน</b></font>", metric_style),
            Paragraph(f"<font color='#64748B'>ชั่วโมงทำงาน</font><br/><font size='16' color='#1D4ED8'><b>{total_hours:,.1f} ชม.</b></font>", metric_style),
            Paragraph(f"<font color='#64748B'>รายได้ก่อนหัก</font><br/><font size='16' color='#C2410C'><b>{money(total_amount)} บาท</b></font>", metric_style),
            Paragraph(f"<font color='#64748B'>ยอดรับสุทธิ</font><br/><font size='16' color='#15803D'><b>{money(total_net)} บาท</b></font>", metric_style),
        ]
        metric_table = Table([metrics], colWidths=[66.75 * mm] * 4, rowHeights=[24 * mm])
        metric_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.4, line_color),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, line_color),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#ECFDF5")),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EFF6FF")),
            ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#FFF7ED")),
            ("BACKGROUND", (3, 0), (3, 0), colors.HexColor("#F0FDF4")),
        ]))
        story.extend([metric_table, Spacer(1, 5 * mm)])

    def add_table(
        title: str,
        headers: list[str],
        rows: list[list],
        widths: list[float] | None = None,
        total_row: list | None = None,
    ):
        if story:
            story.append(Spacer(1, 2 * mm))
        table_rows = [headers] + rows
        if total_row is not None:
            table_rows.append(total_row)
        elif len(table_rows) == 1:
            table_rows.append(["-" for _ in headers])
        col_widths = widths or [(267 / len(headers)) * mm for _ in headers]
        table = Table(table_rows, repeatRows=1, colWidths=col_widths)
        set_pdf_table_style(table, 1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), brand_green),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, pale]),
            ("GRID", (0, 0), (-1, -1), 0.35, line_color),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        if total_row is not None:
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, -1), (-1, -1), mint),
                ("TEXTCOLOR", (0, -1), (-1, -1), dark_green),
                ("FONTNAME", (0, -1), (-1, -1), THAI_FONT_BOLD),
                ("LINEABOVE", (0, -1), (-1, -1), 1, dark_green),
            ]))
        story.extend([Paragraph(title, section), table, Spacer(1, 7 * mm)])

    if options["summary"]:
        add_table(
            "สรุปตามกลุ่ม",
            ["กลุ่ม", "พนักงาน", "รายการ", "เวลาสุทธิ", "ชม.ปกติ", "OT", "ก่อนหัก", "เงินเพิ่ม", "รายการหัก", "สุทธิ"],
            [[
                row.get("pay_group", "-"),
                report_number(group_employee_count(row), 0),
                report_number(row.get("records", 0), 0),
                minutes_text(row.get("net_minutes", 0)),
                report_number(row.get("normal_hours", 0)),
                report_number(row.get("ot_hours", 0)),
                money(row.get("amount", 0)),
                money(row.get("bonus_amount", 0)),
                money(row.get("deduction_amount", 0)),
                money(row.get("net_amount", row.get("amount", 0))),
            ] for row in group_rows],
            [38 * mm, 18 * mm, 18 * mm, 25 * mm, 22 * mm, 18 * mm, 30 * mm, 27 * mm, 27 * mm, 30 * mm],
            [
                "รวมทั้งหมด",
                report_number(total_employees, 0),
                report_number(sum(safe_float(row.get("records")) for row in group_rows), 0),
                minutes_text(sum(safe_float(row.get("net_minutes")) for row in group_rows)),
                report_number(total_normal_hours),
                report_number(total_ot_hours),
                money(total_amount),
                money(total_bonus),
                money(total_deduction),
                money(total_net),
            ],
        )

    if options["employees"]:
        add_table(
            "สรุปรายบุคคล",
            ["กลุ่ม", "รหัส", "ชื่อพนักงาน", "วัน", "เวลาสุทธิ", "ชม.ปกติ", "OT", "ก่อนหัก", "เงินเพิ่ม", "รายการหัก", "สุทธิ"],
            [[
                row.get("pay_group", "-"),
                str(row.get("emp_code") or "-"),
                row.get("fullname", "-"),
                report_number(row.get("records", 0), 0),
                minutes_text(row.get("net_minutes", 0)),
                report_number(row.get("normal_hours", 0)),
                report_number(row.get("ot_hours", 0)),
                money(row.get("amount", 0)),
                money(row.get("bonus_amount", 0)),
                money(row.get("deduction_amount", 0)),
                money(row.get("net_amount", row.get("amount", 0))),
            ] for row in employee_rows],
            [31 * mm, 21 * mm, 34 * mm, 14 * mm, 22 * mm, 19 * mm, 15 * mm, 27 * mm, 25 * mm, 25 * mm, 28 * mm],
        )

    if options["details"]:
        if options["summary"] or options["employees"]:
            story.append(PageBreak())
        add_table(
            "รายละเอียดเวลาเข้า-ออก",
            ["วันที่", "กลุ่ม", "รหัส", "ชื่อพนักงาน", "เข้า", "ออก", "เวลาสุทธิ", "อัตรา OT", "ยอดเงิน", "สถานะ"],
            [[
                format_report_date(record.get("record_date", "")),
                record.get("employee_type_label", ""),
                str(record.get("emp_code") or ""),
                record.get("fullname", ""),
                record.get("clock_in", ""),
                record.get("clock_out", ""),
                minutes_text(record.get("net_minutes", 0)),
                report_number(record.get("ot_hourly_rate", 0), 0),
                money(record.get("total_amount", 0)),
                " / ".join(
                    label
                    for condition, label in [
                        (is_late_time(record.get("clock_in")), "มาสาย"),
                        (is_early_out_time(record.get("clock_out")), "ออกก่อน"),
                    ]
                    if condition
                ) or "ปกติ",
            ] for record in records],
            [25 * mm, 35 * mm, 22 * mm, 38 * mm, 18 * mm, 18 * mm, 24 * mm, 20 * mm, 25 * mm, 24 * mm],
        )

    if not any(options.values()):
        story.append(Paragraph("ไม่ได้เลือกส่วนรายงานสำหรับส่งออก", section))

    def draw_page(canvas_obj, document):
        page_width, page_height = landscape(A4)
        canvas_obj.saveState()
        canvas_obj.setFillColor(dark_green)
        canvas_obj.rect(0, page_height - 22 * mm, page_width, 22 * mm, fill=1, stroke=0)
        logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
        if logo_path.exists():
            try:
                canvas_obj.drawImage(
                    str(logo_path),
                    12 * mm,
                    page_height - 19 * mm,
                    width=15 * mm,
                    height=15 * mm,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass
        canvas_obj.setFont(THAI_FONT_BOLD, 15)
        canvas_obj.setFillColor(colors.white)
        canvas_obj.drawString(31 * mm, page_height - 10 * mm, COMPANY_NAME)
        canvas_obj.setFont(THAI_FONT, 8)
        canvas_obj.setFillColor(mint)
        canvas_obj.drawString(31 * mm, page_height - 16 * mm, "รายงานแบบกลุ่มตามเวลา")
        canvas_obj.setStrokeColor(brand_green)
        canvas_obj.line(15 * mm, 12 * mm, page_width - 15 * mm, 12 * mm)
        canvas_obj.setFont(THAI_FONT, 7.5)
        canvas_obj.setFillColor(colors.HexColor("#475467"))
        canvas_obj.drawString(
            15 * mm,
            7 * mm,
            f"ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)} | {export_meta_text(payload)}",
        )
        canvas_obj.drawRightString(page_width - 15 * mm, 7 * mm, f"หน้า {document.page}")
        canvas_obj.restoreState()

    doc.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
    return buffer.getvalue()


def summarize_time(records: list[dict]) -> tuple[dict, list[dict], list[dict]]:
    summary = {
        "records": len(records),
        "employees": len({record.get("employee_id") or record.get("emp_code") or record.get("fullname") for record in records}),
        "days": len({record.get("record_date") for record in records}),
        "raw_minutes": sum(safe_float(record.get("raw_minutes")) for record in records),
        "break_minutes": sum(safe_float(record.get("break_minutes")) for record in records),
        "net_minutes": sum(safe_float(record.get("net_minutes")) for record in records),
        "late": sum(1 for record in records if is_late_time(record.get("clock_in"))),
        "early": sum(1 for record in records if is_early_out_time(record.get("clock_out"))),
    }
    by_date: dict[str, dict] = {}
    by_employee: dict[str, dict] = {}
    for record in records:
        date_key = record.get("record_date") or "-"
        daily = by_date.setdefault(date_key, {"date": date_key, "records": 0, "employees": set(), "net_minutes": 0.0, "late": 0, "early": 0})
        daily["records"] += 1
        daily["employees"].add(record.get("employee_id") or record.get("emp_code") or record.get("fullname") or "")
        daily["net_minutes"] += safe_float(record.get("net_minutes"))
        daily["late"] += 1 if is_late_time(record.get("clock_in")) else 0
        daily["early"] += 1 if is_early_out_time(record.get("clock_out")) else 0

        employee_key = str(record.get("employee_id") or record.get("emp_code") or record.get("fullname") or "")
        employee = by_employee.setdefault(
            employee_key,
            {
                "emp_code": record.get("emp_code", "-"),
                "fullname": record.get("fullname", "-"),
                "department": record.get("department", "-"),
                "records": 0,
                "days": set(),
                "net_minutes": 0.0,
                "late": 0,
                "early": 0,
                "first_in": record.get("clock_in") or "-",
                "last_out": record.get("clock_out") or "-",
            },
        )
        employee["records"] += 1
        employee["days"].add(date_key)
        employee["net_minutes"] += safe_float(record.get("net_minutes"))
        employee["late"] += 1 if is_late_time(record.get("clock_in")) else 0
        employee["early"] += 1 if is_early_out_time(record.get("clock_out")) else 0
        if record.get("clock_in") and (employee["first_in"] == "-" or record["clock_in"] < employee["first_in"]):
            employee["first_in"] = record["clock_in"]
        if record.get("clock_out") and (employee["last_out"] == "-" or record["clock_out"] > employee["last_out"]):
            employee["last_out"] = record["clock_out"]

    daily_rows = sorted(by_date.values(), key=lambda item: item["date"])
    employee_rows = sorted(by_employee.values(), key=lambda item: (str(item["emp_code"]), str(item["fullname"])))
    return summary, daily_rows, employee_rows


def build_time_summary_excel(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    records = filtered_time_records(payload)
    summary, daily_rows, employee_rows = summarize_time(records)
    department_label = payload.get("department_label") or payload.get("department") or "ทุกแผนก"
    workbook = Workbook()
    overview = workbook.active
    overview.title = "ภาพรวม"
    font_name = "Sarabun"
    dark_green = "075B44"
    brand_green = "0F8A55"
    mint = "D1FAE5"
    pale = "F8FAFC"
    line_color = "D8E2EA"
    thin = Side(style="thin", color=line_color)
    medium = Side(style="medium", color=dark_green)
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def setup_time_sheet(sheet, subtitle: str, last_column: str, freeze_at: str | None = None):
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells("A1:B3")
        sheet["A1"] = "PF"
        sheet["A1"].fill = PatternFill("solid", fgColor=brand_green)
        sheet["A1"].font = Font(name=font_name, bold=True, size=25, color="FFFFFF")
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(f"C1:{last_column}2")
        sheet["C1"] = COMPANY_NAME
        sheet["C1"].fill = PatternFill("solid", fgColor=dark_green)
        sheet["C1"].font = Font(name=font_name, bold=True, size=20, color="FFFFFF")
        sheet["C1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet.merge_cells(f"C3:{last_column}3")
        sheet["C3"] = subtitle
        sheet["C3"].fill = PatternFill("solid", fgColor=dark_green)
        sheet["C3"].font = Font(name=font_name, size=10, color=mint)
        sheet["C3"].alignment = Alignment(horizontal="left", vertical="center")
        add_excel_logo(sheet, "A1")
        sheet.freeze_panes = freeze_at
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.orientation = "landscape"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.oddFooter.left.text = SYSTEM_NAME
        sheet.oddFooter.center.text = "รายงานสรุปเวลาเข้างาน"
        sheet.oddFooter.right.text = "หน้า &P จาก &N"
        for row_index in (1, 2, 3):
            sheet.row_dimensions[row_index].height = 23

    setup_time_sheet(overview, "รายงานสรุปเวลาเข้างาน • Attendance Summary", "L", "A13")
    overview.merge_cells("A5:L5")
    overview["A5"] = (
        f"ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)}"
        f"  |  แผนก {department_label}  |  {export_meta_text(payload)}"
    )
    overview["A5"].fill = PatternFill("solid", fgColor=mint)
    overview["A5"].font = Font(name=font_name, bold=True, size=10, color=dark_green)
    overview["A5"].alignment = Alignment(vertical="center", wrap_text=True)
    overview.row_dimensions[5].height = 28
    cards = [
        ("A7:C7", "A8:C9", "จำนวนพนักงาน", summary["employees"], "ECFDF5", "047857", '0 "คน"'),
        ("D7:F7", "D8:F9", "ชั่วโมงทำงานสุทธิ", summary["net_minutes"] / 60, "EFF6FF", "1D4ED8", '#,##0.0 "ชม."'),
        ("G7:I7", "G8:I9", "จำนวนรายการ", summary["records"], "FFF7ED", "C2410C", '0 "รายการ"'),
        ("J7:L7", "J8:L9", "มาสาย / ออกก่อน", f"{summary['late']} / {summary['early']} ครั้ง", "F0FDF4", "15803D", "@"),
    ]
    for label_range, value_range, label, value, fill_color, font_color, number_format in cards:
        overview.merge_cells(label_range)
        overview.merge_cells(value_range)
        label_cell = overview[label_range.split(":")[0]]
        value_cell = overview[value_range.split(":")[0]]
        label_cell.value = label
        value_cell.value = value
        for row in overview[f"{label_range.split(':')[0]}:{value_range.split(':')[1]}"]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = table_border
        label_cell.font = Font(name=font_name, bold=True, size=10, color="64748B")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.font = Font(name=font_name, bold=True, size=17, color=font_color)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format
    overview.merge_cells("A11:L11")
    overview["A11"] = "สรุปรายวัน"
    overview["A11"].fill = PatternFill("solid", fgColor=dark_green)
    overview["A11"].font = Font(name=font_name, bold=True, size=12, color="FFFFFF")
    overview.append([])
    overview.append(["วันที่", "จำนวนรายการ", "จำนวนพนักงาน", "ชั่วโมงสุทธิ", "มาสาย", "ออกก่อน"])
    for row in daily_rows:
        overview.append([row["date"], row["records"], len(row["employees"]), row["net_minutes"] / 60, row["late"], row["early"]])
    total_row = overview.max_row + 1
    overview.cell(total_row, 1, "รวมทั้งหมด")
    overview.cell(total_row, 2, summary["records"])
    overview.cell(total_row, 3, summary["employees"])
    overview.cell(total_row, 4, summary["net_minutes"] / 60)
    overview.cell(total_row, 5, summary["late"])
    overview.cell(total_row, 6, summary["early"])
    style_excel_report_sheet(overview, [13], [18] * 12)
    for cell in overview[total_row]:
        if cell.value is not None:
            cell.fill = PatternFill("solid", fgColor=mint)
            cell.font = Font(name=font_name, bold=True, color=dark_green)
            cell.border = Border(top=medium, bottom=medium)

    daily = workbook.create_sheet("สรุปรายวัน")
    setup_time_sheet(daily, "สรุปเวลาเข้างานแยกตามวัน", "F", "A6")
    daily.append([])
    daily.append(["วันที่", "จำนวนรายการ", "จำนวนพนักงาน", "ชั่วโมงสุทธิ", "มาสาย", "ออกก่อน"])
    for row in daily_rows:
        daily.append([row["date"], row["records"], len(row["employees"]), row["net_minutes"] / 60, row["late"], row["early"]])
    style_excel_report_sheet(daily, [5], [16, 16, 18, 18, 14, 14])

    employee_sheet = workbook.create_sheet("สรุปพนักงาน")
    setup_time_sheet(employee_sheet, "สรุปเวลาเข้างานรายบุคคล", "J", "A6")
    employee_sheet.append([])
    employee_sheet.append(["รหัสพนักงาน", "ชื่อพนักงาน", "แผนก", "จำนวนวัน", "จำนวนรายการ", "ชั่วโมงสุทธิ", "มาสาย", "ออกก่อน", "เข้าเร็วสุด", "ออกช้าสุด"])
    for row in employee_rows:
        employee_sheet.append([str(row["emp_code"]), row["fullname"], row["department"], len(row["days"]), row["records"], row["net_minutes"] / 60, row["late"], row["early"], row["first_in"], row["last_out"]])
    style_excel_report_sheet(employee_sheet, [5], [17, 31, 18, 12, 15, 17, 13, 13, 15, 15])
    for row_index in range(6, employee_sheet.max_row + 1):
        employee_sheet.cell(row_index, 1).number_format = "@"
        name_cell = employee_sheet.cell(row_index, 2)
        name_cell.font = Font(name=font_name, bold=True, size=10, color=dark_green)
        name_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    detail = workbook.create_sheet("รายละเอียด")
    setup_time_sheet(detail, "รายละเอียดเวลาเข้า-ออก • ชื่อจากทะเบียนพนักงาน", "J", "A6")
    detail.append([])
    detail.append(["วันที่", "รหัสพนักงาน", "ชื่อพนักงาน", "แผนก", "เข้า", "ออก", "พัก(นาที)", "สุทธิ(นาที)", "ชั่วโมง", "ผู้บันทึก"])
    for record in records:
        net_minutes = safe_float(record.get("net_minutes"))
        detail.append([record.get("record_date", ""), str(record.get("emp_code") or ""), record.get("fullname", ""), record.get("department", ""), record.get("clock_in", ""), record.get("clock_out", ""), safe_float(record.get("break_minutes")), net_minutes, net_minutes / 60, record.get("created_by", "")])
    style_excel_report_sheet(detail, [5], [15, 17, 31, 18, 12, 12, 14, 16, 14, 20])
    for row_index in range(6, detail.max_row + 1):
        detail.cell(row_index, 2).number_format = "@"
        name_cell = detail.cell(row_index, 3)
        name_cell.font = Font(name=font_name, bold=True, size=10, color=dark_green)
        name_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for index, width in enumerate([18] * 12, 1):
        overview.column_dimensions[get_column_letter(index)].width = width
    overview.print_area = f"A1:L{overview.max_row}"
    overview.print_title_rows = "1:13"
    overview.sheet_view.zoomScale = 80

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_time_summary_pdf(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    records = filtered_time_records(payload)
    summary, daily_rows, employee_rows = summarize_time(records)
    department_label = payload.get("department_label") or payload.get("department") or "ทุกแผนก"
    _, _, pdf_normal, section = pdf_styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=31 * mm,
        bottomMargin=18 * mm,
    )
    story = []
    dark_green = colors.HexColor("#075B44")
    brand_green = colors.HexColor("#0F8A55")
    mint = colors.HexColor("#D1FAE5")
    pale = colors.HexColor("#F8FAFC")
    line_color = colors.HexColor("#D8E2EA")
    meta = Table(
        [[
            Paragraph(f"<b>ช่วงรายงาน</b><br/>{format_report_date(start_date)} - {format_report_date(end_date)}", pdf_normal),
            Paragraph(f"<b>แผนก</b><br/>{xml_escape(str(department_label))}", pdf_normal),
            Paragraph(f"<b>ข้อมูลการพิมพ์</b><br/>{xml_escape(export_meta_text(payload))}", pdf_normal),
        ]],
        colWidths=[78 * mm, 64 * mm, 127 * mm],
    )
    meta.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), mint),
        ("GRID", (0, 0), (-1, -1), 0.35, line_color),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.extend([meta, Spacer(1, 5 * mm)])
    metric_style = getSampleStyleSheet()["BodyText"]
    metric_style.fontName = THAI_FONT
    metric_style.fontSize = 8
    metric_style.leading = 20
    metric_values = [
        ("พนักงาน", f"{report_number(summary['employees'], 0)} คน", "#047857"),
        ("เวลาทำงานสุทธิ", minutes_text(summary["net_minutes"]), "#1D4ED8"),
        ("จำนวนรายการ", f"{report_number(summary['records'], 0)} รายการ", "#C2410C"),
        ("มาสาย / ออกก่อน", f"{report_number(summary['late'], 0)} / {report_number(summary['early'], 0)} ครั้ง", "#15803D"),
    ]
    metric_cells = [
        Paragraph(
            f"<font color='#64748B'>{label}</font><br/><font size='15' color='{color}'><b>{value}</b></font>",
            metric_style,
        )
        for label, value, color in metric_values
    ]
    metric_table = Table([metric_cells], colWidths=[67.25 * mm] * 4, rowHeights=[23 * mm])
    metric_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.35, line_color),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#ECFDF5")),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#EFF6FF")),
        ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#FFF7ED")),
        ("BACKGROUND", (3, 0), (3, 0), colors.HexColor("#F0FDF4")),
    ]))
    story.extend([metric_table, Spacer(1, 5 * mm)])

    daily_table_rows = [["วันที่", "รายการ", "พนักงาน", "เวลาสุทธิ", "มาสาย", "ออกก่อน"]]
    for row in daily_rows:
        daily_table_rows.append([format_report_date(row["date"]), report_number(row["records"], 0), report_number(len(row["employees"]), 0), minutes_text(row["net_minutes"]), report_number(row["late"], 0), report_number(row["early"], 0)])
    if len(daily_table_rows) == 1:
        daily_table_rows.append(["-", "0", "0", "0 ชม. 0 นาที", "0", "0"])
    daily_table = Table(daily_table_rows, repeatRows=1, colWidths=[38 * mm, 32 * mm, 35 * mm, 55 * mm, 32 * mm, 32 * mm])
    set_pdf_table_style(daily_table, 1)
    daily_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_green),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, pale]),
        ("GRID", (0, 0), (-1, -1), 0.35, line_color),
    ]))
    story += [Paragraph("สรุปรายวัน", section), daily_table, Spacer(1, 7 * mm)]

    employee_table_rows = [["รหัส", "ชื่อพนักงาน", "แผนก", "วัน", "รายการ", "เวลาสุทธิ", "สาย", "ออกก่อน"]]
    for row in employee_rows:
        employee_table_rows.append([
            str(row["emp_code"]),
            Paragraph(
                f"<b><font color='#075B44'>{xml_escape(str(row['fullname'] or '-'))}</font></b>",
                pdf_normal,
            ),
            row["department"],
            report_number(len(row["days"]), 0),
            report_number(row["records"], 0),
            minutes_text(row["net_minutes"]),
            report_number(row["late"], 0),
            report_number(row["early"], 0),
        ])
    if len(employee_table_rows) == 1:
        employee_table_rows.append(["-", "-", "-", "0", "0", "0 ชม. 0 นาที", "0", "0"])
    employee_table = Table(employee_table_rows, repeatRows=1, colWidths=[24 * mm, 58 * mm, 34 * mm, 20 * mm, 25 * mm, 42 * mm, 20 * mm, 24 * mm])
    set_pdf_table_style(employee_table, 3)
    employee_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_green),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, pale]),
        ("GRID", (0, 0), (-1, -1), 0.35, line_color),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story += [Paragraph("สรุปรายพนักงาน", section), employee_table]

    def draw_page(canvas_obj, document):
        page_width, page_height = landscape(A4)
        canvas_obj.saveState()
        canvas_obj.setFillColor(dark_green)
        canvas_obj.rect(0, page_height - 22 * mm, page_width, 22 * mm, fill=1, stroke=0)
        logo_path = Path(__file__).with_name("assets") / "pitsamai-logo.png"
        if logo_path.exists():
            try:
                canvas_obj.drawImage(
                    str(logo_path), 12 * mm, page_height - 19 * mm,
                    width=15 * mm, height=15 * mm, preserveAspectRatio=True, mask="auto",
                )
            except Exception:
                pass
        canvas_obj.setFont(THAI_FONT_BOLD, 15)
        canvas_obj.setFillColor(colors.white)
        canvas_obj.drawString(31 * mm, page_height - 10 * mm, COMPANY_NAME)
        canvas_obj.setFont(THAI_FONT, 8)
        canvas_obj.setFillColor(mint)
        canvas_obj.drawString(31 * mm, page_height - 16 * mm, "รายงานสรุปเวลาเข้างาน")
        canvas_obj.setStrokeColor(brand_green)
        canvas_obj.line(14 * mm, 12 * mm, page_width - 14 * mm, 12 * mm)
        canvas_obj.setFont(THAI_FONT, 7.5)
        canvas_obj.setFillColor(colors.HexColor("#475467"))
        canvas_obj.drawString(
            14 * mm, 7 * mm,
            f"ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)} | {export_meta_text(payload)}",
        )
        canvas_obj.drawRightString(page_width - 14 * mm, 7 * mm, f"หน้า {document.page}")
        canvas_obj.restoreState()

    doc.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
    return buffer.getvalue()


def accounting_payment_rows(payload: dict) -> list[dict]:
    method = str(payload.get("payment_method") or "cash")
    if method not in {"cash", "transfer"}:
        method = "cash"
    rows = []
    for index, raw in enumerate(payload.get("rows") or [], start=1):
        if not isinstance(raw, dict) or str(raw.get("payment_method") or method) != method:
            continue
        rows.append({
            "sequence": index,
            "emp_code": str(raw.get("emp_code") or "-"),
            "fullname": str(raw.get("fullname") or "-"),
            "group_label": str(raw.get("group_label") or "-"),
            "gross_amount": max(0, safe_float(raw.get("gross_amount"))),
            "bonus_amount": max(0, safe_float(raw.get("bonus_amount"))),
            "deduction_amount": max(0, safe_float(raw.get("deduction_amount"))),
            "withholding_tax_amount": max(0, safe_float(raw.get("withholding_tax_amount"))),
            "net_amount": max(0, safe_float(raw.get("net_amount"))),
            "payment_method": method,
        })
    return rows


def accounting_payment_label(method: str) -> str:
    return "เงินโอน" if method == "transfer" else "เงินสด"


def build_accounting_payments_excel(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    method = "transfer" if payload.get("payment_method") == "transfer" else "cash"
    label = accounting_payment_label(method)
    scope_label = str(payload.get("scope_label") or "พนักงานทั้งหมด")
    rows = accounting_payment_rows(payload)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Payment {method.title()}"
    sheet.merge_cells("A1:J1")
    sheet["A1"] = COMPANY_NAME
    sheet["A1"].font = Font(name="Sarabun", bold=True, size=18, color="0F7A3D")
    sheet.merge_cells("A2:J2")
    sheet["A2"] = f"รายการจ่ายค่าแรง - {label}"
    sheet["A2"].font = Font(name="Sarabun", bold=True, size=16, color="111827")
    sheet.merge_cells("A3:J3")
    sheet["A3"] = f"ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)} | ขอบเขต {scope_label}"
    sheet.merge_cells("A4:J4")
    sheet["A4"] = export_meta_text(payload)
    add_excel_logo(sheet, "K1")
    sheet.append([])
    sheet.append(["ลำดับ", "รหัส", "ชื่อพนักงาน", "กลุ่ม", "ค่าแรง", "เงินเพิ่ม", "เงินหัก", "หัก ณ ที่จ่าย", "ยอดสุทธิ", "ลงชื่อรับเงิน/หมายเหตุ"])
    for row in rows:
        sheet.append([row["sequence"], row["emp_code"], row["fullname"], row["group_label"], row["gross_amount"], row["bonus_amount"], row["deduction_amount"], row["withholding_tax_amount"], row["net_amount"], ""])
    total_row = sheet.max_row + 1
    sheet.cell(total_row, 1, "รวม")
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    for column in range(5, 10):
        letter = get_column_letter(column)
        sheet.cell(total_row, column, f"=SUM({letter}7:{letter}{total_row - 1})" if rows else 0)
    style_excel_report_sheet(sheet, [6], [9, 14, 28, 24, 15, 14, 14, 17, 16, 28])
    for cell in sheet[total_row]:
        cell.font = Font(name="Sarabun", bold=True, color="0F7A3D", size=10)
        cell.fill = PatternFill("solid", fgColor="E8F5EE")
    sheet.print_title_rows = "1:6"
    sheet.print_area = f"A1:J{total_row}"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_accounting_payments_pdf(payload: dict) -> bytes:
    start_date, end_date = normalized_range(payload)
    method = "transfer" if payload.get("payment_method") == "transfer" else "cash"
    label = accounting_payment_label(method)
    scope_label = str(payload.get("scope_label") or "พนักงานทั้งหมด")
    rows = accounting_payment_rows(payload)
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=16 * mm)
    _, _, normal, section = pdf_styles()
    small = normal.clone("PfPaymentSmall")
    small.fontSize = 7
    small.leading = 9
    story = report_header_story(
        f"รายการจ่ายค่าแรง - {label}",
        f"ช่วงวันที่ {format_report_date(start_date)} - {format_report_date(end_date)} | ขอบเขต {xml_escape(scope_label)}",
        payload,
    )
    story += [Spacer(1, 5 * mm), Paragraph(f"จำนวน {len(rows)} คน | ยอดสุทธิรวม {report_number(sum(row['net_amount'] for row in rows))} บาท", section), Spacer(1, 3 * mm)]
    table_rows = [["ลำดับ", "รหัส", "ชื่อพนักงาน", "กลุ่ม", "ค่าแรง", "เงินเพิ่ม", "เงินหัก", "หัก ณ ที่จ่าย", "ยอดสุทธิ", "ลงชื่อ/หมายเหตุ"]]
    for row in rows:
        table_rows.append([
            row["sequence"], row["emp_code"], Paragraph(xml_escape(row["fullname"]), small), Paragraph(xml_escape(row["group_label"]), small),
            report_number(row["gross_amount"]), report_number(row["bonus_amount"]), report_number(row["deduction_amount"]), report_number(row["withholding_tax_amount"]), report_number(row["net_amount"]), "",
        ])
    table_rows.append(["รวม", "", "", "", report_number(sum(row["gross_amount"] for row in rows)), report_number(sum(row["bonus_amount"] for row in rows)), report_number(sum(row["deduction_amount"] for row in rows)), report_number(sum(row["withholding_tax_amount"] for row in rows)), report_number(sum(row["net_amount"] for row in rows)), ""])
    table = Table(table_rows, repeatRows=1, colWidths=[12 * mm, 20 * mm, 44 * mm, 37 * mm, 25 * mm, 22 * mm, 22 * mm, 25 * mm, 27 * mm, 35 * mm])
    set_pdf_table_style(table, 3)
    table.setStyle(TableStyle([
        ("ALIGN", (0, 1), (1, -1), "CENTER"), ("ALIGN", (4, 1), (8, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("TOPPADDING", (0, 1), (-1, -2), 6), ("BOTTOMPADDING", (0, 1), (-1, -2), 6),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5EE")), ("FONTNAME", (0, -1), (-1, -1), THAI_FONT_BOLD),
    ]))
    story.append(table)

    def draw_page(canvas_obj, doc):
        page_width, _ = landscape(A4)
        canvas_obj.saveState()
        canvas_obj.setStrokeColor(colors.HexColor(BRAND_GREEN))
        canvas_obj.line(12 * mm, 12 * mm, page_width - 12 * mm, 12 * mm)
        canvas_obj.setFont(THAI_FONT, 7)
        canvas_obj.setFillColor(colors.HexColor("#475467"))
        canvas_obj.drawString(12 * mm, 7 * mm, f"PF Accounting | {label} | {format_report_date(start_date)} - {format_report_date(end_date)}")
        canvas_obj.drawRightString(page_width - 12 * mm, 7 * mm, f"หน้า {doc.page}")
        canvas_obj.restoreState()

    document.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
    return buffer.getvalue()


def inbound_export_rows(payload: dict) -> list[dict]:
    rows = payload.get("rows", [])
    if not isinstance(rows, list):
        return []
    cleaned = []
    for row in rows[:200]:
        if not isinstance(row, dict):
            continue
        cleaned.append(row)
    return cleaned


def inbound_export_datetime(value: object) -> str:
    try:
        parsed = datetime.fromisoformat(str(value or "").replace("Z", "+00:00"))
        return parsed.astimezone(timezone(timedelta(hours=7))).strftime("%d/%m/%Y %H:%M")
    except (TypeError, ValueError):
        return str(value or "-")


def build_inbound_selected_excel(payload: dict) -> bytes:
    rows = inbound_export_rows(payload)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "รายการรับเข้า"
    headers = ["ลำดับ", "วันเวลา", "เลขที่รายการ", "ผู้ส่ง", "ผลไม้", "น้ำหนัก (กก.)", "ราคา/กก.", "ยอดเงิน (บาท)", "หมายเหตุ", "ผู้บันทึก"]
    sheet.merge_cells("A1:J1")
    sheet["A1"] = COMPANY_NAME
    sheet["A1"].font = Font(name="Sarabun", bold=True, size=18, color="0F7A3D")
    sheet.merge_cells("A2:J2")
    sheet["A2"] = "ฟอร์มรวมรายการรับเข้าที่เลือก"
    sheet["A2"].font = Font(name="Sarabun", bold=True, size=15, color="111827")
    sheet.merge_cells("A3:J3")
    sheet["A3"] = f"จำนวน {len(rows)} รายการ | พิมพ์เมื่อ {datetime.now(timezone(timedelta(hours=7))).strftime('%d/%m/%Y %H:%M')} | ผู้พิมพ์ {payload.get('printed_by') or '-'}"
    sheet.append([])
    sheet.append(headers)
    header_row = 5
    for index, row in enumerate(rows, 1):
        sheet.append([
            index, inbound_export_datetime(row.get("received_at")), f"IN-{int(row.get('id') or 0):06d}",
            row.get("supplier_name") or "-", row.get("fruit_name") or "-", safe_float(row.get("weight_kg")),
            safe_float(row.get("price_per_kg")), safe_float(row.get("total_amount")), row.get("note") or "-", row.get("created_by") or "-",
        ])
    total_row = sheet.max_row + 1
    sheet.cell(total_row, 1, "รวมรายการที่เลือก")
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    sheet.cell(total_row, 6, sum(safe_float(row.get("weight_kg")) for row in rows))
    sheet.cell(total_row, 8, sum(safe_float(row.get("total_amount")) for row in rows))
    green, white, line = "0F766E", "FFFFFF", "D0D5DD"
    for cell in sheet[header_row]:
        cell.font = Font(name="Sarabun", bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=green)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color=line)
    for row_cells in sheet.iter_rows(min_row=header_row, max_row=total_row, min_col=1, max_col=10):
        for cell in row_cells:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row > header_row:
                cell.font = Font(name="Sarabun", bold=cell.row == total_row, size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in (6, 7, 8):
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=header_row + 1, max_row=total_row):
            cell[0].number_format = '#,##0.00'
    widths = [8, 20, 16, 28, 18, 16, 14, 18, 32, 18]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = f"A{header_row}:J{max(header_row, total_row - 1)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_inbound_selected_pdf(payload: dict) -> bytes:
    rows = inbound_export_rows(payload)
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10 * mm, leftMargin=10 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = THAI_FONT
    styles["Title"].fontName = THAI_FONT_BOLD
    story = [Paragraph("ฟอร์มรวมรายการรับเข้าที่เลือก", styles["Title"]), Paragraph(f"{COMPANY_NAME} | {len(rows)} รายการ | ผู้พิมพ์ {xml_escape(str(payload.get('printed_by') or '-'))}", styles["Normal"]), Spacer(1, 5 * mm)]
    table_rows = [["ลำดับ", "วันเวลา", "เลขที่", "ผู้ส่ง", "ผลไม้", "น้ำหนัก", "ราคา/กก.", "ยอดเงิน", "หมายเหตุ", "ผู้บันทึก"]]
    for index, row in enumerate(rows, 1):
        table_rows.append([
            index, inbound_export_datetime(row.get("received_at")), f"IN-{int(row.get('id') or 0):06d}",
            Paragraph(xml_escape(str(row.get("supplier_name") or "-")), styles["Normal"]), str(row.get("fruit_name") or "-"),
            f"{safe_float(row.get('weight_kg')):,.2f}", f"{safe_float(row.get('price_per_kg')):,.2f}", f"{safe_float(row.get('total_amount')):,.2f}",
            Paragraph(xml_escape(str(row.get("note") or "-")), styles["Normal"]), str(row.get("created_by") or "-"),
        ])
    table_rows.append(["รวม", "", "", "", f"{len(rows)} รายการ", f"{sum(safe_float(row.get('weight_kg')) for row in rows):,.2f}", "", f"{sum(safe_float(row.get('total_amount')) for row in rows):,.2f}", "", ""])
    table = Table(table_rows, colWidths=[12*mm, 27*mm, 20*mm, 35*mm, 20*mm, 22*mm, 19*mm, 24*mm, 43*mm, 24*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), THAI_FONT_BOLD), ("FONTNAME", (0, 1), (-1, -1), THAI_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 7), ("GRID", (0, 0), (-1, -1), .4, colors.HexColor("#D0D5DD")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (5, 1), (7, -1), "RIGHT"), ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#ECFDF3")),
        ("FONTNAME", (0, -1), (-1, -1), THAI_FONT_BOLD),
    ]))
    story.append(table)
    document.build(story)
    return buffer.getvalue()


class ReportHandler(BaseHTTPRequestHandler):
    def end_headers(self) -> None:
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, X-Backup-Code, X-Session-Token")
        super().end_headers()

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.end_headers()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        payload = self.read_json()

        if parsed.path == "/api/inbound/fruits":
            actor = inbound_authorized_actor(self)
            if not actor:
                self.send_json({"error": "C5 or higher is required for inbound receiving."}, 403)
                return
            name = inbound_clean_text(payload.get("name"), 100)
            if not name:
                self.send_json({"error": "กรุณากรอกชื่อผลไม้"}, 400)
                return
            row = {"name": name, "normalized_name": name.casefold(), "status": "Active", "created_by": inbound_actor_name(actor)}
            status, body = supabase_request("POST", "inbound_fruits", row, prefer="return=representation")
            if status >= 400:
                self.send_json({"error": "ชื่อผลไม้นี้มีอยู่แล้ว" if status == 409 else body, "migration": "supabase_inbound_receiving_migration.sql"}, status)
                return
            saved = body[0] if isinstance(body, list) and body else body
            inbound_audit(self, actor, "CREATE_INBOUND_FRUIT", f"เพิ่มผลไม้รับเข้า: {name}", {"fruit_id": saved.get("id") if isinstance(saved, dict) else None})
            self.send_json({"data": saved}, 201)
            return

        if parsed.path == "/api/inbound/prices":
            actor = inbound_authorized_actor(self)
            if not actor:
                self.send_json({"error": "C5 or higher is required for inbound receiving."}, 403)
                return
            try:
                fruit_id = int(payload.get("fruit_id") or 0)
                price = round(float(payload.get("price_per_kg") or 0), 2)
                effective_date = str(payload.get("effective_date") or "")
                datetime.strptime(effective_date, "%Y-%m-%d")
            except (TypeError, ValueError):
                self.send_json({"error": "ข้อมูลผลไม้ ราคา หรือวันที่เริ่มใช้ไม่ถูกต้อง"}, 400)
                return
            if fruit_id <= 0 or price <= 0 or price > 1000000:
                self.send_json({"error": "ราคาต่อกิโลต้องมากกว่า 0"}, 400)
                return
            fruit_status, fruits = supabase_request("GET", f"inbound_fruits?id=eq.{fruit_id}&status=eq.Active&select=id,name&limit=1")
            if fruit_status >= 400 or not isinstance(fruits, list) or not fruits:
                self.send_json({"error": "ไม่พบผลไม้ที่เปิดใช้งาน"}, 404)
                return
            row = {"fruit_id": fruit_id, "price_per_kg": price, "effective_date": effective_date, "note": inbound_clean_text(payload.get("note"), 500), "created_by": inbound_actor_name(actor)}
            status, body = supabase_request("POST", "inbound_fruit_prices", row, prefer="return=representation")
            if status >= 400:
                self.send_json({"error": body, "migration": "supabase_inbound_receiving_migration.sql"}, status)
                return
            saved = body[0] if isinstance(body, list) and body else body
            inbound_audit(self, actor, "SET_INBOUND_PRICE", f"ตั้งราคาแนะนำ {fruits[0]['name']} {price:.2f} บาท/กก.", {"fruit_id": fruit_id, "price_per_kg": price, "effective_date": effective_date})
            self.send_json({"data": saved}, 201)
            return

        if parsed.path == "/api/inbound/receipts":
            actor = inbound_authorized_actor(self)
            if not actor:
                self.send_json({"error": "C5 or higher is required for inbound receiving."}, 403)
                return
            try:
                fruit_id = int(payload.get("fruit_id") or 0)
                weight = round(float(payload.get("weight_kg") or 0), 2)
                price = round(float(payload.get("price_per_kg") or 0), 2)
                received_at = datetime.fromisoformat(str(payload.get("received_at") or "").replace("Z", "+00:00"))
            except (TypeError, ValueError):
                self.send_json({"error": "ข้อมูลวันที่ ผลไม้ น้ำหนัก หรือราคาไม่ถูกต้อง"}, 400)
                return
            supplier = inbound_clean_text(payload.get("supplier_name"), 160)
            client_uid = inbound_clean_text(payload.get("client_uid"), 160)
            if not supplier or not re.fullmatch(r"[A-Za-z0-9._:-]{8,160}", client_uid) or fruit_id <= 0 or weight <= 0 or price <= 0:
                self.send_json({"error": "กรุณากรอกผู้ส่ง ผลไม้ น้ำหนัก และราคาให้ครบ"}, 400)
                return
            if weight > 100000000 or price > 1000000:
                self.send_json({"error": "น้ำหนักหรือราคาเกินขอบเขตที่รองรับ"}, 400)
                return
            fruit_status, fruits = supabase_request("GET", f"inbound_fruits?id=eq.{fruit_id}&status=eq.Active&select=id,name&limit=1")
            if fruit_status >= 400 or not isinstance(fruits, list) or not fruits:
                self.send_json({"error": "ไม่พบผลไม้ที่เปิดใช้งาน"}, 404)
                return
            total = round(weight * price + 1e-9, 2)
            row = {
                "client_uid": client_uid, "received_at": received_at.astimezone(timezone.utc).isoformat(),
                "supplier_name": supplier, "fruit_id": fruit_id, "fruit_name": fruits[0]["name"],
                "weight_kg": weight, "price_per_kg": price, "total_amount": total,
                "note": inbound_clean_text(payload.get("note"), 1000), "created_by": inbound_actor_name(actor),
            }
            status, body = supabase_request("POST", "inbound_receipts", row, prefer="return=representation")
            if status >= 400:
                existing_status, existing = supabase_request("GET", f"inbound_receipts?client_uid=eq.{quote(client_uid)}&select=*&limit=1")
                if existing_status < 400 and isinstance(existing, list) and existing:
                    self.send_json({"data": existing[0], "idempotent": True})
                    return
                self.send_json({"error": body, "migration": "supabase_inbound_receiving_migration.sql"}, status)
                return
            saved = body[0] if isinstance(body, list) and body else body
            audit_status, audit_body = inbound_audit(self, actor, "CREATE_INBOUND_RECEIPT", f"รับเข้า {fruits[0]['name']} จาก {supplier} {weight:.2f} กก. @ {price:.2f} = {total:.2f} บาท", {"receipt_id": saved.get("id") if isinstance(saved, dict) else None, "client_uid": client_uid})
            if audit_status >= 400 and isinstance(saved, dict) and saved.get("id"):
                supabase_request("DELETE", f"inbound_receipts?id=eq.{saved['id']}", prefer="return=minimal")
                self.send_json({"error": "บันทึก Audit Log ไม่สำเร็จ รายการจึงถูกย้อนกลับ", "audit_error": audit_body}, 500)
                return
            self.send_json({"data": saved}, 201)
            return

        if parsed.path == "/api/online-users":
            self.send_json({"data": register_online_user(payload)})
            return

        if parsed.path == "/api/issue-reports":
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            report, validation_error = validate_issue_report_payload(payload)
            if validation_error:
                self.send_json({"error": validation_error}, 400)
                return
            actor_username = str(actor.get("username") or "").strip()
            account_status, accounts = supabase_request(
                "GET",
                f"account_users?username=eq.{quote(actor_username)}&select=fullname,role,user_level&limit=1",
            )
            account = accounts[0] if account_status < 400 and isinstance(accounts, list) and accounts else {}
            row = {
                **report,
                "status": "received",
                "reporter_username": actor_username,
                "reporter_fullname": str(account.get("fullname") or actor_username),
                "reporter_role": str(account.get("role") or actor.get("role") or ""),
            }
            status, body = supabase_request("POST", "issue_reports", row, prefer="return=representation")
            self.send_json({"data": body[0] if status < 400 and isinstance(body, list) and body else body}, status)
            return

        if parsed.path == "/api/issue-reports/status":
            actor = secret_room_actor(self)
            if not actor or account_level_number(actor.get("level")) < 5:
                self.send_json({"error": "Administrator access is required."}, 403)
                return
            try:
                report_id = int(payload.get("id") or 0)
            except (TypeError, ValueError):
                report_id = 0
            next_status = str(payload.get("status") or "").strip()
            if report_id < 1 or next_status not in ISSUE_REPORT_STATUSES:
                self.send_json({"error": "Report id or status is invalid."}, 400)
                return
            now = datetime.now(timezone.utc).isoformat()
            update = {
                "status": next_status,
                "updated_at": now,
                "resolved_at": now if next_status == "resolved" else None,
                "assigned_to": str(actor.get("username") or ""),
            }
            status, body = supabase_request(
                "PATCH",
                f"issue_reports?id=eq.{report_id}",
                update,
                prefer="return=representation",
            )
            self.send_json({"data": body[0] if status < 400 and isinstance(body, list) and body else body}, status)
            return

        if parsed.path.startswith("/api/secret-room/"):
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "Session is missing or expired."}, 401)
                return

            actor_username = str(actor.get("username") or "").strip()
            if parsed.path == "/api/secret-room/posts":
                content = str(payload.get("content") or "").strip()
                if not content or len(content) > 2000:
                    self.send_json({"error": "Post content must contain 1-2,000 characters."}, 400)
                    return
                account_status, accounts = supabase_request(
                    "GET",
                    f"account_users?username=eq.{quote(actor_username)}&select=fullname&limit=1",
                )
                fullname = actor_username
                if account_status < 400 and isinstance(accounts, list) and accounts:
                    fullname = str(accounts[0].get("fullname") or actor_username)
                status, body = supabase_request(
                    "POST",
                    "community_posts",
                    {"author_username": actor_username, "author_fullname": fullname, "content": content},
                    prefer="return=representation",
                )
                self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
                return

            if parsed.path == "/api/secret-room/messages":
                recipient = str(payload.get("recipient_username") or "").strip()
                content = str(payload.get("content") or "").strip()
                if not recipient or recipient.lower() == actor_username.lower():
                    self.send_json({"error": "Please choose another coworker."}, 400)
                    return
                if not content or len(content) > 4000:
                    self.send_json({"error": "Message must contain 1-4,000 characters."}, 400)
                    return
                account_status, accounts = supabase_request(
                    "GET",
                    f"account_users?username=eq.{quote(recipient)}&status=eq.Active&select=username&limit=1",
                )
                if account_status >= 400:
                    self.send_json({"error": accounts}, account_status)
                    return
                if not isinstance(accounts, list) or not accounts:
                    self.send_json({"error": "Coworker account was not found."}, 404)
                    return
                status, body = supabase_request(
                    "POST",
                    "secret_messages",
                    {"sender_username": actor_username, "recipient_username": recipient, "content": content},
                    prefer="return=representation",
                )
                self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
                return

            if parsed.path == "/api/secret-room/messages/read":
                sender = str(payload.get("username") or "").strip()
                if not sender:
                    self.send_json({"error": "username is required."}, 400)
                    return
                status, body = supabase_request(
                    "PATCH",
                    f"secret_messages?sender_username=eq.{quote(sender)}&recipient_username=eq.{quote(actor_username)}&is_read=eq.false",
                    {"is_read": True},
                    prefer="return=minimal",
                )
                self.send_json({"data": {"updated": status < 400}, "error": body if status >= 400 else None}, status)
                return

            self.send_error(404, "Not found")
            return

        if parsed.path == "/api/accounting/journals":
            actor = accounting_actor(self)
            if not actor:
                self.send_json({"error": "Accounting session is missing or expired."}, 401)
                return
            lines = payload.get("lines", [])
            if not isinstance(lines, list):
                self.send_json({"error": "Journal lines must be an array."}, 400)
                return
            rpc_payload = {
                "p_company_key": ACCOUNTING_COMPANY_KEY,
                "p_entry_date": payload.get("date"),
                "p_journal_type": payload.get("journal_type", "general"),
                "p_reference": payload.get("reference", ""),
                "p_description": payload.get("description", ""),
                "p_document_no": payload.get("document_no", ""),
                "p_lines": [
                    {
                        "account_id": line.get("accountId") or line.get("account_id"),
                        "description": line.get("memo") or line.get("description", ""),
                        "debit": line.get("debit", 0),
                        "credit": line.get("credit", 0),
                        "partner_id": line.get("partner_id"),
                        "due_date": line.get("due_date"),
                        "tax_code": line.get("tax_code"),
                        "cost_center": line.get("cost_center"),
                        "production_batch": line.get("production_batch"),
                        "project_code": line.get("project_code"),
                    }
                    for line in lines if isinstance(line, dict)
                ],
                "p_actor": actor.get("username", "unknown"),
                "p_submit": payload.get("intent") != "draft",
            }
            status, body = supabase_request("POST", "rpc/ac_create_journal", rpc_payload)
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path in {"/api/accounting/journals/approve", "/api/accounting/journals/reject"}:
            actor = accounting_actor(self, 5)
            if not actor:
                self.send_json({"error": "C5 or higher accounting session is required."}, 403)
                return
            journal_id = str(payload.get("journal_id", "")).strip()
            if not journal_id:
                self.send_json({"error": "journal_id is required."}, 400)
                return
            if parsed.path.endswith("/approve"):
                rpc_name = "rpc/ac_approve_journal"
                rpc_payload = {"p_journal_id": journal_id, "p_actor": actor.get("username"), "p_actor_level": actor.get("level")}
            else:
                rpc_name = "rpc/ac_reject_journal"
                rpc_payload = {"p_journal_id": journal_id, "p_actor": actor.get("username"), "p_actor_level": actor.get("level"), "p_reason": payload.get("reason", "")}
            status, body = supabase_request("POST", rpc_name, rpc_payload)
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/accounting/periods/close":
            actor = accounting_actor(self, 5)
            if not actor:
                self.send_json({"error": "C5 or higher accounting session is required."}, 403)
                return
            status, body = supabase_request("POST", "rpc/ac_close_period", {"p_period_id": payload.get("period_id"), "p_actor": actor.get("username")})
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/accounting/accounts":
            actor = accounting_actor(self, 5)
            if not actor:
                self.send_json({"error": "C5 or higher accounting session is required."}, 403)
                return
            company_status, companies = supabase_request("GET", f"ac_companies?company_key=eq.{quote(ACCOUNTING_COMPANY_KEY)}&select=id&limit=1")
            if company_status >= 400 or not isinstance(companies, list) or not companies:
                self.send_json({"error": companies if company_status >= 400 else "Accounting company is not initialized."}, company_status if company_status >= 400 else 404)
                return
            account_type = str(payload.get("type", ""))
            if account_type not in {"asset", "liability", "equity", "revenue", "expense"}:
                self.send_json({"error": "Invalid account type."}, 422)
                return
            normal_side = "debit" if account_type in {"asset", "expense"} else "credit"
            if bool(payload.get("contra")):
                normal_side = "credit" if normal_side == "debit" else "debit"
            row = {"company_id": companies[0]["id"], "code": str(payload.get("code", "")).strip(), "name_th": str(payload.get("name", "")).strip(), "account_type": account_type, "normal_side": normal_side, "is_contra": bool(payload.get("contra")), "active": True, "system_account": False}
            if not row["code"] or not row["name_th"]:
                self.send_json({"error": "Account code and name are required."}, 400)
                return
            status, body = supabase_request("POST", "ac_accounts", row, prefer="return=representation")
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/reports/sync":
            save_data(payload)
            self.send_json({"status": "ok"})
            return

        if parsed.path == "/api/cleanup-test-data":
            # These codes belong solely to the original demo dataset. They
            # were never created as central employee records and must not be
            # mixed into payroll or production history.
            demo_codes = "10001,10002,10003,10021,10025,10031,10044,10052,10067"
            results = {}
            for table in ("production_records", "time_records"):
                status, body = supabase_request(
                    "DELETE",
                    f"{table}?emp_code=in.({demo_codes})",
                    prefer="return=representation",
                )
                if status >= 400:
                    self.send_json({"error": body, "table": table}, status)
                    return
                results[table] = len(body) if isinstance(body, list) else 0
            self.send_json({"deleted": results})
            return

        if parsed.path == "/api/production-records/delete-by-date":
            record_date = str(payload.get("record_date", "")).strip()
            try:
                datetime.strptime(record_date, "%Y-%m-%d")
            except ValueError:
                self.send_json({"error": "record_date must be YYYY-MM-DD"}, 400)
                return
            status, body = supabase_request(
                "DELETE",
                f"production_records?record_date=eq.{quote(record_date)}",
                prefer="return=representation",
            )
            if status >= 400:
                self.send_json({"error": body}, status)
                return
            self.send_json({"deleted": len(body) if isinstance(body, list) else 0, "record_date": record_date})
            return

        if parsed.path == "/api/admin/remove-superseded-production-records":
            if not backup_authorized(self):
                self.send_json({"error": "Backup code is required."}, 403)
                return
            try:
                record_ids = sorted({int(value) for value in payload.get("record_ids", []) if int(value) > 0})
            except (TypeError, ValueError):
                self.send_json({"error": "record_ids must contain positive integers."}, 400)
                return
            expected_count = int(payload.get("expected_count", 0) or 0)
            expected_date = str(payload.get("expected_date", "")).strip()
            expected_creators = {
                str(value).strip()
                for value in payload.get("expected_creators", [])
                if str(value).strip()
            }
            if not record_ids or expected_count != len(record_ids) or not expected_date or not expected_creators:
                self.send_json({"error": "Complete deletion safeguards are required."}, 400)
                return

            id_filter = ",".join(str(value) for value in record_ids)
            status, existing = supabase_request(
                "GET",
                f"production_records?id=in.({id_filter})&select=id,record_date,created_by&order=id.asc",
            )
            if status >= 400:
                self.send_json({"error": existing}, status)
                return
            invalid_rows = [
                row
                for row in (existing if isinstance(existing, list) else [])
                if row.get("record_date") != expected_date
                or str(row.get("created_by") or "") not in expected_creators
            ]
            if not isinstance(existing, list) or len(existing) != expected_count or invalid_rows:
                self.send_json(
                    {
                        "error": "Deletion safeguards did not match the live rows.",
                        "found_count": len(existing) if isinstance(existing, list) else 0,
                        "invalid_rows": invalid_rows,
                    },
                    409,
                )
                return

            status, deleted = supabase_request(
                "DELETE",
                f"production_records?id=in.({id_filter})",
                prefer="return=representation",
            )
            if status >= 400:
                self.send_json({"error": deleted}, status)
                return
            self.send_json(
                {
                    "deleted": len(deleted) if isinstance(deleted, list) else 0,
                    "record_ids": record_ids,
                }
            )
            return

        if parsed.path == "/api/production-save-queue/enqueue":
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            records = payload.get("records", [])
            queue_uid = str(payload.get("queue_uid") or "").strip()
            if not isinstance(records, list) or not (1 <= len(records) <= 100):
                self.send_json({"error": "records must contain 1-100 items."}, 400)
                return
            if not re.fullmatch(r"[A-Za-z0-9._:-]{8,160}", queue_uid):
                self.send_json({"error": "queue_uid is invalid."}, 400)
                return
            converted = []
            actor_username = str(actor.get("username") or "unknown").strip()
            for record in records:
                if not isinstance(record, dict):
                    self.send_json({"error": "all records must be objects."}, 400)
                    return
                row = live_state_row("production_records", record)
                row.pop("id", None)
                row["queue_dedupe_key"] = production_queue_dedupe_key(row)
                row["created_by"] = actor_username
                row["updated_by"] = actor_username
                raw = row.get("raw_payload") if isinstance(row.get("raw_payload"), dict) else {}
                raw["created_by"] = actor_username
                raw["updated_by"] = actor_username
                row["raw_payload"] = raw
                converted.append(row)
            client_uids = [production_record_client_uid(row) for row in converted]
            if any(not uid for uid in client_uids) or len(set(client_uids)) != len(client_uids):
                self.send_json({"error": "Every queued record requires a unique client_uid."}, 400)
                return
            identity_keys = {
                (
                    str(row.get("record_date") or ""),
                    str(row.get("emp_code") or ""),
                    str(row.get("fruit_type") or "mangosteen"),
                )
                for row in converted
            }
            if len(identity_keys) != 1:
                self.send_json({"error": "One queue must contain one employee, date, and fruit type."}, 400)
                return
            record_date, emp_code, fruit_type = next(iter(identity_keys))
            try:
                datetime.strptime(record_date, "%Y-%m-%d")
            except ValueError:
                self.send_json({"error": "record_date must be YYYY-MM-DD."}, 400)
                return
            payload_hash = production_queue_payload_hash(converted)
            first = converted[0]
            queue_row = {
                "queue_uid": queue_uid,
                "batch_uid": production_record_batch_uid(first) or None,
                "payload": converted,
                "payload_hash": payload_hash,
                "record_count": len(converted),
                "fruit_type": fruit_type,
                "record_date": record_date,
                "employee_id": first.get("employee_id"),
                "emp_code": emp_code,
                "employee_name": str(first.get("employee_name") or ""),
                "total_weight": round(sum(production_record_weight_total(row) for row in converted), 3),
                "total_amount": round(sum(safe_float(row.get("amount", row.get("raw_payload", {}).get("total_amount", 0))) for row in converted), 2),
                "status": "queued",
                "created_by": actor_username,
            }
            status, body = supabase_request(
                "POST",
                "production_save_queue",
                queue_row,
                prefer="return=representation",
                timeout_seconds=5,
            )
            if status >= 400:
                existing_status, existing = supabase_request(
                    "GET",
                    f"production_save_queue?queue_uid=eq.{quote(queue_uid)}&select=*&limit=1",
                    timeout_seconds=5,
                )
                if existing_status < 400 and isinstance(existing, list) and existing:
                    existing_row = existing[0]
                    if hmac.compare_digest(str(existing_row.get("payload_hash") or ""), payload_hash):
                        production_save_queue_wakeup.set()
                        self.send_json({"data": production_queue_to_client(existing_row), "idempotent": True}, 200)
                        return
                    self.send_json({"error": "queue_uid already exists with different data."}, 409)
                    return
                self.send_json({"error": body, "migration": "supabase_production_save_queue_migration.sql"}, status)
                return
            saved = body[0] if isinstance(body, list) and body else None
            if not isinstance(saved, dict):
                self.send_json({"error": "Queue insert returned no row."}, 500)
                return
            production_save_queue_wakeup.set()
            self.send_json({"data": production_queue_to_client(saved), "accepted": True}, 202)
            return

        queue_action_match = re.fullmatch(r"/api/production-save-queue/(\d+)/(verify|retry|cancel|link-existing)", parsed.path)
        if queue_action_match:
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            queue_id = int(queue_action_match.group(1))
            action = queue_action_match.group(2)
            status, rows = supabase_request(
                "GET",
                f"production_save_queue?id=eq.{queue_id}&select=*&limit=1",
                timeout_seconds=6,
            )
            if status >= 400:
                self.send_json({"error": rows}, status)
                return
            if not isinstance(rows, list) or not rows:
                self.send_json({"error": "Queue item was not found."}, 404)
                return
            job = rows[0]
            actor_username = str(actor.get("username") or "unknown").strip()
            is_owner = str(job.get("created_by") or "").casefold() == actor_username.casefold()
            is_supervisor = account_level_number(actor.get("level")) >= 4

            if action == "verify":
                if job.get("status") == "succeeded":
                    self.send_json({"data": production_queue_to_client(job, include_payload=True)})
                    return
                if job.get("status") in {"queued", "processing"}:
                    self.send_json({"data": production_queue_to_client(job), "processing": True}, 202)
                    return
                if job.get("status") == "cancelled":
                    self.send_json({"error": "Cancelled queues cannot be verified or restored."}, 409)
                    return
                payload_rows = job.get("payload") if isinstance(job.get("payload"), list) else []
                match_type, matched_rows, details = inspect_production_queue_rows(payload_rows)
                if match_type == "error":
                    self.send_json({"error": details}, int(details.get("status") or 500))
                    return
                if match_type == "exact":
                    values = production_queue_success_values(matched_rows)
                    update_status, updated = update_production_queue(queue_id, values)
                    if update_status >= 400 or not updated:
                        self.send_json({"error": "Could not confirm the queue result."}, update_status if update_status >= 400 else 500)
                        return
                    production_queue_event(queue_id, "manual_verified", "succeeded", "Existing production records matched the queue.", actor_username, details)
                    self.send_json({"data": production_queue_to_client(updated, include_payload=True), "match": "exact"})
                    return
                error_code = "possible_duplicate" if match_type == "near" else "not_found"
                error_message = (
                    str(details.get("error") or "พบข้อมูลใกล้เคียง ต้องตรวจสอบก่อนสร้างรายการใหม่")
                    if match_type == "near"
                    else "ไม่พบข้อมูลชุดนี้ในฐานข้อมูล"
                )
                update_status, updated = update_production_queue(
                    queue_id,
                    {
                        "status": "needs_review",
                        "duplicate_details": details,
                        "error_code": error_code,
                        "error_message": error_message,
                        "finished_at": datetime.now(timezone.utc).isoformat(),
                        "locked_at": None,
                        "locked_by": None,
                    },
                    expected_status=str(job.get("status") or "needs_review"),
                )
                if update_status >= 400 or not updated:
                    self.send_json({"error": "Could not update queue verification."}, update_status if update_status >= 400 else 500)
                    return
                production_queue_event(queue_id, "manual_checked", "needs_review", error_message, actor_username, details)
                self.send_json({"data": production_queue_to_client(updated, include_payload=True), "match": match_type})
                return

            if not (is_owner or is_supervisor):
                self.send_json({"error": "Only the queue owner or C4 and higher may perform this action."}, 403)
                return
            if action == "retry":
                if job.get("status") not in {"needs_review", "failed"}:
                    self.send_json({"error": "Only failed or review queues can be submitted again."}, 409)
                    return
                if str(job.get("error_code") or "") == "possible_duplicate":
                    self.send_json({"error": "Similar data already exists. Verify and link the existing records or cancel this queue."}, 409)
                    return
                update_status, updated = update_production_queue(
                    queue_id,
                    {
                        "status": "queued",
                        "attempt_count": 0,
                        "next_attempt_at": datetime.now(timezone.utc).isoformat(),
                        "finished_at": None,
                        "error_code": None,
                        "error_message": None,
                        "locked_at": None,
                        "locked_by": None,
                    },
                    expected_status=str(job.get("status") or "queued"),
                )
                if update_status >= 400 or not updated:
                    self.send_json({"error": "Could not return the item to the queue."}, update_status if update_status >= 400 else 500)
                    return
                production_queue_event(queue_id, "manual_retry", "queued", "Queue was submitted again with the original payload.", actor_username)
                production_save_queue_wakeup.set()
                self.send_json({"data": production_queue_to_client(updated)})
                return
            if action == "cancel":
                if job.get("status") in {"processing", "succeeded", "cancelled"}:
                    self.send_json({"error": "Processing, completed, or cancelled queues cannot be cancelled."}, 409)
                    return
                update_status, updated = update_production_queue(
                    queue_id,
                    {
                        "status": "cancelled",
                        "cancelled_by": actor_username,
                        "cancelled_at": datetime.now(timezone.utc).isoformat(),
                        "finished_at": datetime.now(timezone.utc).isoformat(),
                        "locked_at": None,
                        "locked_by": None,
                    },
                    expected_status=str(job.get("status") or "queued"),
                )
                if update_status >= 400 or not updated:
                    self.send_json({"error": "Could not cancel the queue."}, update_status if update_status >= 400 else 500)
                    return
                production_queue_event(queue_id, "cancelled", "cancelled", str(payload.get("reason") or "Cancelled by user."), actor_username)
                self.send_json({"data": production_queue_to_client(updated)})
                return
            if not is_supervisor:
                self.send_json({"error": "C4 or higher is required to link an existing record."}, 403)
                return
            if job.get("status") != "needs_review" or str(job.get("error_code") or "") == "not_found":
                self.send_json({"error": "Only verified duplicate candidates can be linked."}, 409)
                return
            candidate_ids = {
                int(value)
                for value in (job.get("duplicate_details") or {}).get("existing_ids", [])
                if str(value).isdigit() and int(value) > 0
            }
            requested_ids = {
                int(value)
                for value in payload.get("record_ids", [])
                if str(value).isdigit() and int(value) > 0
            }
            if not requested_ids or not requested_ids.issubset(candidate_ids):
                self.send_json({"error": "Selected records are not verified duplicate candidates."}, 400)
                return
            record_status, existing_records = supabase_request(
                "GET",
                f"production_records?id=in.({','.join(str(value) for value in sorted(requested_ids))})&select=*&order=id.asc",
                timeout_seconds=6,
            )
            if record_status >= 400 or not isinstance(existing_records, list) or len(existing_records) != len(requested_ids):
                self.send_json({"error": existing_records if record_status >= 400 else "Existing records are incomplete."}, record_status if record_status >= 400 else 409)
                return
            update_status, updated = update_production_queue(
                queue_id,
                production_queue_success_values(existing_records),
                expected_status=str(job.get("status") or "needs_review"),
            )
            if update_status >= 400 or not updated:
                self.send_json({"error": "Could not link existing records."}, update_status if update_status >= 400 else 500)
                return
            production_queue_event(queue_id, "linked_existing", "succeeded", "C4 linked verified existing production records.", actor_username, {"record_ids": sorted(requested_ids)})
            self.send_json({"data": production_queue_to_client(updated)})
            return

        if parsed.path == "/api/production-records/bulk-sync":
            records = payload.get("records", [])
            if not isinstance(records, list) or not records:
                self.send_json({"error": "records must be a non-empty list"}, 400)
                return
            mode = str(payload.get("mode") or "upsert").strip().lower()
            converted = [
                live_state_row("production_records", record)
                for record in records
                if isinstance(record, dict)
            ]
            if len(converted) != len(records):
                self.send_json({"error": "all records must be objects"}, 400)
                return
            if mode == "insert":
                insert_rows = []
                for row in converted:
                    insert_row = dict(row)
                    insert_row.pop("id", None)
                    insert_rows.append(insert_row)
                with production_record_insert_lock:
                    status, body = insert_production_records_compatible(insert_rows)
                if status >= 400:
                    self.send_json({"data": None, "error": body}, status)
                    return
                synced_rows = [
                    live_state_to_client("production_records", row)
                    for row in body
                    if isinstance(row, dict)
                ] if isinstance(body, list) else []
                self.send_json({"data": synced_rows, "error": None}, status)
                return
            if mode != "upsert":
                self.send_json({"error": "mode must be insert or upsert"}, 400)
                return
            with live_state_sync_lock:
                status, body = sync_rows_by_id("production_records", converted)
            if status >= 400:
                self.send_json({"data": None, "error": body}, status)
                return
            synced_rows = [
                live_state_to_client("production_records", row)
                for row in body.get("synced", [])
                if isinstance(row, dict)
            ]
            self.send_json({"data": synced_rows, "error": None}, status)
            return

        if parsed.path == "/api/time-records/bulk-sync":
            mode = str(payload.get("mode") or "upsert").strip().lower()
            if mode == "delete":
                record_ids = payload.get("record_ids", [])
                if not isinstance(record_ids, list) or not record_ids:
                    self.send_json({"error": "record_ids must be a non-empty list"}, 400)
                    return
                clean_ids = []
                for value in record_ids:
                    try:
                        record_id = int(value)
                    except (TypeError, ValueError):
                        continue
                    if record_id > 0 and record_id not in clean_ids:
                        clean_ids.append(record_id)
                if not clean_ids:
                    self.send_json({"error": "record_ids must include valid ids"}, 400)
                    return
                id_list = ",".join(str(record_id) for record_id in clean_ids)
                status, body = supabase_request(
                    "DELETE",
                    f"time_records?id=in.({id_list})",
                    prefer="return=representation",
                )
                if status >= 400:
                    self.send_json({"data": None, "error": body}, status)
                    return
                self.send_json({"data": body if isinstance(body, list) else [], "error": None}, status)
                return

            records = payload.get("records", [])
            if not isinstance(records, list) or not records:
                self.send_json({"error": "records must be a non-empty list"}, 400)
                return
            converted = [
                live_state_row("time_records", record)
                for record in records
                if isinstance(record, dict)
            ]
            if len(converted) != len(records):
                self.send_json({"error": "all records must be objects"}, 400)
                return
            if mode == "insert":
                insert_rows = []
                for row in converted:
                    insert_row = dict(row)
                    insert_row.pop("id", None)
                    raw_payload = insert_row.get("raw_payload")
                    if isinstance(raw_payload, dict):
                        insert_row["raw_payload"] = {key: value for key, value in raw_payload.items() if key != "id"}
                    insert_rows.append(insert_row)
                with time_record_insert_lock:
                    status, conflict = validate_time_record_conflicts(insert_rows)
                    if status >= 400:
                        self.send_json({"data": None, **(conflict or {"error": "time record conflict"})}, status)
                        return
                    status, body = insert_time_records_compatible(insert_rows)
                if status >= 400:
                    self.send_json({"data": None, "error": body}, status)
                    return
                synced_rows = [
                    live_state_to_client("time_records", row)
                    for row in body
                    if isinstance(row, dict)
                ] if isinstance(body, list) else []
                self.send_json({"data": synced_rows, "error": None}, status)
                return
            if mode != "upsert":
                self.send_json({"error": "mode must be insert, upsert, or delete"}, 400)
                return
            with live_state_sync_lock:
                status, conflict = validate_time_record_conflicts(converted)
                if status >= 400:
                    self.send_json({"data": None, **(conflict or {"error": "time record conflict"})}, status)
                    return
                status, body = update_time_records_compatible(converted)
            if status >= 400:
                self.send_json({"data": None, "error": body}, status)
                return
            synced_rows = [
                live_state_to_client("time_records", row)
                for row in body
                if isinstance(row, dict)
            ] if isinstance(body, list) else []
            self.send_json({"data": synced_rows, "error": None}, status)
            return

        if parsed.path == "/api/time-save-queue/enqueue":
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            queue_uid = str(payload.get("queue_uid") or "").strip()
            operation = str(payload.get("operation") or "insert").strip().lower()
            records = payload.get("records")
            if not queue_uid or operation not in {"insert", "update"}:
                self.send_json({"error": "queue_uid and a valid operation are required."}, 400)
                return
            if not isinstance(records, list) or not records or len(records) > 31:
                self.send_json({"error": "records must contain 1-31 items."}, 400)
                return
            converted = [live_state_row("time_records", row) for row in records if isinstance(row, dict)]
            if len(converted) != len(records):
                self.send_json({"error": "all records must be objects"}, 400)
                return
            if operation == "update" and any(row.get("id") in (None, "") for row in converted):
                self.send_json({"error": "record id is required for queued updates."}, 400)
                return
            if operation == "insert":
                for index, row in enumerate(converted):
                    row.pop("id", None)
                    row["queue_dedupe_key"] = time_queue_row_key(queue_uid, index)
                    raw = row.get("raw_payload") if isinstance(row.get("raw_payload"), dict) else {}
                    row["raw_payload"] = {key: value for key, value in raw.items() if key != "id"}

            identities = [time_record_identity(row) for row in converted]
            employee_codes = {emp_code for _date, emp_code in identities}
            work_dates = sorted({work_date for work_date, _emp_code in identities})
            if len(employee_codes) != 1 or not work_dates or any(not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value) for value in work_dates):
                self.send_json({"error": "one queue must contain one employee and valid work dates."}, 400)
                return
            payload_hash = production_queue_payload_hash(converted)
            emp_code = next(iter(employee_codes))
            first_raw = converted[0].get("raw_payload") if isinstance(converted[0].get("raw_payload"), dict) else {}
            queue_row = {
                "queue_uid": queue_uid, "operation": operation, "payload": converted,
                "payload_hash": payload_hash, "record_count": len(converted), "emp_code": emp_code,
                "employee_name": str(converted[0].get("employee_name") or first_raw.get("fullname") or ""),
                "first_work_date": work_dates[0], "last_work_date": work_dates[-1],
                "status": "queued", "created_by": str(actor.get("username") or actor.get("fullname") or "unknown"),
            }
            status, body = supabase_request(
                "POST", "time_save_queue", queue_row, prefer="return=representation", timeout_seconds=5,
            )
            if status >= 400:
                lookup_status, existing = supabase_request(
                    "GET", f"time_save_queue?queue_uid=eq.{quote(queue_uid)}&select=*&limit=1", timeout_seconds=5,
                )
                if lookup_status < 400 and isinstance(existing, list) and existing:
                    if hmac.compare_digest(str(existing[0].get("payload_hash") or ""), payload_hash):
                        schedule_time_queue_job(int(existing[0].get("id") or 0))
                        self.send_json({"data": time_queue_to_client(existing[0]), "idempotent": True})
                        return
                    self.send_json({"error": "queue_uid already exists with different data."}, 409)
                    return
                self.send_json({"error": body, "migration": "supabase_time_save_queue_migration.sql"}, status)
                return
            saved = body[0] if isinstance(body, list) and body else None
            if not isinstance(saved, dict):
                self.send_json({"error": "Queue insert returned no row."}, 500)
                return
            schedule_time_queue_job(int(saved.get("id") or 0))
            self.send_json({"data": time_queue_to_client(saved), "accepted": True}, 202)
            return

        time_queue_action_match = re.fullmatch(r"/api/time-save-queue/(\d+)/(verify|retry|edit-retry)", parsed.path)
        if time_queue_action_match:
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            queue_id = int(time_queue_action_match.group(1))
            action = time_queue_action_match.group(2)
            status, queue_rows = supabase_request(
                "GET", f"time_save_queue?id=eq.{queue_id}&select=*&limit=1", timeout_seconds=6,
            )
            if status >= 400:
                self.send_json({"error": queue_rows}, status)
                return
            if not isinstance(queue_rows, list) or not queue_rows:
                self.send_json({"error": "Time queue item was not found."}, 404)
                return
            job = queue_rows[0]
            if str(job.get("status") or "") == "succeeded":
                self.send_json({"data": time_queue_to_client(job, include_payload=True), "already_succeeded": True})
                return
            actor_name = str(actor.get("username") or actor.get("fullname") or "unknown")

            if action == "edit-retry":
                records = payload.get("records")
                if not isinstance(records, list) or not records or len(records) > 31:
                    self.send_json({"error": "records must contain 1-31 items."}, 400)
                    return
                operation = str(job.get("operation") or "insert")
                converted = [live_state_row("time_records", row) for row in records if isinstance(row, dict)]
                if len(converted) != len(records):
                    self.send_json({"error": "all records must be objects"}, 400)
                    return
                if operation == "update" and any(row.get("id") in (None, "") for row in converted):
                    self.send_json({"error": "record id is required for queued updates."}, 400)
                    return
                if operation == "insert":
                    for index, row in enumerate(converted):
                        row.pop("id", None)
                        row["queue_dedupe_key"] = time_queue_row_key(str(job.get("queue_uid") or ""), index)
                        raw = row.get("raw_payload") if isinstance(row.get("raw_payload"), dict) else {}
                        row["raw_payload"] = {key: value for key, value in raw.items() if key != "id"}
                identities = [time_record_identity(row) for row in converted]
                employee_codes = {emp_code for _date, emp_code in identities}
                dates = sorted({work_date for work_date, _emp_code in identities})
                if len(employee_codes) != 1 or not dates or any(not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value) for value in dates):
                    self.send_json({"error": "one queue must contain one employee and valid work dates."}, 400)
                    return
                previous_payload = job.get("payload") if isinstance(job.get("payload"), list) else []
                values = {
                    "payload": converted,
                    "payload_hash": production_queue_payload_hash(converted),
                    "record_count": len(converted),
                    "emp_code": next(iter(employee_codes)),
                    "employee_name": str(converted[0].get("employee_name") or ""),
                    "first_work_date": dates[0], "last_work_date": dates[-1],
                    "status": "queued", "attempt_count": 0,
                    "next_attempt_at": datetime.now(timezone.utc).isoformat(),
                    "error_code": None, "error_message": None, "finished_at": None,
                    "locked_at": None, "locked_by": None,
                }
                update_status, updated = update_time_queue(queue_id, values, expected_status=str(job.get("status") or "needs_review"))
                if update_status >= 400 or not updated:
                    self.send_json({"error": "Could not update the time queue for retry."}, update_status if update_status >= 400 else 409)
                    return
                time_queue_event(queue_id, "edited_retry", "queued", "Queue data was edited and submitted again.", actor_name,
                                 {"previous_payload": previous_payload})
                schedule_time_queue_job(queue_id)
                self.send_json({"data": time_queue_to_client(updated, include_payload=True)})
                return

            if action == "verify":
                rows = job.get("payload") if isinstance(job.get("payload"), list) else []
                existing_status, existing = time_queue_existing_rows(rows)
                if existing_status >= 400:
                    self.send_json({"error": existing}, existing_status)
                    return
                if str(job.get("operation") or "insert") == "insert" and isinstance(existing, list) and len(existing) == len(rows):
                    update_status, processing = update_time_queue(
                        queue_id,
                        {"status": "processing", "locked_at": datetime.now(timezone.utc).isoformat(),
                         "locked_by": time_save_queue_worker_id},
                        expected_status=str(job.get("status") or "needs_review"),
                    )
                    if update_status < 400 and processing:
                        finish_time_queue({**processing, "payload": rows}, existing, "manual_verified")
                        refreshed_status, refreshed = supabase_request("GET", f"time_save_queue?id=eq.{queue_id}&select=*&limit=1")
                        result = refreshed[0] if refreshed_status < 400 and isinstance(refreshed, list) and refreshed else processing
                        self.send_json({"data": time_queue_to_client(result, include_payload=True), "match": "saved"})
                        return
                rows_to_check = rows
                if str(job.get("operation") or "insert") == "insert" and isinstance(existing, list):
                    existing_keys = {str(row.get("queue_dedupe_key") or "") for row in existing}
                    rows_to_check = [
                        row for row in rows
                        if str(row.get("queue_dedupe_key") or "") not in existing_keys
                    ]
                conflict_status, conflict = validate_time_record_conflicts(rows_to_check) if rows_to_check else (200, None)
                if conflict_status >= 400:
                    reason = str((conflict or {}).get("error") or "Time queue verification failed.")
                    update_status, updated = update_time_queue(
                        queue_id,
                        {"status": "needs_review", "error_code": "time_overlap" if conflict_status == 409 else "invalid_time",
                         "error_message": reason, "finished_at": datetime.now(timezone.utc).isoformat(),
                         "locked_at": None, "locked_by": None},
                    )
                    time_queue_event(queue_id, "manual_verify_failed", "needs_review", reason, actor_name)
                    self.send_json({"data": time_queue_to_client(updated or job, include_payload=True), "match": "conflict"})
                    return

            update_status, updated = update_time_queue(
                queue_id,
                {"status": "queued", "attempt_count": 0, "next_attempt_at": datetime.now(timezone.utc).isoformat(),
                 "error_code": None, "error_message": None, "finished_at": None,
                 "locked_at": None, "locked_by": None},
                expected_status=str(job.get("status") or "needs_review"),
            )
            if update_status >= 400 or not updated:
                self.send_json({"error": "Could not submit the time queue again."}, update_status if update_status >= 400 else 409)
                return
            time_queue_event(queue_id, "manual_verify_retry" if action == "verify" else "manual_retry", "queued",
                             "Queue was checked and submitted again." if action == "verify" else "Queue was submitted again with the original data.", actor_name)
            schedule_time_queue_job(queue_id)
            self.send_json({"data": time_queue_to_client(updated, include_payload=True)})
            return

        if parsed.path == "/api/production-records/verify":
            client_uids = payload.get("client_uids", [])
            if not isinstance(client_uids, list) or not client_uids:
                self.send_json({"error": "client_uids must be a non-empty list"}, 400)
                return
            clean_uids = []
            for value in client_uids:
                uid = str(value or "").strip()
                if uid and uid not in clean_uids:
                    clean_uids.append(uid)
            if not clean_uids:
                self.send_json({"error": "client_uids must include at least one value"}, 400)
                return
            if len(clean_uids) > 100:
                self.send_json({"error": "client_uids limit is 100 per verification"}, 400)
                return
            verified_rows = []
            uid_filter_field = quote("raw_payload->>client_uid", safe="")
            for uid in clean_uids:
                status, body = supabase_request(
                    "GET",
                    f"production_records?{uid_filter_field}=eq.{quote(uid)}&select=*&order=id.asc",
                )
                if status >= 400:
                    self.send_json({"data": None, "error": body, "client_uid": uid}, status)
                    return
                if isinstance(body, list):
                    verified_rows.extend(
                        live_state_to_client("production_records", row)
                        for row in body
                        if isinstance(row, dict)
                    )
            self.send_json({"data": verified_rows, "error": None})
            return

        production_delete_match = re.fullmatch(r"/api/production-records/(\d+)/delete", parsed.path)
        if production_delete_match:
            actor = accounting_actor(self, 4)
            if not actor:
                self.send_json({"error": "C4 or higher session is required."}, 403)
                return
            record_id = int(production_delete_match.group(1))
            reason = str(payload.get("reason", "")).strip()
            if len(reason) < 3:
                self.send_json({"error": "Delete reason is required."}, 400)
                return

            status, existing_rows = supabase_request(
                "GET",
                f"production_records?id=eq.{record_id}&select=*&limit=1",
            )
            if status >= 400:
                self.send_json({"error": existing_rows}, status)
                return
            if not isinstance(existing_rows, list) or not existing_rows:
                self.send_json({"error": "Production record was not found."}, 404)
                return

            existing_row = existing_rows[0]
            before = live_state_to_client("production_records", existing_row)
            record_client_uid = production_record_client_uid(existing_row)
            if not record_client_uid:
                self.send_json({"error": "รายการนี้ไม่มี client_uid จึงหยุดการลบเพื่อป้องกันการลบผิดรายการ"}, 409)
                return
            uid_filter_field = quote("raw_payload->>client_uid", safe="")
            uid_status, uid_rows = supabase_request(
                "GET",
                f"production_records?{uid_filter_field}=eq.{quote(record_client_uid)}&select=id&order=id.asc&limit=3",
            )
            if uid_status >= 400:
                self.send_json({"error": uid_rows}, uid_status)
                return
            uid_ids = [int(row.get("id")) for row in uid_rows if isinstance(row, dict) and str(row.get("id") or "").isdigit()] if isinstance(uid_rows, list) else []
            if uid_ids != [record_id]:
                self.send_json(
                    {
                        "error": "พบรายการซ้ำก่อนแก้ไข ระบบหยุดไว้เพื่อไม่ให้แก้เพียงบางรายการ",
                        "client_uid": record_client_uid,
                        "existing_ids": uid_ids,
                    },
                    409,
                )
                return
            expected_updated_at = str(payload.get("expected_updated_at") or "")
            current_updated_at = str(before.get("updated_at") or before.get("created_at") or "")
            if expected_updated_at and expected_updated_at != current_updated_at:
                self.send_json(
                    {"error": "ข้อมูลรายการนี้ถูกแก้ไขจากเครื่องอื่นแล้ว กรุณาโหลดข้อมูลใหม่ก่อนลบอีกครั้ง"},
                    409,
                )
                return

            actor_status, actor_rows = supabase_request(
                "GET",
                f"account_users?id=eq.{quote(str(actor.get('sub', '')))}&select=username,fullname,user_level&limit=1",
            )
            actor_account = actor_rows[0] if actor_status < 400 and isinstance(actor_rows, list) and actor_rows else {}
            actor_name = str(actor_account.get("fullname") or actor.get("username") or "System")
            actor_username = str(actor_account.get("username") or actor.get("username") or "")

            delete_status, deleted_rows = supabase_request(
                "DELETE",
                f"production_records?id=eq.{record_id}",
                prefer="return=representation",
            )
            if delete_status >= 400:
                self.send_json({"error": deleted_rows}, delete_status)
                return
            if not isinstance(deleted_rows, list) or not deleted_rows:
                self.send_json({"error": "Production record deletion returned no row."}, 500)
                return

            audit_created_at = datetime.utcnow().isoformat() + "Z"
            audit_description = f"ลบผลผลิต #{record_id} รหัส {before.get('emp_code', '-')} เหตุผล: {reason}"
            audit_row = {
                "action": "DELETE_PRODUCTION",
                "module": "production",
                "description": audit_description,
                "created_by": actor_username,
                "user_fullname": actor_name,
                "ip_address": self.client_address[0] if self.client_address else None,
                "created_at": audit_created_at,
                "metadata": {
                    "action": "DELETE_PRODUCTION",
                    "module": "production",
                    "detail": audit_description,
                    "description": audit_description,
                    "created_by": actor_username,
                    "user_fullname": actor_name,
                    "username": actor_username,
                    "role": actor_account.get("user_level") or actor.get("level", "C4"),
                    "created_at": audit_created_at,
                    "record_id": record_id,
                    "reason": reason,
                    "before": before,
                    "after": None,
                    "actor_level": actor.get("level", "C4"),
                },
            }
            audit_status, audit_result = insert_audit_log_compatible(audit_row)
            if audit_status >= 400:
                restore_status, restore_result = supabase_request(
                    "POST",
                    "production_records",
                    existing_row,
                    prefer="resolution=merge-duplicates,return=representation",
                )
                if restore_status >= 400:
                    self.send_json(
                        {"error": "Audit log and automatic restore both failed.", "restore_error": restore_result},
                        500,
                    )
                    return
                self.send_json(
                    {
                        "error": "Audit log failed; deleted production record was restored.",
                        "audit_error": audit_result,
                    },
                    500,
                )
                return

            self.send_json({"deleted": before, "audit": audit_result})
            return

        if parsed.path == "/api/production-records/batch-date":
            actor = accounting_actor(self, 1)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 403)
                return
            selections = payload.get("records")
            target_date = str(payload.get("record_date") or "").strip()
            reason = str(payload.get("reason") or "").strip()
            if (
                not isinstance(selections, list)
                or not 1 <= len(selections) <= 50
                or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", target_date)
                or target_date > datetime.now(timezone(timedelta(hours=7))).date().isoformat()
                or len(reason) < 3
            ):
                self.send_json({"error": "Select 1-50 records, a valid date, and an edit reason."}, 400)
                return

            expected_by_id: dict[int, str] = {}
            for selection in selections:
                if not isinstance(selection, dict) or not str(selection.get("id") or "").isdigit():
                    self.send_json({"error": "Every selected record must include a numeric id."}, 400)
                    return
                record_id = int(selection["id"])
                if record_id in expected_by_id:
                    self.send_json({"error": "Duplicate record ids are not allowed."}, 400)
                    return
                expected_by_id[record_id] = str(selection.get("expected_updated_at") or "")

            record_ids = sorted(expected_by_id)
            id_filter = ",".join(str(record_id) for record_id in record_ids)
            status, existing_rows = supabase_request(
                "GET",
                f"production_records?id=in.({id_filter})&select=*&order=id.asc",
            )
            if status >= 400:
                self.send_json({"error": existing_rows}, status)
                return
            existing_rows = existing_rows if isinstance(existing_rows, list) else []
            existing_by_id = {
                int(row.get("id")): row
                for row in existing_rows
                if isinstance(row, dict) and str(row.get("id") or "").isdigit()
            }
            if sorted(existing_by_id) != record_ids:
                missing = [record_id for record_id in record_ids if record_id not in existing_by_id]
                self.send_json({"error": "Some production records were not found.", "missing_ids": missing}, 404)
                return

            actor_status, actor_rows = supabase_request(
                "GET",
                f"account_users?id=eq.{quote(str(actor.get('sub', '')))}&select=username,fullname,user_level&limit=1",
            )
            actor_account = actor_rows[0] if actor_status < 400 and isinstance(actor_rows, list) and actor_rows else {}
            actor_name = str(actor_account.get("fullname") or actor.get("username") or "System")
            actor_username = str(actor_account.get("username") or actor.get("username") or "")
            actor_level = account_level_number(actor_account.get("user_level") or actor.get("level"))

            before_rows: list[dict] = []
            for record_id in record_ids:
                existing_row = existing_by_id[record_id]
                before = live_state_to_client("production_records", existing_row)
                if actor_level < 4 and not production_record_within_self_edit_window(
                    before,
                    actor,
                    actor_account,
                ):
                    self.send_json(
                        {
                            "error": "C1-C3 may batch edit only their own production records within 5 minutes of creation.",
                            "record_id": record_id,
                        },
                        403,
                    )
                    return
                current_updated_at = str(before.get("updated_at") or before.get("created_at") or "")
                if expected_by_id[record_id] and expected_by_id[record_id] != current_updated_at:
                    self.send_json(
                        {
                            "error": "ข้อมูลบางรายการถูกแก้ไขจากเครื่องอื่นแล้ว กรุณาโหลดข้อมูลใหม่ก่อนแก้ไขแบบชุด",
                            "record_id": record_id,
                        },
                        409,
                    )
                    return
                client_uid = production_record_client_uid(existing_row)
                if not client_uid:
                    self.send_json({"error": "Selected record has no client_uid.", "record_id": record_id}, 409)
                    return
                uid_filter_field = quote("raw_payload->>client_uid", safe="")
                uid_status, uid_rows = supabase_request(
                    "GET",
                    f"production_records?{uid_filter_field}=eq.{quote(client_uid)}&select=id&order=id.asc&limit=3",
                )
                if uid_status >= 400:
                    self.send_json({"error": uid_rows}, uid_status)
                    return
                uid_ids = [
                    int(row.get("id")) for row in uid_rows
                    if isinstance(row, dict) and str(row.get("id") or "").isdigit()
                ] if isinstance(uid_rows, list) else []
                if uid_ids != [record_id]:
                    self.send_json(
                        {"error": "พบรายการซ้ำก่อนแก้ไขแบบชุด", "record_id": record_id, "existing_ids": uid_ids},
                        409,
                    )
                    return
                before_rows.append(before)

            if all(str(before.get("record_date") or before.get("date") or "") == target_date for before in before_rows):
                self.send_json({"error": "Selected records already use this production date."}, 400)
                return

            updated_rows: list[dict] = []
            changed_originals: list[dict] = []

            def rollback_batch_date() -> None:
                for original in reversed(changed_originals):
                    original_id = int(original.get("id") or 0)
                    rollback = {key: value for key, value in original.items() if key != "id"}
                    supabase_request(
                        "PATCH",
                        f"production_records?id=eq.{original_id}",
                        rollback,
                        prefer="return=minimal",
                    )

            updated_at = datetime.utcnow().isoformat() + "Z"
            for record_id in record_ids:
                original = existing_by_id[record_id]
                raw_payload = original.get("raw_payload") if isinstance(original.get("raw_payload"), dict) else {}
                next_raw_payload = {
                    **raw_payload,
                    "record_date": target_date,
                    "date": target_date,
                    "updated_by": actor_name,
                    "updated_at": updated_at,
                }
                client_uid = production_record_client_uid(original)
                uid_filter_field = quote("raw_payload->>client_uid", safe="")
                update_status, update_result = supabase_request(
                    "PATCH",
                    f"production_records?id=eq.{record_id}&{uid_filter_field}=eq.{quote(client_uid)}",
                    {
                        "record_date": target_date,
                        "updated_by": actor_name,
                        "updated_at": updated_at,
                        "raw_payload": next_raw_payload,
                    },
                    prefer="return=representation",
                )
                updated_row = update_result[0] if update_status < 400 and isinstance(update_result, list) and update_result else None
                if (
                    not updated_row
                    or int(updated_row.get("id") or 0) != record_id
                    or production_record_client_uid(updated_row) != client_uid
                ):
                    rollback_batch_date()
                    self.send_json(
                        {"error": "Batch date update failed; all changed records were rolled back.", "record_id": record_id},
                        update_status if update_status >= 400 else 409,
                    )
                    return
                changed_originals.append(original)
                updated_rows.append(updated_row)

            after_rows = [live_state_to_client("production_records", row) for row in updated_rows]
            audit_created_at = datetime.utcnow().isoformat() + "Z"
            audit_description = f"แก้วันที่ผลผลิต {len(record_ids)} รายการเป็น {target_date} เหตุผล: {reason}"
            audit_row = {
                "action": "BATCH_UPDATE_PRODUCTION_DATE",
                "module": "production",
                "description": audit_description,
                "created_by": actor_username,
                "user_fullname": actor_name,
                "ip_address": self.client_address[0] if self.client_address else None,
                "created_at": audit_created_at,
                "metadata": {
                    "action": "BATCH_UPDATE_PRODUCTION_DATE",
                    "module": "production",
                    "detail": audit_description,
                    "description": audit_description,
                    "created_by": actor_username,
                    "user_fullname": actor_name,
                    "username": actor_username,
                    "role": actor_account.get("user_level") or actor.get("level", "C1"),
                    "created_at": audit_created_at,
                    "record_ids": record_ids,
                    "reason": reason,
                    "changed_fields": ["record_date"],
                    "before": before_rows,
                    "after": after_rows,
                    "actor_level": actor.get("level", "C1"),
                },
            }
            audit_status, audit_result = insert_audit_log_compatible(audit_row)
            if audit_status >= 400:
                rollback_batch_date()
                self.send_json(
                    {"error": "Audit log failed; all batch date changes were rolled back.", "audit_error": audit_result},
                    500,
                )
                return
            self.send_json({"data": after_rows, "audit": audit_result, "updated_count": len(after_rows)})
            return

        production_record_match = re.fullmatch(r"/api/production-records/(\d+)", parsed.path)
        if production_record_match:
            actor = accounting_actor(self, 1)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 403)
                return
            record_id = int(production_record_match.group(1))
            incoming = payload.get("record")
            reason = str(payload.get("reason", "")).strip()
            if not isinstance(incoming, dict) or len(reason) < 3:
                self.send_json({"error": "Record and edit reason are required."}, 400)
                return

            status, existing_rows = supabase_request(
                "GET",
                f"production_records?id=eq.{record_id}&select=*&limit=1",
            )
            if status >= 400:
                self.send_json({"error": existing_rows}, status)
                return
            if not isinstance(existing_rows, list) or not existing_rows:
                self.send_json({"error": "Production record was not found."}, 404)
                return

            existing_row = existing_rows[0]
            before = live_state_to_client("production_records", existing_row)
            record_client_uid = production_record_client_uid(existing_row)
            if not record_client_uid:
                self.send_json({"error": "รายการนี้ไม่มี client_uid จึงหยุดการแก้ไขเพื่อป้องกันการสร้างรายการใหม่"}, 409)
                return
            uid_filter_field = quote("raw_payload->>client_uid", safe="")
            uid_status, uid_rows = supabase_request(
                "GET",
                f"production_records?{uid_filter_field}=eq.{quote(record_client_uid)}&select=id&order=id.asc&limit=3",
            )
            if uid_status >= 400:
                self.send_json({"error": uid_rows}, uid_status)
                return
            uid_ids = [int(row.get("id")) for row in uid_rows if isinstance(row, dict) and str(row.get("id") or "").isdigit()] if isinstance(uid_rows, list) else []
            if uid_ids != [record_id]:
                self.send_json(
                    {
                        "error": "พบรายการซ้ำก่อนแก้ไข ระบบหยุดไว้เพื่อไม่ให้แก้เพียงบางรายการ",
                        "client_uid": record_client_uid,
                        "existing_ids": uid_ids,
                    },
                    409,
                )
                return
            expected_updated_at = str(payload.get("expected_updated_at") or "")
            current_updated_at = str(before.get("updated_at") or before.get("created_at") or "")
            if expected_updated_at and expected_updated_at != current_updated_at:
                self.send_json(
                    {"error": "ข้อมูลรายการนี้ถูกแก้ไขจากเครื่องอื่นแล้ว กรุณาโหลดข้อมูลใหม่ก่อนแก้ไขอีกครั้ง"},
                    409,
                )
                return
            actor_status, actor_rows = supabase_request(
                "GET",
                f"account_users?id=eq.{quote(str(actor.get('sub', '')))}&select=username,fullname,user_level&limit=1",
            )
            actor_account = actor_rows[0] if actor_status < 400 and isinstance(actor_rows, list) and actor_rows else {}
            actor_name = str(actor_account.get("fullname") or actor.get("username") or "System")
            actor_username = str(actor_account.get("username") or actor.get("username") or "")
            actor_level = account_level_number(actor_account.get("user_level") or actor.get("level"))
            if actor_level < 4 and not production_record_within_self_edit_window(
                before,
                actor,
                actor_account,
            ):
                self.send_json(
                    {"error": "C1-C3 may edit only their own production records within 5 minutes of creation."},
                    403,
                )
                return

            protected_record = {
                **incoming,
                "id": record_id,
                "fruit_type": before.get("fruit_type") or "mangosteen",
                "created_by": before.get("created_by"),
                "created_at": before.get("created_at"),
                "updated_by": actor_name,
                "updated_at": datetime.utcnow().isoformat() + "Z",
            }
            record_date = str(protected_record.get("record_date") or protected_record.get("date") or "")
            pile_number = production_pile_number(protected_record)
            if (
                not re.fullmatch(r"\d{4}-\d{2}-\d{2}", record_date)
                or record_date > datetime.utcnow().date().isoformat()
                or not str(protected_record.get("emp_code") or "").strip()
                or pile_number not in [1, 2, 3, 4, 5]
            ):
                self.send_json({"error": "Invalid production date, employee, or pile."}, 400)
                return

            def valid_edit_weight(value) -> bool:
                numeric = safe_float(value)
                return numeric >= 0 and abs(numeric * 10 - round(numeric * 10)) < 0.000001

            if protected_record["fruit_type"] == "durian":
                durian_weight = production_total_weight(protected_record)
                if not valid_edit_weight(durian_weight) or durian_weight <= 0:
                    self.send_json({"error": "Durian weight must be positive with one decimal place."}, 400)
                    return
                previous_weight = production_total_weight(before)
                durian_rate = safe_float(before.get("durian_rate"))
                if durian_rate <= 0 and previous_weight > 0:
                    durian_rate = safe_float(before.get("total_amount", before.get("grand_total"))) / previous_weight
                if durian_rate <= 0:
                    durian_rate = safe_float((before.get("grade_rates") or {}).get("A"))
                total_amount = round(durian_weight * durian_rate, 2)
                grade_weights = {"A": durian_weight, "B": 0, "C": 0, "D": 0, "E": 0}
                grade_rates = {"A": durian_rate, "B": 0, "C": 0, "D": 0, "E": 0}
                grade_amounts = {"A": total_amount, "B": 0, "C": 0, "D": 0, "E": 0}
                protected_record.update({
                    "durian_weight": durian_weight,
                    "durian_rate": durian_rate,
                    "durian_amount": total_amount,
                    "grade_weights": grade_weights,
                    "grade_rates": grade_rates,
                    "grade_amounts": grade_amounts,
                    "total_weight": round(durian_weight, 2),
                    "total_amount": total_amount,
                    "grand_total": total_amount,
                    "water_weight": 0,
                    "flower_weight": 0,
                })
            else:
                water_weight = safe_float(protected_record.get("water_weight", protected_record.get("water")))
                flower_weight = safe_float(protected_record.get("flower_weight", protected_record.get("flower")))
                if not valid_edit_weight(water_weight) or not valid_edit_weight(flower_weight):
                    self.send_json({"error": "Production weights must be non-negative with one decimal place."}, 400)
                    return
                water_rate = safe_float(before.get("water_rate"))
                flower_rate = safe_float(before.get("flower_rate"))
                water_amount = round(water_weight * water_rate, 2)
                flower_amount = round(flower_weight * flower_rate, 2)
                protected_record.update({
                    "water_weight": water_weight,
                    "flower_weight": flower_weight,
                    "water": water_weight,
                    "flower": flower_weight,
                    "water_rate": water_rate,
                    "flower_rate": flower_rate,
                    "water_amount": water_amount,
                    "flower_amount": flower_amount,
                    "water_total": water_amount,
                    "flower_total": flower_amount,
                    "total_weight": round(water_weight + flower_weight, 2),
                    "total_amount": round(water_amount + flower_amount, 2),
                    "grand_total": round(water_amount + flower_amount, 2),
                })
            converted = live_state_row("production_records", protected_record)
            converted.pop("id", None)
            status, updated_rows = supabase_request(
                "PATCH",
                f"production_records?id=eq.{record_id}&{uid_filter_field}=eq.{quote(record_client_uid)}",
                converted,
                prefer="return=representation",
            )
            if status >= 400:
                self.send_json({"error": updated_rows}, status)
                return
            updated_row = updated_rows[0] if isinstance(updated_rows, list) and updated_rows else None
            if not updated_row:
                self.send_json({"error": "Production record update returned no row."}, 500)
                return
            if int(updated_row.get("id") or 0) != record_id or production_record_client_uid(updated_row) != record_client_uid:
                rollback = {key: value for key, value in existing_row.items() if key != "id"}
                supabase_request("PATCH", f"production_records?id=eq.{record_id}", rollback, prefer="return=minimal")
                self.send_json({"error": "การแก้ไขเปลี่ยนตัวตนรายการ ระบบย้อนข้อมูลเดิมกลับแล้ว"}, 409)
                return
            after = live_state_to_client("production_records", updated_row)
            tracked_fields = [
                "record_date", "emp_code", "employee_name", "fruit_type", "pile_no",
                "water_weight", "flower_weight", "durian_weight", "grade_weights", "total_weight", "total_amount",
            ]
            changed_fields = [field for field in tracked_fields if before.get(field) != after.get(field)]
            if not changed_fields:
                rollback = {key: value for key, value in existing_row.items() if key != "id"}
                supabase_request(
                    "PATCH",
                    f"production_records?id=eq.{record_id}",
                    rollback,
                    prefer="return=minimal",
                )
                self.send_json({"error": "No production values were changed."}, 400)
                return

            audit_created_at = datetime.utcnow().isoformat() + "Z"
            audit_description = f"แก้ไขผลผลิต #{record_id} รหัส {after.get('emp_code', '-')} เหตุผล: {reason}"
            audit_row = {
                "action": "UPDATE_PRODUCTION",
                "module": "production",
                "description": audit_description,
                "created_by": actor_username,
                "user_fullname": actor_name,
                "ip_address": self.client_address[0] if self.client_address else None,
                "created_at": audit_created_at,
                "metadata": {
                    "action": "UPDATE_PRODUCTION",
                    "module": "production",
                    "detail": audit_description,
                    "description": audit_description,
                    "created_by": actor_username,
                    "user_fullname": actor_name,
                    "username": actor_username,
                    "role": actor_account.get("user_level") or actor.get("level", "C4"),
                    "created_at": audit_created_at,
                    "record_id": record_id,
                    "reason": reason,
                    "changed_fields": changed_fields,
                    "before": before,
                    "after": after,
                    "actor_level": actor.get("level", "C4"),
                },
            }
            audit_status, audit_result = insert_audit_log_compatible(audit_row)
            if audit_status >= 400:
                rollback = {key: value for key, value in existing_row.items() if key != "id"}
                supabase_request(
                    "PATCH",
                    f"production_records?id=eq.{record_id}",
                    rollback,
                    prefer="return=minimal",
                )
                self.send_json(
                    {
                        "error": "Audit log failed; production change was rolled back.",
                        "audit_error": audit_result,
                    },
                    500,
                )
                return
            self.send_json({"data": after, "audit": audit_result})
            return

        if parsed.path == "/api/state":
            table = str(payload.get("table", "")).strip()
            rows = payload.get("rows", [])
            if table not in LIVE_STATE_TABLES or not isinstance(rows, list):
                self.send_json({"error": "Invalid live-state table or rows."}, 400)
                return
            converted = [live_state_row(table, row) for row in rows if isinstance(row, dict)]
            with live_state_sync_lock:
                status, body = sync_rows_by_id(table, converted)
            if status < 400:
                # Never delete rows missing from one browser's local snapshot.
                # Multiple stations submit concurrently and each may only know
                # about its own most recent records.
                read_status, read_body = supabase_get_all(f"{table}?select=*&order=id.asc")
                if read_status >= 400:
                    self.send_json({"error": read_body, "table": table}, read_status)
                    return
                client_rows = [live_state_to_client(table, row) for row in read_body]
                self.send_json({"data": client_rows, "error": None}, status)
                return
            self.send_json({"data": None, "error": body}, status)
            return

        if parsed.path == "/api/backup/restore":
            if not backup_authorized(self):
                self.send_json({"error": "Backup code is required."}, 403)
                return
            data = payload.get("data") if isinstance(payload.get("data"), dict) else payload
            if not isinstance(data, dict):
                self.send_json({"error": "Invalid backup payload."}, 400)
                return
            status, body = restore_supabase_backup(data)
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/backup/clear":
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            actor_username = str(actor.get("username") or "").strip()
            actor_status, actor_rows = supabase_request(
                "GET",
                f"account_users?username=eq.{quote(actor_username)}&select=username,fullname,user_level,status&limit=1",
                timeout_seconds=6,
            )
            actor_account = actor_rows[0] if actor_status < 400 and isinstance(actor_rows, list) and actor_rows else None
            if not actor_account or str(actor_account.get("status") or "") != "Active":
                self.send_json({"error": "The account is missing, inactive, or could not be verified."}, 403)
                return
            if account_level_number(actor_account.get("user_level")) < 4:
                self.send_json({"error": "C4 or higher is required for Backup / Clear."}, 403)
                return
            if not backup_authorized(self):
                self.send_json({"error": "Backup code is required."}, 403)
                return
            scope = str(payload.get("scope") or "").strip().lower()
            if scope not in {"queue", "main"}:
                self.send_json({"error": "scope must be queue or main."}, 400)
                return
            if str(payload.get("confirmation") or "") != "BACKUP_CLEAR":
                self.send_json({"error": "Backup / Clear confirmation is missing."}, 400)
                return
            if not backup_clear_lock.acquire(blocking=False):
                self.send_json({"error": "Another Backup / Clear operation is already running."}, 409)
                return
            try:
                tables = QUEUE_BACKUP_TABLES if scope == "queue" else BACKUP_TABLES
                snapshot_status, snapshot_data = read_supabase_backup(tables)
                if snapshot_status >= 400:
                    self.send_json({"error": snapshot_data, "stage": "snapshot"}, snapshot_status)
                    return
                backup = backup_snapshot_payload(scope, actor_username, snapshot_data)
                archive_content = backup_archive_bytes(backup)
                checksum = backup_archive_checksum(archive_content)
                timestamp = datetime.now(timezone.utc).strftime("%Y/%m/%Y%m%dT%H%M%SZ")
                object_path = f"{scope}/{timestamp}-{checksum[:12]}.json"
                archive_status, archive_result = supabase_storage_request("POST", object_path, archive_content)
                if archive_status >= 400:
                    self.send_json({
                        "error": archive_result,
                        "stage": "archive_upload",
                        "migration": "supabase_backup_archive_migration.sql",
                    }, archive_status)
                    return
                verify_status, verified_content = supabase_storage_request("GET", object_path)
                verified_checksum = backup_archive_checksum(verified_content) if isinstance(verified_content, bytes) else ""
                if verify_status >= 400 or not hmac.compare_digest(checksum, verified_checksum):
                    self.send_json({
                        "error": "The private archive could not be verified. No data was cleared.",
                        "stage": "archive_verify",
                        "archive": object_path,
                    }, 500)
                    return
                clear_complete, cleared, clear_error = delete_backup_snapshot_rows(snapshot_data, scope)
                audit_action = "BACKUP_CLEAR_QUEUE" if scope == "queue" else "BACKUP_CLEAR_MAIN"
                audit_description = (
                    f"Archived {backup['total_rows']} rows to {object_path}; "
                    f"cleared {sum(cleared.values())} rows; checksum {checksum}; "
                    f"complete={str(clear_complete).lower()}"
                )
                audit_status, audit_result = insert_audit_log_compatible({
                    "action": audit_action,
                    "module": "backup",
                    "description": audit_description,
                    "created_by": actor_username,
                    "user_fullname": str(actor_account.get("fullname") or actor_username),
                    "ip_address": self.client_address[0] if self.client_address else None,
                    "metadata": {
                        "scope": scope,
                        "archive_path": object_path,
                        "checksum": checksum,
                        "snapshot_counts": backup["row_counts"],
                        "cleared_counts": cleared,
                        "clear_complete": clear_complete,
                        "clear_error": clear_error,
                    },
                })
                self.send_json({
                    "backup": backup,
                    "archive": {
                        "bucket": BACKUP_ARCHIVE_BUCKET,
                        "path": object_path,
                        "checksum": checksum,
                        "bytes": len(archive_content),
                        "verified": True,
                    },
                    "cleared": cleared,
                    "clear_complete": clear_complete,
                    "clear_error": clear_error,
                    "audit_saved": audit_status < 400,
                    "audit_error": audit_result if audit_status >= 400 else None,
                })
            finally:
                backup_clear_lock.release()
            return

        if parsed.path == "/api/auth/login":
            ensure_system_accounts()
            username = str(payload.get("username", "")).strip()
            password = str(payload.get("password", ""))
            if not username or not password:
                self.send_json({"error": "Username and password are required."}, 400)
                return
            status, body = supabase_request(
                "GET",
                f"account_users?username=eq.{quote(username)}&select=*&limit=1",
            )
            if status >= 400:
                self.send_json({"error": body}, status)
                return
            account = body[0] if isinstance(body, list) and body else None
            if not account:
                self.send_json({"error": "ไม่พบบัญชีนี้ในฐานข้อมูลกลาง"}, 404)
                return
            if account.get("status") != "Active":
                self.send_json({"error": "บัญชีนี้ถูกปิดใช้งาน กรุณาติดต่อผู้ดูแลระบบ"}, 403)
                return
            if not verify_password(password, str(account.get("password_hash", ""))):
                self.send_json({"error": "รหัสผ่านไม่ถูกต้อง กรุณาตรวจสอบอีกครั้ง"}, 401)
                return
            supabase_request(
                "PATCH",
                f"account_users?id=eq.{account.get('id')}",
                {"last_login_at": datetime.utcnow().isoformat() + "Z"},
                prefer="return=minimal",
            )
            self.send_json({"user": account_to_client(account), "token": session_token(account)})
            return

        if parsed.path == "/api/accounts/sync":
            accounts = payload.get("accounts", [])
            if not isinstance(accounts, list):
                self.send_json({"error": "accounts must be a list"}, 400)
                return
            cloud_accounts = []
            for account_payload in accounts:
                if not isinstance(account_payload, dict):
                    continue
                account = account_from_payload(account_payload)
                if account.get("user_level") == "C7" or account.get("role") == "developer":
                    continue
                if account["username"] and account["fullname"] and account.get("password_hash"):
                    cloud_accounts.append(account)
            if not cloud_accounts:
                self.send_json({"data": []})
                return
            status, body = supabase_account_bulk_write(
                "POST",
                "account_users?on_conflict=username",
                cloud_accounts,
                prefer="resolution=merge-duplicates,return=representation",
            )
            self.send_json({"data": body if status < 400 else [], "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/accounts":
            account = account_from_payload(payload)
            if not account["username"] or not account["fullname"] or not account.get("password_hash"):
                self.send_json({"error": "Username, fullname, and password are required."}, 400)
                return
            if account.get("user_level") == "C7" or account.get("role") == "developer":
                self.send_json({"error": "C7/developer accounts can only be managed by the system."}, 403)
                return
            status, body = supabase_account_write(
                "POST",
                "account_users",
                account,
                prefer="return=representation",
            )
            data = [account_to_client(item) for item in body] if status < 400 and isinstance(body, list) else None
            self.send_json({"data": data, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/employees/sync":
            employees_payload = payload.get("employees", [])
            if not isinstance(employees_payload, list):
                self.send_json({"error": "employees must be a list"}, 400)
                return
            employees = []
            for employee_payload in employees_payload:
                if not isinstance(employee_payload, dict):
                    continue
                employee = employee_from_payload(employee_payload)
                required = ["emp_code", "fullname", "department", "pay_group"]
                if all(employee.get(key) not in [None, ""] for key in required):
                    employees.append(employee)
            if not employees:
                self.send_json({"data": []})
                return
            status, body = sync_rows_by_id("employees", employees)
            synced = body.get("synced", []) if status < 400 and isinstance(body, dict) else []
            self.send_json({"data": synced, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/time-employees/sync":
            employees_payload = payload.get("employees", [])
            if not isinstance(employees_payload, list):
                self.send_json({"error": "employees must be a list"}, 400)
                return
            employees = []
            for employee_payload in employees_payload:
                if not isinstance(employee_payload, dict):
                    continue
                employee = time_employee_from_payload(employee_payload)
                required = ["emp_code", "fullname", "employee_type", "daily_wage"]
                if all(employee.get(key) not in [None, ""] for key in required):
                    employees.append(employee)
            if not employees:
                self.send_json({"data": []})
                return
            status, body = sync_rows_by_id("time_employees", employees)
            synced = body.get("synced", []) if status < 400 and isinstance(body, dict) else []
            self.send_json({"data": synced, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/employees":
            employee = ensure_row_id("employees", employee_from_payload(payload))
            required = ["emp_code", "fullname", "department", "pay_group"]
            missing = [key for key in required if not employee[key]]
            if missing:
                self.send_json({"error": f"Missing required fields: {', '.join(missing)}"}, 400)
                return
            status, body = supabase_request(
                "POST",
                "employees",
                employee,
                prefer="return=representation",
            )
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/time-employees":
            employee = ensure_row_id("time_employees", time_employee_from_payload(payload))
            required = ["emp_code", "fullname", "employee_type", "daily_wage"]
            missing = [key for key in required if employee.get(key) in [None, ""]]
            if missing:
                self.send_json({"error": f"Missing required fields: {', '.join(missing)}"}, 400)
                return
            status, body = supabase_request(
                "POST",
                "time_employees",
                employee,
                prefer="return=representation",
            )
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/deductions":
            deduction = deduction_from_payload(payload)
            required = ["employee_kind", "employee_id", "emp_code", "employee_name", "start_date", "end_date", "deduction_type", "deduction_label", "amount"]
            missing = [key for key in required if deduction.get(key) in [None, ""]]
            if missing:
                self.send_json({"error": f"Missing required fields: {', '.join(missing)}"}, 400)
                return
            with deduction_record_insert_lock:
                status, body = insert_deduction_record_compatible(deduction)
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/deduction-applications/apply":
            applied_date = str(payload.get("applied_date", "")).strip()
            items = payload.get("items", [])
            if not applied_date or not isinstance(items, list) or not items:
                self.send_json({"error": "applied_date and items are required."}, 400)
                return
            clean_items = [deduction_application_from_payload(item) for item in items if isinstance(item, dict)]
            if not clean_items:
                self.send_json({"error": "No valid deduction items."}, 400)
                return
            status, body = supabase_request(
                "POST",
                "rpc/apply_deduction_batch",
                {
                    "p_applied_date": applied_date,
                    "p_created_by": str(payload.get("created_by", "")).strip(),
                    "p_items": clean_items,
                },
                prefer="return=representation",
            )
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/wage-rates":
            actor = accounting_actor(self, 4)
            if not actor:
                self.send_json({"error": "C4 or higher session is required to manage wage rates."}, 403)
                return
            wage_rate = {
                "item_type": str(payload.get("item_type", "")).strip(),
                "rate": payload.get("rate", 0),
                "effective_date": str(payload.get("effective_date", "")).strip(),
                "note": str(payload.get("note", "")).strip() or None,
                "created_by": str(payload.get("created_by", "")).strip() or None,
            }
            if not wage_rate["item_type"] or not wage_rate["effective_date"]:
                self.send_json({"error": "item_type and effective_date are required."}, 400)
                return
            status, body = supabase_request(
                "POST",
                "wage_rates",
                wage_rate,
                prefer="return=representation",
            )
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path in {"/reports/accounting-payments-excel", "/reports/accounting-payments-pdf"}:
            actor = accounting_actor(self, 4)
            if not actor:
                self.send_json({"error": "C4 or higher accounting session is required."}, 403)
                return
            payload["printed_by"] = payload.get("printed_by") or actor.get("fullname") or actor.get("username")
            start_date, end_date = normalized_range(payload)
            method = "transfer" if payload.get("payment_method") == "transfer" else "cash"
            if parsed.path.endswith("-excel"):
                content = build_accounting_payments_excel(payload)
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                extension = "xlsx"
            else:
                content = build_accounting_payments_pdf(payload)
                content_type = "application/pdf"
                extension = "pdf"
            self.send_file(content, content_type, f"PF_Payment_{method}_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.{extension}")
            return

        if parsed.path in {"/reports/inbound-selected-excel", "/reports/inbound-selected-pdf"}:
            actor = inbound_authorized_actor(self)
            if not actor:
                self.send_json({"error": "C5 or higher is required for inbound receiving."}, 403)
                return
            try:
                receipt_ids = list(dict.fromkeys(int(value) for value in payload.get("receipt_ids", []) if int(value) > 0))
            except (TypeError, ValueError):
                receipt_ids = []
            if not receipt_ids or len(receipt_ids) > 200:
                self.send_json({"error": "Select 1-200 inbound receipts for export."}, 400)
                return
            status, body = supabase_request("GET", f"inbound_receipts?id=in.({','.join(str(value) for value in receipt_ids)})&select=*")
            if status >= 400:
                self.send_json({"error": body}, status)
                return
            by_id = {int(row.get("id") or 0): row for row in body if isinstance(row, dict)} if isinstance(body, list) else {}
            rows = [by_id[value] for value in receipt_ids if value in by_id]
            if len(rows) != len(receipt_ids):
                self.send_json({"error": "Some selected inbound receipts no longer exist."}, 409)
                return
            export_payload = {"rows": rows, "printed_by": actor.get("username") or "-"}
            timestamp = datetime.now(timezone(timedelta(hours=7))).strftime("%Y%m%d-%H%M")
            if parsed.path.endswith("-excel"):
                self.send_file(build_inbound_selected_excel(export_payload), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"Inbound_Selected_{timestamp}.xlsx")
            else:
                self.send_file(build_inbound_selected_pdf(export_payload), "application/pdf", f"Inbound_Selected_{timestamp}.pdf")
            return

        if parsed.path == "/reports/selected-employees-pdf":
            data = {
                "employees": payload.get("employees", []),
                "production_records": payload.get("production_records", []),
            }
            date = payload.get("date", datetime.now().date().isoformat())
            employee_ids = [int(value) for value in payload.get("employee_ids", [])]
            content = build_pdf(data, date, employee_ids)
            self.send_file(
                content,
                "application/pdf",
                f"selected-employees-{date}.pdf",
            )
            return

        if parsed.path == "/reports/employee-range-pdf":
            data = {
                "employees": payload.get("employees", []),
                "production_records": payload.get("production_records", []),
                "deduction_records": payload.get("deduction_records", []),
                "printed_by": payload.get("printed_by", ""),
                "printed_by_position": payload.get("printed_by_position", ""),
            }
            start_date = payload.get("start_date", datetime.now().date().isoformat())
            end_date = payload.get("end_date", start_date)
            employee_id = int(payload.get("employee_id", 0))
            fruit_type = payload.get("fruit_type", "all")
            content = build_employee_range_pdf(data, start_date, end_date, employee_id, fruit_type)
            self.send_file(
                content,
                "application/pdf",
                f"employee-{employee_id}-{start_date}-to-{end_date}.pdf",
            )
            return

        if parsed.path == "/reports/employee-range-excel":
            data = {
                "employees": payload.get("employees", []),
                "production_records": payload.get("production_records", []),
                "deduction_records": payload.get("deduction_records", []),
            }
            start_date = payload.get("start_date", datetime.now().date().isoformat())
            end_date = payload.get("end_date", start_date)
            employee_id = int(payload.get("employee_id", 0))
            fruit_type = payload.get("fruit_type", "all")
            content = build_employee_range_excel(data, start_date, end_date, employee_id, fruit_type)
            self.send_file(
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"employee-{employee_id}-{start_date}-to-{end_date}.xlsx",
            )
            return

        if parsed.path == "/reports/summary-by-pile-pdf":
            data = {
                "employees": payload.get("employees", []),
                "production_records": payload.get("production_records", []),
            }
            start_date = payload.get("start_date", datetime.now().date().isoformat())
            content = build_summary_by_pile_pdf(data, payload)
            self.send_file(
                content,
                "application/pdf",
                f"Summary_ByPile_{clean_filename_date(start_date)}.pdf",
            )
            return

        if parsed.path == "/reports/summary-by-pile-excel":
            data = {
                "employees": payload.get("employees", []),
                "production_records": payload.get("production_records", []),
            }
            start_date = payload.get("start_date", datetime.now().date().isoformat())
            content = build_summary_by_pile_excel(data, payload)
            self.send_file(
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"Summary_ByPile_{clean_filename_date(start_date)}.xlsx",
            )
            return

        if parsed.path == "/reports/time-full-export-excel":
            start_date = payload.get("start_date", datetime.now().date().isoformat())
            end_date = payload.get("end_date", start_date)
            content = build_time_full_export_excel(payload)
            self.send_file(
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"Full_Details_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.xlsx",
            )
            return

        if parsed.path == "/reports/time-receipts-pdf":
            start_date = payload.get("start_date", datetime.now().date().isoformat())
            end_date = payload.get("end_date", start_date)
            content = build_time_receipts_pdf(payload)
            self.send_file(
                content,
                "application/pdf",
                f"Time_Receipts_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.pdf",
            )
            return

        if parsed.path == "/reports/production-summary-pdf":
            start_date, end_date = normalized_range(payload)
            content = build_production_summary_pdf(payload)
            self.send_file(
                content,
                "application/pdf",
                f"Production_Summary_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.pdf",
            )
            return

        if parsed.path == "/reports/production-summary-excel":
            start_date, end_date = normalized_range(payload)
            content = build_production_summary_excel(payload)
            self.send_file(
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"Production_Summary_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.xlsx",
            )
            return

        if parsed.path == "/reports/group-report-pdf":
            start_date, end_date = normalized_range(payload)
            content = build_group_report_pdf(payload)
            self.send_file(
                content,
                "application/pdf",
                f"Group_Report_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.pdf",
            )
            return

        if parsed.path == "/reports/group-report-excel":
            start_date, end_date = normalized_range(payload)
            content = build_group_report_excel(payload)
            self.send_file(
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"Group_Report_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.xlsx",
            )
            return

        if parsed.path == "/reports/time-group-report-pdf":
            start_date, end_date = normalized_range(payload)
            content = build_time_group_report_pdf(payload)
            self.send_file(
                content,
                "application/pdf",
                f"Time_Group_Report_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.pdf",
            )
            return

        if parsed.path == "/reports/time-group-report-excel":
            start_date, end_date = normalized_range(payload)
            content = build_time_group_report_excel(payload)
            self.send_file(
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"Time_Group_Report_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.xlsx",
            )
            return

        if parsed.path == "/reports/time-summary-pdf":
            start_date, end_date = normalized_range(payload)
            content = build_time_summary_pdf(payload)
            self.send_file(
                content,
                "application/pdf",
                f"Time_Summary_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.pdf",
            )
            return

        if parsed.path == "/reports/time-summary-excel":
            start_date, end_date = normalized_range(payload)
            content = build_time_summary_excel(payload)
            self.send_file(
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"Time_Summary_{clean_filename_date(start_date)}_to_{clean_filename_date(end_date)}.xlsx",
            )
            return

        self.send_error(404, "Not found")

    def do_PUT(self) -> None:
        parsed = urlparse(self.path)
        payload = self.read_json()

        if parsed.path == "/api/accounting/payment-allocations":
            actor = accounting_actor(self, 4)
            if not actor:
                self.send_json({"error": "C4 or higher accounting session is required."}, 403)
                return
            start_date, end_date = normalized_range(payload)
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", start_date) or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", end_date):
                self.send_json({"error": "A valid payment date range is required."}, 400)
                return
            clean_rows = []
            for raw in payload.get("allocations") or []:
                if not isinstance(raw, dict):
                    continue
                employee_key = str(raw.get("employee_key") or "").strip()[:160]
                employee_kind = str(raw.get("employee_kind") or "").strip()
                payment_method = str(raw.get("payment_method") or "").strip()
                if not employee_key or employee_kind not in {"production", "time"} or payment_method not in {"cash", "transfer"}:
                    continue
                clean_rows.append({
                    "company_key": ACCOUNTING_COMPANY_KEY,
                    "week_start": start_date,
                    "week_end": end_date,
                    "employee_key": employee_key,
                    "employee_kind": employee_kind,
                    "employee_id": str(raw.get("employee_id") or "")[:100] or None,
                    "emp_code": str(raw.get("emp_code") or "-")[:80],
                    "fullname": str(raw.get("fullname") or "-")[:200],
                    "group_label": str(raw.get("group_label") or "-")[:200],
                    "payment_method": payment_method,
                    "net_amount": max(0, round(safe_float(raw.get("net_amount")), 2)),
                    "updated_by": str(actor.get("username") or "unknown")[:120],
                })
            if not clean_rows:
                self.send_json({"data": []})
                return
            status, body = supabase_request(
                "POST",
                "ac_payment_allocations?on_conflict=company_key,week_start,week_end,employee_key",
                clean_rows,
                prefer="resolution=merge-duplicates,return=representation",
            )
            if status >= 400:
                self.send_json({"error": body, "migration": "supabase_pf_accounting_payments.sql"}, status)
                return
            self.send_json({"data": body})
            return

        if parsed.path == "/api/accounting/workspace":
            actor = accounting_actor(self)
            if not actor:
                self.send_json({"error": "Accounting session is missing or expired."}, 401)
                return
            workspace = payload.get("workspace")
            validation_error = validate_accounting_workspace(workspace)
            if validation_error:
                self.send_json({"error": validation_error}, 422)
                return
            expected_revision = int(payload.get("revision") or 0)
            status, existing_rows = supabase_request(
                "GET",
                f"accounting_workspaces?company_key=eq.{quote(ACCOUNTING_COMPANY_KEY)}&select=*",
            )
            if status >= 400:
                self.send_json({"error": existing_rows}, status)
                return
            existing = existing_rows[0] if isinstance(existing_rows, list) and existing_rows else None
            current_revision = int(existing.get("revision", 0)) if existing else 0
            if expected_revision != current_revision:
                self.send_json({"error": "Accounting workspace revision conflict.", "data": existing}, 409)
                return
            if existing:
                old_workspace = existing.get("workspace") if isinstance(existing.get("workspace"), dict) else {}
                old_periods = old_workspace.get("periods", {}) if isinstance(old_workspace.get("periods"), dict) else {}
                new_periods = workspace.get("periods", {}) if isinstance(workspace.get("periods"), dict) else {}
                for period, period_row in old_periods.items():
                    if not isinstance(period_row, dict) or period_row.get("status") != "closed":
                        continue
                    old_journals = [row for row in old_workspace.get("journals", []) if str(row.get("date", "")).startswith(period)]
                    new_journals = [row for row in workspace.get("journals", []) if str(row.get("date", "")).startswith(period)]
                    reopened = isinstance(new_periods.get(period), dict) and new_periods[period].get("status") == "open"
                    actor_level = int("".join(filter(str.isdigit, str(actor.get("level", "C1")))) or "1")
                    if old_journals != new_journals and not (reopened and actor_level >= 5):
                        self.send_json({"error": f"Closed accounting period {period} is immutable."}, 423)
                        return
            next_revision = current_revision + 1
            row = {
                "company_key": ACCOUNTING_COMPANY_KEY,
                "revision": next_revision,
                "workspace": workspace,
                "updated_by": actor.get("username", "unknown"),
                "updated_at": datetime.utcnow().isoformat() + "Z",
            }
            if existing:
                save_status, saved = supabase_request(
                    "PATCH",
                    f"accounting_workspaces?company_key=eq.{quote(ACCOUNTING_COMPANY_KEY)}&revision=eq.{current_revision}",
                    row,
                    prefer="return=representation",
                )
                if save_status < 400 and (not isinstance(saved, list) or not saved):
                    self.send_json({"error": "Accounting workspace revision conflict."}, 409)
                    return
            else:
                save_status, saved = supabase_request("POST", "accounting_workspaces", row, prefer="return=representation")
            if save_status >= 400:
                self.send_json({"error": saved}, save_status)
                return
            workspace_hash = hashlib.sha256(json.dumps(workspace, sort_keys=True, separators=(",", ":")).encode()).hexdigest()
            supabase_request(
                "POST",
                "accounting_change_log",
                {"company_key": ACCOUNTING_COMPANY_KEY, "revision": next_revision, "action": "WORKSPACE_SAVE", "actor_username": actor.get("username", "unknown"), "actor_level": actor.get("level", "C1"), "workspace_hash": workspace_hash, "metadata": {"journal_count": len(workspace.get("journals", [])), "account_count": len(workspace.get("accounts", []))}},
                prefer="return=minimal",
            )
            self.send_json({"data": {"revision": next_revision, "updated_at": row["updated_at"]}})
            return

        if parsed.path == "/api/employees":
            employee_id = payload.get("id")
            if employee_id in [None, ""]:
                self.send_json({"error": "id is required."}, 400)
                return
            employee = employee_from_payload(payload)
            employee["updated_at"] = datetime.utcnow().isoformat() + "Z"
            status, body = supabase_request(
                "PATCH",
                f"employees?id=eq.{quote(str(employee_id))}",
                employee,
                prefer="return=representation",
            )
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/accounts":
            account_id = payload.get("id")
            if account_id in [None, ""]:
                self.send_json({"error": "id is required."}, 400)
                return
            existing_status, existing_rows = supabase_request(
                "GET",
                f"account_users?id=eq.{quote(str(account_id))}&select=id,username,fullname,phone,role,user_level&limit=1",
            )
            if existing_status >= 400:
                self.send_json({"error": existing_rows}, existing_status)
                return
            if not isinstance(existing_rows, list) or not existing_rows:
                self.send_json({"error": "Account not found."}, 404)
                return
            existing = existing_rows[0]
            existing_username = str(existing.get("username", "")).lower()
            existing_level = str(existing.get("user_level", "")).upper()
            existing_role = str(existing.get("role", ""))
            if existing_username in SYSTEM_ACCOUNT_USERNAMES or existing_level == "C7" or existing_role == "developer":
                self.send_json({"error": "C7/developer account cannot be edited."}, 403)
                return
            account_payload = dict(payload)
            if not str(account_payload.get("phone", "")).strip() and str(existing.get("phone", "")).strip():
                account_payload["phone"] = existing.get("phone")
            account = account_from_payload(account_payload, include_password=bool(str(payload.get("password", ""))))
            if account.get("user_level") == "C7" or account.get("role") == "developer":
                self.send_json({"error": "C7/developer accounts can only be managed by the system."}, 403)
                return
            if not account["username"] or not account["fullname"]:
                self.send_json({"error": "Username and fullname are required."}, 400)
                return
            account["updated_at"] = datetime.utcnow().isoformat() + "Z"
            status, body = supabase_account_write(
                "PATCH",
                f"account_users?id=eq.{quote(str(account_id))}",
                account,
                prefer="return=representation",
            )
            data = [account_to_client(item) for item in body] if status < 400 and isinstance(body, list) else None
            self.send_json({"data": data, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/time-employees":
            employee_id = payload.get("id")
            if employee_id in [None, ""]:
                self.send_json({"error": "id is required."}, 400)
                return
            employee = time_employee_from_payload(payload)
            employee["updated_at"] = datetime.utcnow().isoformat() + "Z"
            status, body = supabase_request(
                "PATCH",
                f"time_employees?id=eq.{quote(str(employee_id))}",
                employee,
                prefer="return=representation",
            )
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/deductions":
            deduction_id = payload.get("id")
            if deduction_id in [None, ""]:
                self.send_json({"error": "id is required."}, 400)
                return
            existing_status, existing_rows = supabase_request(
                "GET",
                f"deduction_records?id=eq.{quote(str(deduction_id))}&select=employee_id,start_date,deduction_type,amount&limit=1",
            )
            if existing_status >= 400 or not isinstance(existing_rows, list) or not existing_rows:
                self.send_json({"error": existing_rows if existing_status >= 400 else "Deduction not found."}, existing_status if existing_status >= 400 else 404)
                return
            status_check, applications = supabase_request(
                "GET",
                f"deduction_applications?deduction_id=eq.{quote(str(deduction_id))}&status=eq.Applied&select=amount",
            )
            if status_check >= 400:
                self.send_json({"error": applications}, status_check)
                return
            applied_total = sum(float(row.get("amount") or 0) for row in applications) if isinstance(applications, list) else 0
            requested_amount = float(payload.get("amount") or 0)
            if requested_amount + 0.00001 < applied_total:
                self.send_json({"error": "Amount cannot be lower than the amount already deducted."}, 409)
                return
            existing = existing_rows[0]
            immutable_changed = (
                abs(requested_amount - float(existing.get("amount") or 0)) > 0.00001
                or str(payload.get("start_date") or "") != str(existing.get("start_date") or "")
                or str(payload.get("employee_id") or "") != str(existing.get("employee_id") or "")
                or str(payload.get("deduction_type") or "") != str(existing.get("deduction_type") or "")
            )
            if applied_total > 0 and immutable_changed:
                self.send_json({"error": "A deduction with payment history cannot change its original details."}, 409)
                return
            deduction = deduction_from_payload(payload)
            if applied_total > 0 and deduction.get("status") not in ["Cancelled"]:
                deduction["status"] = "Completed" if requested_amount <= applied_total + 0.00001 else "Pending"
            deduction["updated_at"] = datetime.utcnow().isoformat() + "Z"
            status, body = supabase_request(
                "PATCH",
                f"deduction_records?id=eq.{quote(str(deduction_id))}",
                deduction,
                prefer="return=representation",
            )
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/wage-rates":
            actor = accounting_actor(self, 4)
            if not actor:
                self.send_json({"error": "C4 or higher session is required to manage wage rates."}, 403)
                return
            wage_rate_id = payload.get("id")
            if wage_rate_id in [None, ""]:
                self.send_json({"error": "id is required."}, 400)
                return
            wage_rate = {
                "item_type": str(payload.get("item_type", "")).strip(),
                "rate": payload.get("rate", 0),
                "effective_date": str(payload.get("effective_date", "")).strip(),
                "note": str(payload.get("note", "")).strip() or None,
            }
            if not wage_rate["item_type"] or not wage_rate["effective_date"]:
                self.send_json({"error": "item_type and effective_date are required."}, 400)
                return
            try:
                if float(wage_rate["rate"] or 0) <= 0:
                    self.send_json({"error": "rate must be greater than 0."}, 400)
                    return
            except (TypeError, ValueError):
                self.send_json({"error": "rate must be a number."}, 400)
                return
            status, body = supabase_request(
                "PATCH",
                f"wage_rates?id=eq.{quote(str(wage_rate_id))}",
                wage_rate,
                prefer="return=representation",
            )
            self.send_json({"data": body if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        self.send_error(404, "Not found")

    def do_DELETE(self) -> None:
        parsed = urlparse(self.path)
        payload = self.read_json()

        if parsed.path == "/api/accounts":
            account_id = payload.get("id")
            username = str(payload.get("username", "")).strip()
            protected_username = username.lower()
            if account_id not in [None, ""]:
                status, existing = supabase_request(
                    "GET",
                    f"account_users?id=eq.{quote(str(account_id))}&select=username,user_level&limit=1",
                )
                if status >= 400:
                    self.send_json({"error": existing}, status)
                    return
                existing_account = existing[0] if isinstance(existing, list) and existing else {}
                protected_username = str(existing_account.get("username", "")).lower()
                protected_level = str(existing_account.get("user_level", "")).upper()
                filter_path = f"account_users?id=eq.{quote(str(account_id))}"
            elif username:
                status, existing = supabase_request(
                    "GET",
                    f"account_users?username=eq.{quote(username)}&select=username,user_level&limit=1",
                )
                if status >= 400:
                    self.send_json({"error": existing}, status)
                    return
                existing_account = existing[0] if isinstance(existing, list) and existing else {}
                protected_level = str(existing_account.get("user_level", "")).upper()
                filter_path = f"account_users?username=eq.{quote(username)}"
            else:
                self.send_json({"error": "id or username is required."}, 400)
                return
            if protected_username in SYSTEM_ACCOUNT_USERNAMES or protected_level == "C7":
                self.send_json({"error": "C7/developer account cannot be deleted."}, 403)
                return
            status, body = supabase_request("DELETE", filter_path, prefer="return=minimal")
            self.send_json({"data": {"deleted": True} if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/employees":
            employee_id = payload.get("id")
            if employee_id in [None, ""]:
                self.send_json({"error": "id is required."}, 400)
                return
            status, body = supabase_request(
                "DELETE",
                f"employees?id=eq.{quote(str(employee_id))}",
                prefer="return=minimal",
            )
            self.send_json({"data": {"deleted": True} if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/time-employees":
            employee_id = payload.get("id")
            if employee_id in [None, ""]:
                self.send_json({"error": "id is required."}, 400)
                return
            status, body = supabase_request(
                "DELETE",
                f"time_employees?id=eq.{quote(str(employee_id))}",
                prefer="return=minimal",
            )
            self.send_json({"data": {"deleted": True} if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/deductions":
            deduction_id = payload.get("id")
            if deduction_id in [None, ""]:
                self.send_json({"error": "id is required."}, 400)
                return
            status_check, applications = supabase_request(
                "GET",
                f"deduction_applications?deduction_id=eq.{quote(str(deduction_id))}&status=eq.Applied&select=id&limit=1",
            )
            if status_check >= 400:
                self.send_json({"error": applications}, status_check)
                return
            if isinstance(applications, list) and applications:
                self.send_json({"error": "A deduction with payment history cannot be deleted."}, 409)
                return
            status, body = supabase_request(
                "DELETE",
                f"deduction_records?id=eq.{quote(str(deduction_id))}",
                prefer="return=minimal",
            )
            self.send_json({"data": {"deleted": True} if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        self.send_error(404, "Not found")

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)

        if parsed.path == "/api/inbound/bootstrap":
            actor = inbound_authorized_actor(self)
            if not actor:
                self.send_json({"error": "C5 or higher is required for inbound receiving."}, 403)
                return
            results = {}
            for key, request_path in {
                "fruits": "inbound_fruits?select=*&order=name.asc",
                "prices": "inbound_fruit_prices?select=*&order=effective_date.desc,created_at.desc&limit=1000",
                "receipts": "inbound_receipts?select=*&order=received_at.desc&limit=5000",
            }.items():
                status, body = supabase_get_all(request_path) if key == "receipts" else supabase_request("GET", request_path)
                if status >= 400:
                    self.send_json({"error": body, "migration": "supabase_inbound_receiving_migration.sql"}, status)
                    return
                results[key] = body if isinstance(body, list) else []
            self.send_json({"data": results})
            return

        if parsed.path == "/api/online-users":
            self.send_json({"data": online_user_snapshot()})
            return

        if parsed.path == "/api/storage-usage":
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            force = query.get("refresh", ["0"])[0] == "1"
            status, result = read_database_storage_usage(force)
            self.send_json({"data": result if status < 400 else None, "error": result.get("error") if status >= 400 else None}, status)
            return

        if parsed.path == "/api/issue-reports":
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            actor_username = str(actor.get("username") or "").strip()
            filters = ["select=*"]
            if account_level_number(actor.get("level")) < 5:
                filters.append(f"reporter_username=eq.{quote(actor_username)}")
            requested_status = query.get("status", [""])[0].strip()
            if requested_status in ISSUE_REPORT_STATUSES:
                filters.append(f"status=eq.{quote(requested_status)}")
            filters.extend(["order=created_at.desc", "limit=200"])
            status, body = supabase_request("GET", f"issue_reports?{'&'.join(filters)}")
            self.send_json({"data": body if status < 400 and isinstance(body, list) else [], "error": body if status >= 400 else None}, status)
            return

        if parsed.path.startswith("/api/secret-room/"):
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "Session is missing or expired."}, 401)
                return
            actor_username = str(actor.get("username") or "").strip()

            if parsed.path == "/api/secret-room/coworkers":
                status, body = supabase_request(
                    "GET",
                    "account_users?status=eq.Active&select=id,username,fullname,role,user_level,status&order=fullname.asc",
                )
                if status >= 400:
                    self.send_json({"error": body}, status)
                    return
                online_names = {
                    str(item.get("username") or "").lower()
                    for item in online_user_snapshot().get("users", [])
                    if item.get("username")
                }
                coworkers = []
                for account in body if isinstance(body, list) else []:
                    item = account_to_client(account)
                    item.pop("password", None)
                    item["is_online"] = str(account.get("username") or "").lower() in online_names
                    item["is_self"] = str(account.get("username") or "").lower() == actor_username.lower()
                    coworkers.append(item)
                self.send_json({"data": coworkers})
                return

            if parsed.path == "/api/secret-room/posts":
                status, body = supabase_request(
                    "GET",
                    "community_posts?select=id,author_username,author_fullname,content,created_at&order=created_at.desc&limit=100",
                )
                self.send_json({"data": body if status < 400 else [], "error": body if status >= 400 else None}, status)
                return

            if parsed.path == "/api/secret-room/notifications":
                message_status, messages = supabase_request(
                    "GET",
                    f"secret_messages?recipient_username=eq.{quote(actor_username)}&is_read=eq.false&select=id,created_at&order=created_at.desc&limit=1000",
                )
                if message_status >= 400:
                    self.send_json({"error": messages}, message_status)
                    return
                try:
                    last_read_post_id = max(0, int(query.get("after_post_id", ["0"])[0]))
                except (TypeError, ValueError):
                    last_read_post_id = 0
                post_status, posts = supabase_request(
                    "GET",
                    f"community_posts?id=gt.{last_read_post_id}&author_username=neq.{quote(actor_username)}&select=id,created_at&order=id.desc&limit=1000",
                )
                if post_status >= 400:
                    self.send_json({"error": posts}, post_status)
                    return
                unread_messages = messages if isinstance(messages, list) else []
                unread_posts = posts if isinstance(posts, list) else []
                self.send_json({
                    "data": {
                        "unread_count": len(unread_messages) + len(unread_posts),
                        "unread_message_count": len(unread_messages),
                        "unread_post_count": len(unread_posts),
                        "latest_message_at": unread_messages[0].get("created_at") if unread_messages else None,
                        "latest_post_id": unread_posts[0].get("id") if unread_posts else last_read_post_id,
                    }
                })
                return

            if parsed.path in {"/api/secret-room/chats", "/api/secret-room/messages"}:
                status, body = supabase_request(
                    "GET",
                    f"secret_messages?or=(sender_username.eq.{quote(actor_username)},recipient_username.eq.{quote(actor_username)})&select=id,sender_username,recipient_username,content,is_read,created_at&order=created_at.asc&limit=1000",
                )
                if status >= 400:
                    self.send_json({"error": body}, status)
                    return
                messages = body if isinstance(body, list) else []
                if parsed.path == "/api/secret-room/messages":
                    peer = query.get("with", [""])[0].strip()
                    if not peer:
                        self.send_json({"error": "with is required."}, 400)
                        return
                    selected = [
                        {**message, "is_mine": str(message.get("sender_username", "")).lower() == actor_username.lower()}
                        for message in messages
                        if {str(message.get("sender_username", "")).lower(), str(message.get("recipient_username", "")).lower()}
                        == {actor_username.lower(), peer.lower()}
                    ]
                    self.send_json({"data": selected})
                    return

                peer_names = sorted({
                    str(message.get("recipient_username") if str(message.get("sender_username", "")).lower() == actor_username.lower() else message.get("sender_username") or "")
                    for message in messages
                } - {""})
                account_map = {}
                if peer_names:
                    encoded_names = ",".join(quote(name) for name in peer_names)
                    account_status, accounts = supabase_request(
                        "GET",
                        f"account_users?username=in.({encoded_names})&select=username,fullname",
                    )
                    if account_status < 400 and isinstance(accounts, list):
                        account_map = {str(item.get("username")): item for item in accounts}
                online_names = {
                    str(item.get("username") or "").lower()
                    for item in online_user_snapshot().get("users", [])
                }
                chats = []
                for peer in peer_names:
                    peer_messages = [
                        message for message in messages
                        if peer.lower() in {
                            str(message.get("sender_username", "")).lower(),
                            str(message.get("recipient_username", "")).lower(),
                        }
                    ]
                    last = peer_messages[-1] if peer_messages else {}
                    unread = sum(
                        1 for message in peer_messages
                        if str(message.get("recipient_username", "")).lower() == actor_username.lower()
                        and not bool(message.get("is_read"))
                    )
                    account = account_map.get(peer, {})
                    chats.append({
                        "username": peer,
                        "fullname": account.get("fullname") or peer,
                        "last_message": last.get("content") or "",
                        "last_message_at": last.get("created_at"),
                        "unread_count": unread,
                        "is_online": peer.lower() in online_names,
                    })
                chats.sort(key=lambda item: str(item.get("last_message_at") or ""), reverse=True)
                self.send_json({"data": chats})
                return

            self.send_error(404, "Not found")
            return

        if parsed.path == "/api/accounting/payment-allocations":
            actor = accounting_actor(self, 4)
            if not actor:
                self.send_json({"error": "C4 or higher accounting session is required."}, 403)
                return
            start_date = str(query.get("start_date", [""])[0])
            end_date = str(query.get("end_date", [start_date])[0])
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", start_date) or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", end_date):
                self.send_json({"error": "A valid payment date range is required."}, 400)
                return
            status, rows = supabase_request(
                "GET",
                f"ac_payment_allocations?company_key=eq.{quote(ACCOUNTING_COMPANY_KEY)}&week_start=eq.{quote(start_date)}&week_end=eq.{quote(end_date)}&select=*&order=group_label.asc,emp_code.asc",
            )
            if status >= 400:
                self.send_json({"error": rows, "migration": "supabase_pf_accounting_payments.sql"}, status)
                return
            self.send_json({"data": rows if isinstance(rows, list) else []})
            return

        if parsed.path == "/api/accounting/workspace":
            actor = accounting_actor(self)
            if not actor:
                self.send_json({"error": "Accounting session is missing or expired."}, 401)
                return
            status, rows = supabase_request(
                "GET",
                f"accounting_workspaces?company_key=eq.{quote(ACCOUNTING_COMPANY_KEY)}&select=revision,workspace,updated_by,updated_at&limit=1",
            )
            if status >= 400:
                self.send_json({"error": rows}, status)
                return
            if not isinstance(rows, list) or not rows:
                self.send_json({"error": "Accounting workspace has not been initialized."}, 404)
                return
            self.send_json({"data": rows[0]})
            return

        if parsed.path == "/api/accounting/bootstrap":
            actor = accounting_actor(self)
            if not actor:
                self.send_json({"error": "Accounting session is missing or expired."}, 401)
                return
            company_status, companies = supabase_request("GET", f"ac_companies?company_key=eq.{quote(ACCOUNTING_COMPANY_KEY)}&select=*&limit=1")
            if company_status >= 400 or not isinstance(companies, list) or not companies:
                self.send_json({"error": companies if company_status >= 400 else "Accounting company is not initialized."}, company_status if company_status >= 400 else 404)
                return
            company = companies[0]
            company_id = quote(str(company.get("id")))
            datasets = {}
            queries = {
                "accounts": f"ac_accounts?company_id=eq.{company_id}&select=*&order=code.asc",
                "periods": f"ac_periods?company_id=eq.{company_id}&select=*&order=start_date.desc",
                "journals": f"ac_journal_entries?company_id=eq.{company_id}&select=*&order=entry_date.desc,created_at.desc&limit=500",
                "partners": f"ac_partners?company_id=eq.{company_id}&select=*&active=eq.true&order=partner_code.asc",
            }
            for key, path in queries.items():
                status, rows = supabase_request("GET", path)
                if status >= 400:
                    self.send_json({"error": rows, "dataset": key}, status)
                    return
                datasets[key] = rows if isinstance(rows, list) else []
            journal_ids = [str(row.get("id")) for row in datasets["journals"] if row.get("id")]
            if journal_ids:
                status, lines = supabase_request("GET", f"ac_journal_lines?journal_id=in.({','.join(quote(value) for value in journal_ids)})&select=*&order=journal_id.asc,line_no.asc")
                if status >= 400:
                    self.send_json({"error": lines, "dataset": "journal_lines"}, status)
                    return
                datasets["journal_lines"] = lines if isinstance(lines, list) else []
            else:
                datasets["journal_lines"] = []
            self.send_json({"data": {"company": company, **datasets}})
            return

        if parsed.path == "/api/health":
            status, body = supabase_request("GET", "employees?select=id&limit=1")
            time_status, time_body = supabase_request("GET", "time_employees?select=id&limit=1")
            self.send_json(
                {
                    "status": "ok" if status < 400 and time_status < 400 else "error",
                    "server": "ready",
                    "supabase_configured": supabase_configured(),
                    "supabase_status": status,
                    "supabase_response": body,
                    "time_employees_status": time_status,
                    "time_employees_response": time_body,
                },
                200 if status < 500 and time_status < 500 else max(status, time_status),
            )
            return

        if parsed.path == "/api/production-save-queue/lookup":
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            queue_uid = query.get("queue_uid", [""])[0].strip()
            if not queue_uid:
                self.send_json({"error": "queue_uid is required."}, 400)
                return
            status, body = supabase_request(
                "GET",
                f"production_save_queue?queue_uid=eq.{quote(queue_uid)}&select=*&limit=1",
                timeout_seconds=6,
            )
            if status >= 400:
                self.send_json({"error": body}, status)
                return
            row = body[0] if isinstance(body, list) and body else None
            self.send_json({"data": production_queue_to_client(row, include_payload=True) if isinstance(row, dict) else None})
            return

        if parsed.path == "/api/time-save-queue/lookup":
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            queue_uid = query.get("queue_uid", [""])[0].strip()
            status, body = supabase_request(
                "GET", f"time_save_queue?queue_uid=eq.{quote(queue_uid)}&select=*&limit=1", timeout_seconds=6,
            )
            if status >= 400:
                self.send_json({"error": body}, status)
                return
            row = body[0] if isinstance(body, list) and body else None
            self.send_json({"data": time_queue_to_client(row, include_payload=True) if isinstance(row, dict) else None})
            return

        if parsed.path == "/api/time-save-queue":
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            try:
                limit = min(max(int(query.get("limit", ["60"])[0]), 1), 150)
            except (TypeError, ValueError):
                limit = 60
            fields = (
                "id,queue_uid,operation,record_count,emp_code,employee_name,first_work_date,last_work_date,"
                "status,attempt_count,max_attempts,started_at,finished_at,result_record_ids,result_payload,"
                "error_code,error_message,created_by,created_at,updated_at"
            )
            status, body = supabase_request(
                "GET", f"time_save_queue?select={fields}&order=created_at.desc,id.desc&limit={limit}", timeout_seconds=6,
            )
            if status >= 400:
                self.send_json({"error": body, "migration": "supabase_time_save_queue_migration.sql"}, status)
                return
            self.send_json({"data": [time_queue_to_client(row) for row in body if isinstance(row, dict)]})
            return

        time_queue_item_match = re.fullmatch(r"/api/time-save-queue/(\d+)", parsed.path)
        if time_queue_item_match:
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            queue_id = int(time_queue_item_match.group(1))
            status, body = supabase_request(
                "GET", f"time_save_queue?id=eq.{queue_id}&select=*&limit=1", timeout_seconds=6,
            )
            if status >= 400:
                self.send_json({"error": body}, status)
                return
            if not isinstance(body, list) or not body:
                self.send_json({"error": "Time queue item was not found."}, 404)
                return
            self.send_json({"data": time_queue_to_client(body[0], include_payload=True)})
            return

        if parsed.path == "/api/production-save-queue":
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            try:
                limit = min(max(int(query.get("limit", ["80"])[0]), 1), 200)
            except (TypeError, ValueError):
                limit = 80
            fields = (
                "id,queue_uid,batch_uid,record_count,fruit_type,record_date,employee_id,emp_code,"
                "employee_name,total_weight,total_amount,status,attempt_count,max_attempts,started_at,"
                "finished_at,result_record_ids,duplicate_details,error_code,error_message,"
                "created_by,cancelled_by,cancelled_at,created_at,updated_at"
            )
            status, body = supabase_request(
                "GET",
                f"production_save_queue?select={fields}&order=created_at.desc,id.desc&limit={limit}",
                timeout_seconds=6,
            )
            if status >= 400:
                self.send_json({"error": body, "migration": "supabase_production_save_queue_migration.sql"}, status)
                return
            items = [production_queue_to_client(row) for row in body if isinstance(row, dict)] if isinstance(body, list) else []
            self.send_json({"data": items})
            return

        queue_item_match = re.fullmatch(r"/api/production-save-queue/(\d+)", parsed.path)
        if queue_item_match:
            actor = secret_room_actor(self)
            if not actor:
                self.send_json({"error": "A signed-in session is required."}, 401)
                return
            queue_id = int(queue_item_match.group(1))
            status, body = supabase_request(
                "GET",
                f"production_save_queue?id=eq.{queue_id}&select=*&limit=1",
                timeout_seconds=6,
            )
            if status >= 400:
                self.send_json({"error": body}, status)
                return
            if not isinstance(body, list) or not body:
                self.send_json({"error": "Queue item was not found."}, 404)
                return
            event_status, events = supabase_request(
                "GET",
                f"production_save_queue_events?queue_id=eq.{queue_id}&select=id,event_type,status,message,actor,metadata,created_at&order=created_at.asc,id.asc",
                timeout_seconds=6,
            )
            self.send_json({
                "data": production_queue_to_client(body[0], include_payload=True),
                "events": events if event_status < 400 and isinstance(events, list) else [],
            })
            return

        if parsed.path == "/api/state":
            state = {}
            for table in LIVE_STATE_TABLES:
                status, body = supabase_get_all(f"{table}?select=*&order=id.asc")
                if status >= 400:
                    self.send_json({"error": body, "table": table}, status)
                    return
                state[table] = [live_state_to_client(table, row) for row in body]
            self.send_json({"data": state})
            return

        if parsed.path == "/api/employees":
            search = query.get("search", [""])[0].strip()
            params = "select=*&order=emp_code.asc"
            if search:
                escaped = search.replace("*", "").replace(",", " ")
                params += (
                    "&or=("
                    f"emp_code.ilike.*{escaped}*,"
                    f"fullname.ilike.*{escaped}*,"
                    f"department.ilike.*{escaped}*,"
                    f"position.ilike.*{escaped}*,"
                    f"pay_group.ilike.*{escaped}*"
                    ")"
                )
            status, body = supabase_get_all(f"employees?{params}")
            self.send_json({"data": body if status < 400 else [], "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/wage-rates/sync":
            rows = payload.get("rows", [])
            if not isinstance(rows, list):
                self.send_json({"error": "rows must be a list"}, 400)
                return
            rates = []
            for item in rows:
                if not isinstance(item, dict):
                    continue
                row = {
                    "item_type": str(item.get("item_type", "")).strip(),
                    "rate": item.get("rate", 0),
                    "effective_date": item.get("effective_date"),
                    "note": item.get("note") or None,
                    "created_by": item.get("created_by") or None,
                }
                if item.get("id") not in [None, ""]:
                    row["id"] = item.get("id")
                if row["item_type"] and row["effective_date"]:
                    rates.append(row)
            status, body = sync_rows_by_id("wage_rates", rates)
            self.send_json({"data": body.get("synced", []) if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/deductions/sync":
            rows = payload.get("rows", [])
            if not isinstance(rows, list):
                self.send_json({"error": "rows must be a list"}, 400)
                return
            deductions = [deduction_from_payload(item) for item in rows if isinstance(item, dict)]
            status, body = sync_rows_by_id("deduction_records", deductions)
            self.send_json({"data": body.get("synced", []) if status < 400 else None, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/time-employees":
            search = query.get("search", [""])[0].strip()
            params = "select=*&order=emp_code.asc"
            if search:
                escaped = search.replace("*", "").replace(",", " ")
                params += (
                    "&or=("
                    f"emp_code.ilike.*{escaped}*,"
                    f"fullname.ilike.*{escaped}*,"
                    f"employee_type.ilike.*{escaped}*"
                    ")"
                )
            status, body = supabase_get_all(f"time_employees?{params}")
            self.send_json({"data": body if status < 400 else [], "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/deductions":
            employee_kind = query.get("employee_kind", [""])[0].strip()
            start_date = query.get("start_date", [""])[0].strip()
            end_date = query.get("end_date", [""])[0].strip()
            params = "select=*&order=start_date.desc,emp_code.asc,created_at.desc"
            if employee_kind:
                params += f"&employee_kind=eq.{quote(employee_kind)}"
            if start_date:
                params += f"&end_date=gte.{quote(start_date)}"
            if end_date:
                params += f"&start_date=lte.{quote(end_date)}"
            status, body = supabase_get_all(f"deduction_records?{params}")
            self.send_json({"data": body if status < 400 else [], "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/deduction-applications":
            employee_kind = query.get("employee_kind", [""])[0].strip()
            start_date = query.get("start_date", [""])[0].strip()
            end_date = query.get("end_date", [""])[0].strip()
            params = "select=*&status=eq.Applied&order=applied_date.desc,created_at.desc"
            if employee_kind:
                params += f"&employee_kind=eq.{quote(employee_kind)}"
            if start_date:
                params += f"&applied_date=gte.{quote(start_date)}"
            if end_date:
                params += f"&applied_date=lte.{quote(end_date)}"
            status, body = supabase_get_all(f"deduction_applications?{params}")
            self.send_json({"data": body if status < 400 else [], "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/accounts":
            ensure_system_accounts()
            search = query.get("search", [""])[0].strip()
            params = "select=*&order=id.asc"
            if search:
                escaped = quote(search.replace("*", "").replace(",", " "))
                params += (
                    "&or=("
                    f"username.ilike.*{escaped}*,"
                    f"fullname.ilike.*{escaped}*,"
                    f"role.ilike.*{escaped}*,"
                    f"user_level.ilike.*{escaped}*"
                    ")"
                )
            status, body = supabase_request("GET", f"account_users?{params}")
            data = [account_to_client(item) for item in body] if status < 400 and isinstance(body, list) else []
            self.send_json({"data": data, "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/wage-rates":
            item_type = query.get("item_type", ["all"])[0].strip()
            params = "select=*&order=effective_date.desc,created_at.desc"
            if item_type and item_type != "all":
                params += f"&item_type=eq.{quote(item_type)}"
            status, body = supabase_request("GET", f"wage_rates?{params}")
            self.send_json({"data": body if status < 400 else [], "error": body if status >= 400 else None}, status)
            return

        if parsed.path == "/api/backup/verify":
            if not backup_authorized(self):
                self.send_json({"error": "Backup code is required."}, 403)
                return
            self.send_json({"authorized": True})
            return

        if parsed.path in {"/api/backup", "/api/backup/queue"}:
            if not backup_authorized(self):
                self.send_json({"error": "Backup code is required."}, 403)
                return
            scope = "queue" if parsed.path.endswith("/queue") else "main"
            tables = QUEUE_BACKUP_TABLES if scope == "queue" else BACKUP_TABLES
            status, body = read_supabase_backup(tables)
            actor = secret_room_actor(self) or {}
            response = backup_snapshot_payload(scope, str(actor.get("username") or "backup-user"), body) if status < 400 else {
                "data": None,
                "error": body,
            }
            self.send_json(response, status)
            return

        if parsed.path == "/reports/employee-daily-pdf":
            data = load_data()
            date = query.get("date", [datetime.now().date().isoformat()])[0]
            employee_id = int(query.get("employee_id", ["0"])[0])
            content = build_pdf(data, date, [employee_id])
            self.send_file(content, "application/pdf", f"employee-{employee_id}-{date}.pdf")
            return

        if parsed.path == "/reports/daily-excel":
            data = load_data()
            date = query.get("date", [datetime.now().date().isoformat()])[0]
            content = build_daily_excel(data, date)
            self.send_file(
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"daily-production-{date}.xlsx",
            )
            return

        if self.send_static_file(parsed.path):
            return

        self.send_error(404, "Not found")

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        if not length:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def send_json(self, payload: dict, status: int = 200) -> None:
        content = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def send_file(self, content: bytes, content_type: str, filename: str) -> None:
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def send_static_file(self, request_path: str) -> bool:
        safe_path = request_path.split("?", 1)[0].lstrip("/")
        if not safe_path:
            safe_path = "index.html"
        file_path = (BASE_DIR / safe_path).resolve()
        if not str(file_path).startswith(str(BASE_DIR.resolve())):
            return False
        if not file_path.is_file():
            return False

        content_type = mimetypes.guess_type(str(file_path))[0] or "application/octet-stream"
        content = file_path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        if file_path.name in {"index.html", "app.js", "styles.css"}:
            self.send_header("Cache-Control", "no-store, max-age=0")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)
        return True


def main() -> None:
    queue_worker = threading.Thread(
        target=production_save_queue_worker,
        name="production-save-queue-worker",
        daemon=True,
    )
    queue_worker.start()
    production_save_queue_wakeup.set()
    time_queue_worker = threading.Thread(
        target=time_save_queue_worker,
        name="time-save-queue-worker",
        daemon=True,
    )
    time_queue_worker.start()
    time_save_queue_wakeup.set()
    server = ThreadingHTTPServer((HOST, PORT), ReportHandler)
    try:
        print(f"Report server running at http://{HOST}:{PORT}", flush=True)
    except Exception:
        pass
    server.serve_forever()


if __name__ == "__main__":
    main()
