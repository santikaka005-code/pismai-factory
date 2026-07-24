from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import report_server


RECORDS = [
    {
        "record_date": "2026-07-20",
        "record_time": "08:00:00",
        "fruit_type": "mangosteen",
        "emp_code": "01",
        "pile_no": "2",
        "water_weight": 10,
        "flower_weight": 5,
        "total_amount": 90,
    },
    {
        "record_date": "2026-07-20",
        "record_time": "08:05:00",
        "fruit_type": "durian",
        "emp_code": "02",
        "pile_no": 3,
        "grade_weights": {"A": 8, "B": 2},
        "total_amount": 40,
    },
]


def payload(fruit_type):
    return {
        "start_date": "2026-07-20",
        "end_date": "2026-07-20",
        "fruit_type": fruit_type,
        "production_records": RECORDS,
        "export_sections": {"overview": True, "piles": True, "details": True},
    }


def detail_headers(workbook):
    return [cell.value for cell in workbook["Details"][1]]


def test_fruit_filter_and_numeric_piles():
    mangosteen = report_server.filtered_production_records(payload("mangosteen"))
    assert [record["emp_code"] for record in mangosteen] == ["01"]
    assert report_server.pile_summary_rows(mangosteen)[0]["pile"] == 2

    invalid_pile = [{**mangosteen[0], "pile_no": "bad"}]
    invalid_rows = report_server.pile_summary_rows(invalid_pile)
    assert invalid_rows[0]["pile"] == "-"
    assert invalid_rows[0]["total_weight"] == 15
    assert report_server.production_total_weight(RECORDS[1]) == 10
    assert report_server.production_grade_text(RECORDS[1]) == "10"
    assert report_server.production_total_weight({
        "fruit_type": "durian",
        "durian_weight": 12.5,
        "grade_weights": {"A": 1},
    }) == 12.5


def test_excel_columns_follow_selected_fruit():
    mangosteen_book = load_workbook(BytesIO(report_server.build_production_summary_excel(payload("mangosteen"))))
    mangosteen_headers = detail_headers(mangosteen_book)
    assert "น้ำหนักทุเรียน" not in mangosteen_headers
    assert "น้ำหนักน้ำ (กก.)" in mangosteen_headers
    pile_column = mangosteen_headers.index("กอง") + 1
    assert isinstance(mangosteen_book["Details"].cell(2, pile_column).value, int)

    durian_book = load_workbook(BytesIO(report_server.build_production_summary_excel(payload("durian"))))
    durian_headers = detail_headers(durian_book)
    assert "น้ำหนักทุเรียน" in durian_headers
    assert not any(str(header or "").startswith("เกรด ") for header in durian_headers)
    assert "น้ำหนักน้ำ (กก.)" not in durian_headers


if __name__ == "__main__":
    test_fruit_filter_and_numeric_piles()
    test_excel_columns_follow_selected_fruit()
    assert report_server.production_report_weight_labels(payload("mango")) == (
        "\u0e21\u0e30\u0e21\u0e48\u0e27\u0e07\u0e1d\u0e32",
        "\u0e21\u0e30\u0e21\u0e48\u0e27\u0e07\u0e2b\u0e31\u0e48\u0e19\u0e40\u0e15\u0e4b\u0e32",
    )
    print("Production summary fruit and pile tests passed.")
