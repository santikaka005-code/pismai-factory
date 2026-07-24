from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import report_server


BASE_PAYLOAD = {
    "start_date": "2026-07-20",
    "end_date": "2026-07-20",
    "pay_group": "all",
    "employees": [
        {"id": 1, "emp_code": "01", "fullname": "Mangosteen Worker", "pay_group": "กลุ่มใหม่"},
        {"id": 2, "emp_code": "02", "fullname": "Durian Worker", "pay_group": "เหมาโรงงาน"},
    ],
    "production_records": [
        {
            "employee_id": 1,
            "emp_code": "01",
            "record_date": "2026-07-20",
            "fruit_type": "mangosteen",
            "water_weight": 10,
            "flower_weight": 5,
            "total_weight": 15,
            "total_amount": 90,
        },
        {
            "employee_id": 2,
            "emp_code": "02",
            "record_date": "2026-07-20",
            "fruit_type": "durian",
            "durian_weight": 12,
            "grade_weights": {"A": 12},
            "total_weight": 12,
            "total_amount": 48,
        },
    ],
    "deduction_records": [],
    "export_options": {"summary": True, "fruit": True, "employees": True, "details": True},
}


def workbook_for(fruit_type):
    payload = {**BASE_PAYLOAD, "fruit_type": fruit_type}
    return load_workbook(BytesIO(report_server.build_group_report_excel(payload)))


def test_custom_group_is_not_dropped():
    rows = report_server.group_report_records({**BASE_PAYLOAD, "fruit_type": "all"})
    assert {row["pay_group"] for row in rows} == {"กลุ่มใหม่", "เหมาโรงงาน"}


def test_group_excel_columns_follow_fruit_filter():
    mangosteen = workbook_for("mangosteen")
    mangosteen_headers = [cell.value for cell in mangosteen["Summary By Group"][1]]
    assert "น้ำหนักน้ำ" in mangosteen_headers
    assert "น้ำหนักดอก" in mangosteen_headers
    assert "น้ำหนักทุเรียน" not in mangosteen_headers

    durian = workbook_for("durian")
    durian_headers = [cell.value for cell in durian["Summary By Group"][1]]
    assert "น้ำหนักทุเรียน" in durian_headers
    assert "น้ำหนักน้ำ" not in durian_headers
    assert "น้ำหนักดอก" not in durian_headers

    all_fruits = workbook_for("all")
    employee_sheet = all_fruits["Employees"]
    employee_headers = [cell.value for cell in employee_sheet[1]]
    name_column = employee_headers.index("ชื่อพนักงาน") + 1
    name_letter = employee_sheet.cell(1, name_column).column_letter
    assert employee_sheet.column_dimensions[name_letter].width >= 30
    assert employee_sheet.page_setup.fitToWidth == 1


def test_group_pdf_builds_for_each_fruit():
    for fruit_type in ("all", "mangosteen", "durian"):
        content = report_server.build_group_report_pdf({**BASE_PAYLOAD, "fruit_type": fruit_type})
        assert content.startswith(b"%PDF")


if __name__ == "__main__":
    test_custom_group_is_not_dropped()
    test_group_excel_columns_follow_fruit_filter()
    test_group_pdf_builds_for_each_fruit()
    print("Group report fruit filtering tests passed.")
