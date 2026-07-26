from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import report_server


PAYLOAD = {
    "start_date": "2026-07-10",
    "end_date": "2026-07-16",
    "department": "all",
    "department_label": "ทุกแผนก",
    "printed_by": "System Admin",
    "time_records": [
        {
            "record_date": f"2026-07-{day:02d}",
            "emp_code": "0005",
            "fullname": "Ei Shwe Zin",
            "department": "ฝ่ายผลิต",
            "clock_in": "08:00",
            "clock_out": "17:00",
            "raw_minutes": 540,
            "break_minutes": 60,
            "net_minutes": 480,
            "created_by": "System Admin",
        }
        for day in range(10, 17)
    ],
}


def test_time_summary_exports_use_consistent_layout():
    excel = report_server.build_time_summary_excel(PAYLOAD)
    workbook = load_workbook(BytesIO(excel))
    assert workbook.sheetnames == ["ภาพรวม", "สรุปรายวัน", "สรุปพนักงาน", "รายละเอียด"]
    assert workbook["ภาพรวม"]["A11"].value == "สรุปรายวัน"
    assert workbook["สรุปพนักงาน"]["B6"].value == "Ei Shwe Zin"
    assert workbook["รายละเอียด"]["C6"].value == "Ei Shwe Zin"
    assert workbook["สรุปพนักงาน"].column_dimensions["B"].width >= 30

    pdf = report_server.build_time_summary_pdf(PAYLOAD)
    assert pdf.startswith(b"%PDF")


if __name__ == "__main__":
    test_time_summary_exports_use_consistent_layout()
    print("Time summary export tests passed.")
